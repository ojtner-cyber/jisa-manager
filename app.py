"""
매장 관리 시스템 v1.0
Flask + SQLite | 본사 전용 지사 관리 플랫폼
"""
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
import sqlite3, json, os, csv, io
from datetime import date, datetime
from functools import wraps
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "enfix-manager-secret-2025")
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB 허용

# Render Disk 마운트 경로 우선 사용 → 없으면 스크립트 폴더
_data_dir = "/data" if os.path.isdir("/data") else os.path.dirname(os.path.abspath(__file__))
DB_FILE   = os.path.join(_data_dir, "jisa.db")
GOAL_FILE = os.path.join(_data_dir, "sales_goals.json")

REGIONS = ["서울","경기","인천","강원","충북","충남","대전","세종","경북","경남","대구","부산","울산","전북","전남","광주","제주"]

# ── DB 초기화 ─────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.execute("""CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        email TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        name TEXT NOT NULL,
        role TEXT DEFAULT 'user',
        created_at TEXT DEFAULT CURRENT_TIMESTAMP)""")
    conn.execute("""CREATE TABLE IF NOT EXISTS branches (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        ceo TEXT,
        ceo_phone TEXT,
        store_manager TEXT,
        store_manager_phone TEXT,
        region TEXT,
        manager TEXT,
        phone TEXT,
        email TEXT,
        address TEXT,
        status TEXT DEFAULT '운영중',
        contract_date TEXT,
        fee_rate REAL DEFAULT 0,
        note TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP)""")
    conn.execute("""CREATE TABLE IF NOT EXISTS sales (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        branch_id INTEGER NOT NULL,
        year INTEGER NOT NULL,
        month INTEGER NOT NULL,
        target INTEGER DEFAULT 0,
        actual INTEGER DEFAULT 0,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(branch_id, year, month),
        FOREIGN KEY(branch_id) REFERENCES branches(id))""")
    conn.execute("""CREATE TABLE IF NOT EXISTS sales_data (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sale_date TEXT,
        seller_name TEXT,
        item_code TEXT,
        item_name TEXT,
        item_group TEXT,
        quantity INTEGER DEFAULT 1,
        unit_price INTEGER DEFAULT 0,
        supply_price INTEGER DEFAULT 0,
        vat INTEGER DEFAULT 0,
        total INTEGER DEFAULT 0,
        buyer TEXT,
        buyer_phone TEXT,
        real_seller TEXT,
        upload_batch TEXT,
        note TEXT DEFAULT '',
        trade_code TEXT DEFAULT '',
        created_at TEXT DEFAULT CURRENT_TIMESTAMP)""")
    conn.execute("""CREATE TABLE IF NOT EXISTS sellers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        real_name TEXT,
        first_seen TEXT,
        total_sales INTEGER DEFAULT 0,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP)""")

    # ── 기존 DB 마이그레이션 (컬럼 누락 시 자동 추가) ──
    try:
        existing_cols = [r[1] for r in conn.execute("PRAGMA table_info(sales_data)").fetchall()]
        if 'note' not in existing_cols:
            conn.execute("ALTER TABLE sales_data ADD COLUMN note TEXT DEFAULT ''")
        if 'upload_batch' not in existing_cols:
            conn.execute("ALTER TABLE sales_data ADD COLUMN upload_batch TEXT DEFAULT ''")
        if 'channel' not in existing_cols:
            conn.execute("ALTER TABLE sales_data ADD COLUMN channel TEXT DEFAULT '오프라인'")
        if 'trade_code' not in existing_cols:
            conn.execute("ALTER TABLE sales_data ADD COLUMN trade_code TEXT DEFAULT ''")
        # branches 테이블 마이그레이션
        branch_cols = [r[1] for r in conn.execute("PRAGMA table_info(branches)").fetchall()]
        for col in ['ceo','ceo_phone','store_manager','store_manager_phone','branch_code']:
            if col not in branch_cols:
                conn.execute(f"ALTER TABLE branches ADD COLUMN {col} TEXT DEFAULT ''")
    except Exception:
        pass

    # 거래처코드 매핑 테이블
    conn.execute("""CREATE TABLE IF NOT EXISTS seller_code (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        code TEXT UNIQUE,
        group_name TEXT DEFAULT '',
        orig_name TEXT DEFAULT '',
        display_name TEXT DEFAULT '',
        real_seller TEXT DEFAULT ''
    )""")

    # 행사/진열 신청 관리 (기한 기반 점수)
    conn.execute("""CREATE TABLE IF NOT EXISTS display_campaign (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        campaign_name TEXT NOT NULL,
        brand TEXT DEFAULT '',
        event_type TEXT DEFAULT 'display',
        period_start TEXT DEFAULT '',
        period_end TEXT DEFAULT '',
        score_in_period INTEGER DEFAULT 5,
        score_out_period INTEGER DEFAULT 2,
        created_at TEXT DEFAULT ''
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS display_upload (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        campaign_id INTEGER NOT NULL,
        sheet_name TEXT NOT NULL,
        product_name TEXT NOT NULL,
        upload_seq INTEGER DEFAULT 1,
        upload_date TEXT DEFAULT '',
        upload_at TEXT DEFAULT ''
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS display_record (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        campaign_id INTEGER NOT NULL,
        upload_id INTEGER NOT NULL,
        seller_name TEXT NOT NULL,
        seller_code TEXT DEFAULT '',
        product_name TEXT NOT NULL,
        quantity INTEGER DEFAULT 0,
        has_display INTEGER DEFAULT 0,
        score INTEGER DEFAULT 0,
        upload_date TEXT DEFAULT '',
        is_manual INTEGER DEFAULT 0,
        memo TEXT DEFAULT '',
        updated_at TEXT DEFAULT '',
        color_detail TEXT DEFAULT '',
        visit_done INTEGER DEFAULT 0,
        call_done INTEGER DEFAULT 0,
        note TEXT DEFAULT '',
        applied_date TEXT DEFAULT '',
        UNIQUE(campaign_id, seller_name, product_name)
    )""")
    # 마이그레이션: 기존 DB에 신규 컬럼 추가
    try:
        dr_cols = [r[1] for r in conn.execute("PRAGMA table_info(display_record)").fetchall()]
        for col, typ in [('color_detail','TEXT DEFAULT \'\''), ('visit_done','INTEGER DEFAULT 0'),
                          ('call_done','INTEGER DEFAULT 0'), ('note','TEXT DEFAULT \'\''),
                          ('applied_date','TEXT DEFAULT \'\'')]:
            if col not in dr_cols:
                conn.execute(f"ALTER TABLE display_record ADD COLUMN {col} {typ}")
    except Exception:
        pass
    # 마이그레이션: 구버전 테이블도 유지 (기존 데이터 보호)
    conn.execute("""CREATE TABLE IF NOT EXISTS display_event (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        event_name TEXT NOT NULL,
        event_type TEXT DEFAULT 'display',
        product_name TEXT DEFAULT '',
        upload_order INTEGER DEFAULT 1,
        upload_date TEXT DEFAULT '',
        created_at TEXT DEFAULT ''
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS display_score (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        seller_name TEXT NOT NULL,
        event_id INTEGER NOT NULL,
        score INTEGER DEFAULT 0,
        quantity INTEGER DEFAULT 0,
        is_manual INTEGER DEFAULT 0,
        memo TEXT DEFAULT '',
        updated_at TEXT DEFAULT '',
        UNIQUE(seller_name, event_id)
    )""")

    # 방문 일정 테이블
    conn.execute("""CREATE TABLE IF NOT EXISTS visit_schedule (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        visit_date TEXT NOT NULL,
        seller_name TEXT NOT NULL,
        visit_type TEXT DEFAULT 'auto',
        reason TEXT DEFAULT '',
        priority INTEGER DEFAULT 2,
        status TEXT DEFAULT 'planned',
        check_points TEXT DEFAULT '',
        result_memo TEXT DEFAULT '',
        is_manual INTEGER DEFAULT 0,
        created_at TEXT DEFAULT '',
        updated_at TEXT DEFAULT ''
    )""")
    # 재고 현황 테이블
    conn.execute("""CREATE TABLE IF NOT EXISTS stock_data (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        item_name TEXT NOT NULL,
        item_group TEXT DEFAULT '',
        quantity INTEGER DEFAULT 0,
        upload_date TEXT DEFAULT '',
        upload_batch TEXT DEFAULT ''
    )""")
    # 보고서 양식 테이블
    conn.execute("""CREATE TABLE IF NOT EXISTS report_template (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        template_name TEXT NOT NULL,
        template_type TEXT DEFAULT 'weekly',
        columns TEXT DEFAULT '',
        uploaded_at TEXT DEFAULT ''
    )""")
    # 타사/자사 제품 단위 리서치 데이터 (리뷰+설명 기반 장단점 축적)
    conn.execute("""CREATE TABLE IF NOT EXISTS product_research (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        side TEXT NOT NULL,              -- 'ours' 또는 'competitor'
        category TEXT NOT NULL,
        brand TEXT NOT NULL,
        product_name TEXT NOT NULL,
        price_text TEXT DEFAULT '',
        pros TEXT DEFAULT '',            -- JSON 배열
        cons TEXT DEFAULT '',            -- JSON 배열
        description TEXT DEFAULT '',
        review_snippets TEXT DEFAULT '', -- JSON 배열 (원문 리뷰 조각)
        source_titles TEXT DEFAULT '',
        source_urls TEXT DEFAULT '',
        fetched_at TEXT DEFAULT '',
        product_type TEXT DEFAULT '',    -- 유모차: 디럭스/절충형/휴대용/쌍둥이, 카시트: 컨버터블/주니어/토들러
        UNIQUE(side, category, brand, product_name)
    )""")
    try:
        pr_cols = [r[1] for r in conn.execute("PRAGMA table_info(product_research)").fetchall()]
        if 'product_type' not in pr_cols:
            conn.execute("ALTER TABLE product_research ADD COLUMN product_type TEXT DEFAULT ''")
    except Exception:
        pass
    conn.execute("""CREATE TABLE IF NOT EXISTS competitor_comparison (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        category TEXT NOT NULL,
        our_product TEXT NOT NULL,
        competitor_brand TEXT NOT NULL,
        competitor_product TEXT DEFAULT '',
        comparison_text TEXT DEFAULT '',
        created_at TEXT DEFAULT ''
    )""")
    try:
        cc_cols = [r[1] for r in conn.execute("PRAGMA table_info(competitor_comparison)").fetchall()]
        if 'competitor_product' not in cc_cols:
            conn.execute("ALTER TABLE competitor_comparison ADD COLUMN competitor_product TEXT DEFAULT ''")
    except Exception:
        pass
    conn.execute("""CREATE TABLE IF NOT EXISTS store_visit_report (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        visit_date TEXT NOT NULL,
        store_name TEXT NOT NULL,
        brand TEXT DEFAULT '',
        region TEXT DEFAULT '',
        manager TEXT DEFAULT '',
        author TEXT DEFAULT '',
        store_rank TEXT DEFAULT '',
        staff_info TEXT DEFAULT '',
        store_size TEXT DEFAULT '',
        content_json TEXT DEFAULT '',
        request_json TEXT DEFAULT '',
        followup_text TEXT DEFAULT '',
        memo_text TEXT DEFAULT '',
        source_filename TEXT DEFAULT '',
        uploaded_at TEXT DEFAULT '',
        raw_grid_json TEXT DEFAULT '',
        merged_cells_json TEXT DEFAULT '',
        sheet_title TEXT DEFAULT '',
        raw_xlsx_b64 TEXT DEFAULT '',
        UNIQUE(visit_date, store_name)
    )""")
    try:
        svr_cols = [r[1] for r in conn.execute("PRAGMA table_info(store_visit_report)").fetchall()]
        for col, typ in [('raw_grid_json',"TEXT DEFAULT ''"), ('merged_cells_json',"TEXT DEFAULT ''"),
                          ('sheet_title',"TEXT DEFAULT ''"), ('raw_xlsx_b64',"TEXT DEFAULT ''")]:
            if col not in svr_cols:
                conn.execute(f"ALTER TABLE store_visit_report ADD COLUMN {col} {typ}")
    except Exception:
        pass
    conn.execute("""CREATE TABLE IF NOT EXISTS store_communication (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        seller_name TEXT NOT NULL,
        comm_date TEXT NOT NULL,
        comm_type TEXT DEFAULT '방문',
        memo TEXT DEFAULT '',
        raw_memo TEXT DEFAULT '',
        created_by TEXT DEFAULT '',
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )""")
    try:
        sc_cols = [r[1] for r in conn.execute("PRAGMA table_info(store_communication)").fetchall()]
        if 'raw_memo' not in sc_cols:
            conn.execute("ALTER TABLE store_communication ADD COLUMN raw_memo TEXT DEFAULT ''")
    except Exception:
        pass
    conn.execute("""CREATE TABLE IF NOT EXISTS sns_info (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        seller_name TEXT UNIQUE,
        blog_url TEXT DEFAULT '',
        blog_name TEXT DEFAULT '',
        blog_platform TEXT DEFAULT '',
        blog_total_posts INTEGER DEFAULT 0,
        blog_latest_date TEXT DEFAULT '',
        blog_recent_30d INTEGER DEFAULT 0,
        blog_recent_titles TEXT DEFAULT '',
        blog_recent_links TEXT DEFAULT '',
        blog_keywords TEXT DEFAULT '',
        blog_has_product_post INTEGER DEFAULT 0,
        blog_score INTEGER DEFAULT 0,
        blog_grade TEXT DEFAULT '',
        last_searched TEXT DEFAULT '',
        memo TEXT DEFAULT '',
        updated_at TEXT DEFAULT ''
    )""")
    # 마이그레이션: 기존 테이블에 컬럼 추가
    try:
        sns_cols = [r[1] for r in conn.execute("PRAGMA table_info(sns_info)").fetchall()]
        new_cols = {
            'blog_platform': 'TEXT DEFAULT ""',
            'blog_total_posts': 'INTEGER DEFAULT 0',
            'blog_latest_date': 'TEXT DEFAULT ""',
            'blog_recent_30d': 'INTEGER DEFAULT 0',
            'blog_recent_titles': 'TEXT DEFAULT ""',
            'blog_recent_links': 'TEXT DEFAULT ""',
            'blog_keywords': 'TEXT DEFAULT ""',
            'blog_has_product_post': 'INTEGER DEFAULT 0',
            'blog_grade': 'TEXT DEFAULT ""',
            'last_searched': 'TEXT DEFAULT ""',
        }
        for col, typ in new_cols.items():
            if col not in sns_cols:
                conn.execute(f"ALTER TABLE sns_info ADD COLUMN {col} {typ}")
    except: pass

    # 인스타그램 피드/릴스 인증 게시물 (매장이 자사 제품 업로드 시 담당자가 등록 → 가산점 부여)
    conn.execute("""CREATE TABLE IF NOT EXISTS instagram_post (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        seller_name TEXT NOT NULL,
        post_url TEXT NOT NULL,
        post_type TEXT DEFAULT '피드',
        brand TEXT DEFAULT '',
        product_name TEXT DEFAULT '',
        score INTEGER DEFAULT 5,
        note TEXT DEFAULT '',
        registered_by TEXT DEFAULT '',
        posted_date TEXT DEFAULT '',
        created_at TEXT DEFAULT '',
        view_count INTEGER DEFAULT NULL,
        view_count_raw TEXT DEFAULT '',
        campaign_name TEXT DEFAULT '',
        UNIQUE(seller_name, post_url)
    )""")
    try:
        ip_cols = [r[1] for r in conn.execute("PRAGMA table_info(instagram_post)").fetchall()]
        for col, typ in [('view_count','INTEGER DEFAULT NULL'), ('view_count_raw',"TEXT DEFAULT ''"),
                          ('campaign_name',"TEXT DEFAULT ''")]:
            if col not in ip_cols:
                conn.execute(f"ALTER TABLE instagram_post ADD COLUMN {col} {typ}")
    except Exception:
        pass

    # ── 업무 탭 관련 테이블 ──────────────────────────
    conn.execute("""CREATE TABLE IF NOT EXISTS work_visit_coverage (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        year INTEGER NOT NULL,
        month INTEGER NOT NULL,
        manager TEXT NOT NULL,
        region TEXT NOT NULL,
        visited_count INTEGER DEFAULT 0,
        note TEXT DEFAULT '',
        updated_at TEXT DEFAULT '',
        UNIQUE(year, month, manager, region)
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS work_promotion (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        year INTEGER NOT NULL,
        month INTEGER NOT NULL,
        period_start TEXT DEFAULT '',
        period_end TEXT DEFAULT '',
        brand TEXT DEFAULT '',
        event_name TEXT DEFAULT '',
        target_channel TEXT DEFAULT '',
        prep_items TEXT DEFAULT '',
        status TEXT DEFAULT '예정',
        source_filename TEXT DEFAULT '',
        created_at TEXT DEFAULT '',
        detail_json TEXT DEFAULT ''
    )""")
    try:
        wp_cols = [r[1] for r in conn.execute("PRAGMA table_info(work_promotion)").fetchall()]
        if 'detail_json' not in wp_cols:
            conn.execute("ALTER TABLE work_promotion ADD COLUMN detail_json TEXT DEFAULT ''")
    except Exception:
        pass
    conn.execute("""CREATE TABLE IF NOT EXISTS work_kpi_manual (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        year INTEGER NOT NULL,
        month INTEGER NOT NULL,
        manager TEXT NOT NULL,
        item_key TEXT NOT NULL,
        target REAL DEFAULT NULL,
        actual REAL DEFAULT NULL,
        updated_at TEXT DEFAULT '',
        UNIQUE(year, month, manager, item_key)
    )""")

    # 수정1: 판매실적 업로드 시 원본 엑셀 파일을 그대로 보관 (기초데이터 시트 재현용)
    conn.execute("""CREATE TABLE IF NOT EXISTS sales_upload_file (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        upload_batch TEXT,
        year INTEGER,
        months TEXT DEFAULT '',
        filename TEXT DEFAULT '',
        file_b64 TEXT,
        uploaded_at TEXT DEFAULT ''
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS work_retro (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        year INTEGER NOT NULL,
        month INTEGER NOT NULL,
        manager TEXT NOT NULL,
        keep_text TEXT DEFAULT '',
        problem_text TEXT DEFAULT '',
        try_text TEXT DEFAULT '',
        updated_at TEXT DEFAULT '',
        UNIQUE(year, month, manager)
    )""")

    # 기본 계정 — 없으면 생성, 있으면 비밀번호 보장
    conn.execute("INSERT OR IGNORE INTO users(email,password,name,role) VALUES(?,?,?,?)",
        ("hwkim@enfix.com","hwkim123!","관리자","admin"))
    conn.execute("INSERT OR IGNORE INTO users(email,password,name,role) VALUES(?,?,?,?)",
        ("user@visang.com","hwkim123!","일반사용자","user"))
    # Railway 재배포 후 비밀번호가 달라졌을 경우를 위해 강제 보장
    conn.execute("UPDATE users SET password=? WHERE email=? AND role='admin'",
        ("hwkim123!", "hwkim@enfix.com"))
    conn.commit(); conn.close()

# ── 인증 ──────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            if request.path.startswith("/api/"):
                return jsonify({"error":"unauthorized"}), 401
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("user",{}).get("role") != "admin":
            return jsonify({"error":"forbidden"}), 403
        return f(*args, **kwargs)
    return login_required(decorated)

# ── 페이지 라우트 ──────────────────────────────
@app.route("/")
@login_required
def index():
    return redirect("/dashboard")

@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        d = request.json
        conn = get_db()
        user = conn.execute("SELECT * FROM users WHERE email=? AND password=?",
            (d.get("email",""), d.get("password",""))).fetchone()
        conn.close()
        if user:
            session["user"] = dict(user)
            return jsonify({"ok":True, "role": user["role"]})
        return jsonify({"ok":False, "msg":"이메일 또는 비밀번호가 올바르지 않습니다."}), 401
    return render_template("index.html", regions=REGIONS)

@app.route("/api/reset-admin", methods=["POST"])
def api_reset_admin():
    """비상 관리자 계정 리셋 — 배포 후 로그인 불가 시 사용"""
    secret = request.json.get("secret","") if request.is_json else request.form.get("secret","")
    if secret != "enfix2024reset":
        return jsonify({"ok": False}), 403
    conn = get_db()
    conn.execute("UPDATE users SET password=? WHERE email=?", ("hwkim123!", "hwkim@enfix.com"))
    conn.execute("INSERT OR IGNORE INTO users(email,password,name,role) VALUES(?,?,?,?)",
        ("hwkim@enfix.com","hwkim123!","관리자","admin"))
    conn.commit(); conn.close()
    return jsonify({"ok": True, "msg": "관리자 비밀번호가 hwkim123! 으로 초기화됐습니다"})


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

# ── 대시보드 API ───────────────────────────────
@app.route("/api/dashboard")
@login_required
def api_dashboard():
    y = request.args.get("year", str(datetime.now().year))
    conn = get_db()
    total_branches = conn.execute("SELECT COUNT(*) FROM branches").fetchone()[0]
    active = conn.execute("SELECT COUNT(*) FROM branches WHERE status='운영중'").fetchone()[0]

    # 판매현황(sales_data) 기반 실적
    sd_total = conn.execute("""SELECT SUM(total) t, COUNT(*) c FROM sales_data
        WHERE sale_date LIKE ?""", (f"{y}%",)).fetchone()

    # 월별: sales_data 기반
    monthly_sd = [dict(r) for r in conn.execute("""
        SELECT CAST(strftime('%m', sale_date) AS INTEGER) month, SUM(total) actual, COUNT(*) cnt
        FROM sales_data WHERE sale_date LIKE ? AND sale_date != ''
        GROUP BY month ORDER BY month""", (f"{y}%",)).fetchall()]
    # 목표는 항상 15억으로 고정 (수정4)
    MONTHLY_TARGET_FIXED = 1_500_000_000
    monthly = []
    for m in range(1, 13):
        sd_row = next((r for r in monthly_sd if r["month"]==m), None)
        monthly.append({"month": m, "target": MONTHLY_TARGET_FIXED,
                        "actual": sd_row["actual"] if sd_row else 0})

    # TOP5 판매처 (real_seller 기준)
    top5 = [dict(r) for r in conn.execute("""
        SELECT real_seller name, SUM(total) total FROM sales_data
        WHERE sale_date LIKE ? AND real_seller != ''
        GROUP BY real_seller ORDER BY total DESC LIMIT 5""",
        (f"{y}%",)).fetchall()]

    # 지역별 (branches + sales_data 조인 — real_seller 기준)
    region_stats = [dict(r) for r in conn.execute("""
        SELECT b.region, SUM(sd.total) total
        FROM sales_data sd JOIN branches b ON sd.real_seller=b.name
        WHERE sd.sale_date LIKE ?
        GROUP BY b.region ORDER BY total DESC""", (f"{y}%",)).fetchall()]

    conn.close()
    total_actual = int(sd_total["t"] or 0)
    total_count  = int(sd_total["c"] or 0)
    return jsonify({
        "total_branches": total_branches,
        "active_branches": active,
        "total_target": 0,
        "total_actual": total_actual,
        "total_count": total_count,
        "achievement": 0,
        "monthly": monthly,
        "top5": top5,
        "region_stats": region_stats,
    })

# ── 판매처(지사) API ───────────────────────────
@app.route("/api/branches")
@login_required
def api_branches():
    region = request.args.get("region","")
    status = request.args.get("status","")
    q_str  = request.args.get("q","").strip()
    conn = get_db()
    q = "SELECT * FROM branches WHERE 1=1"
    params = []
    if region: q += " AND region=?"; params.append(region)
    if status: q += " AND status=?"; params.append(status)
    if q_str:  q += " AND (name LIKE ? OR manager LIKE ? OR address LIKE ?)"; params+=[f"%{q_str}%"]*3
    q += " ORDER BY name"
    rows = [dict(r) for r in conn.execute(q, params).fetchall()]
    y = datetime.now().year
    for row in rows:
        # 판매현황(sales_data)에서 실적 연동 — real_seller 기준
        sd = conn.execute("""
            SELECT SUM(total) total FROM sales_data
            WHERE real_seller=? AND sale_date LIKE ?""",
            (row["name"], f"{y}%")).fetchone()
        row["year_actual"] = int(sd["total"] or 0)
    conn.close()
    return jsonify(rows)

@app.route("/api/branches", methods=["POST"])
@login_required
def api_branches_add():
    d = request.json
    conn = get_db()
    conn.execute("""INSERT INTO branches(name,ceo,ceo_phone,store_manager,store_manager_phone,
        region,manager,phone,email,address,status,note) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
        (d["name"],d.get("ceo",""),d.get("ceo_phone",""),d.get("store_manager",""),
         d.get("store_manager_phone",""),d.get("region",""),d.get("manager",""),
         d.get("phone",""),d.get("email",""),d.get("address",""),d.get("status","운영중"),d.get("note","")))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/branches/<int:bid>", methods=["GET"])
@login_required
def api_branch_get(bid):
    conn = get_db()
    row = conn.execute("SELECT * FROM branches WHERE id=?", (bid,)).fetchone()
    conn.close()
    return jsonify(dict(row) if row else {})

@app.route("/api/branches/<int:bid>", methods=["PUT"])
@login_required
def api_branches_update(bid):
    d = request.json
    conn = get_db()
    conn.execute("""UPDATE branches SET name=?,ceo=?,ceo_phone=?,store_manager=?,
        store_manager_phone=?,region=?,manager=?,phone=?,email=?,address=?,status=?,note=?
        WHERE id=?""",
        (d["name"],d.get("ceo",""),d.get("ceo_phone",""),d.get("store_manager",""),
         d.get("store_manager_phone",""),d.get("region",""),d.get("manager",""),
         d.get("phone",""),d.get("email",""),d.get("address",""),d.get("status","운영중"),
         d.get("note",""),bid))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

# ── 매장 정보 xlsx 업로드 ─────────────────────
def parse_region_from_address(addr):
    """주소에서 지역 추출"""
    addr = addr or ''
    region_map = [
        ('서울', '서울'), ('경기', '경기'), ('인천', '인천'), ('강원', '강원'),
        ('충북', '충북'), ('충남', '충남'), ('대전', '대전'), ('세종', '세종'),
        ('경북', '경북'), ('경남', '경남'), ('대구', '대구'), ('부산', '부산'),
        ('울산', '울산'), ('전북', '전북'), ('전남', '전남'), ('광주', '광주'),
        ('제주', '제주'),
    ]
    for key, region in region_map:
        if key in addr:
            return region
    return ''

def detect_region_from_name(name):
    """매장명에서 지역 자동 추출 (경기는 북부/남부로 세분화)"""
    name = name or ''

    # 경기북부 (10개 시·군): 고양, 남양주, 파주, 의정부, 양주, 구리, 포천, 동두천, 가평, 연천
    GG_NORTH = ['고양', '남양주', '파주', '의정부', '양주', '구리', '포천', '동두천', '가평', '연천', '일산']
    # 경기남부 (21개 시·군): 수원, 용인, 성남, 부천, 화성, 안산, 평택, 안양, 시흥, 광명,
    #   군포, 광주, 이천, 안성, 하남, 오산, 여주, 양평, 김포, 의왕, 과천
    GG_SOUTH = ['수원', '용인', '성남', '부천', '화성', '안산', '평택', '안양', '시흥', '광명',
                '군포', '이천', '안성', '하남', '오산', '여주', '양평', '김포', '의왕', '과천',
                '영통', '동탄', '판교', '분당', '서수원', '다산', '미사']

    # 인천 (검단/청라 등 서구 포함) — 경기와 혼동 주의, 최우선 체크
    if any(k in name for k in ['인천', '부평', '송도', '계양', '검단', '청라', '연수', '남동']):
        return '인천'

    # 경기광주(경기도 광주시)는 광주광역시와 다르므로 최우선 체크
    if '경기광주' in name or '경기 광주' in name:
        return '경기남부'

    # 경기북부/남부 우선 체크 (일반 '경기' 이전에)
    for k in GG_NORTH:
        if k in name: return '경기북부'
    for k in GG_SOUTH:
        if k in name: return '경기남부'

    region_keywords = [
        # 특별시/광역시
        ('서울', '서울'), ('강남', '서울'), ('강북', '서울'), ('강서', '서울'),
        ('강동', '서울'), ('마포', '서울'), ('용산', '서울'), ('성북', '서울'),
        ('송파', '서울'), ('노원', '서울'), ('은평', '서울'), ('도봉', '서울'),
        ('관악', '서울'), ('동작', '서울'), ('영등포', '서울'), ('구로', '서울'),
        ('금천', '서울'), ('양천', '서울'), ('마곡', '서울'), ('목동', '서울'),
        ('부산', '부산'), ('해운대', '부산'), ('동래', '부산'), ('사하', '부산'),
        ('연제', '부산'), ('수영', '부산'), ('금정', '부산'), ('남구', '부산'),
        ('대구', '대구'), ('달성', '대구'), ('수성', '대구'), ('달서', '대구'),
        ('광주', '광주'), ('북구', '광주'), ('서구', '광주'),
        ('대전', '대전'), ('유성', '대전'), ('서대전', '대전'),
        ('울산', '울산'),
        ('세종', '세종'),
        # 경기 (북부/남부 키워드에 없는 나머지 — 광범위 매칭)
        ('경기', '경기남부'),
        # 강원
        ('강원', '강원'), ('춘천', '강원'), ('원주', '강원'), ('강릉', '강원'),
        ('속초', '강원'), ('동해', '강원'), ('삼척', '강원'), ('태백', '강원'),
        # 충청
        ('청주', '충북'), ('충주', '충북'), ('제천', '충북'),
        ('천안', '충남'), ('아산', '충남'), ('서산', '충남'), ('당진', '충남'),
        ('홍성', '충남'), ('공주', '충남'), ('보령', '충남'),
        # 전라
        ('전주', '전북'), ('익산', '전북'), ('군산', '전북'), ('완주', '전북'),
        ('목포', '전남'), ('여수', '전남'), ('순천', '전남'), ('나주', '전남'),
        ('광양', '전남'),
        # 경상
        ('포항', '경북'), ('경주', '경북'), ('구미', '경북'), ('안동', '경북'),
        ('영천', '경북'), ('경산', '경북'),
        ('창원', '경남'), ('진주', '경남'), ('김해', '경남'), ('양산', '경남'),
        ('거제', '경남'), ('통영', '경남'), ('밀양', '경남'),
        # 제주
        ('제주', '제주'), ('서귀포', '제주'),
    ]
    for keyword, region in region_keywords:
        if keyword in name:
            return region
    return ''

@app.route("/api/upload/stores", methods=["POST"])
@login_required
def upload_stores():
    """매장 정보 xlsx 업로드
    B열:업체구분, D열:거래처명, E열:실적용거래처명, F열:매장전화,
    G열:사장님이름, H열:사장연락처, I열:점장이름, J열:점장연락처,
    M열:담당자, N열:주소, O열:이메일
    """
    import zipfile as zf2, xml.etree.ElementTree as ET2
    f = request.files.get("file")
    if not f: return jsonify({"error": "파일이 없습니다"}), 400

    file_bytes = f.read()
    stores = []
    try:
        with zf2.ZipFile(io.BytesIO(file_bytes)) as z:
            strings = []
            if 'xl/sharedStrings.xml' in z.namelist():
                sst = z.read('xl/sharedStrings.xml').decode('utf-8')
                sr = ET2.fromstring(sst)
                ns2 = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
                for si in sr.findall(f'{{{ns2}}}si'):
                    strings.append(''.join(t.text or '' for t in si.findall(f'.//{{{ns2}}}t')))

            sheet_xml = z.read('xl/worksheets/sheet1.xml').decode('utf-8')
            root = ET2.fromstring(sheet_xml)
            ns2 = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'

            def cell_val(cell, ns2=ns2, strings=strings):
                t = cell.get('t', '')
                is_el = cell.find(f'{{{ns2}}}is')
                v_el  = cell.find(f'{{{ns2}}}v')
                if is_el is not None:
                    return ''.join(x.text or '' for x in is_el.findall(f'.//{{{ns2}}}t'))
                if t == 's' and v_el is not None:
                    idx = int(v_el.text)
                    return strings[idx] if idx < len(strings) else ''
                return v_el.text or '' if v_el is not None else ''

            current_group = ''
            for row in root.findall(f'.//{{{ns2}}}row'):
                rnum = int(row.get('r', 0))
                if rnum < 5: continue  # 헤더 스킵 (4행이 헤더)

                vals = {}
                for c in row.findall(f'{{{ns2}}}c'):
                    ref = c.get('r', '')
                    col = ''.join(x for x in ref if x.isalpha())
                    v = cell_val(c)
                    if v: vals[col] = v

                # B열에 업체구분이 있으면 그룹 업데이트
                if 'B' in vals and vals['B'] not in ('업체구분', '※ 오프라인 거래처별 리스트'):
                    current_group = vals['B']

                # E열: 실적용거래처명이 기준 (없으면 D열 사용)
                name = vals.get('E', '').strip() or vals.get('D', '').strip()
                if not name: continue

                # 이름 정제: "이정현사장님" → "이정현", "이준석점장님" → "이준석"
                def clean_name(s):
                    return s.replace('사장님','').replace('점장님','').replace('매니저님','').replace('실장','').replace('과장','').strip()

                address = vals.get('N', '').strip()
                region  = parse_region_from_address(address) or detect_region_from_name(name)

                stores.append({
                    'name':                 name.replace('_', ' '),
                    'group':                current_group,
                    'phone':                vals.get('F', '').strip(),       # 매장 전화
                    'ceo':                  clean_name(vals.get('G', '')),   # 사장님 이름
                    'ceo_phone':            vals.get('H', '').strip(),       # 사장 연락처
                    'store_manager':        clean_name(vals.get('I', '') or vals.get('K', '')),  # 점장
                    'store_manager_phone':  vals.get('J', '').strip() or vals.get('L', '').strip(),  # 점장 연락처
                    'manager':              vals.get('M', '').strip(),       # 담당자(본사)
                    'address':              address,
                    'region':               region,
                    'email':                vals.get('O', '').strip(),
                    'note':                 current_group,
                })
    except Exception as e:
        return jsonify({"error": f"파일 파싱 오류: {str(e)}"}), 400

    # preview 모드
    if request.args.get('preview') == '1':
        return jsonify({"stores": stores, "count": len(stores)})

    # 저장
    conn = get_db()
    added, updated = 0, 0
    for s in stores:
        existing = conn.execute("SELECT id FROM branches WHERE name=?", (s['name'],)).fetchone()
        if existing:
            conn.execute("""UPDATE branches SET phone=?,ceo=?,ceo_phone=?,store_manager=?,
                store_manager_phone=?,manager=?,address=?,region=?,email=?,note=?
                WHERE id=?""", (s['phone'],s['ceo'],s['ceo_phone'],s['store_manager'],
                s['store_manager_phone'],s['manager'],s['address'],s['region'],s['email'],
                s['note'],existing['id']))
            updated += 1
        else:
            conn.execute("""INSERT INTO branches(name,region,ceo,ceo_phone,store_manager,
                store_manager_phone,manager,phone,address,email,status,note)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                (s['name'],s['region'],s['ceo'],s['ceo_phone'],s['store_manager'],
                 s['store_manager_phone'],s['manager'],s['phone'],s['address'],
                 s['email'],'운영중',s['note']))
            added += 1
    conn.commit(); conn.close()
    return jsonify({"ok": True, "added": added, "updated": updated, "total": len(stores)})

# ── 판매부수 페이지용 — 매장별 실적 ──────────────
@app.route("/api/sales-by-store")
@login_required
def api_sales_by_store():
    year   = request.args.get("year",   str(datetime.now().year))
    seller = request.args.get("seller", "").strip()
    month  = request.args.get("month",  "").strip()
    channel = request.args.get("channel", "").strip()  # '오프라인' / '백화점' / '' (전체)
    conn   = get_db()

    if month:
        date_cond = f"{year}-{month.zfill(2)}%"
    else:
        date_cond = f"{year}%"

    channel_sql = " AND channel=?" if channel else ""
    channel_params = [channel] if channel else []

    if seller:
        # aliases 포함 조회 (하남미시점 → 하남미사점 등)
        db_names = get_all_real_sellers_for(seller)
        placeholders = ','.join('?' for _ in db_names)
        rows = [dict(r) for r in conn.execute(f"""
            SELECT ? AS seller_name,
                   CAST(strftime('%m', sale_date) AS INTEGER) AS month,
                   COUNT(*) cnt, SUM(total) total, SUM(quantity) qty
            FROM sales_data
            WHERE real_seller IN ({placeholders}) AND sale_date LIKE ? AND sale_date != ''{channel_sql}
            GROUP BY month ORDER BY month""", [seller] + db_names + [date_cond] + channel_params).fetchall()]
        conn.close()
        return jsonify(rows)
    else:
        rows = [dict(r) for r in conn.execute(f"""
            SELECT real_seller AS seller_name,
                   COUNT(*) cnt, SUM(total) total, SUM(quantity) qty
            FROM sales_data
            WHERE sale_date LIKE ? AND real_seller != '' AND real_seller IS NOT NULL{channel_sql}
            GROUP BY real_seller ORDER BY real_seller""", [date_cond] + channel_params).fetchall()]
        conn.close()

        # 표시명으로 그룹화 (베이비하우스 하남미시점 + 하남미사점 → 하남미사점으로 합산)
        from collections import defaultdict
        merged = defaultdict(lambda: {'cnt':0,'total':0,'qty':0})
        for r in rows:
            display_nm = display_seller(r['seller_name'])
            if is_hidden_seller(r['seller_name']): continue
            merged[display_nm]['cnt']   += r['cnt']
            merged[display_nm]['total'] += r['total']
            merged[display_nm]['qty']   += r['qty']
        rows = [{'seller_name':k, **v} for k,v in merged.items()]

        def brand_sort_key(r):
            nm = (r['seller_name'] or '').replace('_', ' ').lower()
            if '베이비하우스' in nm: return (0, -r['total'])
            if '링크맘' in nm:      return (1, -r['total'])
            if '베이비파크' in nm:  return (2, -r['total'])
            if '베네피아' in nm:    return (3, -r['total'])
            return (9, -r['total'])

        rows.sort(key=brand_sort_key)
        # 화면 표시명 변환 + 숨김 처리
        rows = [r for r in rows if not is_hidden_seller(r['seller_name'])]
        for r in rows:
            r['seller_name'] = display_seller(r['seller_name'])
        return jsonify(rows)


@app.route("/api/sales/trend")
@login_required
def api_sales_trend():
    """매장별 최근 월별 매출 흐름 분석 — 추세 계산"""
    year  = request.args.get("year",  str(datetime.now().year))
    conn  = get_db()

    # 매장별 월별 매출
    monthly = conn.execute("""
        SELECT real_seller, CAST(strftime('%m', sale_date) AS INTEGER) mo,
               SUM(total) total, SUM(quantity) qty
        FROM sales_data
        WHERE real_seller!='' AND sale_date LIKE ? AND sale_date!=''
        GROUP BY real_seller, mo ORDER BY real_seller, mo
    """, (f"{year}%",)).fetchall()

    # 전체 연매출
    totals = {r[0]:r[1] for r in conn.execute("""
        SELECT real_seller, SUM(total) FROM sales_data
        WHERE real_seller!='' AND sale_date LIKE ? AND sale_date!=''
        GROUP BY real_seller
    """, (f"{year}%",)).fetchall()}
    conn.close()

    # 매장별 월 데이터 집계
    from collections import defaultdict
    seller_months = defaultdict(lambda: defaultdict(lambda: {'total':0,'qty':0}))
    for r in monthly:
        seller_months[r[0]][r[1]]['total'] = r[2]
        seller_months[r[0]][r[1]]['qty']   = r[3]

    def calc_trend(monthly_vals):
        """최근 3개월 평균 대비 직전 3개월 평균으로 추세 계산"""
        if len(monthly_vals) < 2: return 0, 'none'
        # 선형 회귀 기울기 (간단 버전)
        n = len(monthly_vals)
        xs = list(range(n))
        ys = [v for v in monthly_vals]
        x_mean = sum(xs)/n; y_mean = sum(ys)/n
        num = sum((x-x_mean)*(y-y_mean) for x,y in zip(xs,ys))
        den = sum((x-x_mean)**2 for x in xs)
        slope = num/den if den else 0
        # 최근 2개월 변화
        pct = (ys[-1]-ys[-2])/ys[-2]*100 if ys[-2] else 0
        # 방향 판정
        if slope > y_mean * 0.05: direction = 'up'
        elif slope < -y_mean * 0.05: direction = 'down'
        else: direction = 'stable'
        return round(slope), direction, round(pct, 1)

    result = []
    for seller, months_data in seller_months.items():
        sorted_months = sorted(months_data.items())
        monthly_totals = [v['total'] for _, v in sorted_months]
        monthly_list   = [{'mo': m, 'total': v['total'], 'qty': v['qty']}
                          for m, v in sorted_months]

        if len(monthly_totals) < 2:
            direction = 'none'; slope = 0; pct = 0
        else:
            slope, direction, pct = calc_trend(monthly_totals)

        # 최고/최저 달
        peak_mo   = max(sorted_months, key=lambda x:x[1]['total'], default=(0,{'total':0}))
        trough_mo = min(sorted_months, key=lambda x:x[1]['total'], default=(0,{'total':0}))

        # 연속 상승/하락 스트릭
        streak = 0; streak_dir = 'none'
        if len(monthly_totals) >= 2:
            streak_dir = 'up' if monthly_totals[-1] > monthly_totals[-2] else 'down'
            for i in range(len(monthly_totals)-1, 0, -1):
                if streak_dir == 'up' and monthly_totals[i] > monthly_totals[i-1]: streak += 1
                elif streak_dir == 'down' and monthly_totals[i] < monthly_totals[i-1]: streak += 1
                else: break

        result.append({
            'seller_name': seller,
            'year_total':  totals.get(seller, 0),
            'monthly':     monthly_list,
            'direction':   direction,
            'slope':       slope,
            'pct_last':    pct,
            'peak_mo':     peak_mo[0],
            'peak_total':  peak_mo[1]['total'],
            'trough_mo':   trough_mo[0],
            'trough_total':trough_mo[1]['total'],
            'streak':      streak,
            'streak_dir':  streak_dir,
            'month_cnt':   len(monthly_list),
        })

    # 연매출 내림차순 정렬
    result.sort(key=lambda x: -x['year_total'])
    # 화면 표시명 변환 + 숨김 처리
    result = [r for r in result if not is_hidden_seller(r['seller_name'])]
    for r in result:
        r['seller_name'] = display_seller(r['seller_name'])
    return jsonify(result)

# ── 판매현황 — 판매처 수 전체 반환 ──────────────
# ── 판매실적 엑셀 내보내기 ────────────────────────
@app.route("/api/export/sales-monthly")
@login_required
def export_sales_monthly():
    year   = request.args.get("year",   str(datetime.now().year))
    month  = request.args.get("month",  "")
    seller = request.args.get("seller", "").strip()
    date_cond = f"{year}-{month.zfill(2)}%" if month else f"{year}%"
    conn = get_db()
    if seller:
        rows = [dict(r) for r in conn.execute("""
            SELECT ? AS seller_name, CAST(strftime('%m', sale_date) AS INTEGER) AS month,
                   COUNT(*) cnt, SUM(total) total, SUM(quantity) qty
            FROM sales_data WHERE real_seller=? AND sale_date LIKE ? AND sale_date != ''
            GROUP BY month ORDER BY month""", (seller, seller, date_cond)).fetchall()]
    else:
        rows = [dict(r) for r in conn.execute("""
            SELECT real_seller AS seller_name, COUNT(*) cnt, SUM(total) total, SUM(quantity) qty
            FROM sales_data WHERE sale_date LIKE ? AND real_seller != ''
            GROUP BY real_seller ORDER BY total DESC""", (date_cond,)).fetchall()]
    conn.close()
    buf = io.StringIO()
    w = csv.writer(buf)
    if seller:
        w.writerow(['매장명', '월', '판매건수', '판매수량', '판매금액'])
        for r in rows:
            w.writerow([r['seller_name'], f"{r['month']}월", r['cnt'], r['qty'], r['total']])
    else:
        w.writerow(['매장명', '판매건수', '판매수량', '판매금액'])
        for r in rows:
            w.writerow([r['seller_name'], r['cnt'], r['qty'], r['total']])
    buf.seek(0)
    fname = f"월별실적_{year}{'_'+month+'월' if month else ''}.csv"
    return send_file(io.BytesIO(buf.getvalue().encode('utf-8-sig')), mimetype='text/csv',
                     as_attachment=True, download_name=fname)

@app.route("/api/export/sales-weekly")
@login_required
def export_sales_weekly():
    year   = request.args.get("year",   str(datetime.now().year))
    month  = request.args.get("month",  "")
    seller = request.args.get("seller", "").strip()
    from datetime import datetime as dt2, timedelta
    qp = ["sale_date != ''"]
    pp = []
    if month: qp.append("sale_date LIKE ?"); pp.append(f"{year}-{month.zfill(2)}%")
    else:     qp.append("sale_date LIKE ?"); pp.append(f"{year}%")
    if seller: qp.append("real_seller = ?"); pp.append(seller)
    conn = get_db()
    rows = [dict(r) for r in conn.execute(f"""
        SELECT strftime('%Y-%W', sale_date) AS week_key,
               COUNT(*) cnt, SUM(quantity) qty, SUM(total) total, MIN(sale_date) AS min_date
        FROM sales_data WHERE {' AND '.join(qp)} AND sale_date != ''
        GROUP BY week_key ORDER BY week_key""", pp).fetchall()]
    conn.close()
    def wr(ds):
        d = dt2.strptime(ds, "%Y-%m-%d"); wd=d.weekday()
        sun=d-timedelta(days=(wd+1)%7)
        return sun.strftime("%Y-%m-%d"), (sun+timedelta(days=6)).strftime("%Y-%m-%d")
    for r in rows:
        try: r['ws'],r['we']=wr(r['min_date'])
        except: r['ws']=r['we']=''
    buf = io.StringIO(); w = csv.writer(buf)
    w.writerow(['주차','기간 시작','기간 종료','판매건수','판매수량','판매금액'])
    for i,r in enumerate(rows):
        w.writerow([f"{i+1}주차",r['ws'],r['we'],r['cnt'],r['qty'],r['total']])
    buf.seek(0)
    fname = f"주별실적_{year}{'_'+month+'월' if month else ''}.csv"
    return send_file(io.BytesIO(buf.getvalue().encode('utf-8-sig')), mimetype='text/csv',
                     as_attachment=True, download_name=fname)

@app.route("/api/export/sales-ranking")
@login_required
def export_sales_ranking():
    year  = request.args.get("year",  str(datetime.now().year))
    month = request.args.get("month", "")
    date_cond = f"{year}-{month.zfill(2)}%" if month else f"{year}%"
    conn = get_db()
    rows = [dict(r) for r in conn.execute("""
        SELECT real_seller AS seller_name, COUNT(*) cnt, SUM(total) total, SUM(quantity) qty
        FROM sales_data WHERE sale_date LIKE ? AND real_seller != ''
        GROUP BY real_seller ORDER BY total DESC""", (date_cond,)).fetchall()]
    conn.close()
    buf = io.StringIO(); w = csv.writer(buf)
    w.writerow(['순위','매장명','판매금액','판매건수','판매수량'])
    for i,r in enumerate(rows):
        w.writerow([i+1,r['seller_name'],r['total'],r['cnt'],r['qty']])
    buf.seek(0)
    fname = f"매출순위_{year}{'_'+month+'월' if month else ''}.csv"
    return send_file(io.BytesIO(buf.getvalue().encode('utf-8-sig')), mimetype='text/csv',
                     as_attachment=True, download_name=fname)

# ── xlsx 엑셀 내보내기 헬퍼 ──────────────────────

# 브랜드 순서 (엑셀 열 순서)
# ── 매장명 별칭 매핑 ─────────────────────────────
SELLER_ALIAS = {
    # 관악점
    '주식회사 위드에이컴퍼니': '베이비하우스 관악점',
    '위드에이컴퍼니':          '베이비하우스 관악점',
    # 청라점
    '주식회사 티에스엘컴퍼니': '베이비하우스 청라점',
    '티에스엘컴퍼니':          '베이비하우스 청라점',
    # 대구점 통합
    '베이비하우스 대구':        '베이비하우스 대구점',
    # 동대전점 통합 (발육/용품 점포 → 동대전점)
    '주식회사 에스엘컴퍼니':               '베이비하우스 동대전점',
    '베이비하우스_동대전점_발육':           '베이비하우스 동대전점',
    '베이비하우스_동대전점_용품':           '베이비하우스 동대전점',
    '베이비하우스 동대전점 발육':           '베이비하우스 동대전점',
    '베이비하우스 동대전점 용품':           '베이비하우스 동대전점',
    # 대전점(대전세종점) — 동대전점과 별개 매장
    '베이비하우스 대전점':      '베이비하우스 대전세종점',
    '베이비하우스_대전점':      '베이비하우스 대전세종점',
    '베이비하우스 대전세종':    '베이비하우스 대전세종점',
    # 하남미사점
    '베이비하우스 하남미시점':  '베이비하우스 하남미사점',
    '베이비하우스 하남미사':    '베이비하우스 하남미사점',
    # 수정9: 부산동래점 → 동래점
    '베이비하우스 부산동래점':  '베이비하우스 동래점',
    '베이비하우스 동래':        '베이비하우스 동래점',
    '베이비하우스동래':         '베이비하우스 동래점',
    # 수정9: 동탄오산점 → 동탄점
    '베이비하우스 동탄오산점':  '베이비하우스 동탄점',
    # 군포점 (안양점 → 군포점)
    '베이비하우스 안양점':      '베이비하우스 군포점',
    # 부천점
    '베이비하우스 부천':        '베이비하우스 부천점',
    '부천 베이비하우스':        '베이비하우스 부천점',
    '베네피아 부천점':          '링크맘 부천점',
    # 일산점
    '베이비하우스 일산화정점':  '베이비하우스 일산점',
    # 광주점
    '광주베이비하우스':         '베이비하우스 광주점',
    '베이비하우스광주':         '베이비하우스 광주점',
    # 다산 (수정5: 베네피아 다산 = 베이비스토어 다산 → "베이비스토어 다산"으로 통일)
    '베네피아 다산':            '베이비스토어 다산',
    '베네피아다산':             '베이비스토어 다산',
    # 에이블베이비 → 링크맘 부산점
    '주식회사 에이블베이비':    '링크맘 부산점',
    '에이블베이비':             '링크맘 부산점',
    '에이블 부산':              '링크맘 부산점',
    # 수정9: 베투키 (DB에서 이미 '베이비 투 키즈'로 통합 완료)
    '주식회사 베이비투키즈':    '베이비 투 키즈',
    '베이비투키즈':             '베이비 투 키즈',
    '주식회사 베이비 투 키즈': '베이비 투 키즈',
    '베투키':                   '베이비 투 키즈',
    # 수정9: 베네피아 창원2호점 → 링크맘 창원2호점
    '베네피아 창원2호점(링크맘)': '링크맘 창원2호점',
    '베네피아 창원2호점':       '링크맘 창원2호점',
    # 수정3-13: 베네피아 창원점 → 링크맘 창원점
    '베네피아 창원점':          '링크맘 창원점',
    # 수정3-14: 링크맘 하남초일점 → 링크맘 하남점
    '링크맘 하남초일점':        '링크맘 하남점',
    # 신규: 링크맘 부산점 송정 → 링크맘 부산점
    '링크맘 부산점 송정':       '링크맘 부산점',
    '링크맘 부산점송정':        '링크맘 부산점',
    # 신규: 링크맘 판교 → 베이비스토리 판교 → 베이비스토리(판교점)
    '링크맘 판교':              '베이비스토리 판교점',
    '베이비스토리 판교':        '베이비스토리 판교점',
    # 신규: 링크맘 대전점 → 베이비페어365 대전점
    '링크맘 대전점':            '베이비페어365 대전점',
    # 신규: 부천디에스컴퍼니 → 베이비하우스 부천점
    '부천디에스컴퍼니':          '베이비하우스 부천점',
    '주식회사 부천디에스컴퍼니': '베이비하우스 부천점',
    # 은평점 (베이비파크 은평 = 베이비하우스 은평점, 같은 매장)
    '베이비파크 은평':          '베이비하우스 은평점',
    '베이비파크 은평점':        '베이비하우스 은평점',
    # 수정3-9: 베이비하우스 도봉 → 베이비하우스 도봉점
    '베이비하우스 도봉':        '베이비하우스 도봉점',
    # 베이비스토리
    '베이비스토리':             '베이비스토리 판교점',
    # 파주점 고객 데이터 → 삭제
    '베이비하우스 파주점/신성준고객': '',
}

# 화면 표시 전용 변환 (DB 값은 그대로, 보이는 이름만 변경)
DISPLAY_NAME = {
    # 하남미시점 오타 (DB에 남아있을 경우)
    '베이비하우스 하남미시점':           '베이비하우스 하남미사점',
    # 동대전점 (발육/용품 모두 동대전점으로 — DB에 남아있을 경우)
    '베이비하우스_동대전점_발육':        '베이비하우스 동대전점',
    '베이비하우스_동대전점_용품':        '베이비하우스 동대전점',
    '베이비하우스 동대전점 발육':        '베이비하우스 동대전점',
    '베이비하우스 동대전점 용품':        '베이비하우스 동대전점',
    '주식회사 에스엘컴퍼니':            '베이비하우스 동대전점',
    # 대전점 → 대전세종점 (동대전점과 별개)
    '베이비하우스 대전점':              '베이비하우스 대전세종점',
    '베이비하우스_대전점':              '베이비하우스 대전세종점',
    # 군포점 → 안양점 표시
    '베이비하우스 군포점':              '베이비하우스 안양점',
    # 부천, 일산
    '베네피아 부천점':                  '링크맘 부천점',
    '베이비하우스 일산화정점':          '베이비하우스 일산점',
    # 울산점
    '베이비하우스 울산점  엔픽스. 타프토이즈,카오스': '베이비하우스 울산점',
    '베이비하우스 울산점 엔픽스. 타프토이즈,카오스':  '베이비하우스 울산점',
    # 베이비스토리
    '베이비스토리':                     '베이비스토리 판교점',
    # 창원, 하남
    '베네피아 창원점':                  '링크맘 창원점',
    '링크맘 하남초일점':                '링크맘 하남점',
    # 도봉
    '베이비하우스 도봉':                '베이비하우스 도봉점',
    # 다산 (DB에서 이미 베이비스토어 다산으로 통합됐으나, 혹시 남아있을 경우)
    '베네피아 다산':                    '베이비스토어 다산',
    # 숨김 처리 (None = 표시 안 함)
    '베이비하우스 파주점/신성준고객':   None,
    # DB에 남아있을 경우 대비
    '베이비하우스 부산동래점':          '베이비하우스 동래점',
    '베이비하우스 동탄오산점':          '베이비하우스 동탄점',
    '베네피아 창원2호점(링크맘)':       '링크맘 창원2호점',
    '베네피아 창원2호점':              '링크맘 창원2호점',
}

def display_seller(name):
    """화면 표시용 매장명 변환 (DB 저장값 → 표시명)"""
    if not name: return name
    return DISPLAY_NAME.get(name, name)

def is_hidden_seller(name):
    """숨겨야 할 매장인지 확인"""
    return DISPLAY_NAME.get(name) is None and name in DISPLAY_NAME

def get_all_real_sellers_for(seller_name):
    """표시명에 해당하는 모든 DB real_seller 값 반환 (하남미시→미사 등 aliases 포함)"""
    # 표시명 → DB 저장명 역매핑
    reverse_map = {}
    for raw, display in DISPLAY_NAME.items():
        if display and display != raw:
            if display not in reverse_map:
                reverse_map[display] = []
            reverse_map[display].append(raw)
    # 해당 표시명에 매핑된 모든 DB 이름
    aliases = reverse_map.get(seller_name, [])
    # DB 저장명이 display_seller로 변환된 결과와 일치하는 것들
    db_names = [seller_name] + aliases
    return db_names
def resolve_seller(name):
    """매장명 정규화 — 별칭 매핑 + 불필요한 텍스트 제거"""
    import re as _re
    if not name: return name
    # 1. 별칭 매핑
    if name in SELLER_ALIAS: return SELLER_ALIAS[name]
    # 2. 매장명 뒤에 브랜드명/품목이 붙은 경우 제거
    #    예: "베이비하우스 울산점  엔픽스. 타프토이즈,카오스" → "베이비하우스 울산점"
    cleaned = _re.split(r'\s{2,}|\s*[,./]+\s*(?:엔픽스|줄즈|레카로|타프토이즈|카오스|원더폴드|ABC)', name)[0].strip()
    # 3. 앞뒤 공백 제거
    cleaned = cleaned.strip()
    return cleaned if cleaned else name

# 백화점 운영사(거래처명 기준) — 오프라인 매장과 별도 채널로 구분
DEPARTMENT_STORE_COMPANIES = [
    '서양네트웍스', '서양네트웤스',
    '가이아코퍼레이션', '가이아 코퍼레이션', '가이아코포레이션',
]

def detect_channel(seller_name_raw, real_seller=''):
    """거래처명(원본 계약법인명) 또는 매장명 기준으로 채널(오프라인/백화점) 판별"""
    combined = f"{seller_name_raw or ''} {real_seller or ''}"
    for company in DEPARTMENT_STORE_COMPANIES:
        if company in combined:
            return '백화점'
    return '오프라인'

BRAND_ORDER = ['줄즈', '레카로', 'ABC디자인', '원더폴드', '카오스', '엔픽스', '타프토이즈']

def remap_group(group, item_name=''):
    """품목그룹을 브랜드명으로 정규화 — 브랜드 태그 우선 적용"""
    import re
    g    = (group or '').strip()
    item = (item_name or '')

    # 제품명에서 [브랜드] 태그 추출
    brand_match = re.match(r'\[([^\]]+)\]', item)
    brand_tag   = brand_match.group(1) if brand_match else ''
    bt_lower    = brand_tag.lower()

    # ── 브랜드 태그 우선 판단 ──────────────────────
    if '줄즈' in brand_tag:                          return '줄즈'
    if '레카로' in brand_tag:                        return '레카로'
    if 'abc' in bt_lower:                            return 'ABC디자인'
    if '원더폴드' in brand_tag:                      return '원더폴드'
    if '카오스' in brand_tag:                        return '카오스'
    if '엔픽스' in brand_tag:                        return '엔픽스'
    if '타프토이즈' in brand_tag or 'taft' in bt_lower: return '타프토이즈'

    # ── 그룹명 기반 매핑 (태그 없는 경우) ──────────
    GROUP_MAP = {
        '유모차':          '줄즈',      # 태그 없으면 줄즈 기본
        '웨건':            '원더폴드',
        '컨버터블카시트':  '레카로',
        '주니어카시트':    '레카로',
        '토들러카시트':    '레카로',
        '카시트':          '레카로',
        '식탁의자':        '카오스',
        '하이체어':        '엔픽스',
        '보행기':          '엔픽스',
        '쏘서':            '엔픽스',
        '점퍼루':          '엔픽스',
        '휴대용부스터':    '엔픽스',
        'TAFTOYS':         '타프토이즈',
        '유아섬유류':      'ABC디자인',
    }
    return GROUP_MAP.get(g, g or '기타')

def normalize_item_name(name):
    """제품명에서 색상/옵션 완전 제거
    [줄즈]에어2_샌디타프 → [줄즈]에어2
    [레카로]토론1_엘레강트베이지_캐노피형 → [레카로]토론1
    [카오스]클랩하이체어_비치_내츄럴 → [카오스]클랩하이체어
    [원더폴드]W시리즈 엘리트프로_2인승_제트블랙 → [원더폴드]W시리즈 엘리트프로
    """
    if not name: return name
    import re
    # 언더바 이후 모든 내용 제거 (색상, 옵션, 한정판 등)
    cleaned = re.sub(r'_.*$', '', name).strip()
    # 괄호 내용도 제거 (예: "다크", "한정판")
    cleaned = re.sub(r'\s*\([^)]*\)\s*$', '', cleaned).strip()
    # 원더폴드 W시리즈: 인승 정보도 제거
    cleaned = re.sub(r'\s+\d+인승.*$', '', cleaned).strip()
    # 카오스: 클랩 베이비시트 → 클랩하이체어로 통합
    if '[카오스]클랩' in cleaned and '베이비시트' in cleaned:
        cleaned = '[카오스]클랩하이체어'
    return cleaned if cleaned else name


# ── 엑셀 리포트용 브랜드별 제품 라벨/정렬 커스텀 규칙 (수정2,4,5,6) ──
PRODUCT_LABEL_ORDER = {
    '줄즈':      ['지오3', '데이5', '허브2', '에어2'],
    '레카로':    ['토론1', '제논1', '액시언1', '벨릭스'],
    'ABC디자인': ['버디', '폴디', '태그', '루프트', '슈타트듀오'],
}

def get_custom_product_label(brand, item_name):
    """브랜드별 특수 제품 라벨 규칙. 매칭되면 라벨 문자열, 아니면 None(기본 규칙 사용)"""
    if brand == '원더폴드':
        # 수정5: L2/L4/W2/W4/W2슈퍼맨/W4슈퍼맨/W2폭스/W4폭스
        base = 'L' if 'L시리즈' in item_name else ('W' if 'W시리즈' in item_name else None)
        if not base:
            return None
        seater = '2' if '2인승' in item_name else ('4' if '4인승' in item_name else '')
        suffix = ''
        if '슈퍼맨' in item_name: suffix = '슈퍼맨'
        elif '폭스바겐' in item_name: suffix = '폭스'
        return f"{base}{seater}{suffix}" if seater else None
    if brand == '카오스':
        # 수정6: 오크/비치/리사이클 소재 기준으로 구분
        for material in ['오크', '비치', '리사이클']:
            if material in item_name:
                return material
        return None
    return None


def get_product_display_label(brand, norm_name):
    """제품 표시 라벨 — ABC디자인 등은 대괄호 브랜드 태그 완전 제거 (수정4)"""
    import re as _re_lbl
    return _re_lbl.sub(r'^\[[^\]]+\]', '', norm_name).strip()


def sort_product_labels(brand, labels):
    """브랜드별 지정 순서로 정렬 (수정2,4). 지정 안 된 브랜드/라벨은 알파벳 순 뒤에 배치"""
    order = PRODUCT_LABEL_ORDER.get(brand)
    if not order:
        # 원더폴드: L2,L4,W2,W4,W2슈퍼맨,W4슈퍼맨,W2폭스,W4폭스 순
        if brand == '원더폴드':
            wf_order = ['L2','L4','W2','W4','W2슈퍼맨','W4슈퍼맨','W2폭스','W4폭스']
            return sorted(labels, key=lambda x: (wf_order.index(x) if x in wf_order else 99, x))
        # 카오스: 오크,비치,리사이클 순
        if brand == '카오스':
            ks_order = ['오크','비치','리사이클']
            return sorted(labels, key=lambda x: (ks_order.index(x) if x in ks_order else 99, x))
        return sorted(labels)
    return sorted(labels, key=lambda x: (order.index(x) if x in order else 99, x))



def get_group_sort_key(group):
    """브랜드 정렬 순서"""
    try:
        return BRAND_ORDER.index(group)
    except ValueError:
        return 99
# ── 타프토이즈 제품 카탈로그 ─────────────────────
TAFTOYS_CATALOG = {
    '[타프토이즈]드라이브&디스커버트래블토이': {'price':26900,'category':'트래블토이','desc':'이동 중 아이 집중도 UP, 유모차 부착 가능'},
    '[타프토이즈]사바나 어드벤쳐 아치':       {'price':28600,'category':'아치/모빌','desc':'바닥 놀이 필수템, 감각 발달 + 인테리어 효과'},
    '[타프토이즈]라이드타임비지북':            {'price':23600,'category':'비지북','desc':'유모차·카시트 부착, 0-3세 인지발달'},
    '[타프토이즈]트로피컬 오케스트라 아치 모빌':{'price':27800,'category':'아치/모빌','desc':'뮤지컬 아치, 터미타임 필수'},
    '[타프토이즈]코알라 카 휠 토이':          {'price':30900,'category':'카시트 장난감','desc':'카시트 부착, 지루한 이동 시간 해결사'},
    '[타프토이즈]어반가든 팝업 티슈 박스':    {'price':22500,'category':'감각 장난감','desc':'무한 반복 놀이, 소근육 발달 최고'},
    '[타프토이즈]어반가든 유모차 모빌':       {'price':15600,'category':'아치/모빌','desc':'유모차 클립형, 시각 자극 + 휴대성'},
    '[타프토이즈]피크 앤 플레이 큐브':        {'price':18400,'category':'큐브','desc':'6면 다기능, 0-2세 전방위 발달'},
    '[타프토이즈]마이홈비지북':               {'price':18400,'category':'비지북','desc':'집 모양 비지북, 역할놀이 시작'},
    '[타프토이즈]어반가든 액티비티 큐브':     {'price':15300,'category':'큐브','desc':'4면 액티비티, 혼자 놀기 최적'},
    '[타프토이즈]사바나 360 액티비티짐':      {'price':89000,'category':'액티비티짐','desc':'360도 회전 아치, 신생아부터 12개월'},
    '[타프토이즈]사바나 터미타임 북':         {'price':19800,'category':'터미타임','desc':'엎드려 놀기 훈련, 목 근육 강화'},
    '[타프토이즈]어반가든 뮤지컬 버니':       {'price':28900,'category':'인형/뮤지컬','desc':'뮤지컬 봉제 인형, 수면 루틴 도움'},
    '[타프토이즈]어반가든 터미타임 스피닝북': {'price':22000,'category':'터미타임','desc':'스피닝 기능, 아이 주의 집중'},
    '[타프토이즈]사바나 디스커버리 큐브':     {'price':32900,'category':'큐브','desc':'프리미엄 큐브, 1-3세 탐색놀이'},
    '[타프토이즈]북극 액티비티 북':           {'price':21000,'category':'비지북','desc':'천 소재 액티비티 북, 감촉 자극'},
    '[타프토이즈]아이스크림 베어 워터매트':   {'price':35000,'category':'워터매트','desc':'여름 필수템, 터미타임 + 시각자극'},
    '[타프토이즈]팬더 블룸 워터매트':         {'price':35000,'category':'워터매트','desc':'실내 물놀이, 감각 자극 극대화'},
    '[타프토이즈]팝앤플레이스테이션':         {'price':45000,'category':'액티비티짐','desc':'팝업 텐트형, 실내 놀이공간 완성'},
    '[타프토이즈]파멜라 레인스틱':            {'price':16000,'category':'감각 장난감','desc':'청각 자극, 비 소리 감각놀이'},
    '[타프토이즈]코알라 액티비티 스파이럴':   {'price':14500,'category':'트래블토이','desc':'유모차/카시트 나선형, 다양한 질감'},
    '[타프토이즈]미니문 유모차 모빌':         {'price':13500,'category':'아치/모빌','desc':'초소형 모빌, 어디든 클립 부착'},
    '[타프토이즈]베어 허그 스파이럴':         {'price':14500,'category':'트래블토이','desc':'곰돌이 스파이럴, 촉감+색상 자극'},
}

@app.route("/api/script/analysis")
@login_required
def api_script_analysis():
    seller_raw = request.args.get("seller","").strip()
    year       = request.args.get("year", str(datetime.now().year))
    seller     = resolve_seller(seller_raw)
    conn       = get_db()

    sold_items=[dict(r) for r in conn.execute("""
        SELECT item_group,item_name,SUM(quantity) qty,SUM(total) total,COUNT(*) cnt,
               MIN(sale_date) first_sale,MAX(sale_date) last_sale
        FROM sales_data WHERE (real_seller=? OR real_seller=?) AND sale_date LIKE ? AND sale_date!=''
        GROUP BY item_name ORDER BY total DESC""",(seller,seller_raw,f"{year}%")).fetchall()]

    # 색상 통합 — 같은 제품 합산
    norm_items = {}
    for r in sold_items:
        brand = remap_group(r['item_group'], r['item_name'])
        norm  = normalize_item_name(r['item_name'])
        key   = (brand, norm)
        if key not in norm_items:
            norm_items[key] = dict(r); norm_items[key]['item_name'] = norm; norm_items[key]['item_group'] = brand
        else:
            norm_items[key]['qty']   += r['qty']
            norm_items[key]['total'] += r['total']
            norm_items[key]['cnt']   += r['cnt']
    sold_items = sorted(norm_items.values(), key=lambda x: -x['total'])

    brand_summary={}
    for r in sold_items:
        b=remap_group(r['item_group'],r['item_name'])
        if b not in brand_summary: brand_summary[b]={'qty':0,'total':0}
        brand_summary[b]['qty']+=r['qty']; brand_summary[b]['total']+=r['total']

    sold_taft=set(normalize_item_name(r['item_name']) for r in sold_items
                  if remap_group(r['item_group'],r['item_name'])=='타프토이즈')

    # 재고 데이터 로드 (10개 이상인 타프토이즈만 추천)
    stock_rows = conn.execute(
        "SELECT item_name, SUM(quantity) qty FROM stock_data WHERE quantity>0 GROUP BY item_name"
    ).fetchall()
    stock_map = {}
    for sr in stock_rows:
        norm_s = normalize_item_name(sr[0])
        stock_map[norm_s] = stock_map.get(norm_s, 0) + sr[1]

    # 전체 타프토이즈 인기 순위 (전체 판매 데이터 기반)
    taft_popularity = {
        normalize_item_name(r[0]): r[1]
        for r in conn.execute("""
            SELECT item_name, SUM(quantity) qty FROM sales_data
            WHERE item_name LIKE '%타프토이즈%' OR item_group='TAFTOYS'
            GROUP BY item_name ORDER BY qty DESC
        """).fetchall()
    }

    unsold_taft = []
    for k, v in TAFTOYS_CATALOG.items():
        norm_k = normalize_item_name(k)
        if norm_k in sold_taft: continue  # 이미 취급 중
        # 재고 확인 — 10개 이상인 것만
        stock_qty = stock_map.get(norm_k, 0)
        if stock_map and stock_qty < 10: continue  # 재고 있을 때만 필터링
        popularity = taft_popularity.get(norm_k, 0)
        unsold_taft.append({
            'name': norm_k, 'category': v['category'],
            'price': v['price'], 'desc': v['desc'],
            'stock': stock_qty, 'popularity': popularity,
        })

    # 인기순 정렬 (전체 판매량 기준)
    unsold_taft.sort(key=lambda x: -x['popularity'])

    daily=[dict(r) for r in conn.execute("""
        SELECT sale_date,SUM(total) total,SUM(quantity) qty,COUNT(*) cnt
        FROM sales_data WHERE (real_seller=? OR real_seller=?) AND sale_date LIKE ? AND sale_date!=''
        GROUP BY sale_date ORDER BY sale_date""",(seller,seller_raw,f"{year}%")).fetchall()]

    weekly_raw=conn.execute("""
        SELECT strftime('%Y-%W',sale_date) wk,MIN(sale_date) md,SUM(total) total,SUM(quantity) qty
        FROM sales_data WHERE (real_seller=? OR real_seller=?) AND sale_date LIKE ? AND sale_date!=''
        GROUP BY wk ORDER BY wk""",(seller,seller_raw,f"{year}%")).fetchall()

    from datetime import datetime as dt2,timedelta
    weekly=[]
    for r in weekly_raw:
        try:
            d=dt2.strptime(r[1],"%Y-%m-%d"); sun=d-timedelta(days=(d.weekday()+1)%7)
            weekly.append({'week':r[0],'week_start':sun.strftime("%Y-%m-%d"),
                           'week_end':(sun+timedelta(days=6)).strftime("%Y-%m-%d"),'total':r[2],'qty':r[3]})
        except: pass

    total_all=conn.execute(f"SELECT SUM(total) FROM sales_data WHERE sale_date LIKE '{year}%'").fetchone()[0] or 1
    seller_total=sum(r['total'] for r in sold_items)
    conn.close()

    return jsonify({
        'seller':seller,'year':year,'total':seller_total,
        'total_pct':round(seller_total/total_all*100,1),
        'brand_summary':[{'brand':k,'qty':v['qty'],'total':v['total'],
                          'pct':round(v['total']/seller_total*100,1) if seller_total else 0}
                         for k,v in sorted(brand_summary.items(),key=lambda x:-x[1]['total'])],
        'sold_items':sold_items,'top5':sold_items[:5],
        'unsold_taft':unsold_taft,'daily':daily,'weekly':weekly,
    })

@app.route("/api/script/generate", methods=["POST"])
@login_required
def api_script_generate():
    """데이터 기반 영업 스크립트 — 매장 패턴별 분기 + 매번 다른 각도"""
    import random, hashlib
    from datetime import datetime as dt2

    data        = request.json or {}
    seller      = data.get('seller', '')
    analysis    = data.get('analysis', {})
    gen_count   = data.get('gen_count', 0)

    year        = analysis.get('year', str(dt2.now().year))
    total       = analysis.get('total', 0)
    total_pct   = analysis.get('total_pct', 0.0)
    brands      = analysis.get('brand_summary', [])
    top5        = analysis.get('top5', [])
    sold_items  = analysis.get('sold_items', [])
    unsold_taft = analysis.get('unsold_taft', [])
    weekly      = analysis.get('weekly', [])

    # 시드: 매번 다른 결과
    seed_str = f"{seller}{gen_count}{dt2.now().strftime('%H%M%S')}"
    rng = random.Random(int(hashlib.md5(seed_str.encode()).hexdigest()[:8], 16))
    def w(n): return f"{n:,}"
    def pick(lst): return rng.choice(lst)

    # ── 데이터 분석 ────────────────────────────────
    top_brand   = brands[0]['brand'] if brands else ''
    top_pct     = brands[0]['pct']   if brands else 0
    top2_brand  = brands[1]['brand'] if len(brands) > 1 else ''
    top2_pct    = brands[1]['pct']   if len(brands) > 1 else 0
    top_item    = normalize_item_name(top5[0].get('item_name','')) if top5 else ''
    top_item_qty= top5[0].get('qty',0) if top5 else 0
    top_item_tot= top5[0].get('total',0) if top5 else 0

    all_brands  = set(b['brand'] for b in brands)
    missing_brs = [b for b in BRAND_ORDER if b not in all_brands and b != '타프토이즈']
    weak_brs    = [b for b in brands if b['pct'] < 5 and b['brand'] != '타프토이즈']

    taft_sold   = [r for r in sold_items if remap_group(r.get('item_group',''), r.get('item_name',''))=='타프토이즈']
    taft_total  = sum(r.get('total',0) for r in taft_sold)
    taft_cnt    = len(set(normalize_item_name(r.get('item_name','')) for r in taft_sold))
    taft_pct    = round(taft_total/total*100,1) if total else 0

    week_avg   = int(sum(wk.get('total',0) for wk in weekly)/len(weekly)) if weekly else 0
    week_trend = ''
    if len(weekly) >= 3:
        recent = [wk.get('total',0) for wk in weekly[-3:]]
        if recent[-1] > recent[0]*1.15:   week_trend = '강한상승'
        elif recent[-1] > recent[0]*1.05: week_trend = '상승'
        elif recent[-1] < recent[0]*0.85: week_trend = '하락'
        elif recent[-1] < recent[0]*0.95: week_trend = '약한하락'
        else: week_trend = '안정'

    CAT_PRI = {'아치/모빌':1,'액티비티짐':2,'트래블토이':3,'비지북':4,'큐브':5,'워터매트':6,'터미타임':7}
    rec_taft = sorted(unsold_taft, key=lambda x: CAT_PRI.get(x.get('category',''),9))

    month_now = dt2.now().month
    season_map = [(range_k, v) for range_k, v in [((3,4,5),'봄'),((6,7,8),'여름'),((9,10,11),'가을'),((12,1,2),'겨울')]]
    season = next((v for k,v in season_map if month_now in k), '봄')

    # 매장 등급 / 패턴
    if total_pct >= 10:    store_tier = 'VIP'
    elif total_pct >= 5:   store_tier = 'A'
    elif total_pct >= 2:   store_tier = 'B'
    else:                  store_tier = 'C'

    if top_pct >= 70:          store_pattern = 'mono'
    elif top_pct >= 45:        store_pattern = 'dominant'
    elif len(brands) >= 4 and top_pct < 40: store_pattern = 'balanced'
    elif len(brands) <= 2:     store_pattern = 'narrow'
    else:                      store_pattern = 'duo'

    taft_pattern = 'none' if taft_pct==0 else ('low' if taft_pct<5 else ('mid' if taft_pct<15 else 'high'))

    def rank_expr(tier, pct):
        if tier=='VIP': return f"상위 {pct}% 핵심 거래처"
        if tier=='A':   return f"상위권 거래처 (전체 대비 {pct}%)"
        if tier=='B':   return f"중요 거래처 (전체의 {pct}%)"
        return f"성장 가능성 높은 거래처 ({pct}%)"

    # ── 섹션 1: 오프닝 ────────────────────────────
    opening_pool = [
        f'''영업사원: "사장님, 안녕하세요! 오늘 오기 전에 {seller} 데이터 뽑아봤는데 숫자가 좋아서 오는 길에 기분이 좋았어요.\n잠깐 같이 보실 수 있으세요?"\n\n(태블릿/자료 꺼내며)\n\n영업사원: "저희 전체 거래처 중에서 {rank_expr(store_tier, total_pct)}이에요. {top_brand} 비중이 {top_pct}%로 탄탄하게 잡혀 있어요."''',
        f'''영업사원: "사장님, 들어오면서 {top_item} 진열이 눈에 잘 띄더라고요. 역시 매장 동선을 잘 잡고 계신 것 같아요."\n\n사장님: (반응)\n\n영업사원: "실제로 {top_item}이 {w(top_item_qty)}개 나갔거든요. 저희 거래처 중에서도 상위권이에요. {year}년 데이터 정리해서 가져왔는데, 같이 보실까요?"''',
        f'''영업사원: "사장님! 요즘 {top_brand} 어떠세요? 저희 다른 매장들이 {top_brand} 문의가 {pick(['많이 늘었다','꾸준하다','올해 특히 좋다'])}고 하더라고요."\n\n사장님: (반응)\n\n영업사원: "{seller}도 비슷한 흐름이에요. {year}년 데이터 분석해서 가져왔어요. 꼭 공유드리고 싶었습니다."''',
        f'''영업사원: "사장님, {{"봄":"신학기 시즌이라","여름":"여름이라","가을":"가을 나들이 시즌이라","겨울":"연말이라"}}[season] 매장 분위기 어떠세요?"\n\n사장님: (반응)\n\n영업사원: "맞아요. 저도 이 시즌에 딱 맞는 제안 드리려고 왔어요. {year}년 데이터 기반으로 준비했거든요."''',
    ]
    s1 = pick(opening_pool)

    # ── 섹션 2: 실적 공유 ────────────────────────
    brand_lines = '\n'.join(f"  · {b['brand']}: {w(b['total'])}원 ({b['pct']}%)" for b in brands[:5])
    trend_map = {'강한상승':'최근 3주 추이가 강하게 올라가고 있어요! 이 흐름 놓치면 안 됩니다.',
                 '상승':'최근 추이도 상승 중이라 지금이 발주 타이밍이에요.',
                 '안정':'판매가 꾸준히 안정적으로 유지되고 있어요. 탄탄한 베이스가 있는 거예요.',
                 '약한하락':'최근 3주가 살짝 내려갔는데, 진열 변화로 충분히 잡을 수 있어요.',
                 '하락':'최근 흐름이 빠졌는데, 오늘 원인 같이 찾아봐요. 해결책이 있어요.','':''}
    trend_comment = trend_map.get(week_trend,'')

    pattern_comments = {
        'mono': f"{top_brand} 하나에 {top_pct}% 집중하고 계신데, 이걸로 {w(total)}원을 만드신 게 대단해요. 근데 한 브랜드 의존도가 높으면 리스크가 있어요. 오늘 그 다음 전략 얘기해봐요.",
        'dominant': f"{top_brand}({top_pct}%)가 압도적이고, {top2_brand}({top2_pct}%)가 받쳐주는 구조예요. {top2_brand} 비중을 더 키우면 전체 매출이 쑥 올라가요.",
        'balanced': "브랜드 구성이 다양하게 잡혀 있어요. 각 브랜드가 역할 분담하는 구조인데, 조금만 최적화하면 같은 방문객 수로 매출을 더 올릴 수 있어요.",
        'narrow': f"지금 {len(brands)}개 브랜드 취급하고 계신데, 1-2개 추가하면 고객 이탈을 줄일 수 있어요.",
        'duo': f"{top_brand}와 {top2_brand}의 2강 구도예요. 이 구조 자체는 좋은데, 세 번째 기둥이 생기면 더 안정적이에요.",
    }
    real_talk = pattern_comments.get(store_pattern, '')

    motivation = pick([
        "분명히 고객들이 사장님 매장을 신뢰한다는 거예요. 추천을 잘 해주시니까요.",
        "제품을 그냥 파는 게 아니라 제대로 설명해서 파신다는 게 느껴져요.",
        "이 매출은 그냥 나오는 게 아니에요. 사장님이 만들어 낸 거예요.",
    ])
    pushback = pick([
        "다른 매장들이랑 비교하면 여기가 훨씬 잘하고 있어요. 체감이 안 될 뿐이에요.",
        "힘들다고 느끼실 때가 도약 직전인 경우가 많아요. 오늘 같이 방법 찾아봐요.",
        "그래도 이 숫자는 시장 평균보다 위에 있어요. 기반이 탄탄하다는 뜻이에요.",
    ])

    s2 = f'''영업사원: "사장님, {year}년 {seller} 전체 데이터예요.\n\n총 {w(total)}원 — {rank_expr(store_tier, total_pct)}입니다.\n{"주간 평균 " + w(week_avg) + "원으로 " if week_avg else ""}꾸준히 판매되고 있고요.\n\n브랜드별로 보면:\n{brand_lines}\n\n{real_talk}\n{"  ※ " + trend_comment if trend_comment else ""}\n\n{motivation}\n\n💡 힘들다 하시면:\n→ "{pushback}"'''

    # ── 섹션 3: 베스트 제품 ───────────────────────
    brand_insights = {
        '레카로': {'reason':'카시트는 안전 민감 제품 → 전문점 신뢰도 핵심. 설명 잘 해주시니까 팔림', 'upsell':'카시트 구매자에게 타프토이즈 카시트 장난감 추가 제안', 'risk':'재고 소진 시 이탈 위험 높음 — 안전재고 3개 권장'},
        '줄즈': {'reason':'SNS 바이럴 강함 + 색상 다양성 → 엄마 커뮤니티 추천 1위', 'upsell':'에어2 구매자에게 데이5 신색상 미리 예약 유도', 'risk':'시즌별 신색상 → 구색 부족 시 기회 손실'},
        '원더폴드': {'reason':'웨건 카테고리 독점적 포지션 → 비교 구매 없이 결정', 'upsell':'웨건 구매자에게 타프토이즈 트래블토이 번들 제안', 'risk':'전시 필수 — 실물 못 보면 구매 주저'},
        '엔픽스': {'reason':'국내 브랜드 신뢰 + 합리적 가격 → 재구매율 높음', 'upsell':'보행기 구매자에게 비바체(하이체어) 또는 쏘서 연계', 'risk':'시즌 수요 집중 — 봄여름 전 선발주 중요'},
        '카오스': {'reason':'하이체어 프리미엄 포지션 + 디자인 감성 → 인테리어 중시 부모층 강함', 'upsell':'하이체어 구매 후 이유식 용품 연계', 'risk':'높은 단가 → 충분한 설명과 체험 필수'},
        'ABC디자인': {'reason':'유럽 감성 디자인 → 20-30대 부모층 강함', 'upsell':'유모차+카시트 패밀리 세트 구성 제안', 'risk':'인지도 낮음 → 설명력이 판매 좌우'},
        '타프토이즈': {'reason':'완구 시장 최고 성장 브랜드 + 선물 수요 높음', 'upsell':'아치→모빌→비지북→큐브 시리즈 업셀', 'risk':'전시 위치가 판매 좌우 — 눈에 잘 띄는 곳 배치'},
    }

    best_blocks = []
    for i, r in enumerate(top5[:3]):
        nm    = normalize_item_name(r.get('item_name',''))
        qty   = r.get('qty', 0)
        tot   = r.get('total', 0)
        brand = remap_group(r.get('item_group',''), r.get('item_name',''))
        ins   = brand_insights.get(brand, {'reason':'검증된 베스트셀러','upsell':'','risk':''})
        if qty >= 20:   qty_comment = f"{w(qty)}개는 전국 상위권 판매량이에요."
        elif qty >= 10: qty_comment = f"{w(qty)}개, 이 브랜드 기준 잘 나가는 편이에요."
        else:           qty_comment = f"{w(qty)}개인데, 여기서 더 올릴 여지가 충분해요."
        block = f"  {i+1}위. {nm} — {w(qty)}개 / {w(tot)}원\n  {qty_comment}\n  잘 팔리는 이유: {ins['reason']}\n  연계 제안: {ins['upsell']}\n  주의: {ins['risk']}\n"
        best_blocks.append(block)

    stock_q = pick([
        f'"{top_item} 재고 지금 몇 개 남아계세요? 이 제품은 품절 나면 고객이 바로 온라인으로 가거든요."',
        f'"{top_item} 다음 발주 언제 생각하고 계세요? 제가 미리 물량 잡아드릴게요."',
        f'"{top_item} 재고 체크해보실 수 있어요? 이번 달 소진 속도 보고 발주량 같이 정해드릴게요."',
    ])
    pushback2 = pick([
        "혹시 최근에 진열 위치가 바뀌셨어요? 위치가 판매량에 정말 크게 영향 주거든요.",
        "고객 문의는 있는데 구매로 안 이어지나요? 어떤 제품과 비교하시는지 여쭤봐도 될까요?",
        "그럴 때일수록 재고 줄이고 다른 제품 비중 늘리는 게 맞을 수 있어요. 같이 봐요.",
    ])
    s3 = f'''영업사원: "이 매장 베스트 TOP3 분석해봤어요.\n\n{chr(10).join(best_blocks)}\n{stock_q}\n\n💡 \"요즘 그 제품 잘 안 나가요\" 하시면:\n→ "{pushback2}"'''

    # ── 섹션 4: 타프토이즈 ───────────────────────
    if taft_pattern == 'none':
        taft_approaches = [
            f'''영업사원: "사장님, 타프토이즈 아세요? 유아 완구 브랜드인데 베이비페어에서 카시트보다 줄 서는 브랜드가 됐어요.\n\n이 매장에 없는 이유가 있을 것 같아서요. 혹시 완구는 취급 안 하시는 정책인가요?\n\n사실 완구가 매장 객단가 올리는 데 효과적이에요. 카시트 하나 사러 온 고객이 {w(rec_taft[0].get("price",25000) if rec_taft else 25000)}원짜리 완구 하나 더 집어가거든요.\n\n💡 \"마진이 낮지 않나요?\" 하시면:\n→ \"오히려 반대예요. 카시트보다 완구 마진이 높아요. 재방문 효과도 있어요.\""''',
            f'''영업사원: "사장님, 솔직하게 여쭤볼게요. 지금 고객 한 분당 평균 구매 금액이 얼마인 것 같으세요?"\n\n사장님: (반응)\n\n영업사원: "저희 데이터로 {seller} 평균 객단가가 {w(int(total/len(sold_items)) if sold_items else 0)}원 정도예요. 근데 타프토이즈 취급 매장들은 평균 25,000원씩 더 나와요.\n카시트 사면서 완구 하나 더 집어가는 거거든요. 그 역할 할 제품이 지금 이 매장엔 없어요."''',
        ]
        s4 = pick(taft_approaches)
        if rec_taft:
            r0 = rec_taft[0]; nm0 = r0.get("name","").replace("[타프토이즈]","").strip()
            s4 += f'\n\n제가 이 매장에 맞는 제품 골라봤어요:\n  ◆ {nm0} ({r0.get("category","")}) — {w(r0.get("price",0))}원\n    "{r0.get("desc","")}"\n    처음엔 3종 소량으로 시작해보세요. 한 달 후에 반응 보고 확대해드릴게요.'

    elif taft_pattern == 'low':
        taft_names = [normalize_item_name(r.get('item_name','')) for r in taft_sold[:2]]
        untracked = [u for u in rec_taft if normalize_item_name(u.get('name','')) not in taft_names][:2]
        ut_txt = ""
        if untracked:
            u = untracked[0]
            ut_txt = f'\n  ◆ {u.get("name","").replace("[타프토이즈]","").strip()} ({u.get("category","")}) — {w(u.get("price",0))}원\n    "{u.get("desc","")}"'
        s4 = f'''영업사원: "타프토이즈 {taft_cnt}종에서 {w(taft_total)}원 나왔어요. 비중이 {taft_pct}%인데, 이걸 10%로만 올려도 전체 매출이 달라져요.\n\n지금 취급 중인 제품 고객들에게 시리즈 연결이 잘 안 되고 있을 가능성이 높아요.\n아치 산 고객한테 3주 후 \"아이가 자라면 이 제품이 딱이에요\" 연락하면 재방문이 돼요.\n\n이번에 추가 추천 제품:{ut_txt}\n\n💡 \"관리하기 어려워요\" 하시면:\n→ \"이 브랜드는 팔고 나면 고객이 알아서 찾아와요. 설명이 필요 없는 브랜드예요.\""''' 

    elif taft_pattern == 'mid':
        s4 = f'''영업사원: "타프토이즈가 {taft_pct}%까지 올라왔는데, 여기가 중간 고비예요. 이 브랜드가 20% 이상 되면 매장 이미지 자체가 바뀌거든요.\n\n{pick(["트래블토이는 카시트/유모차 옆에 두면 번들 구매가 자연스럽게 일어나요.","액티비티짐 하나만 놔도 인스타 감성 사진이 나와서 매장이 SNS에 올라가요.","비지북 시리즈는 선물용 수요가 강해서 돌잔치 코너 옆에 두면 효과적이에요."])}\n\n이번에 2종 추가해보시고, 한 달 후에 반응 체크해드릴게요."'''

    else:
        top_taft_item = normalize_item_name(taft_sold[0].get('item_name','')) if taft_sold else ''
        s4 = f'''영업사원: "타프토이즈 {taft_pct}%면 저희 거래처 중 최상위권이에요. {top_taft_item}을 중심으로 정말 잘 운영하고 계세요.\n\n이제 다음 레벨 얘기를 해도 될 것 같아요. 신상 독점 전시를 해보시는 건 어때요?\n전시한 매장들은 한 달 만에 타프 매출이 평균 {pick(["40%","35%","28%"])} 올랐어요."'''

    # ── 섹션 5: 구조 개선 ────────────────────────
    if missing_brs:
        miss1 = missing_brs[0]
        miss_advice = {'원더폴드':'웨건은 유모차와 겹치지 않아요. 오히려 유모차 사고 웨건도 사는 가정이 많아요.','ABC디자인':'ABC는 유럽 감성이라 줄즈와 다른 고객층이에요. 경쟁이 아니라 보완이에요.','카오스':'하이체어는 이유식 시작 6개월 필수품이에요. 카시트 구매 후 타이밍 맞게 제안하면 돼요.','엔픽스':'보행기/쏘서는 6-12개월 집중 수요예요. 카시트 구매 3-4개월 후 제안하면 재방문이 돼요.'}.get(miss1, f'{miss1}은 이 매장 고객층에 맞는 브랜드예요.')
        s5 = f'''영업사원: "솔직히 아쉬운 게 있어요. {", ".join(missing_brs[:2])} 쪽이 빠져있거든요.\n\n{miss_advice}\n\n지금 오는 고객들이 {miss1} 때문에 다른 매장 가는 경우가 있을 수 있어요.\n처음에 전시용 1개만 두고 반응 보세요.\n\n💡 \"그 브랜드 잘 모르는데요\" 하시면:\n→ \"제가 직접 설명 드리고, 첫 고객 상담도 같이 해드릴 수 있어요.\""''' 

    elif weak_brs:
        wb1 = weak_brs[0]; wb_name = wb1['brand']; wb_pct = wb1['pct']
        advice = pick([f'{wb_name}이 {wb_pct}%인데, 진열 위치만 바꿔도 달라져요.',f'{wb_name}은 {top_brand} 구매 고객에게 추가 제안하는 방식이 더 효과적이에요.',f'{wb_name} 단독보다 세트 구성으로 팔면 부담이 줄어요.'])
        s5 = f'''영업사원: "브랜드 구성은 좋은데, {wb_name} 비중이 {wb_pct}%로 낮아요.\n\n{advice}\n\n{pick(["제가 다음 방문 때 진열 레이아웃 같이 봐드릴게요.","이 브랜드 잘 파는 다른 매장 사례 공유해드릴게요.","한 달에 2배 올린 매장도 있어요. 비결 알려드릴게요."])}"''' 

    else:
        deep = pick([f'{top_brand}에서 {top_item} 잘 파시는데, 같은 브랜드 2-3종 더 깊이 파는 라인업 확장 전략이 있어요. 7종 이상 취급하면 전문 매장 이미지가 생겨요.','브랜드 구성은 완성 단계예요. 이제 각 브랜드에서 프리미엄 라인 하나씩 추가하면 객단가가 올라가요.','이 정도면 다음 스텝은 고객 관계 관리 시스템화예요. 매출이 한 단계 더 올라갈 수 있어요.'])
        s5 = f'''영업사원: "브랜드 구성은 정말 잘 잡혀 있어요. 진심으로 칭찬이에요.\n\n{deep}"''' 

    # ── 섹션 6: 시즌 전략 ────────────────────────
    season_data = {
        '봄':  {'items':['줄즈 에어2 봄 신색상 (3-4월 출시)','타프토이즈 어반가든 아치 (야외 테마)','엔픽스 보행기 (신학기 선물)'],'insight':'3-5월은 출생아 수 피크 + 어린이날 선물 수요 집중. 이 시기 발주가 1년 매출을 좌우해요.','action':'어린이날 전 선물 포장 세트 구성하면 객단가가 올라가요.'},
        '여름':{'items':['타프토이즈 워터매트 (6-8월 한정)','타프토이즈 팝앤플레이스테이션 (실내 놀이)','레카로 카시트 (여름 휴가 이동 수요)'],'insight':'워터매트는 7월까지가 발주 골든타임. 8월엔 재고 소진 빠르고 보충 어려워요.','action':'에어컨 켠 여름에 실내 액티비티짐이 의외로 잘 나가요.'},
        '가을':{'items':['레카로 카시트 (추석 선물 수요)','원더폴드 웨건 (가을 나들이)','타프토이즈 비지북 (독서의 계절)'],'insight':'추석 전후 2주가 선물 수요 피크. 재고 부족은 기회 손실이 커요.','action':'선물 포장 서비스 앞에 내세우면 입소문이 나요.'},
        '겨울':{'items':['타프토이즈 실내놀이 세트 (겨울 실내)','줄즈 크리스마스 에디션','엔픽스 점퍼루 (실내 활동)'],'insight':'12월 크리스마스 + 1월 설 선물로 더블 피크. 11월 말까지 발주 완료가 핵심이에요.','action':'크리스마스 패키지 구성이 있으면 인스타 바이럴이 잘 돼요.'},
    }
    sd = season_data.get(season, season_data['봄'])
    wait_pushback = pick(['사장님, 지켜보다가 타이밍 놓치면 다음 시즌까지 기다려야 해요.','주변 매장들이 지금 발주 넣고 있어요. 같이 움직이시는 게 유리해요.','소량으로라도 먼저 들여놓고 반응 보세요. 안 팔리면 제가 어떻게든 해결해드릴게요.'])
    s6 = f'''영업사원: "지금 {season}이잖아요. {sd["insight"]}\n\n이 시기 집중 제품:\n{chr(10).join(f"  · {p}" for p in sd["items"])}\n\n{sd["action"]}\n\n지금 발주 넣으시면 이번 주 안으로 납품 가능해요.\n시즌 물량은 한정이라 이번에 같이 넣어두시죠."\n\n💡 \"일단 지켜볼게요\" 하시면:\n→ "{wait_pushback}"'''

    # ── 섹션 7: 클로징 ───────────────────────────
    checklist = []
    if top5: checklist.append(f"{normalize_item_name(top5[0].get('item_name',''))} 재고 확보")
    if taft_pattern in ('none','low'): checklist.append("타프토이즈 3종 소량 시작")
    elif taft_pattern == 'mid': checklist.append("타프토이즈 2종 추가")
    if missing_brs: checklist.append(f"{missing_brs[0]} 전시용 1종 시작")
    checklist.append(f"{season} 시즌 집중 제품 발주")

    closing_variants = [
        f'''영업사원: "오늘 이야기 나눈 거 정리할게요:\n\n{chr(10).join(f"  ✓ {item}" for item in checklist)}\n\n다 한꺼번에 하시기 부담스러우시면, 오늘은 {checklist[0]}만 먼저 해도 돼요.\n어느 쪽부터 시작하실래요?"''',
        f'''영업사원: "오늘 제안드린 것들 다 하시면 {w(int(total*0.15))}~{w(int(total*0.25))}원 추가 매출이 가능해요.\n\n한 번에 다 하실 필요 없고요, 오늘 {checklist[0]}부터 시작해볼까요?\n발주서 바로 뽑아드릴게요."''',
        f'''영업사원: "사장님, 6개월 후 이 매장 그림을 그려봤어요. {top_brand}는 지금보다 {pick(["20%","15%","25%"])} 더 올리고, 타프토이즈가 10%를 차지하면 연간 {w(int(total*1.3))}원이 충분히 가능해요.\n\n그 첫 걸음을 오늘 같이 내딛어볼까요? 발주 구성 최적화해서 바로 올려드릴게요."''',
    ]

    next_visit = pick([
        f'3주 후에 오늘 발주한 제품들 반응 들으러 올게요. 사장님 목소리 기다려요.',
        f'다음 달 초에 다시 방문드릴게요. 그때 신규 제품 판매 현황 같이 봐요.',
        f'2주 후에 들를게요. 그때까지 신규 제품 첫 반응 꼭 알려주세요.',
    ])
    s7 = pick(closing_variants) + f'''\n\n────────────────────────\n다음 방문 약속:\n"{next_visit}"'''

    # ── 최종 조합 ─────────────────────────────────
    def section(title, content):
        return f"{'━'*52}\n【{title}】\n{'━'*52}\n{content}\n"

    now_str = dt2.now().strftime('%Y.%m.%d %H:%M')
    script = f"""{'='*57}
  매장 영업 방문 스크립트 — {seller}
  분석: {year}년 / 생성: {now_str} / 유형: [{store_tier}/{store_pattern}/타프{taft_pattern}]
{'='*57}

{section('1. 오프닝 — 첫 60초가 전체를 결정한다', s1)}
{section('2. 실적 공유 — 숫자로 신뢰를 만든다', s2)}
{section('3. 베스트 제품 심층 분석 — 왜 팔리는가', s3)}
{section('4. 타프토이즈 전략 — 성장 레버 잡기', s4)}
{section('5. 구조 개선 — 빈틈을 기회로', s5)}
{section('6. 시즌 전략 — 지금이 골든타임', s6)}
{section('7. 클로징 — 오늘 결정을 이끌어낸다', s7)}

{'─'*57}
  ※ {seller} / {year}년 실판매 데이터 기반 자동 생성
  ※ 재생성 시마다 다른 각도의 스크립트가 나옵니다
{'─'*57}"""

    return jsonify({'text': script, 'ok': True, 'seller': seller})

@app.route("/api/script/report", methods=["POST"])
@login_required
def api_script_report():
    """매장 분석 리포트 생성"""
    import random, hashlib
    from datetime import datetime as dt2

    data     = request.json or {}
    seller   = data.get('seller', '')
    analysis = data.get('analysis', {})

    year        = analysis.get('year', str(dt2.now().year))
    total       = analysis.get('total', 0)
    total_pct   = analysis.get('total_pct', 0.0)
    brands      = analysis.get('brand_summary', [])
    top5        = analysis.get('top5', [])
    sold_items  = analysis.get('sold_items', [])
    unsold_taft = analysis.get('unsold_taft', [])
    weekly      = analysis.get('weekly', [])

    now    = dt2.now()
    rng    = random.Random(int(hashlib.md5(f"{seller}{now.strftime('%H%M')}".encode()).hexdigest()[:8],16))
    def pick(lst): return rng.choice(lst)
    def w(n): return f"{int(n):,}" if n else '0'

    # ── 기본 지표 ─────────────────────────────────────
    total_qty   = sum(r.get('qty',0) for r in sold_items)
    total_cnt   = sum(r.get('cnt',0) for r in sold_items)
    item_cnt    = len(sold_items)
    brand_cnt   = len(brands)
    avg_per_tx  = int(total/total_cnt) if total_cnt else 0

    # 타프토이즈
    taft_items  = [r for r in sold_items if remap_group(r.get('item_group',''),r.get('item_name',''))=='타프토이즈']
    taft_total  = sum(r.get('total',0) for r in taft_items)
    taft_cnt_k  = len(set(normalize_item_name(r.get('item_name','')) for r in taft_items))
    taft_pct    = round(taft_total/total*100,1) if total else 0

    # 주별 분석
    week_avg    = int(sum(wk.get('total',0) for wk in weekly)/len(weekly)) if weekly else 0
    week_max    = max(weekly, key=lambda x:x.get('total',0)) if weekly else {}
    week_min    = min(weekly, key=lambda x:x.get('total',0)) if weekly else {}
    week_range  = week_max.get('total',0)-week_min.get('total',0)

    trend_label = ''; trend_detail = ''; growth_rate = 0
    if len(weekly) >= 3:
        recent = [wk.get('total',0) for wk in weekly[-4:]]
        growth_rate = (recent[-1]-recent[0])/recent[0]*100 if recent[0] else 0
        if   growth_rate > 20:  trend_label='강한 상승세'; trend_detail=f"최근 {len(recent)}주 {growth_rate:.1f}% 증가"
        elif growth_rate > 8:   trend_label='상승세';       trend_detail=f"최근 {len(recent)}주 {growth_rate:.1f}% 증가"
        elif growth_rate > 2:   trend_label='완만한 상승';  trend_detail=f"최근 {len(recent)}주 {growth_rate:.1f}% 증가"
        elif growth_rate < -20: trend_label='급격한 하락';  trend_detail=f"최근 {len(recent)}주 {abs(growth_rate):.1f}% 감소"
        elif growth_rate < -8:  trend_label='하락세';       trend_detail=f"최근 {len(recent)}주 {abs(growth_rate):.1f}% 감소"
        elif growth_rate < -2:  trend_label='완만한 하락';  trend_detail=f"최근 {len(recent)}주 {abs(growth_rate):.1f}% 감소"
        else:                   trend_label='안정';         trend_detail=f"최근 {len(recent)}주 ±2% 내외 유지"

    # ── 등급 산정 (매출 기준 — DB에서 전체 매장 분포 조회) ──
    try:
        conn_g = get_db()
        all_totals = sorted([r[0] for r in conn_g.execute(
            "SELECT SUM(total) FROM sales_data WHERE real_seller!='' GROUP BY real_seller").fetchall()
        ], reverse=True)
        conn_g.close()
        n = len(all_totals)
        thresh_a = all_totals[max(0,int(n*0.10)-1)] if n >= 10 else 0
        thresh_b = all_totals[max(0,int(n*0.30)-1)] if n >= 4  else 0
        thresh_c = all_totals[max(0,int(n*0.70)-1)] if n >= 2  else 0
        if total >= thresh_a:   grade='A'; grade_basis='전체 거래처 상위 10% 이내'
        elif total >= thresh_b: grade='B'; grade_basis='전체 거래처 상위 30% 이내'
        elif total >= thresh_c: grade='C'; grade_basis='전체 거래처 상위 70% 이내'
        else:                   grade='D'; grade_basis='전체 거래처 하위 30%'
    except:
        if   total_pct >= 10: grade='A'; grade_basis='전체 비중 10% 이상'
        elif total_pct >= 5:  grade='B'; grade_basis='전체 비중 5~10%'
        elif total_pct >= 2:  grade='C'; grade_basis='전체 비중 2~5%'
        else:                  grade='D'; grade_basis='전체 비중 2% 미만'

    # 브랜드 패턴
    top_brand = brands[0]['brand'] if brands else '-'
    top_pct_v = brands[0]['pct']   if brands else 0
    top2_brand= brands[1]['brand'] if len(brands)>1 else ''
    top2_pct_v= brands[1]['pct']   if len(brands)>1 else 0
    all_brand_set = set(b['brand'] for b in brands)
    missing_brs   = [b for b in BRAND_ORDER if b not in all_brand_set and b!='타프토이즈']
    weak_brs      = [b for b in brands if b['pct']<5 and b['brand']!='타프토이즈']

    if top_pct_v >= 70:          concentration='단일 브랜드 집중형'
    elif top_pct_v >= 45:        concentration='1강 중심형'
    elif brand_cnt >= 4 and top_pct_v < 40: concentration='다브랜드 균형형'
    elif brand_cnt <= 2:         concentration='소수 브랜드형'
    else:                         concentration='2강 구도형'

    # ── 총괄 현황 — 수백 가지 경우의 수 ──────────────
    # A등급별 코멘트 풀
    grade_comments = {
        'A': [
            f"저희 전체 거래처 {len([1])}개 중 상위 10%에 해당하는 핵심 거래처입니다. {top_brand}를 중심으로 안정적인 매출 구조를 갖추고 있으며, 지속적인 관리가 중요합니다.",
            f"매출 규모와 브랜드 구성 면에서 우수한 성과를 내고 있는 거래처입니다. {top_brand} 판매 역량이 특히 두드러지며, 추가 브랜드 확대 시 더 큰 성과가 기대됩니다.",
            f"연간 {w(total)}원의 매출을 기록한 상위권 거래처로, 사장님의 적극적인 영업 활동이 실적에 반영된 결과로 판단됩니다.",
        ],
        'B': [
            f"전체 거래처 중 상위 30% 이내에 위치한 성장형 거래처입니다. 현재의 판매 흐름이 지속된다면 A등급 진입도 충분히 가능합니다.",
            f"{w(total)}원 매출로 중상위권을 유지하고 있습니다. {top_brand} 비중({top_pct_v}%)이 높아 해당 브랜드 재고 관리가 실적에 직접 영향을 미칩니다.",
            f"안정적인 판매 기반을 갖추고 있으나, 취급 브랜드 다양화를 통해 매출 성장의 여지가 있습니다.",
        ],
        'C': [
            f"중간 수준의 매출을 유지하고 있으며, 집중 관리를 통한 성장이 기대되는 거래처입니다. 방문 빈도를 높이고 제품 구성 개선이 우선 과제입니다.",
            f"현재 매출 수준에서 브랜드 구성을 보완하고 핵심 제품의 재고 관리를 강화한다면 단기간 내 성과 개선이 가능합니다.",
            f"{top_brand} 중심의 판매 구조를 갖추고 있으나, 추가 브랜드 도입과 타프토이즈 확대를 통해 객단가 향상이 필요합니다.",
        ],
        'D': [
            f"매출 성장이 필요한 거래처로, 기본적인 제품 구성 점검과 함께 사장님과의 밀착 소통이 요구됩니다. 방문 주기를 단축하고 원인 분석이 선행되어야 합니다.",
            f"현재 매출이 목표 대비 낮은 수준입니다. 취급 제품 라인업 검토, 진열 방식 개선, 사장님 교육을 통한 판매 역량 강화가 필요합니다.",
            f"거래처 유지를 위한 집중 지원이 필요한 시점입니다. 방문 시 애로사항을 파악하고 단기 실행 가능한 개선 방안을 함께 마련해야 합니다.",
        ],
    }

    # 추이별 코멘트
    trend_comments = {
        '강한 상승세': [
            f"최근 판매 추이가 가파르게 상승하고 있어 매우 고무적입니다. 이 흐름을 유지하기 위한 재고 선제 확보가 중요합니다.",
            f"주간 매출이 {growth_rate:.0f}% 이상 증가하는 강한 성장세를 보이고 있습니다. 현 상황을 적극 활용해야 합니다.",
        ],
        '상승세': [
            f"꾸준한 상승 흐름이 확인됩니다. 성장 모멘텀을 유지하면서 취약 브랜드 보완을 병행하는 것이 효과적입니다.",
            f"안정적인 성장세를 보이고 있으며, 방문 시 성장 요인을 파악하여 타 거래처에도 적용할 수 있는 사례 발굴이 필요합니다.",
        ],
        '완만한 상승': [
            f"소폭의 성장이 지속되고 있습니다. 계절 수요와 신제품 도입을 통해 성장 속도를 높일 수 있을 것으로 판단됩니다.",
        ],
        '안정': [
            f"판매가 안정적으로 유지되고 있습니다. 현 수준의 유지와 함께 새로운 성장 동력 발굴이 필요합니다.",
            f"고른 판매 흐름이 지속되고 있으나, 안정세가 장기화되면 성장 둔화로 이어질 수 있어 신제품 도입을 검토할 시점입니다.",
        ],
        '완만한 하락': [
            f"소폭의 매출 감소가 감지됩니다. 진열 위치 점검, 재고 상황 확인, 경쟁 매장 동향 파악이 필요합니다.",
            f"완만한 하락 추세로, 현 시점에서 원인을 파악하고 조기 대응하는 것이 중요합니다.",
        ],
        '하락세': [
            f"판매 하락이 지속되고 있어 즉각적인 원인 분석과 대응이 필요합니다. 방문 빈도를 높이고 사장님과 심층 면담을 권고합니다.",
        ],
        '급격한 하락': [
            f"단기간 급격한 매출 감소가 확인되어 긴급 점검이 필요합니다. 경쟁사 진입, 매장 운영 변화, 재고 문제 등 원인을 즉시 파악해야 합니다.",
        ],
        '': ["판매 추이 분석을 위한 추가 데이터 축적이 필요합니다."],
    }

    # 브랜드 집중도 코멘트
    concentration_comments = {
        '단일 브랜드 집중형': [
            f"{top_brand} 의존도({top_pct_v}%)가 매우 높아 단일 브랜드 리스크가 존재합니다. 보완 브랜드 도입을 통한 포트폴리오 다변화가 권고됩니다.",
            f"{top_brand} 한 브랜드로 전체 매출의 {top_pct_v}%를 차지하고 있습니다. 해당 브랜드의 재고 관리가 전체 실적에 직결됩니다.",
        ],
        '1강 중심형': [
            f"{top_brand}({top_pct_v}%)가 주력이며 {top2_brand}({top2_pct_v}%)가 보조 역할을 하는 구조입니다. 2위 브랜드 성장이 전체 매출 확대의 핵심입니다.",
        ],
        '다브랜드 균형형': [
            f"다양한 브랜드를 고르게 취급하고 있어 안정적인 매출 구조를 갖추고 있습니다. 각 브랜드의 시너지 효과를 극대화하는 전략이 필요합니다.",
        ],
        '소수 브랜드형': [
            f"취급 브랜드가 {brand_cnt}개로 적어 고객 선택의 폭이 제한됩니다. 1-2개 브랜드 추가를 통한 구색 확장이 시급합니다.",
        ],
        '2강 구도형': [
            f"{top_brand}({top_pct_v}%)와 {top2_brand}({top2_pct_v}%)의 2강 구도가 형성되어 있습니다. 세 번째 핵심 브랜드를 육성하면 더 탄탄한 매출 기반을 만들 수 있습니다.",
        ],
    }

    # 한줄 평 — 매장 상황별 다양한 표현
    one_liners = {
        ('A','강한 상승세'): f"핵심 거래처로서 성장 가속도가 붙어 있는 이상적인 상태로, 적극적인 지원과 재고 확보로 상승세를 극대화해야 할 시점이다.",
        ('A','상승세'):      f"상위권 거래처로서 성장이 이어지고 있으며, 현재의 판매 방식과 구성을 유지하면서 추가 브랜드 도입을 검토할 적기다.",
        ('A','안정'):        f"핵심 거래처의 안정적인 매출을 유지하고 있으나, 성장 정체를 극복하기 위한 새로운 모멘텀 발굴이 필요하다.",
        ('A','하락세'):      f"핵심 거래처임에도 하락세가 감지되어 원인 파악과 즉각적인 대응이 요구되는 상황이다.",
        ('B','강한 상승세'): f"중상위권에서 강한 성장세를 보이고 있어 A등급 진입 가능성이 매우 높다. 지금이 집중 지원의 적기다.",
        ('B','상승세'):      f"안정적인 성장을 거듭하고 있는 거래처로, 지속적인 지원과 브랜드 다양화를 통해 상위 등급으로 도약이 기대된다.",
        ('B','안정'):        f"중상위권의 안정적인 매출을 유지하고 있으며, 집중 관리를 통해 A등급 진입을 목표로 해야 한다.",
        ('B','완만한 하락'): f"중상위권이지만 소폭 하락세가 감지되어 원인 파악과 함께 회복 전략이 필요하다.",
        ('B','하락세'):      f"잠재력 있는 거래처에서 하락이 이어지고 있어 즉각적인 현장 점검과 사장님 밀착 소통이 필요하다.",
        ('C','강한 상승세'): f"중간 수준이지만 강한 성장세가 확인되어 집중 지원 시 단기간 내 B등급 이상으로 성장이 가능하다.",
        ('C','안정'):        f"중간 수준의 매출을 유지하고 있으며, 브랜드 구성 개선과 방문 빈도 강화를 통한 성장 전략이 필요하다.",
        ('C','하락세'):      f"매출 수준과 하락세가 동시에 나타나 집중 관리가 시급한 거래처다.",
        ('D','강한 상승세'): f"저매출이지만 성장 신호가 감지되고 있어, 집중 지원을 통해 빠른 회복이 기대된다.",
        ('D','안정'):        f"매출 성장이 정체된 거래처로, 근본적인 판매 환경 개선과 사장님과의 긴밀한 협력이 필요하다.",
        ('D','하락세'):      f"즉각적인 원인 파악과 집중 지원이 필요한 거래처다. 이탈 방지를 위한 적극적인 관계 관리가 요구된다.",
    }
    one_liner_key = (grade, trend_label)
    one_liner = one_liners.get(one_liner_key,
        one_liners.get((grade,'안정'),
        f"{seller}은(는) {year}년 기준 [{grade}등급] 거래처로, {concentration}이며 매출 추이는 {trend_label if trend_label else '분석 중'}이다."))

    # 코멘트 선택
    grade_comment     = pick(grade_comments.get(grade, grade_comments['C']))
    trend_comment     = pick(trend_comments.get(trend_label, trend_comments['']))
    conc_comment      = pick(concentration_comments.get(concentration, ['포트폴리오 검토가 필요합니다.']))

    # ── 브랜드별 실적 표 (정렬 고정) ─────────────────
    # 숫자 오른쪽 정렬, 고정폭 폰트 기준
    brand_table = f"  {'브랜드':<10}  {'비율':>6}  {'판매금액':>15}  {'수량':>6}  {'평가'}\n"
    brand_table += f"  {'─'*10}  {'─'*6}  {'─'*15}  {'─'*6}  {'─'*8}\n"
    for b in brands:
        bar    = '■'*int(b['pct']/5) + '□'*(20-int(b['pct']/5))
        eval_k = '◎ 핵심' if b['pct']>=35 else ('○ 주력' if b['pct']>=15 else ('△ 보조' if b['pct']>=5 else '▽ 소량'))
        brand_table += f"  {b['brand']:<10}  {b['pct']:>5.1f}%  {w(b['total']):>15}원  {b['qty']:>5}개  {eval_k}\n"
        brand_table += f"  {'':10}  {bar}\n"
    brand_table += f"  {'─'*10}  {'─'*6}  {'─'*15}  {'─'*6}\n"
    brand_table += f"  {'합 계':<10}  {'100.0':>5}%  {w(total):>15}원  {w(total_qty):>5}개\n"

    # ── TOP5 제품 ──────────────────────────────────
    top_table = ''
    top_trend_block = ''
    for i, r in enumerate(top5[:5], 1):
        nm    = normalize_item_name(r.get('item_name',''))
        br    = remap_group(r.get('item_group',''), r.get('item_name',''))
        qty   = r.get('qty',0)
        tot   = r.get('total',0)
        share = round(tot/total*100,1) if total else 0
        top_table += f"  {i}위. [{br}] {nm}\n"
        top_table += f"       판매 {w(qty)}개 / {w(tot)}원 / 비중 {share}%\n\n"

    # ── 수정3: 주요 판매 제품 최근 흐름 분석 ────────────
    try:
        conn_trend = get_db()
        top_trend_block = "\n  [ 주요 제품 최근 흐름 분석 ]\n"
        MONTHS_KR = {1:'1월',2:'2월',3:'3월',4:'4월',5:'5월',6:'6월',
                     7:'7월',8:'8월',9:'9월',10:'10월',11:'11월',12:'12월'}
        for i, r in enumerate(top5[:5], 1):
            nm = normalize_item_name(r.get('item_name',''))
            br = remap_group(r.get('item_group',''), r.get('item_name',''))
            # nm에서 브랜드 제거 후 모델명만 추출
            base_name = nm.replace(f'[{br}]','').strip().split('_')[0].strip()

            # 해당 매장 + 해당 제품 월별 흐름
            m_rows = conn_trend.execute("""
                SELECT CAST(strftime('%m',sale_date) AS INTEGER) mo,
                       SUM(total) total, SUM(quantity) qty
                FROM sales_data
                WHERE real_seller=? AND item_name LIKE ? AND sale_date LIKE ? AND sale_date!=''
                GROUP BY mo ORDER BY mo
            """, (seller, f'%{base_name}%', f'{year}%')).fetchall()

            if not m_rows:
                continue

            monthly_vals = [row[1] for row in m_rows]
            monthly_mos  = [row[0] for row in m_rows]

            # 흐름 판단 (최근 2개월 비교)
            if len(monthly_vals) >= 2:
                last   = monthly_vals[-1]
                prev   = monthly_vals[-2]
                diff   = last - prev
                pct_ch = round(diff / prev * 100, 1) if prev else 0
                if pct_ch >= 10:
                    direction = f"▲ 상승 (+{pct_ch}%)"
                    comment   = "증가 추세 — 재고 확보 및 적극 추천 필요"
                elif pct_ch <= -10:
                    direction = f"▼ 하락 ({pct_ch}%)"
                    comment   = "감소 추세 — 원인 파악 및 대체 제품 제안 검토"
                else:
                    direction = f"→ 안정 ({pct_ch:+}%)"
                    comment   = "안정적 흐름 — 유지 전략 지속"
            elif len(monthly_vals) == 1:
                direction = "— 단일 월 데이터"
                comment   = "추세 판단을 위해 추가 데이터 필요"
            else:
                continue

            # 월별 바 시각화
            max_v = max(monthly_vals) or 1
            bar_str = ''
            for mo, mv in zip(monthly_mos, monthly_vals):
                bar_h = int(mv / max_v * 5)
                bar_str += f"    {MONTHS_KR[mo]:>3}: {'█'*bar_h}{'░'*(5-bar_h)}  {w(mv)}원\n"

            top_trend_block += f"\n  {i}위 [{br}] {base_name} — {direction}\n"
            top_trend_block += f"  └ {comment}\n"
            top_trend_block += bar_str

        conn_trend.close()
    except Exception:
        top_trend_block = ''

    # ── 주별 추이 + 품목 상세 ────────────────────────
    weekly_table = ''
    conn_w = get_db()
    for i, wk in enumerate(weekly, 1):
        wk_key = wk.get('week','')
        ws_    = wk.get('week_start','')[:10]
        we_    = wk.get('week_end','')[:10]
        tot_w  = wk.get('total',0)
        max_t  = week_max.get('total',1) or 1
        bar_len= int(tot_w/max_t*15)
        bar    = '▮'*bar_len + '▯'*(15-bar_len)
        weekly_table += f"\n  {i:2}주차 ({ws_}~{we_})\n"
        weekly_table += f"  매출: {bar}  {w(tot_w)}원  ({wk.get('qty',0)}개)\n"
        # 해당 주·해당 매장의 브랜드별 판매 상세
        if wk_key:
            try:
                # 해당 매장 + 해당 주 데이터만
                wk_items = conn_w.execute("""
                    SELECT item_group, item_name, SUM(quantity) qty, SUM(total) total
                    FROM sales_data
                    WHERE strftime('%Y-%W',sale_date)=? AND sale_date!=''
                      AND (real_seller=? OR real_seller=?)
                    GROUP BY item_name ORDER BY total DESC LIMIT 20""",
                    (wk_key, seller, seller_raw)).fetchall()

                # 브랜드별 집계
                brand_wi = {}
                for wi in wk_items:
                    br_n = remap_group(wi[0], wi[1])
                    nm_n = normalize_item_name(wi[1])
                    if br_n not in brand_wi:
                        brand_wi[br_n] = {'qty': 0, 'total': 0, 'items': []}
                    brand_wi[br_n]['qty']   += wi[2]
                    brand_wi[br_n]['total'] += wi[3]
                    brand_wi[br_n]['items'].append(nm_n)

                if brand_wi:
                    weekly_table += f"  ┌ 브랜드별 현황:\n"
                    for br_n, bv in sorted(brand_wi.items(), key=lambda x:-x[1]['total']):
                        pct_w = round(bv['total']/tot_w*100,1) if tot_w else 0
                        top_item = bv['items'][0].replace(f'[{br_n}]','').strip() if bv['items'] else ''
                        weekly_table += f"  │ {br_n:<10} {w(bv['qty']):>5}개  {w(bv['total']):>12}원  ({pct_w}%)  주력:{top_item}\n"
                    weekly_table += f"  └\n"
            except: pass
    conn_w.close()

    # ── 개선 포인트 ────────────────────────────────
    improvements = []
    imp_details  = []

    if missing_brs:
        brs_str = ', '.join(missing_brs[:3])
        improvements.append(f"미취급 브랜드 도입 검토 ({brs_str})")
        imp_details.append(pick([
            f"{missing_brs[0]}는 현재 매장 고객층과 부합하는 브랜드입니다. 전시용 1개부터 시작하여 반응을 확인하는 방식을 권고합니다.",
            f"{brs_str} 도입 시 현재 취급 브랜드와의 시너지 효과가 기대되며, 객단가 향상에도 기여할 수 있습니다.",
        ]))

    if weak_brs:
        wb_str = ', '.join(b['brand'] for b in weak_brs[:2])
        wb1_name = weak_brs[0]['brand']; wb1_pct = weak_brs[0]['pct']
        improvements.append(f"저비중 브랜드 활성화 ({wb_str})")
        imp_details.append(pick([
            f"{wb_str} 진열 위치를 주력 제품 옆으로 변경하고, 함께 구매 시 효과적인 조합을 사장님과 논의하는 것을 권고합니다.",
            f"{wb_str}은 단독 판매보다 기존 인기 제품과의 세트 구성으로 접근하면 판매 활성화에 도움이 됩니다.",
        ]))

    if taft_pct < 3:
        improvements.append(f"타프토이즈 신규 도입 필요 (현재 {taft_pct}%)")
        improvements.append(f"완구 카테고리 공백으로 인한 기회 손실 발생 가능")
        imp_details.append(pick([
            "카시트·유모차 구매 고객에게 타프토이즈를 번들 제안하면 추가 구매를 유도할 수 있습니다. 초기 3~5종으로 시작을 권고합니다.",
            "타프토이즈 미취급으로 인해 고객이 완구 구매 시 타 채널로 이탈하는 상황입니다. 진입 장벽이 낮은 20,000원대 제품부터 시작을 권고합니다.",
        ]))
    elif taft_pct < 8:
        improvements.append(f"타프토이즈 비중 확대 필요 (현재 {taft_pct}% → 목표 10%)")
        imp_details.append(pick([
            f"현재 {taft_cnt_k}종을 취급하고 있습니다. 미취급 카테고리(아치, 비지북, 큐브 등)를 보완하면 타프 매출이 2배 이상 성장 가능합니다.",
            "타프토이즈 판매 고객의 재방문율이 높습니다. 시리즈 구성으로 연속 구매를 유도하는 전략이 효과적입니다.",
        ]))

    if trend_label in ('하락세','급격한 하락'):
        improvements.append("판매 감소 원인 분석 및 즉각적 대응 필요")
        imp_details.append(pick([
            "방문 시 경쟁 매장 동향, 재고 상황, 진열 변화 여부를 파악하고 단기 실행 가능한 개선책을 제시해야 합니다.",
            "판매 하락의 주요 원인(재고 부족, 진열 문제, 경쟁사 진입 등)을 현장에서 직접 확인하고 즉시 대응이 필요합니다.",
        ]))

    if avg_per_tx < 100000:
        improvements.append(f"건당 매출 향상 필요 (현재 {w(avg_per_tx)}원)")
        imp_details.append("고가 제품(레카로 카시트, 원더폴드 웨건 등)과 타프토이즈 번들 구성을 통해 건당 구매액을 높이는 전략이 필요합니다.")

    if not improvements:
        improvements.append("현재 지표 전반 양호 — 현 수준 유지 및 점진적 확대 권고")
        imp_details.append("모든 핵심 지표가 양호한 수준입니다. 현재의 운영 방식을 유지하면서 신규 시즌 제품 선주문을 통한 기회 선점을 권고합니다.")

    imp_block = ''
    for i, (item, detail) in enumerate(zip(improvements, imp_details), 1):
        imp_block += f"\n  {i}. {item}\n     → {detail}\n"

    # ── 향후 관리 방향 — 등급·추이·거리·패턴별 ─────
    # 방문 주기 (100개 매장 관리 현실 반영)
    if grade == 'A' and trend_label in ('하락세','급격한 하락'):  visit_cycle = '1~2주'
    elif grade == 'A':                                              visit_cycle = '3~4주'
    elif grade == 'B' and trend_label in ('강한 상승세','상승세'): visit_cycle = '2~3주'
    elif grade == 'B':                                              visit_cycle = '3~4주'
    elif grade == 'C' and trend_label in ('하락세','급격한 하락'): visit_cycle = '1~2주'
    elif grade == 'C':                                              visit_cycle = '4~6주'
    elif grade == 'D' and trend_label in ('강한 상승세','상승세'): visit_cycle = '2~3주'
    elif grade == 'D':                                              visit_cycle = '2~4주'
    else:                                                           visit_cycle = '4주'

    # 목표 매출 — 항상 현재 대비 +, 잠재력 기반 계산
    # 기준: 등급 + 추세 + 브랜드 다양성 + 현재 매출 수준
    base_growth = 1.10  # 기본 10%
    if grade == 'A':
        if trend_label in ('강한 상승세',): base_growth = 1.15
        else: base_growth = 1.10
    elif grade == 'B':
        if trend_label in ('강한 상승세','상승세'): base_growth = 1.20
        else: base_growth = 1.15
    elif grade == 'C':
        if trend_label in ('강한 상승세','상승세'): base_growth = 1.25
        else: base_growth = 1.20
    else:  # D
        base_growth = 1.30  # 저매출 매장은 더 큰 성장 여지

    # 잠재력: 미취급 브랜드가 많으면 추가 성장 가능
    potential_boost = min(len(missing_brs) * 0.03, 0.10)
    growth_target = base_growth + potential_boost
    # 항상 최소 +5% 보장
    growth_target = max(growth_target, 1.05)
    target_total = int(total * growth_target) if total > 0 else 0

    # 핵심 액션 아이템 (상황별 자동 생성)
    action_items = []
    if grade == 'A':
        action_items.append(pick([
            f"{top_brand} 안전 재고 유지 (2~3주치 상시 확보) — 품절은 고객 이탈로 직결",
            f"시즌별 선주문 체계 구축 — {top_brand} 수요 예측 기반 선제적 재고 관리",
        ]))
        if taft_pct < 5:
            action_items.append("타프토이즈 카테고리 도입으로 객단가 추가 향상")
    elif grade == 'B':
        action_items.append(pick([
            f"{top_brand} 판매량 20% 확대 목표 수립 — 구체적 실행 방안 현장에서 협의",
            "A등급 진입을 위한 분기별 성과 점검 체계 수립",
        ]))
    elif grade == 'C':
        action_items.append(pick([
            "방문 빈도 강화 및 진열 개선 지원 — 현장 점검을 통한 즉각적 개선",
            "핵심 제품 집중 전략 수립 — 잘 팔리는 제품 라인 강화부터 시작",
        ]))
    else:
        action_items.append(pick([
            "긴급 현장 점검 및 사장님 면담 — 운영 현황과 애로사항 파악 선행",
            "단기 성과 목표 설정 — 1개월 내 가시적 개선 지표 공동 설정",
        ]))

    if missing_brs:
        action_items.append(f"{missing_brs[0]} 도입 제안 및 초기 교육 지원")
    if taft_pct > 0 and taft_pct < 10:
        action_items.append("타프토이즈 시리즈 판매 가이드 제공 및 전시 레이아웃 개선")
    if trend_label in ('강한 상승세','상승세'):
        action_items.append("성공 사례 문서화 — 타 거래처 적용 방안 검토")

    # 관리 방향 코멘트
    mgmt_comments = {
        ('A','강한 상승세'): "현재의 상승세를 최대한 활용할 수 있도록 재고 충분 공급과 신제품 우선 배정을 지원합니다.",
        ('A','안정'):        "핵심 거래처로서의 안정적 관계 유지를 최우선으로 하며, 정기 방문을 통한 신뢰 강화에 집중합니다.",
        ('A','하락세'):      "즉각적인 현장 방문을 통해 하락 원인을 파악하고, 핵심 거래처 이탈을 방지하기 위한 적극적 지원이 필요합니다.",
        ('B','상승세'):      "성장 모멘텀을 유지하면서 A등급 진입을 위한 중점 관리 대상으로 선정하여 집중 지원합니다.",
        ('B','안정'):        "안정적 성과를 인정하고 다음 단계 성장을 위한 구체적 방안을 함께 수립합니다.",
        ('C','강한 상승세'): "성장 신호를 놓치지 않도록 방문 빈도를 높이고, 성장 가속화를 위한 집중 지원을 시작합니다.",
        ('C','안정'):        "현상 유지에서 벗어나 성장 단계로 전환하기 위한 체계적인 지원 계획을 수립합니다.",
        ('D','상승세'):      "하위 등급에서의 성장 신호는 중요한 기회입니다. 즉각적인 지원으로 성장세를 가속화합니다.",
        ('D','하락세'):      "최하위 등급의 하락세는 거래처 이탈 위험을 의미합니다. 관계 유지를 최우선으로 집중 지원합니다.",
    }
    mgmt_key = (grade, trend_label)
    mgmt_comment = mgmt_comments.get(mgmt_key,
        mgmt_comments.get((grade,'안정'),
        f"[{grade}등급] 거래처로서 {visit_cycle} 주기의 정기 방문과 지속적인 관계 관리를 권고합니다."))

    # ── 수정4: 행사/진열 등급 조회 ──────────────────
    display_grade_info = ''
    try:
        conn_d = get_db()
        disp_scores = conn_d.execute("""
            SELECT SUM(dr.score) total, COUNT(CASE WHEN dr.has_display=1 THEN 1 END) cnt
            FROM display_record dr WHERE dr.seller_name=?
        """, (seller,)).fetchone()
        if disp_scores and disp_scores[0]:
            disp_total = disp_scores[0] or 0
            disp_cnt   = disp_scores[1] or 0
            all_scores = [r[0] or 0 for r in conn_d.execute(
                "SELECT SUM(score) FROM display_record GROUP BY seller_name ORDER BY 1 DESC"
            ).fetchall()]
            disp_rank = next((i+1 for i,s in enumerate(all_scores) if s<=disp_total), len(all_scores))
            disp_display_grade = 'A' if disp_rank<=25 else 'B' if disp_rank<=50 else 'C' if disp_rank<=75 else 'D' if disp_rank<=100 else 'E'
            # 수정8: 참여한 캠페인 목록
            camp_list = conn_d.execute("""
                SELECT DISTINCT dc.campaign_name, dr.product_name, dr.score
                FROM display_record dr JOIN display_campaign dc ON dr.campaign_id=dc.id
                WHERE dr.seller_name=? AND dr.has_display=1
                ORDER BY dc.id DESC
            """, (seller,)).fetchall()
            camp_str = ''
            if camp_list:
                camp_str = '\n' + '\n'.join(f"    · {r[0]} / {r[1]} (+{r[2]}pt)" for r in camp_list[:6])
            display_grade_info = f"  진열/행사 등급 : {disp_display_grade}등급  (누적점수 {disp_total}점 / 참여 {disp_cnt}건 / 전체 {disp_rank}위){camp_str}"
        conn_d.close()
    except: pass

    # ── 수정2: 연도별 동기 비교 ──────────────────────
    yearly_comparison = ''
    try:
        conn_y = get_db()
        now_m = dt2.now()
        cur_month = now_m.month  # 현재 달까지만 비교
        cur_year  = int(year)
        prev_year = cur_year - 1

        # 현재 연도 — 1월~현재월
        cur_total = conn_y.execute("""
            SELECT SUM(total), SUM(quantity) FROM sales_data
            WHERE real_seller=? AND sale_date LIKE ? AND sale_date!=''
              AND CAST(strftime('%m',sale_date) AS INTEGER) <= ?
        """, (seller, f"{cur_year}%", cur_month)).fetchone()
        cur_t = cur_total[0] or 0; cur_q = cur_total[1] or 0

        # 전년도 — 동기 (1월~현재월)
        prev_total = conn_y.execute("""
            SELECT SUM(total), SUM(quantity) FROM sales_data
            WHERE real_seller=? AND sale_date LIKE ? AND sale_date!=''
              AND CAST(strftime('%m',sale_date) AS INTEGER) <= ?
        """, (seller, f"{prev_year}%", cur_month)).fetchone()
        prev_t = prev_total[0] or 0; prev_q = prev_total[1] or 0

        # 전년도 전체
        prev_full = conn_y.execute("""
            SELECT SUM(total), SUM(quantity) FROM sales_data
            WHERE real_seller=? AND sale_date LIKE ? AND sale_date!=''
        """, (seller, f"{prev_year}%")).fetchone()
        prev_full_t = prev_full[0] or 0

        conn_y.close()

        period_label = f"1~{cur_month}월"
        if cur_t or prev_t:
            yearly_comparison = f"\n  [ 연도별 동기 비교 — {period_label} 기준 ]\n"
            if prev_t:
                diff = cur_t - prev_t
                pct  = diff / prev_t * 100
                arrow = '▲' if diff >= 0 else '▼'
                chg_str = f"{arrow} {abs(pct):.1f}% ({'증가' if diff>=0 else '감소'})"
            else:
                chg_str = '  (전년 데이터 없음)'

            yearly_comparison += f"  {prev_year}년 {period_label}: {w(prev_t)}원  ({prev_q}개)\n"
            yearly_comparison += f"  {cur_year}년 {period_label}: {w(cur_t)}원  ({cur_q}개)  {chg_str}\n"
            if prev_full_t:
                yearly_comparison += f"\n  {prev_year}년 연간 전체: {w(prev_full_t)}원\n"
    except: pass

    # ── 수정5: 브랜드/제품별 최근 흐름 ─────────────────
    brand_trend_block = ''
    try:
        conn_bt = get_db()
        bt_rows = conn_bt.execute("""
            SELECT item_group, item_name, SUM(total) total, SUM(quantity) qty
            FROM sales_data
            WHERE real_seller=? AND sale_date LIKE ? AND sale_date!=''
            GROUP BY item_name ORDER BY total DESC
        """, (seller, f"{year}%")).fetchall()
        conn_bt.close()

        bt_brands = {}
        for r in bt_rows:
            b = remap_group(r[0], r[1])
            if not b or b == '기타': continue
            norm = normalize_item_name(r[1])
            pname = norm.replace(f'[{b}]','').strip()
            if b not in bt_brands:
                bt_brands[b] = {'total':0, 'qty':0, 'items':{}}
            bt_brands[b]['total'] += r[2] or 0
            bt_brands[b]['qty']   += r[3] or 0
            # 같은 제품명으로 합산
            if pname not in bt_brands[b]['items']:
                bt_brands[b]['items'][pname] = {'total':0, 'qty':0}
            bt_brands[b]['items'][pname]['total'] += r[2] or 0
            bt_brands[b]['items'][pname]['qty']   += r[3] or 0

        yr_total = sum(v['total'] for v in bt_brands.values()) or 1
        if bt_brands:
            brand_trend_block = "\n  [ 브랜드·제품별 판매 상세 ]\n"
            for b_name in BRAND_ORDER:
                bv = bt_brands.get(b_name)
                if not bv or not bv['total']: continue
                b_pct = round(bv['total']/yr_total*100, 1)
                brand_trend_block += f"  ┌ {b_name}: {w(bv['total'])}원 ({b_pct}%) · {bv['qty']}개\n"
                for pname, pv in sorted(bt_brands[b_name]['items'].items(), key=lambda x:-x[1]['total'])[:5]:
                    brand_trend_block += f"  │  · {pname}: {w(pv['total'])}원 / {pv['qty']}개\n"
    except Exception as e:
        brand_trend_block = f''  # 오류 시 조용히 스킵

    action_block = '\n'.join(f"  {i+1}. {a}" for i, a in enumerate(action_items)) if action_items else "  · 지속적인 관계 관리 및 정기 방문 유지"

    # ── 최종 보고서 ────────────────────────────────
    sep1 = '─'*60; sep2 = '━'*60; sep3 = '·'*60

    report = f"""{sep2}
  매장 분석 리포트
{sep2}
  거래처명   : {seller}
  분석 기간  : {year}년 / 작성 일자 : {now.strftime('%Y년 %m월 %d일')}
{sep2}

{sep1}
  1. 총괄 현황
{sep1}

  [ 실적 요약 ]
  연간 매출      : {w(total)}원
  거래처 등급    : {grade}등급  ({grade_basis})
  매출 추이      : {trend_label if trend_label else '-'}  ({trend_detail if trend_detail else '-'})
{display_grade_info}

  판매 건수      : {w(total_cnt)}건
  판매 수량      : {w(total_qty)}개
  건당 평균 매출 : {w(avg_per_tx)}원
  취급 브랜드 수 : {brand_cnt}개 / 취급 제품 종류 : {item_cnt}종
{yearly_comparison}
{brand_trend_block}
{sep3}
  [ 종합 평가 ]
  {grade_comment}

  [ 매출 추이 분석 ]
  {trend_comment}

  [ 브랜드 구성 분석 ]
  {conc_comment}

  [ 한 줄 평 ]
  {one_liner}
{sep3}

{sep1}
  2. 브랜드별 판매 실적
{sep1}

{brand_table}
{sep1}
  3. 주요 판매 제품 (TOP 5)
{sep1}

{top_table}
{top_trend_block}
{sep1}
  4. 주별 판매 추이
{sep1}

  주간 평균 : {w(week_avg)}원  /  최고 주 : {w(week_max.get('total',0))}원  /  최저 주 : {w(week_min.get('total',0))}원  /  편차 : {w(week_range)}원
{sep3}
{weekly_table if weekly_table else chr(10)+'  (데이터 없음)'+chr(10)}
{sep3}

{sep1}
  5. 개선 필요 사항
{sep1}
{imp_block}
{sep1}
  6. 향후 관리 방향
{sep1}

  {mgmt_comment}

  방문 권고 주기  : {visit_cycle}
  연간 목표 매출  : {w(target_total)}원  (현재 대비 +{int((growth_target-1)*100)}%  |  {grade}등급 잠재력 반영)

  핵심 실행 항목:
{action_block}
"""

    return jsonify({'report': report, 'ok': True})

@app.route("/api/script/report/excel", methods=["POST"])
@login_required
def api_script_report_excel():
    """매장 분석 리포트 엑셀 다운로드"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data    = request.json or {}
    report  = data.get('report', '')
    seller  = data.get('seller', '매장')

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = '매장분석리포트'

    def mf(hex_): return PatternFill("solid", fgColor=hex_)
    thin = Side(style='thin', color='DDDDDD')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    left = Alignment(horizontal='left', vertical='top', wrap_text=True)

    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 90

    ri = 1
    in_header = False
    for line in report.split('\n'):
        # 구분선 스킵
        if set(line.strip()) <= {'━','─','·','',} and len(line.strip()) > 3:
            continue
        cell = ws.cell(row=ri, column=2, value=line)
        cell.font = Font(size=10, name='맑은 고딕')
        cell.alignment = left

        stripped = line.strip()
        # 제목 줄 (숫자. 포함하거나 [ ] 포함)
        if stripped.startswith('  매장 분석 리포트'):
            cell.font = Font(bold=True, size=14, name='맑은 고딕', color='FFFFFF')
            cell.fill = mf('1E3A5F')
            ws.row_dimensions[ri].height = 30
        elif stripped.startswith(('1.','2.','3.','4.','5.','6.')):
            cell.font = Font(bold=True, size=11, name='맑은 고딕')
            cell.fill = mf('EFF6FF')
            ws.row_dimensions[ri].height = 22
        elif stripped.startswith('[') and stripped.endswith(']'):
            cell.font = Font(bold=True, size=10, name='맑은 고딕')
            cell.fill = mf('F2F4F7')
            ws.row_dimensions[ri].height = 18
        else:
            ws.row_dimensions[ri].height = 15

        ri += 1

    ws.freeze_panes = 'A1'
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    safe = seller.replace('/','_').replace(':','')
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=f'매장분석리포트_{safe}.xlsx')


@app.route("/api/export/xlsx/script")
@login_required
def export_xlsx_script():
    seller_raw=request.args.get("seller","").strip()
    year=request.args.get("year",str(datetime.now().year))
    seller=resolve_seller(seller_raw)
    conn=get_db()
    sold_items=[dict(r) for r in conn.execute("""
        SELECT item_group,item_name,SUM(quantity) qty,SUM(total) total
        FROM sales_data WHERE (real_seller=? OR real_seller=?) AND sale_date LIKE ? AND sale_date!=''
        GROUP BY item_name ORDER BY total DESC""",(seller,seller_raw,f"{year}%")).fetchall()]
    brand_summary={}
    for r in sold_items:
        b=remap_group(r['item_group'],r['item_name'])
        if b not in brand_summary: brand_summary[b]={'qty':0,'total':0}
        brand_summary[b]['qty']+=r['qty']; brand_summary[b]['total']+=r['total']
    sold_taft=set(normalize_item_name(r['item_name']) for r in sold_items
                  if remap_group(r['item_group'],r['item_name'])=='타프토이즈')
    unsold=[{'name':normalize_item_name(k),'category':v['category'],'price':v['price'],'desc':v['desc']}
            for k,v in TAFTOYS_CATALOG.items() if normalize_item_name(k) not in sold_taft]
    weekly_raw=conn.execute("""SELECT strftime('%Y-%W',sale_date) wk,MIN(sale_date) md,SUM(total) total,SUM(quantity) qty
        FROM sales_data WHERE (real_seller=? OR real_seller=?) AND sale_date LIKE ? AND sale_date!=''
        GROUP BY wk ORDER BY wk""",(seller,seller_raw,f"{year}%")).fetchall()
    conn.close()

    from datetime import datetime as dt2,timedelta
    seller_total=sum(v['total'] for v in brand_summary.values())
    wb=openpyxl.Workbook()
    mf=lambda h: PatternFill(start_color=h,end_color=h,fill_type="solid")
    mft=lambda h,b=False,s=10: Font(color=h,bold=b,size=s)
    thin=Side(style='thin',color='E0E0E0')
    bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    ctr=Alignment(horizontal="center",vertical="center"); rgt=Alignment(horizontal="right")

    # 시트1: 브랜드별 실적
    ws1=wb.active; ws1.title="브랜드별 실적"
    ws1.merge_cells("A1:E1")
    c=ws1.cell(row=1,column=1,value=f"{seller} — {year}년 브랜드별 판매 실적")
    c.fill=mf("1E3A5F"); c.font=mft("FFFFFF",True,13); c.alignment=ctr; ws1.row_dimensions[1].height=28
    for ci,h in enumerate(['브랜드','판매수량','판매금액(원)','비율(%)','등급'],1):
        c=ws1.cell(row=2,column=ci,value=h); c.fill=mf("F2F2F2"); c.font=mft("595959",True,10); c.alignment=ctr; c.border=bdr
    for ri,(brand,v) in enumerate(sorted(brand_summary.items(),key=lambda x:-x[1]['total']),3):
        pct=round(v['total']/seller_total*100,1) if seller_total else 0
        grade="★★★ 핵심" if pct>25 else "★★ 주력" if pct>10 else "★ 보조" if pct>3 else "△ 소량"
        for ci,val in enumerate([brand,v['qty'],v['total'],pct,grade],1):
            c=ws1.cell(row=ri,column=ci,value=val); c.border=bdr
            if ri%2==0: c.fill=mf("FAFAFA")
            if ci==3: c.number_format='#,##0'; c.alignment=rgt
            if ci in (4,5): c.alignment=ctr
    for ci,w in zip('ABCDE',[14,10,16,10,14]): ws1.column_dimensions[ci].width=w

    # 시트2: 제품별 상세
    ws2=wb.create_sheet("제품별 상세")
    ws2.merge_cells("A1:F1")
    c=ws2.cell(row=1,column=1,value=f"{seller} — 제품별 판매 상세")
    c.fill=mf("1E3A5F"); c.font=mft("FFFFFF",True,12); c.alignment=ctr; ws2.row_dimensions[1].height=26
    for ci,h in enumerate(['브랜드','제품명','판매수량','판매금액(원)','비율(%)','등급'],1):
        c=ws2.cell(row=2,column=ci,value=h); c.fill=mf("F2F2F2"); c.font=mft("595959",True,10); c.alignment=ctr; c.border=bdr
    for ri,r in enumerate(sold_items,3):
        brand=remap_group(r['item_group'],r['item_name']); norm=normalize_item_name(r['item_name'])
        pct=round(r['total']/seller_total*100,1) if seller_total else 0
        grade="◎ 인기" if pct>10 else "○ 판매중" if pct>3 else "△ 소량"
        for ci,val in enumerate([brand,norm,r['qty'],r['total'],pct,grade],1):
            c=ws2.cell(row=ri,column=ci,value=val); c.border=bdr
            if ri%2==0: c.fill=mf("FAFAFA")
            if ci==4: c.number_format='#,##0'; c.alignment=rgt
            if ci in (5,6): c.alignment=ctr
    ws2.column_dimensions['A'].width=14; ws2.column_dimensions['B'].width=32
    ws2.column_dimensions['C'].width=10; ws2.column_dimensions['D'].width=16
    ws2.column_dimensions['E'].width=10; ws2.column_dimensions['F'].width=12

    # 시트3: 타프토이즈 추천
    ws3=wb.create_sheet("타프토이즈 추천")
    ws3.merge_cells("A1:E1")
    c=ws3.cell(row=1,column=1,value=f"타프토이즈 미취급 제품 추천 — {seller}")
    c.fill=mf("7C3AED"); c.font=mft("FFFFFF",True,12); c.alignment=ctr; ws3.row_dimensions[1].height=26
    SP={'아치/모빌':'인스타 감성↑, 구매 결정 빠름','트래블토이':'유모차/카시트 필수, 재구매율 높음',
        '비지북':'교육적 가치, 선물용 인기','큐브':'다기능 가성비, 1+1 구성 가능',
        '워터매트':'계절성 높음, 여름 전 선주문','액티비티짐':'고마진, 출산선물 1순위',
        '터미타임':'소아과 추천, 안전 강조','감각 장난감':'6개월부터 사용, 반복구매',
        '인형/뮤지컬':'수면 루틴, 감성 구매','카시트 장난감':'카시트 구매 시 ADD-ON'}
    for ci,h in enumerate(['제품명','카테고리','소비자가(원)','제품 특징','영업 포인트'],1):
        c=ws3.cell(row=2,column=ci,value=h); c.fill=mf("F2F2F2"); c.font=mft("595959",True,10); c.alignment=ctr; c.border=bdr
    for ri,u in enumerate(unsold,3):
        for ci,val in enumerate([u['name'],u.get('category',''),u.get('price',0),u.get('desc',''),SP.get(u.get('category',''),'')],1):
            c=ws3.cell(row=ri,column=ci,value=val); c.border=bdr
            if ri%2==0: c.fill=mf("FAF5FF")
            if ci==3: c.number_format='#,##0'; c.alignment=rgt
    ws3.column_dimensions['A'].width=30; ws3.column_dimensions['B'].width=16
    ws3.column_dimensions['C'].width=14; ws3.column_dimensions['D'].width=42
    ws3.column_dimensions['E'].width=32

    # 시트4: 주별 추이
    ws4=wb.create_sheet("주별 추이")
    ws4.merge_cells("A1:D1")
    c=ws4.cell(row=1,column=1,value=f"{seller} — 주별 판매 추이")
    c.fill=mf("1E3A5F"); c.font=mft("FFFFFF",True,12); c.alignment=ctr; ws4.row_dimensions[1].height=26
    for ci,h in enumerate(['주차','기간','판매금액(원)','판매수량'],1):
        c=ws4.cell(row=2,column=ci,value=h); c.fill=mf("F2F2F2"); c.font=mft("595959",True,10); c.alignment=ctr; c.border=bdr
    for ri,r in enumerate(weekly_raw,3):
        try:
            d=dt2.strptime(r[1],"%Y-%m-%d"); sun=d-timedelta(days=(d.weekday()+1)%7)
            sat=sun+timedelta(days=6); period=f"{sun.strftime('%m/%d')}~{sat.strftime('%m/%d')}"
        except: period=r[0]
        for ci,val in enumerate([f"{ri-2}주차",period,r[2],r[3]],1):
            c=ws4.cell(row=ri,column=ci,value=val); c.border=bdr
            if ri%2==0: c.fill=mf("FAFAFA")
            if ci==3: c.number_format='#,##0'; c.alignment=rgt
            if ci in(1,2): c.alignment=ctr
    for ci,w in zip('ABCD',[10,16,16,10]): ws4.column_dimensions[ci].width=w

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,download_name=f"영업스크립트_{seller}_{year}.xlsx")

def make_xlsx(headers, rows_data, sheet_name="데이터"):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = sheet_name
    hdr_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style='thin', color='E5E7EB')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, hdr in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=hdr)
        c.fill=hdr_fill; c.font=hdr_font
        c.alignment=Alignment(horizontal="center",vertical="center"); c.border=bdr
    ws.row_dimensions[1].height = 24
    even_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    for ri, row in enumerate(rows_data, 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = bdr
            if ri % 2 == 0: c.fill = even_fill
            if isinstance(val,(int,float)) and ci > 1:
                c.alignment = Alignment(horizontal="right")
                if '금액' in headers[ci-1]: c.number_format = '#,##0'
    for col in ws.columns:
        ml = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml+4, 40)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

@app.route("/api/export/xlsx/monthly")
@login_required
def export_xlsx_monthly():
    """브랜드별 정리 엑셀 형식 — 매장×월별×브랜드별 금액/수량"""
    year   = request.args.get("year",   str(datetime.now().year))
    month  = request.args.get("month",  "")
    seller = request.args.get("seller", "").strip()
    conn   = get_db()

    # 브랜드 고정 순서 사용
    brands = BRAND_ORDER  # ['줄즈','레카로','ABC디자인','원더폴드','카오스','엔픽스','타프토이즈']

    # 월 목록
    if month:
        months = [int(month)]
    else:
        months_raw = conn.execute(
            f"SELECT DISTINCT CAST(strftime('%m',sale_date) AS INTEGER) m "
            f"FROM sales_data WHERE sale_date LIKE '{year}%' AND sale_date!='' "
            f"ORDER BY m").fetchall()
        months = [r[0] for r in months_raw] or list(range(1,13))

    # 매장 목록
    seller_cond   = "AND real_seller=?" if seller else ""
    seller_params = [seller] if seller else []
    sellers_raw = conn.execute(
        f"SELECT DISTINCT real_seller FROM sales_data "
        f"WHERE real_seller!='' AND sale_date LIKE '{year}%' {seller_cond} "
        f"ORDER BY real_seller", seller_params).fetchall()
    sellers_list = [r[0] for r in sellers_raw]

    def brand_key(nm):
        nm_lower=(nm or '').replace('_',' ').lower()
        if '베이비하우스' in nm_lower: return (0,nm_lower)
        if '링크맘' in nm_lower:       return (1,nm_lower)
        if '베네피아' in nm_lower:     return (1,nm_lower)   # 링크맘과 같은 그룹
        if '베이비플러스' in nm_lower:  return (1,nm_lower)
        if '베이비파크' in nm_lower:   return (2,nm_lower)
        if '베이비스토리' in nm_lower:  return (3,nm_lower)
        if '베이비스토어' in nm_lower:  return (3,nm_lower)
        if '베이비세븐' in nm_lower:   return (4,nm_lower)
        return (9,nm_lower)

    # 매장별 채널(오프라인/백화점) — 백화점(서양네트웍스/가이아코퍼레이션)은 기타 다음 맨 아래 배치
    seller_channel_m = {}
    for r in conn.execute("""
        SELECT real_seller, (SELECT channel FROM sales_data sd2 WHERE sd2.real_seller=sd1.real_seller
               GROUP BY channel ORDER BY COUNT(*) DESC LIMIT 1) ch
        FROM sales_data sd1 WHERE real_seller!='' GROUP BY real_seller""").fetchall():
        seller_channel_m[r[0]] = r[1] or '오프라인'

    def sort_key_m(s):
        if seller_channel_m.get(s, '') == '백화점':
            return (99, (s or '').lower())
        return brand_key(s)

    sellers_list.sort(key=sort_key_m)

    # 데이터 조회 — 매장×월×품목그룹 (item_group, item_name 모두 가져와 remap)
    data_rows = conn.execute(
        f"""SELECT real_seller, CAST(strftime('%m',sale_date) AS INTEGER) mo,
            item_group, item_name, SUM(total) total, SUM(quantity) qty
            FROM sales_data
            WHERE real_seller!='' AND sale_date LIKE '{year}%' AND sale_date!=''
            {seller_cond}
            GROUP BY real_seller, mo, item_group, item_name""",
        seller_params).fetchall()

    # 인덱스: {(seller, month, brand): {total, qty}} — remap_group 적용
    idx = {}
    for r in data_rows:
        brand = remap_group(r[2], r[3])  # item_group, item_name
        if not brand or brand == '기타': continue
        key = (r[0], r[1], brand)
        if key not in idx:
            idx[key] = {'total': 0, 'qty': 0}
        idx[key]['total'] += r[4] or 0
        idx[key]['qty']   += r[5] or 0

    # ── openpyxl 빌드 ──
    wb = openpyxl.Workbook()

    # ── 스타일 팔레트 ──
    WHITE       = "FFFFFF"
    GRAY_LIGHT  = "F2F2F2"   # 행 2~3 (업체구분, 헤더)
    GRAY_HDR    = "E8E8E8"   # 데이터 헤더 행
    BORDER_CLR  = "BFBFBF"   # 매장정보 열 테두리
    FONT_BLACK  = "000000"
    FONT_GRAY   = "595959"

    def mf(h):  return PatternFill(start_color=h,end_color=h,fill_type="solid")
    def mft(h,bold=False,sz=10): return Font(color=h,bold=bold,size=sz)
    thin_bdr = Side(style='thin', color=BORDER_CLR)
    no_bdr   = Side(style=None)
    bdr_left  = Border(left=thin_bdr,right=thin_bdr,top=thin_bdr,bottom=thin_bdr)  # 매장정보 열
    bdr_none  = Border(left=no_bdr,right=no_bdr,top=no_bdr,bottom=no_bdr)          # 브랜드 데이터 열
    center   = Alignment(horizontal="center",vertical="center")
    right    = Alignment(horizontal="right",  vertical="center")
    left     = Alignment(horizontal="left",   vertical="center")
    num_fmt  = '#,##0'

    col_start = 5  # A=여백, B=업체구분, C=거래처명, D=실적용, E부터 데이터

    def month_title_label():
        """단일 월이면 'N월', 여러 달이면 'N월~M월' 형식 라벨"""
        if len(months) == 1:
            return f"{months[0]}월"
        return f"{months[0]}월~{months[-1]}월"

    # 수정2+4: 브랜드별 제품 목록 수집 (제품별관리 양식 참고 — 브랜드 아래 제품별 서브컬럼)
    # 수정1: 제품별관리와 동일한 커스텀 순서/라벨 적용, 수정2: 타프토이즈는 세부 제품 나열 없이 단일 컬럼
    import re as _re_bp
    brand_products = {}  # {brand: [product_label, ...]}
    prod_rows_all = conn.execute(f"""
        SELECT item_group, item_name FROM sales_data
        WHERE sale_date LIKE '{year}%' AND real_seller!='' GROUP BY item_group, item_name""").fetchall()
    for grp, name in prod_rows_all:
        b = remap_group(grp, name)
        if not b or b == '기타': continue
        if b == '타프토이즈':
            # 수정2: 타프토이즈는 단일 카테고리로 취급 (세부 제품 나열 없음)
            brand_products.setdefault(b, ['전체'])
            continue
        custom_label = get_custom_product_label(b, name)
        if custom_label:
            label = custom_label
        else:
            norm = normalize_item_name(name)
            label = _re_bp.sub(r'^\[[^\]]+\]', '', norm).strip()
        if b not in brand_products: brand_products[b] = []
        if label not in brand_products[b]: brand_products[b].append(label)
    for b in brand_products:
        if b == '타프토이즈': continue
        brand_products[b] = sort_product_labels(b, brand_products[b])

    def match_brand_product(brand, item_name):
        if brand == '타프토이즈':
            return '전체'
        custom_label = get_custom_product_label(brand, item_name)
        if custom_label:
            return custom_label if custom_label in brand_products.get(brand, []) else None
        for label in brand_products.get(brand, []):
            if label in item_name:
                return label
        return None

    def build_sheet(wb_ref, title, field):
        """금액 또는 수량 시트 생성 — 브랜드 아래 제품별 서브컬럼 + 업체구분 소계 + 하단 총합계"""
        ws = wb_ref.create_sheet(title) if title != "브랜드별 금액" else wb_ref.active
        if title == "브랜드별 금액": ws.title = title

        HDR_BG2 = mf("F2F2F2")
        dotted = Side(style='hair', color='808080')

        # 브랜드별 블록 너비 (제품수 + 소계1)
        brand_widths = {b: len(brand_products.get(b, [])) + 1 for b in brands if brand_products.get(b)}
        active_brands = [b for b in brands if b in brand_widths]
        month_width = sum(brand_widths.values()) + 1  # +1 = 월 총합계
        total_cols = col_start - 1 + len(months) * month_width

        ws.column_dimensions['A'].width = 2.5

        field_label = '판매금액' if field=='total' else '판매수량'
        ws.merge_cells(f"B2:{get_column_letter(min(9,total_cols))}2")
        c = ws.cell(row=2,column=2,value=f"※ 거래처별 브랜드·제품별 {field_label}_{month_title_label()}")
        c.fill=mf(WHITE); c.font=mft(FONT_BLACK,True,12); c.alignment=Alignment(horizontal='left',vertical='center')
        ws.row_dimensions[2].height=22

        # 행3: 업체구분/거래처명/실적용거래처명 (세로 3행 병합: 3~5행)
        for ci, h in zip(range(2,5), ["업체구분","거래처명","실적용거래처명"]):
            c=ws.cell(row=3,column=ci,value=h)
            c.fill=HDR_BG2; c.font=mft(FONT_BLACK,True,10); c.alignment=center
            ws.merge_cells(start_row=3, start_column=ci, end_row=5, end_column=ci)

        # 행3: 월 헤더
        col=col_start
        for mo in months:
            end_col = col + month_width - 1
            ws.merge_cells(f"{get_column_letter(col)}3:{get_column_letter(end_col)}3")
            c=ws.cell(row=3,column=col,value=f"{year}_{mo:02d}")
            c.fill=HDR_BG2; c.font=mft(FONT_BLACK,True,11); c.alignment=center
            col += month_width
        ws.row_dimensions[3].height=18

        # 행4: 브랜드 헤더, 행5: 제품명+소계
        col=col_start
        for mo in months:
            for b in active_brands:
                bw = brand_widths[b]
                end_col = col + bw - 1
                ws.merge_cells(f"{get_column_letter(col)}4:{get_column_letter(end_col)}4")
                c=ws.cell(row=4,column=col,value=b)
                c.fill=HDR_BG2; c.font=mft(FONT_BLACK,True,9); c.alignment=center
                for pi, plabel in enumerate(brand_products[b]):
                    c2=ws.cell(row=5,column=col+pi,value=plabel)
                    c2.fill=HDR_BG2; c2.font=mft(FONT_BLACK,False,8); c2.alignment=center
                c3=ws.cell(row=5,column=col+bw-1,value="소계")
                c3.fill=HDR_BG2; c3.font=mft(FONT_BLACK,True,8); c3.alignment=center
                col += bw
            c=ws.cell(row=4,column=col,value="월합계")
            ws.merge_cells(f"{get_column_letter(col)}4:{get_column_letter(col)}5")
            c.fill=HDR_BG2; c.font=mft(FONT_BLACK,True,9); c.alignment=center
            col += 1
        ws.row_dimensions[4].height=16; ws.row_dimensions[5].height=16

        # 컬럼 너비
        ws.column_dimensions['B'].width=12
        ws.column_dimensions['C'].width=22
        ws.column_dimensions['D'].width=24
        num_w = 10 if field=='total' else 7
        for ci in range(col_start, total_cols+1):
            ws.column_dimensions[get_column_letter(ci)].width = num_w

        # 업체구분 판별
        def detect_group(seller_name):
            if seller_channel_m.get(seller_name, '') == '백화점':
                return '백화점'
            nm = (seller_name or '').lower().replace(' ','').replace('_','')
            if '베이비하우스' in nm: return '베이비하우스'
            if '링크맘' in nm or '베네피아' in nm or '베이비플러스' in nm: return '링크맘'
            if '베이비파크' in nm: return '베이비파크'
            if '베이비스토리' in nm or '베이비스토어' in nm: return '베이비스토리'
            if '베이비세븐' in nm: return '베이비세븐'
            return '기타'

        # 매장별 (월,브랜드,제품)별 값 인덱스 (item_name 기준 매칭)
        prod_idx = {}  # {(seller, month, brand, product_label): val}
        for r in data_rows:
            b = remap_group(r[2], r[3])
            if not b or b == '기타': continue
            pl = match_brand_product(b, r[3])
            if not pl: continue
            key = (r[0], r[1], b, pl)
            prod_idx[key] = prod_idx.get(key, 0) + (r[4] if field=='total' else r[5])

        # 데이터 행 — 업체구분별로 묶어서 순회, 그룹 바뀔 때 소계 삽입 (수정3)
        ri = 6
        prev_grp = None
        group_start_row = None
        group_totals = {}  # 그룹별 (mo,col_offset)->합

        def write_group_subtotal(end_row_before, grp_name, grp_seller_list):
            nonlocal ri
            ws.cell(row=ri, column=2, value=f"{grp_name} 소계")
            ws.merge_cells(start_row=ri, start_column=2, end_row=ri, end_column=4)
            c = ws.cell(row=ri, column=2)
            c.fill = HDR_BG2; c.font = mft(FONT_BLACK, True, 9); c.alignment = left_local
            col2 = col_start
            for mo in months:
                mo_grp_total = 0
                for b in active_brands:
                    bw = brand_widths[b]
                    b_grp_total = 0
                    for pi, plabel in enumerate(brand_products[b]):
                        pv = sum(prod_idx.get((s, mo, b, plabel), 0) for s in grp_seller_list)
                        c2 = ws.cell(row=ri, column=col2+pi, value=pv if pv else 0)
                        c2.fill = HDR_BG2; c2.font = mft(FONT_BLACK, False, 8); c2.alignment = right
                        if field=='total': c2.number_format = num_fmt
                        b_grp_total += pv
                    cs = ws.cell(row=ri, column=col2+bw-1, value=b_grp_total if b_grp_total else 0)
                    cs.fill = HDR_BG2; cs.font = mft(FONT_BLACK, True, 8); cs.alignment = right
                    if field=='total': cs.number_format = num_fmt
                    mo_grp_total += b_grp_total
                    col2 += bw
                cm = ws.cell(row=ri, column=col2, value=mo_grp_total if mo_grp_total else 0)
                cm.fill = HDR_BG2; cm.font = mft(FONT_BLACK, True, 9); cm.alignment = right
                if field=='total': cm.number_format = num_fmt
                col2 += 1
            ws.row_dimensions[ri].height = 15
            ri += 1

        left_local = Alignment(horizontal='left', vertical='center')
        current_group_sellers = []
        for s in sellers_list:
            grp = detect_group(s)
            if prev_grp is not None and grp != prev_grp and current_group_sellers:
                write_group_subtotal(ri, prev_grp, current_group_sellers)
                current_group_sellers = []
            gv = grp if grp != prev_grp else ''
            prev_grp = grp
            current_group_sellers.append(s)

            for ci, val in enumerate([gv, s, s], 2):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = mft(FONT_BLACK if ci>2 else FONT_GRAY, False, 9); c.alignment = left_local

            col2 = col_start
            for mo in months:
                mo_total = 0
                for b in active_brands:
                    bw = brand_widths[b]
                    b_total = 0
                    for pi, plabel in enumerate(brand_products[b]):
                        val = prod_idx.get((s, mo, b, plabel), 0)
                        c = ws.cell(row=ri, column=col2+pi, value=val if val else 0)
                        c.font = mft(FONT_BLACK, False, 8); c.alignment = right
                        if field=='total': c.number_format = num_fmt
                        b_total += val
                    cs = ws.cell(row=ri, column=col2+bw-1, value=b_total if b_total else 0)
                    cs.font = mft(FONT_BLACK, True, 8); cs.alignment = right
                    if field=='total': cs.number_format = num_fmt
                    mo_total += b_total
                    col2 += bw
                cm = ws.cell(row=ri, column=col2, value=mo_total if mo_total else 0)
                cm.font = mft(FONT_BLACK, True, 9); cm.alignment = right
                if field=='total': cm.number_format = num_fmt
                col2 += 1
            ws.row_dimensions[ri].height = 14
            ri += 1

        if current_group_sellers:
            write_group_subtotal(ri, prev_grp, current_group_sellers)

        # 수정5(동일 적용): 하단 총합계 행
        tot_row = ri
        ws.cell(row=tot_row, column=2, value="총합계")
        ws.merge_cells(start_row=tot_row, start_column=2, end_row=tot_row, end_column=4)
        tc = ws.cell(row=tot_row, column=2)
        tc.fill = HDR_BG2; tc.font = mft(FONT_BLACK, True, 10); tc.alignment = center
        col2 = col_start
        for mo in months:
            mo_grand = 0
            for b in active_brands:
                bw = brand_widths[b]
                b_grand = 0
                for pi, plabel in enumerate(brand_products[b]):
                    pv = sum(prod_idx.get((s, mo, b, plabel), 0) for s in sellers_list)
                    c2 = ws.cell(row=tot_row, column=col2+pi, value=pv if pv else 0)
                    c2.fill = HDR_BG2; c2.font = mft(FONT_BLACK, True, 8); c2.alignment = right
                    if field=='total': c2.number_format = num_fmt
                    b_grand += pv
                cs = ws.cell(row=tot_row, column=col2+bw-1, value=b_grand if b_grand else 0)
                cs.fill = HDR_BG2; cs.font = mft(FONT_BLACK, True, 8); cs.alignment = right
                if field=='total': cs.number_format = num_fmt
                mo_grand += b_grand
                col2 += bw
            cm = ws.cell(row=tot_row, column=col2, value=mo_grand if mo_grand else 0)
            cm.fill = HDR_BG2; cm.font = mft(FONT_BLACK, True, 9); cm.alignment = right
            if field=='total': cm.number_format = num_fmt
            col2 += 1
        ws.row_dimensions[tot_row].height = 18

        # 수정6: 안쪽 점선(hair) 테두리 전체 데이터 영역에 적용
        last_col = col_start + len(months)*month_width - 1
        for r_ in range(3, tot_row+1):
            for c_ in range(2, last_col+1):
                cell = ws.cell(row=r_, column=c_)
                left_s  = dotted if c_ > 2 else None
                right_s = dotted if c_ < last_col else None
                top_s   = dotted if r_ > 3 else None
                bottom_s= dotted if r_ < tot_row else None
                cell.border = Border(left=left_s, right=right_s, top=top_s, bottom=bottom_s)

        ws.freeze_panes = get_column_letter(col_start) + '6'
        return ws

    build_sheet(wb, "브랜드별 금액", "total")
    build_sheet(wb, "브랜드별 수량", "qty")

    # ── 시트3: 제품별 상세 ──
    ws3=wb.create_sheet("제품별 상세")
    params2=[f"{year}%"]; conds2=["sale_date LIKE ?","sale_date!=''"]
    if seller: conds2.append("real_seller=?"); params2.append(seller)
    if month:  conds2.append(f"strftime('%m',sale_date)='{month.zfill(2)}'")
    raw_items=[dict(r) for r in conn.execute(f"""
        SELECT item_group,item_name,SUM(quantity) qty,SUM(total) total,COUNT(*) cnt
        FROM sales_data WHERE {' AND '.join(conds2)}
        GROUP BY item_name ORDER BY item_group,total DESC""",params2).fetchall()]

    # 제품별 상세(월별) 데이터 미리 추출 (conn 닫기 전)
    prod_monthly_cache = {}
    for mo in months:
        for b in BRAND_ORDER:
            if b == '타프토이즈': continue
            if seller:
                prows = conn.execute("""
                    SELECT item_name, SUM(quantity) qty, SUM(total) total
                    FROM sales_data
                    WHERE sale_date LIKE ? AND item_group NOT IN ('','NULL')
                      AND real_seller=?
                    GROUP BY item_name
                """, (f"{year}-{str(mo).zfill(2)}%", seller)).fetchall()
            else:
                prows = conn.execute("""
                    SELECT item_name, SUM(quantity) qty, SUM(total) total
                    FROM sales_data
                    WHERE sale_date LIKE ? AND item_group NOT IN ('','NULL') AND real_seller!=''
                    GROUP BY item_name
                """, (f"{year}-{str(mo).zfill(2)}%",)).fetchall()
            prod_monthly_cache[(mo, b)] = prows

    merged={}
    for r in raw_items:
        nn=normalize_item_name(r['item_name']); ng=remap_group(r['item_group'],r['item_name'])
        if not ng or ng=='기타': continue
        key=(ng,nn)
        if key not in merged: merged[key]={'item_group':ng,'item_name':nn,'qty':0,'total':0,'cnt':0}
        merged[key]['qty']+=r['qty']; merged[key]['total']+=r['total']; merged[key]['cnt']+=r['cnt']
    sorted_items=sorted(merged.values(),key=lambda x:(get_group_sort_key(x['item_group']),-x['total']))

    thin3=Side(style='thin',color='E0E0E0')
    bdr3=Border(left=thin3,right=thin3,top=thin3,bottom=thin3)
    ws3.column_dimensions['A'].width = 2.5
    for ci,h in enumerate(['품목그룹','제품명','판매건수','판매수량','합계금액(원)'],2):
        c=ws3.cell(row=2,column=ci,value=h)
        c.fill=mf(GRAY_LIGHT); c.font=mft(FONT_GRAY,True,10); c.alignment=center; c.border=bdr3
    ws3.row_dimensions[2].height=20
    for ri,r in enumerate(sorted_items,3):
        for ci,val in enumerate([r['item_group'],r['item_name'],r['cnt'],r['qty'],r['total']],2):
            c=ws3.cell(row=ri,column=ci,value=val); c.border=bdr3
            if ri%2==1: c.fill=mf("FAFAFA")
            if ci>3: c.alignment=right
            if ci==6 and isinstance(val,int): c.number_format=num_fmt
    for ci in range(2,7):
        col_letter = get_column_letter(ci)
        ml=max((len(str(ws3.cell(r,ci).value or '')) for r in range(2,ws3.max_row+1)),default=8)
        ws3.column_dimensions[col_letter].width=min(ml+3,35)

    # ── 시트4: 월별 브랜드 요약 (세로형 — 출력 최적화) ──
    ws4 = wb.create_sheet("월별 브랜드 요약")
    ws4.column_dimensions['A'].width = 2.5
    ws4.merge_cells("B2:I2")
    c=ws4.cell(row=2,column=2,value=f"※ 거래처별 브랜드 판매수량 및 금액_{month_title_label()}")
    c.fill=mf(WHITE); c.font=mft(FONT_BLACK,True,12); c.alignment=Alignment(horizontal='left',vertical='center')
    ws4.row_dimensions[2].height=22
    m_hdrs=['월','브랜드','판매금액(원)','판매수량','비율(%)','월합계(원)','누계(원)','전달 대비']
    for ci,h in enumerate(m_hdrs,2):
        c=ws4.cell(row=3,column=ci,value=h)
        c.fill=mf(GRAY_LIGHT); c.font=mft(FONT_GRAY,True,10); c.alignment=center; c.border=bdr_left
    ws4.row_dimensions[3].height=20
    ri4=4; cum_m=0
    # 수정5: 제품별 전달 대비 증감률 계산을 위해 이전 달 제품별 매출을 미리 캐시
    prev_month_prod_totals = {}  # {(brand, product_norm): total}
    for mo in months:
        mo_total = sum(idx.get((s,mo,b),{}).get('total',0) for s in sellers_list for b in brands)
        mo_qty   = sum(idx.get((s,mo,b),{}).get('qty',0)   for s in sellers_list for b in brands)
        cum_m += mo_total
        ws4.cell(row=ri4,column=2,value=f"{mo}월").fill=mf(GRAY_LIGHT)
        ws4.cell(row=ri4,column=3,value="전체 합계").fill=mf(GRAY_LIGHT)
        for ci,val in [(4,mo_total),(5,mo_qty),(6,100.0),(7,mo_total),(8,cum_m)]:
            c4=ws4.cell(row=ri4,column=ci,value=val)
            c4.fill=mf(GRAY_LIGHT); c4.font=mft(FONT_GRAY,True,10)
            c4.border=bdr_none; c4.alignment=right
            if ci in (4,7,8): c4.number_format=num_fmt
            if ci==6: c4.number_format='0.0'
        ws4.cell(row=ri4,column=9,value='').fill=mf(GRAY_LIGHT)
        for ci in range(2,4):
            ws4.cell(row=ri4,column=ci).font=mft(FONT_GRAY,True,10)
            ws4.cell(row=ri4,column=ci).border=bdr_left
        ws4.row_dimensions[ri4].height=18; ri4+=1

        # 이번 달 제품별 매출 (다음 달 비교용으로 갱신 예정)
        this_month_prod_totals = {}

        for b in brands:
            bv=sum(idx.get((s,mo,b),{}).get('total',0) for s in sellers_list)
            bq=sum(idx.get((s,mo,b),{}).get('qty',0)   for s in sellers_list)
            if bv==0: continue
            pct_b=round(bv/mo_total*100,1) if mo_total else 0
            ws4.cell(row=ri4,column=2,value=""); ws4.cell(row=ri4,column=3,value=f"  └ {b}")
            ws4.cell(row=ri4,column=3).font=mft(FONT_BLACK,True,9)
            for ci,val in [(4,bv),(5,bq),(6,pct_b),(7,""),(8,"")]:
                c4=ws4.cell(row=ri4,column=ci,value=val); c4.border=bdr_none; c4.alignment=right
                if ci==4 and isinstance(val,int): c4.number_format=num_fmt
                if ci==6: c4.number_format='0.0'
            ws4.row_dimensions[ri4].height=16; ri4+=1

            # 타프토이즈 제외 브랜드: 제품별 상세 추가
            if b != '타프토이즈':
                # 캐시에서 해당 브랜드·월 데이터 사용
                prows = prod_monthly_cache.get((mo, b), [])
                prod_brand = {}
                for pr in prows:
                    pb = remap_group('', pr[0]) if not pr[0].startswith('[') else remap_group('X', pr[0])
                    if pb != b: continue
                    norm = normalize_item_name(pr[0])
                    if norm not in prod_brand:
                        prod_brand[norm] = {'qty':0,'total':0}
                    prod_brand[norm]['qty']   += pr[1] or 0
                    prod_brand[norm]['total'] += pr[2] or 0
                for pnorm, pvals in sorted(prod_brand.items(), key=lambda x:-x[1]['total']):
                    if pvals['total']==0: continue
                    pname = pnorm.replace('['+b+']','').strip() if '['+b+']' in pnorm else pnorm
                    prod_key = (b, pnorm)
                    this_month_prod_totals[prod_key] = pvals['total']

                    ws4.cell(row=ri4,column=3,value=f"      · {pname}")
                    ws4.cell(row=ri4,column=3).font=mft(FONT_GRAY,False,8)
                    c4=ws4.cell(row=ri4,column=4,value=pvals['total'])
                    c4.font=mft(FONT_GRAY,False,8); c4.border=bdr_none; c4.alignment=right; c4.number_format=num_fmt
                    c4q=ws4.cell(row=ri4,column=5,value=pvals['qty'])
                    c4q.font=mft(FONT_GRAY,False,8); c4q.border=bdr_none; c4q.alignment=right

                    # 수정5: 전달 대비 증감률
                    prev_total = prev_month_prod_totals.get(prod_key)
                    mom_cell = ws4.cell(row=ri4, column=9)
                    if prev_total is None:
                        mom_cell.value = '—'
                        mom_cell.font = mft('999999', False, 8)
                    elif prev_total == 0:
                        mom_cell.value = '신규'
                        mom_cell.font = mft('2563EB', True, 8)
                    else:
                        mom_pct = round((pvals['total'] - prev_total) / prev_total * 100, 1)
                        mom_cell.value = f"{'▲' if mom_pct>=0 else '▼'} {abs(mom_pct)}%"
                        mom_cell.font = mft('16A34A' if mom_pct>=0 else 'DC2626', True, 8)
                    mom_cell.border = bdr_none
                    mom_cell.alignment = right

                    ws4.row_dimensions[ri4].height=13; ri4+=1

        prev_month_prod_totals = this_month_prod_totals
        ws4.row_dimensions[ri4].height=6; ri4+=1
    for ci,ww in enumerate([8,18,16,10,10,16,16,12],2):
        ws4.column_dimensions[get_column_letter(ci)].width=ww

    # 수정1: 기초 데이터 — 업로드하신 원본 엑셀 파일을 그대로 시트로 재현 (요약치 검증용)
    raw_file_q = "SELECT filename, file_b64, months FROM sales_upload_file WHERE year=?"
    raw_file_params = [int(year)]
    if month:
        raw_file_q += " AND months LIKE ?"
        raw_file_params.append(f"%{year}-{month.zfill(2)}%")
    raw_files = conn.execute(raw_file_q, raw_file_params).fetchall()

    raw_sheet_names = []
    if raw_files:
        import base64 as _b64_dl
        for idx, (fname_orig, fb64, months_str) in enumerate(raw_files):
            try:
                src_bytes = _b64_dl.b64decode(fb64)
                src_wb = openpyxl.load_workbook(io.BytesIO(src_bytes))
                for src_sheet_name in src_wb.sheetnames:
                    src_ws = src_wb[src_sheet_name]
                    tab_name = f"기초데이터_{months_str.split(',')[0][5:]}월"[:31] if len(raw_files)==1 and len(src_wb.sheetnames)==1 \
                        else f"기초_{months_str.split(',')[0][5:]}월_{src_sheet_name}"[:31]
                    # 중복 시트명 방지
                    base_tab = tab_name; suf = 1
                    while tab_name in raw_sheet_names:
                        tab_name = f"{base_tab[:28]}_{suf}"; suf += 1
                    ws_raw = wb.create_sheet(tab_name)
                    _copy_sheet_with_style(src_ws, ws_raw)
                    raw_sheet_names.append(tab_name)
            except Exception:
                continue

    if not raw_sheet_names:
        # 원본 파일이 저장되기 전 데이터(과거 업로드분) 대비 — 요약 데이터로 폴백
        ws_raw = wb.create_sheet("기초데이터")
        ws_raw.column_dimensions['A'].width = 2.5
        c0 = ws_raw.cell(row=2, column=2, value="⚠ 원본 파일이 저장되지 않은 기간입니다. 해당 월 데이터를 다시 업로드하면 원본 그대로 표시됩니다.")
        c0.font = mft(FONT_BLACK, True, 10)
        raw_hdrs = ['일자','거래처명','실적용거래처명','거래처코드','품목명','수량','단가','공급가액','부가세','합계','채널']
        for ci, h in enumerate(raw_hdrs, 2):
            c = ws_raw.cell(row=4, column=ci, value=h)
            c.font=mft(FONT_BLACK,True,9); c.fill=mf("F2F2F2"); c.alignment=center
        ws_raw.row_dimensions[4].height=20
        raw_q = "SELECT sale_date,seller_name,real_seller,trade_code,item_name,quantity,unit_price,supply_price,vat,total,channel FROM sales_data WHERE sale_date LIKE ? AND real_seller!=''"
        raw_params=[f"{year}%"]
        if seller: raw_q += " AND real_seller=?"; raw_params.append(seller)
        raw_q += " ORDER BY sale_date, real_seller"
        ri_raw=5
        for r in conn.execute(raw_q, raw_params).fetchall():
            for ci,v in enumerate(r,2):
                c=ws_raw.cell(row=ri_raw,column=ci,value=v); c.font=mft(FONT_BLACK,False,8)
                c.alignment = right if ci in (7,8,9,10) else (center if ci in (2,6) else left)
                if ci in (8,9,10): c.number_format=num_fmt
            ri_raw+=1
        for ci,w in zip(range(2,13),[11,20,20,14,26,8,10,11,10,11,9]):
            ws_raw.column_dimensions[get_column_letter(ci)].width=w
        ws_raw.freeze_panes='B5'
        raw_sheet_names = ["기초데이터"]

    conn.close()

    # 수정1: 시트 순서를 참조 양식과 동일하게 재배열 (월별 브랜드 요약이 맨 앞, 기초데이터 맨 뒤)
    wb._sheets = [wb["월별 브랜드 요약"], wb["브랜드별 금액"], wb["브랜드별 수량"], wb["제품별 상세"]] + [wb[n] for n in raw_sheet_names]

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    fname=f"오프라인_브랜드별정리_{year}{'_'+month+'월' if month else ''}.xlsx"
    return send_file(buf,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,download_name=fname)
@app.route("/api/export/xlsx/weekly")
@login_required
def export_xlsx_weekly():
    from datetime import datetime as dt2, timedelta
    year   = request.args.get("year",   str(datetime.now().year))
    month  = request.args.get("month",  "")
    seller = request.args.get("seller", "").strip()
    date_from = request.args.get("date_from", "").strip()  # 수정1: 날짜 범위 지정 다운로드
    date_to   = request.args.get("date_to", "").strip()
    conn   = get_db()

    qp = ["sale_date != ''"]
    pp = []
    if date_from and date_to:
        qp.append("sale_date >= ? AND sale_date <= ?"); pp.append(date_from); pp.append(date_to)
    elif month:
        qp.append("sale_date LIKE ?"); pp.append(f"{year}-{month.zfill(2)}%")
    else:
        qp.append("sale_date LIKE ?"); pp.append(f"{year}%")
    if seller: qp.append("real_seller = ?"); pp.append(seller)

    # 주차 목록
    week_rows = [dict(r) for r in conn.execute(f"""
        SELECT strftime('%Y-%W',sale_date) wk, MIN(sale_date) md,
               COUNT(*) cnt, SUM(quantity) qty, SUM(total) total
        FROM sales_data WHERE {' AND '.join(qp)} AND sale_date!=''
        GROUP BY wk ORDER BY wk""", pp).fetchall()]

    def wr(ds):
        d = dt2.strptime(ds, "%Y-%m-%d")
        sun = d - timedelta(days=(d.weekday()+1) % 7)
        return sun.strftime("%Y-%m-%d"), (sun+timedelta(days=6)).strftime("%Y-%m-%d")

    for r in week_rows:
        try: r['ws'], r['we'] = wr(r['md'])
        except: r['ws'] = r['we'] = ''

    def month_week_label(ws_str):
        """해당 주(일요일 시작)의 '월 몇째주'와 'M/D~M/D' 형식 라벨 계산
        (예: 2026-08-09 시작 주 → '8월 2주차', 날짜는 '8/9 ~ 8/15')"""
        try:
            sun = dt2.strptime(ws_str, "%Y-%m-%d")
        except Exception:
            return '', ''
        mo = sun.month
        # 해당 월 1일이 속한 주의 일요일부터 세어, 몇 번째 일요일인지 계산
        first_of_month = sun.replace(day=1)
        first_sunday = first_of_month - timedelta(days=(first_of_month.weekday()+1) % 7)
        week_no = (sun - first_sunday).days // 7 + 1
        end = sun + timedelta(days=6)
        date_range = f"{sun.month}/{sun.day} ~ {end.month}/{end.day}"
        return f"{mo}월 {week_no}주차", date_range

    # 단일 주(또는 지정 범위가 1주 이내)인 경우, 참조 양식과 동일한 타이틀 라벨 사용
    single_week_label = ''
    if len(week_rows) >= 1:
        mw, dr = month_week_label(week_rows[0]['ws'])
        if mw:
            single_week_label = f"{mw} ({dr})"

    weeks = week_rows
    brands = BRAND_ORDER

    # 주차 × 브랜드 × 매장 인덱스 조회
    raw = conn.execute(f"""
        SELECT strftime('%Y-%W',sale_date) wk, item_group, item_name,
               SUM(total) total, SUM(quantity) qty, real_seller
        FROM sales_data WHERE {' AND '.join(qp)} AND sale_date!=''
        GROUP BY wk, item_name, real_seller""", pp).fetchall()

    # 매장 목록 (업체구분 순) — 수정2: 전체 조건(qp/pp)을 재사용해야 date_from/date_to 모드에서도 정상 조회됨
    sellers_raw = conn.execute(
        f"SELECT DISTINCT real_seller FROM sales_data WHERE {' AND '.join(qp)} AND real_seller!=''",
        pp).fetchall()
    sellers_list = [r[0] for r in sellers_raw]

    # 매장별 채널(오프라인/백화점) 조회 — 서양네트웍스/가이아코퍼레이션은 맨 아래 배치
    seller_channel = {}
    for r in conn.execute("""
        SELECT real_seller, (SELECT channel FROM sales_data sd2 WHERE sd2.real_seller=sd1.real_seller
               GROUP BY channel ORDER BY COUNT(*) DESC LIMIT 1) ch
        FROM sales_data sd1 WHERE real_seller!='' GROUP BY real_seller""").fetchall():
        seller_channel[r[0]] = r[1] or '오프라인'

    def bk(nm):
        nm=(nm or '').lower()
        if '베이비하우스' in nm: return (0,nm)
        if '링크맘' in nm or '베네피아' in nm or '베이비플러스' in nm: return (1,nm)
        if '베이비파크' in nm: return (2,nm)
        if '베이비스토리' in nm or '베이비스토어' in nm: return (3,nm)
        if '베이비세븐' in nm: return (4,nm)
        return (9,nm)

    def sort_key(s):
        # 백화점 채널은 항상 맨 아래(기타 다음)로 배치
        if seller_channel.get(s, '') == '백화점':
            return (99, (s or '').lower())
        return bk(s)

    sellers_list.sort(key=sort_key)

    # 업체구분 파악 — 매장명 기반 정확한 브랜드/채널 감지 (branches.note는 자유 메모라 사용하지 않음)
    def detect_group_for_export(seller_name):
        if seller_channel.get(seller_name, '') == '백화점':
            return '백화점'
        nm = (seller_name or '').lower().replace(' ','').replace('_','')
        if '베이비하우스' in nm: return '베이비하우스'
        if '링크맘' in nm or '베네피아' in nm or '베이비플러스' in nm: return '링크맘'
        if '베이비파크' in nm: return '베이비파크'
        if '베이비스토리' in nm or '베이비스토어' in nm: return '베이비스토리'
        if '베이비세븐' in nm: return '베이비세븐'
        return '기타'

    branch_group = {s: detect_group_for_export(s) for s in sellers_list}

    # 수정2+4: 브랜드별 제품 목록 수집 (제품별관리 양식 참고 — 브랜드 아래 제품별 서브컬럼)
    # 수정1: 제품별관리와 동일한 커스텀 순서/라벨 적용, 수정2: 타프토이즈는 세부 제품 나열 없이 단일 컬럼
    import re as _re_wp
    brand_products = {}  # {brand: [product_label, ...]}
    seen_products = set()
    for r in raw:
        b = remap_group(r[1], r[2])
        if not b or b == '기타': continue
        if b == '타프토이즈':
            brand_products.setdefault(b, ['전체'])
            continue
        custom_label = get_custom_product_label(b, r[2])
        if custom_label:
            label = custom_label
        else:
            norm = normalize_item_name(r[2])
            label = _re_wp.sub(r'^\[[^\]]+\]', '', norm).strip()
        key = (b, label)
        if key in seen_products: continue
        seen_products.add(key)
        brand_products.setdefault(b, []).append(label)
    for b in brand_products:
        if b == '타프토이즈': continue
        brand_products[b] = sort_product_labels(b, brand_products[b])

    def match_brand_product_wk(brand, item_name):
        if brand == '타프토이즈':
            return '전체'
        custom_label = get_custom_product_label(brand, item_name)
        if custom_label:
            return custom_label if custom_label in brand_products.get(brand, []) else None
        for label in brand_products.get(brand, []):
            if label in item_name:
                return label
        return None

    # {(wk, brand, seller, product_label): val}
    prod_idx_wk = {}
    for r in raw:
        b = remap_group(r[1], r[2])
        if not b or b == '기타': continue
        pl = match_brand_product_wk(b, r[2])
        if not pl: continue
        s_nm = r[5] if len(r) > 5 else ''
        key = (r[0], b, s_nm, pl)
        if key not in prod_idx_wk: prod_idx_wk[key] = {'total':0,'qty':0}
        prod_idx_wk[key]['total'] += r[3] or 0
        prod_idx_wk[key]['qty']   += r[4] or 0

    # ── idx: {(wk, brand, seller): {total, qty}} — 매장별 브랜드별 주차별 집계 ──
    idx_seller = {}  # (wk, brand, seller) → {total, qty}
    for r in raw:
        brand = remap_group(r[1], r[2])
        if not brand or brand == '기타': continue
        # r = (wk, item_group, item_name, total, qty, real_seller)
        s_nm = r[5] if len(r) > 5 else ''
        key3 = (r[0], brand, s_nm)
        if key3 not in idx_seller: idx_seller[key3] = {'total':0,'qty':0}
        idx_seller[key3]['total'] += r[3] or 0
        idx_seller[key3]['qty']   += r[4] or 0

    # {(wk, brand): {total, qty}}
    idx = {}
    for r in raw:
        brand = remap_group(r[1], r[2])
        if not brand or brand == '기타': continue
        key = (r[0], brand)
        if key not in idx: idx[key] = {'total':0,'qty':0}
        idx[key]['total'] += r[3] or 0
        idx[key]['qty']   += r[4] or 0

    # 주차별 제품 상세 (색상 통합)
    items_by_week = {}
    for r in raw:
        wk = r[0]
        brand = remap_group(r[1], r[2])
        norm  = normalize_item_name(r[2])
        k = (brand, norm)
        if wk not in items_by_week: items_by_week[wk] = {}
        if k not in items_by_week[wk]: items_by_week[wk][k] = {'item_group':brand,'item_name':norm,'qty':0,'total':0}
        items_by_week[wk][k]['qty']   += r[4] or 0
        items_by_week[wk][k]['total'] += r[3] or 0

    # ── 스타일 팔레트 (월별과 동일) ──
    WHITE      = "FFFFFF"
    GRAY_LIGHT = "F2F2F2"
    FONT_BLACK = "000000"
    FONT_GRAY  = "595959"
    thin_bdr   = Side(style='thin', color='BFBFBF')
    no_bdr     = Side(style=None)
    bdr_left   = Border(left=thin_bdr,right=thin_bdr,top=thin_bdr,bottom=thin_bdr)
    bdr_none   = Border(left=no_bdr,right=no_bdr,top=no_bdr,bottom=no_bdr)
    center     = Alignment(horizontal="center",vertical="center")
    right      = Alignment(horizontal="right",vertical="center")
    left       = Alignment(horizontal="left",vertical="center")
    num_fmt    = '#,##0'
    mf  = lambda h: PatternFill(start_color=h,end_color=h,fill_type="solid")
    mft = lambda h,b=False,s=10: Font(color=h,bold=b,size=s)

    wb  = openpyxl.Workbook()
    col_start = 5  # A=여백, B=업체구분, C=거래처명, D=실적용, E부터 데이터

    def build_brand_sheet(wb_ref, title, field, is_first=False):
        ws = wb_ref.active if is_first else wb_ref.create_sheet(title)
        if is_first: ws.title = title

        HDR_BG2 = mf("F2F2F2")
        dotted = Side(style='hair', color='808080')

        brand_widths = {b: len(brand_products.get(b, [])) + 1 for b in brands if brand_products.get(b)}
        active_brands = [b for b in brands if b in brand_widths]
        week_width = sum(brand_widths.values()) + 1  # +1 = 주 총합계
        total_cols = col_start - 1 + len(weeks) * week_width

        ws.column_dimensions['A'].width = 2.5

        field_label = '판매금액' if field=='total' else '판매수량'
        ws.merge_cells(f"B2:{get_column_letter(min(9,total_cols))}2")
        title_text = f"※ 거래처별 브랜드·제품별 {field_label}_{single_week_label}" if single_week_label \
            else f"오프라인 주별 브랜드·제품별 {field_label} 정리_{year}"
        c = ws.cell(row=2,column=2,value=title_text)
        c.fill=mf(WHITE); c.font=mft(FONT_BLACK,True,12); c.alignment=left
        ws.row_dimensions[2].height=22

        # 행3: 업체구분/거래처명/실적용거래처명 (3~5행 세로 병합)
        for ci,h in zip(range(2,5), ["업체구분","거래처명","실적용거래처명"]):
            c = ws.cell(row=3,column=ci,value=h)
            c.fill=HDR_BG2; c.font=mft(FONT_BLACK,True,10); c.alignment=center
            ws.merge_cells(start_row=3, start_column=ci, end_row=5, end_column=ci)

        # 행3: 주차 헤더
        col = col_start
        for i,r in enumerate(weeks):
            end_col = col + week_width - 1
            mw, dr = month_week_label(r['ws'])
            ws.merge_cells(f"{get_column_letter(col)}3:{get_column_letter(end_col)}3")
            label = f"{mw} ({dr})" if mw else f"{i+1}주차 ({r['ws']}~{r['we']})"
            c = ws.cell(row=3,column=col,value=label)
            c.fill=HDR_BG2; c.font=mft(FONT_BLACK,True,10); c.alignment=center
            col += week_width
        ws.row_dimensions[3].height=18

        # 행4: 브랜드, 행5: 제품명+소계
        col = col_start
        for i,r in enumerate(weeks):
            for b in active_brands:
                bw = brand_widths[b]
                end_col = col + bw - 1
                ws.merge_cells(f"{get_column_letter(col)}4:{get_column_letter(end_col)}4")
                c = ws.cell(row=4,column=col,value=b)
                c.fill=HDR_BG2; c.font=mft(FONT_BLACK,True,9); c.alignment=center
                for pi, plabel in enumerate(brand_products[b]):
                    c2 = ws.cell(row=5,column=col+pi,value=plabel)
                    c2.fill=HDR_BG2; c2.font=mft(FONT_BLACK,False,8); c2.alignment=center
                c3 = ws.cell(row=5,column=col+bw-1,value="소계")
                c3.fill=HDR_BG2; c3.font=mft(FONT_BLACK,True,8); c3.alignment=center
                col += bw
            c = ws.cell(row=4,column=col,value="주합계")
            ws.merge_cells(f"{get_column_letter(col)}4:{get_column_letter(col)}5")
            c.fill=HDR_BG2; c.font=mft(FONT_BLACK,True,9); c.alignment=center
            col += 1
        ws.row_dimensions[4].height=16; ws.row_dimensions[5].height=16

        # 컬럼 너비
        ws.column_dimensions['B'].width=12; ws.column_dimensions['C'].width=22; ws.column_dimensions['D'].width=24
        num_w = 10 if field=='total' else 7
        for ci in range(col_start, total_cols+1):
            ws.column_dimensions[get_column_letter(ci)].width = num_w

        def write_group_subtotal_wk(grp_name, grp_seller_list):
            nonlocal ri
            ws.cell(row=ri, column=2, value=f"{grp_name} 소계")
            ws.merge_cells(start_row=ri, start_column=2, end_row=ri, end_column=4)
            c = ws.cell(row=ri, column=2)
            c.fill = HDR_BG2; c.font = mft(FONT_BLACK, True, 9); c.alignment = left
            col2 = col_start
            for r in weeks:
                wk = r['wk']
                wk_grp_total = 0
                for b in active_brands:
                    bw = brand_widths[b]
                    b_grp_total = 0
                    for pi, plabel in enumerate(brand_products[b]):
                        pv = sum(prod_idx_wk.get((wk, b, s, plabel), {}).get(field, 0) for s in grp_seller_list)
                        c2 = ws.cell(row=ri, column=col2+pi, value=pv if pv else 0)
                        c2.fill = HDR_BG2; c2.font = mft(FONT_BLACK, False, 8); c2.alignment = right
                        if field=='total': c2.number_format = num_fmt
                        b_grp_total += pv
                    cs = ws.cell(row=ri, column=col2+bw-1, value=b_grp_total if b_grp_total else 0)
                    cs.fill = HDR_BG2; cs.font = mft(FONT_BLACK, True, 8); cs.alignment = right
                    if field=='total': cs.number_format = num_fmt
                    wk_grp_total += b_grp_total
                    col2 += bw
                cm = ws.cell(row=ri, column=col2, value=wk_grp_total if wk_grp_total else 0)
                cm.fill = HDR_BG2; cm.font = mft(FONT_BLACK, True, 9); cm.alignment = right
                if field=='total': cm.number_format = num_fmt
                col2 += 1
            ws.row_dimensions[ri].height = 15
            ri += 1

        # 데이터 행 (6행~) — 업체구분별 묶어서 소계 삽입 (수정3)
        ri = 6
        prev_grp = None
        current_group_sellers = []
        for s in sellers_list:
            grp = branch_group.get(s, '')
            if prev_grp is not None and grp != prev_grp and current_group_sellers:
                write_group_subtotal_wk(prev_grp, current_group_sellers)
                current_group_sellers = []
            gv = grp if grp != prev_grp else ''
            prev_grp = grp
            current_group_sellers.append(s)

            for ci,val in enumerate([gv,s,s],2):
                c=ws.cell(row=ri,column=ci,value=val)
                c.font=mft(FONT_BLACK if ci>2 else FONT_GRAY,False,9); c.alignment=left

            col2 = col_start
            for r in weeks:
                wk = r['wk']
                wk_total = 0
                for b in active_brands:
                    bw = brand_widths[b]
                    b_total = 0
                    for pi, plabel in enumerate(brand_products[b]):
                        val = prod_idx_wk.get((wk, b, s, plabel), {}).get(field, 0)
                        c = ws.cell(row=ri, column=col2+pi, value=val if val else 0)
                        c.font = mft(FONT_BLACK, False, 8); c.alignment = right
                        if field=='total': c.number_format = num_fmt
                        b_total += val
                    cs = ws.cell(row=ri, column=col2+bw-1, value=b_total if b_total else 0)
                    cs.font = mft(FONT_BLACK, True, 8); cs.alignment = right
                    if field=='total': cs.number_format = num_fmt
                    wk_total += b_total
                    col2 += bw
                cm = ws.cell(row=ri, column=col2, value=wk_total if wk_total else 0)
                cm.font = mft(FONT_BLACK, True, 9); cm.alignment = right
                if field=='total': cm.number_format = num_fmt
                col2 += 1
            ws.row_dimensions[ri].height = 14
            ri += 1

        if current_group_sellers:
            write_group_subtotal_wk(prev_grp, current_group_sellers)

        # 수정5: 하단 총합계 행
        tot_row = ri
        ws.cell(row=tot_row, column=2, value="총합계")
        ws.merge_cells(start_row=tot_row, start_column=2, end_row=tot_row, end_column=4)
        tc = ws.cell(row=tot_row, column=2)
        tc.fill = HDR_BG2; tc.font = mft(FONT_BLACK, True, 10); tc.alignment = center
        col2 = col_start
        for r in weeks:
            wk = r['wk']
            wk_grand = 0
            for b in active_brands:
                bw = brand_widths[b]
                b_grand = 0
                for pi, plabel in enumerate(brand_products[b]):
                    pv = sum(prod_idx_wk.get((wk, b, s, plabel), {}).get(field, 0) for s in sellers_list)
                    c2 = ws.cell(row=tot_row, column=col2+pi, value=pv if pv else 0)
                    c2.fill = HDR_BG2; c2.font = mft(FONT_BLACK, True, 8); c2.alignment = right
                    if field=='total': c2.number_format = num_fmt
                    b_grand += pv
                cs = ws.cell(row=tot_row, column=col2+bw-1, value=b_grand if b_grand else 0)
                cs.fill = HDR_BG2; cs.font = mft(FONT_BLACK, True, 8); cs.alignment = right
                if field=='total': cs.number_format = num_fmt
                wk_grand += b_grand
                col2 += bw
            cm = ws.cell(row=tot_row, column=col2, value=wk_grand if wk_grand else 0)
            cm.fill = HDR_BG2; cm.font = mft(FONT_BLACK, True, 9); cm.alignment = right
            if field=='total': cm.number_format = num_fmt
            col2 += 1
        ws.row_dimensions[tot_row].height = 18

        # 수정6: 안쪽 점선(hair) 테두리
        last_col = col_start + len(weeks)*week_width - 1
        for r_ in range(3, tot_row+1):
            for c_ in range(2, last_col+1):
                cell = ws.cell(row=r_, column=c_)
                left_s  = dotted if c_ > 2 else None
                right_s = dotted if c_ < last_col else None
                top_s   = dotted if r_ > 3 else None
                bottom_s= dotted if r_ < tot_row else None
                cell.border = Border(left=left_s, right=right_s, top=top_s, bottom=bottom_s)

        ws.freeze_panes = get_column_letter(col_start) + '6'

    # ── 시트1: 주별 요약 (세로형 — 출력 최적화) ──
    ws_sum = wb.active; ws_sum.title="주별 요약"
    ws_sum.column_dimensions['A'].width = 2.5

    # 타이틀
    ws_sum.merge_cells("B2:H2")
    title_text_sum = f"※ 거래처별 브랜드 판매실적_{single_week_label}" if single_week_label \
        else f"주별 판매 실적 요약_{year}"
    c=ws_sum.cell(row=2,column=2,value=title_text_sum)
    c.fill=mf(WHITE); c.font=mft(FONT_BLACK,True,12); c.alignment=left
    ws_sum.row_dimensions[2].height=22

    # 헤더행
    sum_hdrs=['주차','기간','브랜드','판매금액(원)','판매수량','비율(%)','누계금액(원)']
    for ci,h in enumerate(sum_hdrs,2):
        c=ws_sum.cell(row=3,column=ci,value=h)
        c.fill=mf(GRAY_LIGHT); c.font=mft(FONT_GRAY,True,10); c.alignment=center; c.border=bdr_left
    ws_sum.row_dimensions[3].height=20

    ri=4
    cumulative=0
    for i,r in enumerate(weeks):
        wk=r['wk']
        wk_total=r.get('total',0); wk_qty=r.get('qty',0); wk_cnt=r.get('cnt',0)
        cumulative+=wk_total
        mw, dr = month_week_label(r['ws'])
        wk_label = mw or f"{i+1}주차"
        period_label = dr and f"{r['ws']}~{r['we']}" or f"{r['ws']}~{r['we']}"

        # 주차 소계 행
        ws_sum.cell(row=ri,column=2,value=wk_label).fill=mf(GRAY_LIGHT)
        ws_sum.cell(row=ri,column=3,value=f"{r['ws']}~{r['we']}").fill=mf(GRAY_LIGHT)
        ws_sum.cell(row=ri,column=4,value="전체 합계").fill=mf(GRAY_LIGHT)
        for ci,val in [(5,wk_total),(6,wk_qty),(7,100.0),(8,cumulative)]:
            c=ws_sum.cell(row=ri,column=ci,value=val)
            c.fill=mf(GRAY_LIGHT); c.font=mft(FONT_GRAY,True,10)
            c.border=bdr_none; c.alignment=right
            if ci in (5,8): c.number_format=num_fmt
            if ci==7: c.number_format='0.0'
        for ci in range(2,4):
            ws_sum.cell(row=ri,column=ci).font=mft(FONT_GRAY,True,10)
            ws_sum.cell(row=ri,column=ci).border=bdr_left
        ws_sum.row_dimensions[ri].height=18
        ri+=1

        # 브랜드별 세부 행 (해당 주차)
        for b in brands:
            bv = sum(idx_seller.get((wk,b,s),{}).get('total',0) for s in sellers_list)
            bq = sum(idx_seller.get((wk,b,s),{}).get('qty',0) for s in sellers_list)
            if bv == 0: continue
            pct = round(bv/wk_total*100,1) if wk_total else 0
            ws_sum.cell(row=ri,column=2,value="")
            ws_sum.cell(row=ri,column=3,value="")
            ws_sum.cell(row=ri,column=4,value=f"  └ {b}")
            for ci,val in [(5,bv),(6,bq),(7,pct),(8,"")]:
                c=ws_sum.cell(row=ri,column=ci,value=val)
                c.border=bdr_none; c.alignment=right
                if ci==5 and isinstance(val,int): c.number_format=num_fmt
                if ci==7: c.number_format='0.0'
            ws_sum.row_dimensions[ri].height=16
            ri+=1

        # 구분 공백행
        ws_sum.row_dimensions[ri].height=6; ri+=1

    # 열 너비
    for ci,w in enumerate([10,28,18,16,10,10,16],2):
        ws_sum.column_dimensions[get_column_letter(ci)].width=w

    # ── 시트2: 주별 브랜드 요약 (브랜드 아래 제품 표시) — 순서 조정: 요약 다음 바로 배치 ──
    ws5 = wb.create_sheet("주별 브랜드 요약")
    ws5.column_dimensions['A'].width = 2.5

    # ── 시트2: 브랜드별 금액 ──
    build_brand_sheet(wb, "브랜드별 금액", "total", False)
    # ── 시트3: 브랜드별 수량 ──
    build_brand_sheet(wb, "브랜드별 수량", "qty", False)

    # ── 시트2(먼저 생성해둔 ws5)에 주별 브랜드 요약 내용 채우기 ──
    ws5.merge_cells("B2:G2")
    title_text_5 = f"※ 거래처별 브랜드 판매실적_{single_week_label}" if single_week_label \
        else f"주별 브랜드 요약_{year}"
    c=ws5.cell(row=2,column=2,value=title_text_5)
    c.fill=mf(WHITE); c.font=mft(FONT_BLACK,True,12); c.alignment=left
    ws5.row_dimensions[2].height=22

    # 전체 주별×브랜드×제품 집계
    brand_week_prod = {}   # {wk: {brand: {prod_norm: {qty, total}}}}
    brand_week_total = {}  # {wk: {brand: {qty, total}}}
    for wk_info in weeks:
        wk = wk_info['wk']
        brand_week_prod[wk]  = {}
        brand_week_total[wk] = {}
        for k, item in items_by_week.get(wk, {}).items():
            b = item['item_group']
            norm = normalize_item_name(item['item_name'])
            if b not in brand_week_prod[wk]:
                brand_week_prod[wk][b]  = {}
                brand_week_total[wk][b] = {'qty':0,'total':0}
            if norm not in brand_week_prod[wk][b]:
                brand_week_prod[wk][b][norm] = {'qty':0,'total':0}
            brand_week_prod[wk][b][norm]['qty']   += item['qty']
            brand_week_prod[wk][b][norm]['total'] += item['total']
            brand_week_total[wk][b]['qty']   += item['qty']
            brand_week_total[wk][b]['total'] += item['total']

    # 헤더
    ws5h = ['주차','기간','브랜드/제품','판매금액(원)','판매수량','비율(%)']
    for ci,h in enumerate(ws5h,2):
        c=ws5.cell(row=3,column=ci,value=h)
        c.fill=mf(GRAY_LIGHT); c.font=mft(FONT_GRAY,True,10)
        c.alignment=center; c.border=bdr_none
    ws5.row_dimensions[3].height=20

    ri5=4
    for i, wk_info in enumerate(weeks):
        wk = wk_info['wk']
        wk_total_val = sum(v['total'] for v in brand_week_total.get(wk,{}).values()) or 1
        mw, dr = month_week_label(wk_info['ws'])
        wk_label = mw or f"{i+1}주차"
        period_label = f"{wk_info['ws']}~{wk_info['we']}"

        # 전체 합계행
        wk_sum = sum(v['total'] for v in brand_week_total.get(wk,{}).values())
        wk_qty = sum(v['qty'] for v in brand_week_total.get(wk,{}).values())
        for ci,v in enumerate([wk_label, period_label, '전체 합계', wk_sum, wk_qty, 100.0],2):
            c=ws5.cell(row=ri5,column=ci,value=v)
            c.font=mft(FONT_BLACK,True,10); c.alignment=right if ci>=5 else (center if ci<=3 else left)
            c.fill=mf(GRAY_LIGHT)
            if ci==5: c.number_format=num_fmt
            if ci==7: c.number_format='0.0'
        ws5.row_dimensions[ri5].height=18; ri5+=1

        for b in brands:
            bt = brand_week_total.get(wk,{}).get(b)
            if not bt or bt['total']==0: continue
            pct_b = round(bt['total']/wk_total_val*100,1)

            # 브랜드행
            for ci,v in enumerate(['','', f'  └ {b}', bt['total'], bt['qty'], pct_b],2):
                c=ws5.cell(row=ri5,column=ci,value=v)
                c.font=mft(FONT_BLACK,True,9); c.alignment=right if ci>=5 else left
                if ci==5: c.number_format=num_fmt
                if ci==7: c.number_format='0.0'
            ws5.row_dimensions[ri5].height=16; ri5+=1

            # 타프토이즈 제외 브랜드만 제품 상세
            if b != '타프토이즈':
                prods = brand_week_prod.get(wk,{}).get(b,{})
                for pnorm, pv in sorted(prods.items(), key=lambda x:-x[1]['total']):
                    pname = pnorm.replace(f'[{b}]','').strip()
                    for ci,v in enumerate(['','', f'      · {pname}', pv['total'], pv['qty'],''],2):
                        c=ws5.cell(row=ri5,column=ci,value=v)
                        c.font=mft(FONT_GRAY,False,8); c.alignment=right if ci in (5,6) else left
                        if ci==5: c.number_format=num_fmt
                    ws5.row_dimensions[ri5].height=13; ri5+=1

        ws5.row_dimensions[ri5].height=6; ri5+=1  # 주간 구분

    for ci,w in enumerate([10,26,24,18,12,10],2):
        ws5.column_dimensions[get_column_letter(ci)].width=w

    # ── 시트5: 제품별 상세 (마지막 배치) ──
    ws4=wb.create_sheet("제품별 상세")
    ws4.column_dimensions['A'].width = 2.5
    item_hdrs=['주차','기간','브랜드','제품명','판매수량','판매금액(원)']
    for ci,h in enumerate(item_hdrs,2):
        c=ws4.cell(row=1,column=ci,value=h)
        c.fill=mf(GRAY_LIGHT); c.font=mft(FONT_GRAY,True,10); c.alignment=center; c.border=bdr_left if ci<=3 else bdr_none
    ws4.row_dimensions[1].height=20
    ri=2
    for i,r in enumerate(weeks):
        mw, dr = month_week_label(r['ws'])
        wk_label = mw or f"{i+1}주차"
        for k,item in sorted(items_by_week.get(r['wk'],{}).items(), key=lambda x:-x[1]['total']):
            for ci,v in enumerate([wk_label,f"{r['ws']}~{r['we']}",item['item_group'],item['item_name'],item['qty'],item['total']],2):
                c=ws4.cell(row=ri,column=ci,value=v); c.border=bdr_left if ci<=3 else bdr_none
                if ci>=6: c.alignment=right
                if ci==7 and isinstance(v,int): c.number_format=num_fmt
            ri+=1
    for ci,w in enumerate([10,26,14,36,12,16],2): ws4.column_dimensions[get_column_letter(ci)].width=w

    # 수정1: 기초 데이터 — 업로드하신 원본 엑셀 파일을 그대로 시트로 재현 (요약치 검증용)
    # 이 리포트에 포함된 주차들이 걸치는 연-월 목록 추출
    covered_ym = sorted(set(f"{r['ws'][:7]}" for r in weeks if r.get('ws')) | set(f"{r['we'][:7]}" for r in weeks if r.get('we')))
    raw_files = []
    if covered_ym:
        placeholders = ' OR '.join(['months LIKE ?'] * len(covered_ym))
        raw_files = conn.execute(
            f"SELECT filename, file_b64, months FROM sales_upload_file WHERE {placeholders}",
            [f"%{ym}%" for ym in covered_ym]).fetchall()

    raw_sheet_names = []
    if raw_files:
        import base64 as _b64_dl
        for fname_orig, fb64, months_str in raw_files:
            try:
                src_bytes = _b64_dl.b64decode(fb64)
                src_wb = openpyxl.load_workbook(io.BytesIO(src_bytes))
                for src_sheet_name in src_wb.sheetnames:
                    src_ws = src_wb[src_sheet_name]
                    tab_name = f"기초_{months_str.split(',')[0][5:]}월_{src_sheet_name}"[:31]
                    base_tab = tab_name; suf = 1
                    while tab_name in raw_sheet_names:
                        tab_name = f"{base_tab[:28]}_{suf}"; suf += 1
                    ws_raw = wb.create_sheet(tab_name)
                    _copy_sheet_with_style(src_ws, ws_raw)
                    raw_sheet_names.append(tab_name)
            except Exception:
                continue

    if not raw_sheet_names:
        ws_raw = wb.create_sheet("기초데이터")
        ws_raw.column_dimensions['A'].width = 2.5
        c0 = ws_raw.cell(row=2, column=2, value="⚠ 원본 파일이 저장되지 않은 기간입니다. 해당 월 데이터를 다시 업로드하면 원본 그대로 표시됩니다.")
        c0.font = mft(FONT_BLACK, True, 10)
        raw_hdrs = ['일자','거래처명','실적용거래처명','거래처코드','품목명','수량','단가','공급가액','부가세','합계','채널']
        for ci, h in enumerate(raw_hdrs, 2):
            c = ws_raw.cell(row=4, column=ci, value=h)
            c.font=mft(FONT_BLACK,True,9); c.fill=mf("F2F2F2"); c.alignment=center
        ws_raw.row_dimensions[4].height=20
        raw_q = f"SELECT sale_date,seller_name,real_seller,trade_code,item_name,quantity,unit_price,supply_price,vat,total,channel FROM sales_data WHERE {' AND '.join(qp)} ORDER BY sale_date, real_seller"
        ri_raw=5
        for r in conn.execute(raw_q, pp).fetchall():
            for ci,v in enumerate(r,2):
                c=ws_raw.cell(row=ri_raw,column=ci,value=v); c.font=mft(FONT_BLACK,False,8)
                c.alignment = right if ci in (7,8,9,10) else (center if ci in (2,6) else left)
                if ci in (8,9,10): c.number_format=num_fmt
            ri_raw+=1
        for ci,w in zip(range(2,13),[11,20,20,14,26,8,10,11,10,11,9]):
            ws_raw.column_dimensions[get_column_letter(ci)].width=w
        ws_raw.freeze_panes='B5'

    conn.close()
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    fname=f"주별실적_{year}{'_'+month+'월' if month else ''}.xlsx"
    return send_file(buf,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,download_name=fname)

@app.route("/api/export/xlsx/ranking")
@login_required
def export_xlsx_ranking():
    year=request.args.get("year",str(datetime.now().year)); month=request.args.get("month","")
    date_cond=f"{year}-{month.zfill(2)}%" if month else f"{year}%"
    conn=get_db()
    rows=[dict(r) for r in conn.execute("""SELECT real_seller AS seller_name,
        COUNT(*) cnt,SUM(total) total,SUM(quantity) qty
        FROM sales_data WHERE sale_date LIKE ? AND real_seller!=''
        GROUP BY real_seller ORDER BY total DESC""",(date_cond,)).fetchall()]
    conn.close()
    hdrs=['순위','매장명','판매금액(원)','판매건수','판매수량']
    data=[[i+1,r['seller_name'],r['total'],r['cnt'],r['qty']] for i,r in enumerate(rows)]
    buf=make_xlsx(hdrs,data,"매출순위")
    fname=f"매출순위_{year}{'_'+month+'월' if month else ''}.xlsx"
    return send_file(buf,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,download_name=fname)

# ── 제품별 관리 API ────────────────────────────────
@app.route("/api/products/groups")
@login_required
def api_product_groups():
    conn = get_db()
    raw = conn.execute("""
        SELECT item_group, item_name,
               SUM(quantity) qty, SUM(total) total
        FROM sales_data WHERE item_group != '' AND item_group IS NOT NULL
        GROUP BY item_group, item_name""").fetchall()
    conn.close()

    # remap 후 재집계 (색상 통합하여 실제 종류 수 계산)
    brand_items = {}  # brand → set of normalized names
    brand_totals = {}
    for r in raw:
        brand = remap_group(r[0], r[1])
        if not brand or brand == '기타': continue
        # 흡착식기 제외
        if '흡착' in (r[0] or '') or '흡착' in (r[1] or ''): continue
        norm  = normalize_item_name(r[1])
        if brand not in brand_items:
            brand_items[brand]  = set()
            brand_totals[brand] = {'qty': 0, 'total': 0}
        brand_items[brand].add(norm)
        brand_totals[brand]['qty']   += r[2] or 0
        brand_totals[brand]['total'] += r[3] or 0

    result = [
        {'item_group': b, 'item_cnt': len(brand_items[b]),
         'qty': brand_totals[b]['qty'], 'total': brand_totals[b]['total']}
        for b in brand_items
    ]
    result.sort(key=lambda x: get_group_sort_key(x['item_group']))
    resp = jsonify(result)
    resp.headers['Cache-Control'] = 'no-store'
    return resp

@app.route("/api/products/items")
@login_required
def api_product_items():
    group  = request.args.get("group",  "")
    seller = request.args.get("seller", "").strip()
    year   = request.args.get("year",   str(datetime.now().year))
    month  = request.args.get("month",  "")
    conn   = get_db()
    date_cond = f"{year}-{month.zfill(2)}%" if month else f"{year}%"
    params = [date_cond]; conds = ["sale_date LIKE ?", "sale_date != ''"]
    if seller: conds.append("real_seller=?"); params.append(seller)
    raw = [dict(r) for r in conn.execute(f"""
        SELECT item_name, item_group, SUM(quantity) qty,
               AVG(unit_price) avg_price, SUM(total) total, COUNT(*) cnt
        FROM sales_data WHERE {' AND '.join(conds)}
        GROUP BY item_name ORDER BY total DESC""", params).fetchall()]
    conn.close()
    # 브랜드 필터 + 정규화 + 재집계 (색상 통합)
    merged = {}
    for r in raw:
        brand = remap_group(r['item_group'], r['item_name'])
        if not brand or brand == '기타': continue
        if group and brand != group: continue
        norm = normalize_item_name(r['item_name'])
        key  = (brand, norm)
        if key not in merged:
            merged[key] = {'item_name': norm, 'item_group': brand,
                           'qty': 0, 'avg_price': r['avg_price'], 'total': 0, 'cnt': 0}
        merged[key]['qty']   += r['qty']
        merged[key]['total'] += r['total']
        merged[key]['cnt']   += r['cnt']
    return jsonify(sorted(merged.values(), key=lambda x: -x['total']))

@app.route("/api/products/by-seller")
@login_required
def api_product_by_seller():
    """특정 브랜드/품목의 매장별 판매 현황"""
    group  = request.args.get("group",  "")
    item   = request.args.get("item",   "")
    year   = request.args.get("year",   str(datetime.now().year))
    month  = request.args.get("month",  "")
    conn   = get_db()
    date_cond = f"{year}-{month.zfill(2)}%" if month else f"{year}%"
    params = [date_cond]; conds = ["sale_date LIKE ?", "sale_date != ''", "real_seller != ''"]
    raw = [dict(r) for r in conn.execute(f"""
        SELECT real_seller seller_name, item_group, item_name,
               SUM(quantity) qty, SUM(total) total, COUNT(*) cnt
        FROM sales_data WHERE {' AND '.join(conds)}
        GROUP BY real_seller, item_name ORDER BY total DESC""", params).fetchall()]
    conn.close()
    # 브랜드/아이템 필터 + 매장별 재집계
    merged = {}
    for r in raw:
        brand = remap_group(r['item_group'], r['item_name'])
        if group and brand != group: continue
        if item and normalize_item_name(r['item_name']) != item: continue
        nm = r['seller_name']
        if nm not in merged:
            merged[nm] = {'seller_name': nm, 'qty': 0, 'total': 0, 'cnt': 0}
        merged[nm]['qty']   += r['qty']
        merged[nm]['total'] += r['total']
        merged[nm]['cnt']   += r['cnt']
    return jsonify(sorted(merged.values(), key=lambda x: -x['total']))

def strip_honorific(name):
    """대표자명 뒤의 호칭(님, 사장님, 대표님, 이사님, 점장님 등) 제거"""
    import re as _re4
    if not name: return name
    n = name.strip()
    n = _re4.sub(r'(사장|대표|이사|점장|원장|실장|부장|팀장)?님\s*$', '', n).strip()
    n = _re4.sub(r'(사장|대표|이사|점장|원장)\s*$', '', n).strip()
    return n or name


@app.route("/api/export/xlsx/sellers")
@login_required
def api_export_sellers_xlsx():
    """판매처 관리 엑셀 — 지역별 분류 + 대표자/주소/연락처 + 전년동기대비 + 취급브랜드"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    year = request.args.get('year', str(datetime.now().year))
    prev_year = str(int(year) - 1)
    cur_month = datetime.now().month  # 현재 월까지만 비교 (전년 동기)

    conn = get_db()
    # 올해 매출: 연간 전체 표시용 (채널도 함께 조회 — 매장별 대표 채널로 판정)
    rows = conn.execute("""
        SELECT real_seller, COUNT(*) cnt, SUM(total) total, SUM(quantity) qty, MAX(sale_date) last_date,
               (SELECT channel FROM sales_data sd2 WHERE sd2.real_seller=sd1.real_seller
                GROUP BY channel ORDER BY COUNT(*) DESC LIMIT 1) AS channel
        FROM sales_data sd1 WHERE real_seller != '' AND sale_date LIKE ?
        GROUP BY real_seller ORDER BY real_seller
    """, (f"{year}%",)).fetchall()

    # 올해 동기(1월~현재월) 매출
    cur_period_totals = {r[0]: r[1] or 0 for r in conn.execute("""
        SELECT real_seller, SUM(total) FROM sales_data
        WHERE real_seller != '' AND sale_date LIKE ?
          AND CAST(strftime('%m',sale_date) AS INTEGER) <= ?
        GROUP BY real_seller
    """, (f"{year}%", cur_month)).fetchall()}

    # 전년 동기(1월~현재월) 매출 — 전년 동기 대비 비교용
    prev_totals = {r[0]: r[1] or 0 for r in conn.execute("""
        SELECT real_seller, SUM(total) FROM sales_data
        WHERE real_seller != '' AND sale_date LIKE ?
          AND CAST(strftime('%m',sale_date) AS INTEGER) <= ?
        GROUP BY real_seller
    """, (f"{prev_year}%", cur_month)).fetchall()}

    # 판매처 상세정보 (대표자/주소/연락처/담당자)
    branch_info = {}
    try:
        for r in conn.execute("SELECT name, ceo, ceo_phone, address, manager, phone FROM branches").fetchall():
            branch_info[r[0]] = {'ceo': r[1] or '', 'ceo_phone': r[2] or '',
                                  'address': r[3] or '', 'manager': r[4] or '', 'phone': r[5] or ''}
    except Exception:
        pass

    # 매장별 브랜드+제품명 취급 현황 (item_group, item_name 조합)
    brand_items = {}  # seller -> {brand: set(item_names)}
    for r in conn.execute("""
        SELECT real_seller, item_group, item_name FROM sales_data
        WHERE real_seller != '' AND sale_date LIKE ? AND item_group != ''
        GROUP BY real_seller, item_group, item_name
    """, (f"{year}%",)).fetchall():
        seller, grp, iname = r[0], r[1], r[2]
        brand = remap_group(grp, iname)
        if seller not in brand_items: brand_items[seller] = {}
        if brand not in brand_items[seller]: brand_items[seller][brand] = set()
        brand_items[seller][brand].add(normalize_item_name(iname))

    conn.close()

    MANUAL_REGION = {
        '링크맘 중랑점': '서울', '링크맘 평촌점': '경기남부', '베이비 투 키즈': '서울',
        '베이비하우스 검단점': '인천', '베이비하우스 뚝섬점': '서울',
        '베이비하우스 위례점': '경기남부', '베이비하우스 청라점': '인천', '베이비하우스 향남점': '경기남부',
    }
    BRANDS_ORDER = ['줄즈','원더폴드','레카로','엔픽스','타프토이즈','ABC디자인','카오스']

    def wonderfold_grade(items):
        """원더폴드 취급 제품에서 등급(프리미엄/일반/슈퍼프리미엄) 추출"""
        grades = set()
        for it in items:
            if '슈퍼프리미엄' in it or '슈퍼 프리미엄' in it: grades.add('슈퍼프리미엄')
            elif '프리미엄' in it: grades.add('프리미엄')
            elif '일반' in it: grades.add('일반')
        return ', '.join(sorted(grades)) if grades else 'O'

    seller_data = []
    for r in rows:
        raw_name = r[0]
        if is_hidden_seller(raw_name): continue
        disp_name = display_seller(raw_name) or raw_name
        region = MANUAL_REGION.get(disp_name) or MANUAL_REGION.get(raw_name) or detect_region_from_name(disp_name) or '기타'
        info = branch_info.get(raw_name, branch_info.get(disp_name, {}))

        cur_total = r[2] or 0
        cur_period_total  = cur_period_totals.get(raw_name, 0)
        prev_period_total = prev_totals.get(raw_name, 0)
        if prev_period_total > 0:
            yoy_pct = round((cur_period_total - prev_period_total) / prev_period_total * 100, 1)
        else:
            yoy_pct = None

        items_by_brand = brand_items.get(raw_name, {})
        brand_marks = {}
        for b in BRANDS_ORDER:
            items = items_by_brand.get(b)
            if not items:
                brand_marks[b] = ''
            elif b == '원더폴드':
                brand_marks[b] = wonderfold_grade(items)
            elif b == '카오스':
                # 카오스는 엑셀(제품명)에 적힌 그대로 표시
                brand_marks[b] = ', '.join(sorted(items))[:40]
            else:
                brand_marks[b] = 'O'

        seller_data.append({
            'name': disp_name, 'region': region, 'channel': r[5] or '오프라인',
            'ceo': strip_honorific(info.get('ceo','')), 'ceo_phone': info.get('ceo_phone',''),
            'address': info.get('address',''), 'manager': info.get('manager',''),
            'cnt': r[1], 'total': cur_total, 'qty': r[3] or 0,
            'last': (r[4] or '')[:10], 'yoy': yoy_pct,
            'cur_period_total': cur_period_total, 'prev_period_total': prev_period_total,
            'brands': brand_marks,
        })

    REGION_ORDER = ['서울','경기북부','경기남부','인천','부산','대구','광주','대전','울산','세종',
                    '강원','충북','충남','전북','전남','경북','경남','제주','기타']
    seller_data.sort(key=lambda x: (
        REGION_ORDER.index(x['region']) if x['region'] in REGION_ORDER else 99, x['name']))

    wb = openpyxl.Workbook()
    FNAME = '맑은 고딕'
    def mf(h): return PatternFill("solid", fgColor=h)
    thin = Side(style='thin', color='E5E7EB')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    rgt  = Alignment(horizontal='right', vertical='center')

    period_label = f'1~{cur_month}월'
    hdrs = ['지역','매장명','대표자','대표자 연락처','주소','담당자',
            f'{year}년 매출(원)', f'{year}년 {period_label}', f'{prev_year}년 {period_label}', '전년동기대비',
            '판매건수','판매수량','최근 거래일'] + BRANDS_ORDER
    widths = [8, 24, 10, 14, 32, 12, 16, 15, 15, 11, 9, 9, 12] + [12]*len(BRANDS_ORDER)
    NCOL = len(hdrs)
    BRAND_COL_START = 14  # N열부터 브랜드 컬럼

    def write_channel_sheet(ws, channel_label, data_list):
        ws.merge_cells(f'A1:{get_column_letter(NCOL)}1')
        c = ws.cell(row=1, column=1,
            value=f'판매처 현황 — {channel_label}  ({year}년)   총 {len(data_list)}개 매장   ·   전년동기대비 기준: {period_label}')
        c.font=Font(bold=True, size=13, name=FNAME, color='1F2937'); c.fill=mf('FFFFFF'); c.alignment=ctr
        ws.row_dimensions[1].height = 26

        for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.font=Font(bold=True, size=9, name=FNAME, color='374151')
            c.fill=mf('F3F4F6'); c.border=bdr; c.alignment=ctr
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[2].height = 20

        region_totals = {}
        ri = 3; prev_region = ''

        def write_subtotal(ri, region_name, rt):
            ws.merge_cells(f'A{ri}:{get_column_letter(NCOL)}{ri}')
            c = ws.cell(row=ri, column=1,
                value=f'{region_name} 소계  ({rt["stores"]}개 매장  /  {rt["total"]:,}원)')
            c.font=Font(bold=True, size=9, name=FNAME, color='6B7280')
            c.fill=mf('F9FAFB'); c.alignment=left; c.border=bdr
            ws.row_dimensions[ri].height = 14

        for s in data_list:
            region = s['region']
            if region != prev_region:
                if prev_region and prev_region in region_totals:
                    write_subtotal(ri, prev_region, region_totals[prev_region]); ri += 1
                ws.merge_cells(f'A{ri}:{get_column_letter(NCOL)}{ri}')
                c = ws.cell(row=ri, column=1, value=f'▌ {region}')
                c.font=Font(bold=True, size=10, name=FNAME, color='1F2937')
                c.fill=mf('E5E7EB'); c.alignment=left
                ws.row_dimensions[ri].height = 18; ri += 1
                prev_region = region
                if region not in region_totals:
                    region_totals[region] = {'total': 0, 'stores': 0}

            yoy_str = f"{'+' if s['yoy']>=0 else ''}{s['yoy']}%" if s['yoy'] is not None else '—'
            row_vals = [region, s['name'], s['ceo'], s['ceo_phone'], s['address'], s['manager'],
                        s['total'], s['cur_period_total'], s['prev_period_total'], yoy_str,
                        s['cnt'], s['qty'], s['last']]
            row_vals += [s['brands'][b] for b in BRANDS_ORDER]

            for ci, v in enumerate(row_vals, 1):
                c = ws.cell(row=ri, column=ci, value=v); c.border=bdr
                c.font=Font(size=9, name=FNAME, color='1F2937')
                if   ci == 1: c.alignment=ctr; c.font=Font(size=9, name=FNAME, color='9CA3AF')
                elif ci == 2: c.alignment=left
                elif ci in (3,6): c.alignment=ctr
                elif ci == 4: c.alignment=ctr; c.font=Font(size=8, name=FNAME, color='6B7280')
                elif ci == 5: c.alignment=left; c.font=Font(size=8, name=FNAME, color='6B7280')
                elif ci in (7,8,9): c.alignment=rgt; c.number_format='#,##0'
                elif ci == 10:
                    c.alignment=ctr
                    if s['yoy'] is not None:
                        c.font=Font(size=9, name=FNAME, bold=True,
                                    color='16A34A' if s['yoy']>=0 else 'DC2626')
                elif ci in (11,12): c.alignment=ctr
                elif ci == 13: c.alignment=ctr; c.font=Font(size=8, name=FNAME, color='9CA3AF')
                elif ci >= BRAND_COL_START: c.alignment=ctr
            ws.row_dimensions[ri].height = 15; ri += 1

            region_totals[region]['total']  += s['total']
            region_totals[region]['stores'] += 1

        if prev_region and prev_region in region_totals:
            write_subtotal(ri, prev_region, region_totals[prev_region]); ri += 1

        grand_total = sum(s['total'] for s in data_list)
        ws.merge_cells(f'A{ri}:{get_column_letter(NCOL)}{ri}')
        c = ws.cell(row=ri, column=1,
            value=f'전체 합계  ({len(data_list)}개 매장)     {grand_total:,}원')
        c.font=Font(bold=True, size=10, name=FNAME, color='1F2937')
        c.fill=mf('F3F4F6'); c.alignment=left; c.border=bdr
        ws.row_dimensions[ri].height = 22
        ws.freeze_panes = 'C3'

    # 수정: 오프라인 매장과 백화점(서양네트웍스/가이아코퍼레이션)을 별도 시트로 분리
    offline_data = [s for s in seller_data if s.get('channel','오프라인') != '백화점']
    dept_data    = [s for s in seller_data if s.get('channel','오프라인') == '백화점']

    ws1 = wb.active; ws1.title = '오프라인매장'
    write_channel_sheet(ws1, '오프라인 매장', offline_data)

    if dept_data:
        ws2 = wb.create_sheet('백화점')
        write_channel_sheet(ws2, '백화점', dept_data)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=f'판매처현황_{year}.xlsx')


@app.route("/api/export/xlsx/branches")
@login_required
def export_xlsx_branches():
    """거래처별 브랜드 입점 리스트 형식으로 판매처 내보내기"""
    year = request.args.get("year", str(datetime.now().year))
    conn = get_db()
    BRANDS_ORDER = ['줄즈','원더폴드','레카로','엔픽스','타프토이즈','ABC디자인','카오스']
    actual_groups = set(r[0] for r in conn.execute(
        "SELECT DISTINCT item_group FROM sales_data WHERE item_group!=''").fetchall())
    brands = [b for b in BRANDS_ORDER if b in actual_groups]
    for g in sorted(actual_groups):
        if g not in brands and g: brands.append(g)

    branches = [dict(r) for r in conn.execute("""
        SELECT id,name,ceo,ceo_phone,store_manager,store_manager_phone,
               manager,phone,address,email,status,note,region
        FROM branches ORDER BY note,name""").fetchall()]

    # 매장별 취급 브랜드
    brand_sold = {}
    for r in conn.execute(f"""SELECT real_seller,item_group FROM sales_data
        WHERE sale_date LIKE '{year}%' AND real_seller!='' AND item_group!=''
        GROUP BY real_seller,item_group""").fetchall():
        if r[0] not in brand_sold: brand_sold[r[0]] = set()
        brand_sold[r[0]].add(r[1])

    # 연간 실적
    year_sales_map = {r[0]:r[1] for r in conn.execute(f"""
        SELECT real_seller, SUM(total) FROM sales_data
        WHERE sale_date LIKE '{year}%' AND real_seller!=''
        GROUP BY real_seller""").fetchall()}
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "오프라인 거래처별 리스트"
    def mf(h): return PatternFill(start_color=h,end_color=h,fill_type="solid")
    def mft(h,b=True,s=10): return Font(color=h,bold=b,size=s)
    thin=Side(style='thin',color='D1D5DB')
    bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    ctr=Alignment(horizontal="center",vertical="center")
    rgt=Alignment(horizontal="right",vertical="center")

    total_cols=15+len(brands)
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    c=ws.cell(row=1,column=1,value=f"오프라인 거래처별 브랜드 입점 리스트_{year}")
    c.fill=mf("1E3A5F"); c.font=mft("FFFFFF",True,12); c.alignment=ctr; ws.row_dimensions[1].height=28
    ws.row_dimensions[2].height=6

    hdrs=['업체구분','거래처명','실적용거래처명','전화번호','사장님','사장연락처',
          '점장','점장연락처','담당자','주소','Email','지역','상태','연간실적(원)',''] + brands
    for ci,h in enumerate(hdrs,1):
        c=ws.cell(row=3,column=ci,value=h)
        c.fill=mf("7C3AED") if ci>15 else mf("2563EB")
        c.font=mft("FFFFFF",True,10); c.alignment=ctr; c.border=bdr
    ws.row_dimensions[3].height=22

    cws=[12,24,26,14,12,14,12,14,12,45,26,8,8,14,4]+[7]*len(brands)
    for ci,w in enumerate(cws,1): ws.column_dimensions[get_column_letter(ci)].width=w

    prev_grp=None
    for ri,b in enumerate(branches,4):
        nm=b['name'] or ''; nml=nm.replace('_',' ').lower()
        if '베이비하우스' in nml: rf=mf("FFF7ED")
        elif '링크맘' in nml: rf=mf("F0FDF4")
        elif ri%2==0: rf=mf("F8FAFC")
        else: rf=mf("FFFFFF")
        grp=b.get('note','') or ''
        gv=grp if grp!=prev_grp else ''; prev_grp=grp
        sold=brand_sold.get(nm,set())
        yr_sales=year_sales_map.get(nm,0)
        row_vals=[gv,nm,nm,b.get('phone',''),b.get('ceo',''),b.get('ceo_phone',''),
                  b.get('store_manager',''),b.get('store_manager_phone',''),
                  b.get('manager',''),b.get('address',''),b.get('email',''),
                  b.get('region',''),b.get('status',''),yr_sales,''] + \
                 ['○' if br in sold else '' for br in brands]
        for ci,val in enumerate(row_vals,1):
            c=ws.cell(row=ri,column=ci,value=val); c.fill=rf; c.border=bdr; c.font=Font(size=10)
            if ci==14: c.number_format='#,##0'; c.alignment=rgt
            if ci>15: c.alignment=ctr
    ws.freeze_panes="A4"

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    fname=f"오프라인_거래처별_브랜드_입점_리스트_{year}.xlsx"
    return send_file(buf,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,download_name=fname)

@app.route("/api/products/trend")
@login_required
def api_product_trend():
    """제품별 일별·주별 판매 추이"""
    item   = request.args.get("item",   "")
    group  = request.args.get("group",  "")  # 브랜드명 (줄즈, 레카로 등)
    year   = request.args.get("year",   str(datetime.now().year))
    month  = request.args.get("month",  "")
    conn   = get_db()
    date_cond = f"{year}-{month.zfill(2)}%" if month else f"{year}%"

    # item은 정규화명이므로 모델명만 추출하여 LIKE 검색
    import re as _re
    model = _re.sub(r'^\[[^\]]+\]', '', item).strip() if item else ''

    # 일별 추이 — item_name으로 필터
    conds  = ["sale_date LIKE ?", "sale_date != ''"]
    params = [date_cond]
    if model:
        conds.append("item_name LIKE ?"); params.append(f"%{model}%")

    daily = [dict(r) for r in conn.execute(f"""
        SELECT sale_date, SUM(quantity) qty, SUM(total) total, COUNT(*) cnt
        FROM sales_data WHERE {' AND '.join(conds)}
        GROUP BY sale_date ORDER BY sale_date""", params).fetchall()]

    # 주별 추이
    weekly_raw = conn.execute(f"""
        SELECT strftime('%Y-%W', sale_date) wk, MIN(sale_date) md,
               SUM(quantity) qty, SUM(total) total
        FROM sales_data WHERE {' AND '.join(conds)} AND sale_date!=''
        GROUP BY wk ORDER BY wk""", params).fetchall()

    from datetime import datetime as dt2, timedelta
    weekly = []
    for r in weekly_raw:
        try:
            d = dt2.strptime(r[1], "%Y-%m-%d")
            sun = d - timedelta(days=(d.weekday()+1)%7)
            label = sun.strftime("%m/%d")
        except: label = r[0]
        weekly.append({'wk': r[0], 'label': label, 'qty': r[2], 'total': r[3]})

    # 매장별 판매 현황
    by_seller = [dict(r) for r in conn.execute(f"""
        SELECT real_seller seller_name, SUM(quantity) qty, SUM(total) total, COUNT(*) cnt
        FROM sales_data WHERE {' AND '.join(conds)} AND real_seller!=''
        GROUP BY real_seller ORDER BY total DESC LIMIT 20""", params).fetchall()]

    conn.close()
    return jsonify({'daily': daily, 'weekly': weekly, 'by_seller': by_seller, 'item': item})

@app.route("/api/sales-data/summary")
@login_required
def sales_data_summary():
    year  = request.args.get("year", "")
    conn  = get_db()
    where = f"AND sale_date LIKE '{year}%'" if year else ""
    total = conn.execute(f"SELECT COUNT(*) c, SUM(total) t, SUM(quantity) q FROM sales_data WHERE 1=1 {where}").fetchone()
    by_seller = [dict(r) for r in conn.execute(f"""
        SELECT real_seller seller_name, COUNT(*) cnt, SUM(quantity) qty, SUM(total) total
        FROM sales_data WHERE real_seller != '' {where}
        GROUP BY real_seller ORDER BY total DESC""").fetchall()]
    by_group = [dict(r) for r in conn.execute(f"""
        SELECT item_group, COUNT(*) cnt, SUM(quantity) qty, SUM(total) total
        FROM sales_data WHERE item_group != '' {where} GROUP BY item_group ORDER BY total DESC""").fetchall()]
    by_date = [dict(r) for r in conn.execute(f"""
        SELECT sale_date, COUNT(*) cnt, SUM(total) total
        FROM sales_data WHERE sale_date != '' {where} GROUP BY sale_date ORDER BY sale_date""").fetchall()]
    by_item = [dict(r) for r in conn.execute(f"""
        SELECT item_name, SUM(quantity) qty, SUM(total) total
        FROM sales_data WHERE 1=1 {where} GROUP BY item_name ORDER BY total DESC LIMIT 20""").fetchall()]
    conn.close()
    return jsonify({
        "total_count": total["c"] or 0,
        "total_amount": total["t"] or 0,
        "total_quantity": total["q"] or 0,
        "seller_count": len(by_seller),
        "by_seller": by_seller,
        "by_group": by_group,
        "by_date": by_date,
        "by_item": by_item,
    })

# ── xlsx 판매현황 — real_seller 기준으로 저장 ──
@app.route("/api/branches/from-xlsx", methods=["POST"])
@login_required
def branches_from_xlsx():
    """판매현황 xlsx에서 실적용거래처명(real_seller) 기준으로 판매처 등록"""
    conn = get_db()
    sellers = [dict(r) for r in conn.execute("""
        SELECT real_seller, COUNT(*) cnt, SUM(total) total
        FROM sales_data WHERE real_seller != ''
        GROUP BY real_seller ORDER BY real_seller""").fetchall()]

    added, updated = 0, 0
    for s in sellers:
        name = s["real_seller"]
        region = detect_region_from_name(name)
        existing = conn.execute("SELECT id FROM branches WHERE name=?", (name,)).fetchone()
        if not existing:
            conn.execute("""INSERT INTO branches(name,region,manager,phone,address,status,note)
                VALUES(?,?,?,?,?,?,?)""", (name, region,"","","","운영중",""))
            added += 1
        else:
            # 지역이 비어있으면 자동 채우기
            if region:
                conn.execute("UPDATE branches SET region=? WHERE id=? AND (region='' OR region IS NULL)",
                             (region, existing["id"]))
            updated += 1
    conn.commit(); conn.close()
    return jsonify({"ok": True, "added": added, "updated": updated, "total": len(sellers)})
    return jsonify({"ok": True, "added": added, "updated": updated, "total": len(sellers)})

@app.route("/api/branches/<int:bid>", methods=["DELETE"])
@login_required
def api_branches_delete(bid):
    conn = get_db()
    conn.execute("DELETE FROM sales WHERE branch_id=?", (bid,))
    conn.execute("DELETE FROM branches WHERE id=?", (bid,))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

# ── 판매부수 API ───────────────────────────────
@app.route("/api/sales")
@login_required
def api_sales():
    bid  = request.args.get("branch_id")
    year = request.args.get("year", str(datetime.now().year))
    conn = get_db()
    if bid:
        rows = [dict(r) for r in conn.execute("""
            SELECT s.*, b.name branch_name FROM sales s
            JOIN branches b ON s.branch_id=b.id
            WHERE s.branch_id=? AND s.year=? ORDER BY s.month""", (bid, year)).fetchall()]
    else:
        rows = [dict(r) for r in conn.execute("""
            SELECT s.*, b.name branch_name, b.region FROM sales s
            JOIN branches b ON s.branch_id=b.id
            WHERE s.year=? ORDER BY b.name, s.month""", (year,)).fetchall()]
    conn.close()
    return jsonify(rows)

@app.route("/api/sales", methods=["POST"])
@login_required
def api_sales_save():
    d = request.json  # [{branch_id, year, month, target, actual}, ...]
    conn = get_db()
    for row in d:
        conn.execute("""INSERT INTO sales(branch_id,year,month,target,actual)
            VALUES(?,?,?,?,?)
            ON CONFLICT(branch_id,year,month) DO UPDATE SET target=excluded.target, actual=excluded.actual""",
            (row["branch_id"], row["year"], row["month"], row.get("target",0), row.get("actual",0)))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

# ── 권한 관리 API ──────────────────────────────
@app.route("/api/users")
@login_required
def api_users():
    conn = get_db()
    rows = [dict(r) for r in conn.execute("SELECT id,email,name,role,created_at FROM users ORDER BY name").fetchall()]
    conn.close()
    return jsonify(rows)

@app.route("/api/users", methods=["POST"])
@login_required
def api_users_add():
    d = request.json
    conn = get_db()
    try:
        conn.execute("INSERT INTO users(email,password,name,role) VALUES(?,?,?,?)",
            (d["email"],d["password"],d["name"],d.get("role","user")))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"ok":False,"msg":"이미 존재하는 이메일입니다."}), 400
    conn.close()
    return jsonify({"ok":True})

@app.route("/api/users/<int:uid>", methods=["PUT"])
@login_required
def api_users_update(uid):
    d = request.json
    conn = get_db()
    if d.get("password"):
        conn.execute("UPDATE users SET name=?,role=?,password=? WHERE id=?",
            (d["name"],d["role"],d["password"],uid))
    else:
        conn.execute("UPDATE users SET name=?,role=? WHERE id=?", (d["name"],d["role"],uid))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/users/<int:uid>", methods=["DELETE"])
@login_required
def api_users_delete(uid):
    if uid == session["user"]["id"]:
        return jsonify({"ok":False,"msg":"본인 계정은 삭제할 수 없습니다."}), 400
    conn = get_db()
    conn.execute("DELETE FROM users WHERE id=?", (uid,))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

# ── 엑셀 내보내기 ──────────────────────────────
@app.route("/api/export/branches")
@login_required
def export_branches():
    conn = get_db()
    rows = conn.execute("SELECT * FROM branches ORDER BY name").fetchall()
    conn.close()
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["ID","지사명","지역","담당자","전화","이메일","주소","상태","계약일","수수료율","메모"])
    for r in rows:
        w.writerow([r["id"],r["name"],r["region"],r["manager"],r["phone"],
                    r["email"],r["address"],r["status"],r["contract_date"],r["fee_rate"],r["note"]])
    buf.seek(0)
    return send_file(io.BytesIO(buf.getvalue().encode("utf-8-sig")), mimetype="text/csv",
                     as_attachment=True, download_name=f"지사목록_{date.today()}.csv")

@app.route("/api/export/sales")
@login_required
def export_sales():
    year = request.args.get("year", str(datetime.now().year))
    conn = get_db()
    rows = conn.execute("""
        SELECT b.name,b.region,s.month,s.target,s.actual,
               ROUND(CAST(s.actual AS REAL)/NULLIF(s.target,0)*100,1) pct
        FROM sales s JOIN branches b ON s.branch_id=b.id
        WHERE s.year=? ORDER BY b.name,s.month""", (year,)).fetchall()
    conn.close()
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["지사명","지역","월","목표부수","실적부수","달성률(%)"])
    for r in rows:
        w.writerow([r["name"],r["region"],f"{r['month']}월",r["target"],r["actual"],r["pct"]])
    buf.seek(0)
    return send_file(io.BytesIO(buf.getvalue().encode("utf-8-sig")), mimetype="text/csv",
                     as_attachment=True, download_name=f"판매부수_{year}.csv")

@app.route("/api/me")
@login_required
def api_me():
    return jsonify(session.get("user",{}))

# ── 엑셀(.xlsx) 판매현황 업로드 ───────────────
def parse_xlsx_sales(file_bytes):
    """xlsx 파싱 — 수량 -1 제외, 특이사항 '교환'/'샘플' 제외, 베이비하우스 본사 → 수취인으로 매장 파악"""
    import zipfile, xml.etree.ElementTree as ET, re
    from datetime import datetime as dt

    results = []
    with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
        strings = []
        if 'xl/sharedStrings.xml' in z.namelist():
            sst = z.read('xl/sharedStrings.xml').decode('utf-8')
            sst_root = ET.fromstring(sst)
            ns2 = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
            for si in sst_root.findall(f'{{{ns2}}}si'):
                strings.append(''.join(t.text or '' for t in si.findall(f'.//{{{ns2}}}t')))

        sheet_xml = z.read('xl/worksheets/sheet1.xml').decode('utf-8')
        root = ET.fromstring(sheet_xml)
        ns2 = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'

        for row in root.findall(f'.//{{{ns2}}}row'):
            rnum = int(row.get('r', 0))
            if rnum <= 2: continue

            row_vals = {}
            for cell in row.findall(f'{{{ns2}}}c'):
                ref = cell.get('r', '')
                col = ''.join(c for c in ref if c.isalpha())
                t = cell.get('t', '')
                is_el = cell.find(f'{{{ns2}}}is')
                v_el  = cell.find(f'{{{ns2}}}v')
                val = ''
                if is_el is not None:
                    val = ''.join(x.text or '' for x in is_el.findall(f'.//{{{ns2}}}t'))
                elif t == 's' and v_el is not None:
                    idx = int(v_el.text)
                    val = strings[idx] if idx < len(strings) else ''
                elif v_el is not None:
                    val = v_el.text or ''
                if val:
                    row_vals[col] = val

            if not row_vals.get('C'):
                continue

            # 수량 파싱 및 -1 제외
            try:
                qty = int(float(row_vals.get('I', 0) or 0))
            except:
                qty = 0
            if qty <= 0:
                continue  # 수량 -1 또는 0 제외

            # 특이사항(P열)에 '교환' 또는 '샘플' 포함 시 제외
            note = row_vals.get('P', '').strip()
            if '교환' in note or '샘플' in note:
                continue

            # 일자 파싱
            raw_date = row_vals.get('B', '')
            sale_date = re.sub(r'\s*-\d+$', '', raw_date).strip()
            try:
                dt.strptime(sale_date, '%Y/%m/%d')
                sale_date = sale_date.replace('/', '-')
            except:
                sale_date = ''

            # 실적용거래처명(AE열) 처리
            real_seller = row_vals.get('AE', '').strip()
            buyer       = row_vals.get('D', '').strip()

            # 베이비하우스_본사 → 수취인명으로 대체
            # 단, 수취인에 "고객님"이 포함된 경우는 제외 (개인 고객 주문)
            if '본사' in real_seller:
                if buyer and '고객님' in buyer:
                    continue  # 베이비하우스_본사이고 수취인이 "고객님"인 경우만 제외
                elif buyer:
                    real_seller = buyer

            # 언더바 정규화: "베이비하우스_영통점" → "베이비하우스 영통점"
            real_seller = real_seller.replace('_', ' ')
            # 별칭 처리: 위드에이컴퍼니 → 베이비하우스 관악점
            real_seller = resolve_seller(real_seller)

            # 채널 판별 (오프라인/백화점) — 원본 거래처명(C열) 기준
            seller_name_raw = row_vals.get('C', '').strip()
            channel = detect_channel(seller_name_raw, real_seller)

            results.append({
                'sale_date':    sale_date,
                'seller_name':  seller_name_raw,
                'item_code':    row_vals.get('G', '').strip(),
                'item_name':    row_vals.get('H', '').strip(),
                'item_group':   row_vals.get('AA', '').strip(),
                'quantity':     qty,
                'unit_price':   int(float(row_vals.get('K', 0) or 0)),
                'supply_price': int(float(row_vals.get('L', 0) or 0)),
                'vat':          int(float(row_vals.get('M', 0) or 0)),
                'total':        int(float(row_vals.get('N', 0) or 0)),
                'buyer':        buyer,
                'buyer_phone':  row_vals.get('E', '').strip(),
                'real_seller':  real_seller,
                'note':         note,
                'channel':      channel,
                'trade_code':   row_vals.get('AC', '').strip(),  # 거래처코드 (수정4)
            })
    return results

@app.errorhandler(413)
def too_large(e):
    return jsonify({"error": "파일이 너무 큽니다 (최대 50MB)"}), 413

@app.route("/api/upload/xlsx/preview", methods=["POST"])
@login_required
def upload_xlsx_preview():
    f = request.files.get("file")
    if not f: return jsonify({"error": "파일이 없습니다"}), 400
    if not f.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({"error": "xlsx 파일만 업로드 가능합니다"}), 400
    try:
        data = f.read()
        rows = parse_xlsx_sales(data)
    except Exception as e:
        return jsonify({"error": f"파일 파싱 오류: {str(e)}"}), 400

    # 날짜 범위
    dates = [r['sale_date'] for r in rows if r['sale_date']]
    d_from = min(dates, default='')
    d_to   = max(dates, default='')

    # 기간 파악 (월 단위)
    months = sorted(set(d[:7] for d in dates if d))

    # real_seller 기준 집계
    sellers = {}
    for r in rows:
        name = r['real_seller'] or r['seller_name']
        if name not in sellers:
            sellers[name] = {'count': 0, 'total': 0, 'qty': 0}
        sellers[name]['count'] += 1
        sellers[name]['total'] += r['total']
        sellers[name]['qty']   += r['quantity']

    # 이미 저장된 해당 월 데이터 여부 확인
    conn = get_db()
    existing_months = []
    for m in months:
        cnt = conn.execute("SELECT COUNT(*) FROM sales_data WHERE sale_date LIKE ?",
                           (f"{m}%",)).fetchone()[0]
        if cnt > 0:
            existing_months.append(m)
    conn.close()

    return jsonify({
        "count": len(rows),
        "seller_count": len(sellers),
        "months": months,
        "existing_months": existing_months,
        "sellers": [{"name": k, "count": v['count'], "total": v['total'], "qty": v['qty']}
                    for k, v in sorted(sellers.items(), key=lambda x: -x[1]['total'])],
        "date_range": {"from": d_from, "to": d_to},
    })

@app.route("/api/upload/xlsx/commit", methods=["POST"])
@login_required
def upload_xlsx_commit():
    f = request.files.get("file")
    if not f: return jsonify({"error": "파일이 없습니다"}), 400
    try:
        data = f.read()
        rows = parse_xlsx_sales(data)
    except Exception as e:
        return jsonify({"error": f"파싱 오류: {str(e)}"}), 400

    if not rows:
        return jsonify({"error": "유효한 데이터가 없습니다. 수량이 0 이하이거나 교환 처리된 행만 있을 수 있습니다."}), 400

    overwrite = request.form.get("overwrite", "0") == "1"
    batch = datetime.now().strftime("%Y%m%d%H%M%S")
    conn = get_db()

    # 해당 월 데이터만 교체 (누적 방식)
    dates = [r['sale_date'] for r in rows if r['sale_date']]
    months = sorted(set(d[:7] for d in dates if d))
    for m in months:
        conn.execute("DELETE FROM sales_data WHERE sale_date LIKE ?", (f"{m}%",))

    # 수정1: 원본 업로드 파일 그대로 보관 (기초데이터 시트 재현용) — 겹치는 월의 기존 원본 파일은 교체
    import base64 as _b64
    file_b64 = _b64.b64encode(data).decode('ascii')
    years_covered = sorted(set(d[:4] for d in dates if d))
    for m in months:
        conn.execute("DELETE FROM sales_upload_file WHERE months LIKE ?", (f"%{m}%",))
    conn.execute("""INSERT INTO sales_upload_file (upload_batch, year, months, filename, file_b64, uploaded_at)
        VALUES(?,?,?,?,?,?)""",
        (batch, int(years_covered[0]) if years_covered else datetime.now().year,
         ','.join(months), f.filename or '', file_b64, datetime.now().strftime('%Y-%m-%d %H:%M')))

    # 판매 데이터 저장
    for r in rows:
        conn.execute("""INSERT INTO sales_data
            (sale_date,seller_name,item_code,item_name,item_group,quantity,
             unit_price,supply_price,vat,total,buyer,buyer_phone,real_seller,upload_batch,note,channel,trade_code)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (r['sale_date'], r['seller_name'], r['item_code'], r['item_name'],
             r['item_group'], r['quantity'], r['unit_price'], r['supply_price'],
             r['vat'], r['total'], r['buyer'], r['buyer_phone'],
             r['real_seller'], batch, r.get('note', ''), r.get('channel', '오프라인'), r.get('trade_code', '')))

    conn.commit(); conn.close()
    return jsonify({"ok": True, "rows": len(rows), "months": months, "batch": batch})

@app.route("/api/sellers")
@login_required
def api_sellers():
    conn = get_db()
    rows = [dict(r) for r in conn.execute(
        "SELECT * FROM sellers ORDER BY total_sales DESC").fetchall()]
    conn.close()
    return jsonify(rows)

@app.route("/api/admin/merge-seller", methods=["POST"])
@login_required
def api_merge_seller():
    """두 real_seller를 하나로 통합 (from → to)"""
    d = request.json or {}
    from_seller = d.get('from_seller','').strip()
    to_seller   = d.get('to_seller','').strip()
    if not from_seller or not to_seller or from_seller == to_seller:
        return jsonify({'ok': False, 'msg': '매장명을 확인해주세요'}), 400
    conn = get_db()
    cnt = conn.execute("SELECT COUNT(*) FROM sales_data WHERE real_seller=?", (from_seller,)).fetchone()[0]
    conn.execute("UPDATE sales_data SET real_seller=? WHERE real_seller=?", (to_seller, from_seller))
    conn.execute("UPDATE display_record SET seller_name=? WHERE seller_name=?", (to_seller, from_seller))
    conn.commit(); conn.close()
    return jsonify({'ok': True, 'merged': cnt, 'from': from_seller, 'to': to_seller})

@app.route("/api/admin/delete-seller", methods=["POST"])
@login_required
def api_delete_seller():
    """real_seller 항목 삭제 (판매 데이터 포함)"""
    d = request.json or {}
    seller = d.get('seller_name','').strip()
    if not seller:
        return jsonify({'ok': False, 'msg': '매장명을 입력해주세요'}), 400
    conn = get_db()
    cnt = conn.execute("SELECT COUNT(*) FROM sales_data WHERE real_seller=?", (seller,)).fetchone()[0]
    conn.execute("DELETE FROM sales_data WHERE real_seller=?", (seller,))
    conn.execute("DELETE FROM display_record WHERE seller_name=?", (seller,))
    conn.commit(); conn.close()
    return jsonify({'ok': True, 'deleted': cnt, 'seller': seller})

@app.route("/api/admin/normalize-sellers", methods=["POST"])
@login_required
def normalize_sellers():
    """sales_data의 real_seller 정규화 + display_record 링크맘 공백 통합"""
    LINKMOM_SPACE_MAP = {
        '링크맘 경기 광주점':   '링크맘 경기광주점',
        '링크맘 대구 달성점':   '링크맘 대구달성점',
        '링크맘 대구 성서점':   '링크맘 대구성서점',
        '링크맘 대구 수성점':   '링크맘 대구수성점',
        '링크맘 파주 직영점':   '링크맘 파주직영점',
        '링크맘 의정부 민락점': '링크맘 의정부민락점',
        '링크맘 의정부 직영점': '링크맘 의정부직영점',
    }
    conn = get_db()
    rows = conn.execute("SELECT DISTINCT real_seller FROM sales_data WHERE real_seller != ''").fetchall()
    updated = 0
    deleted = 0

    for r in rows:
        old = r[0]
        if '고객' in old:
            conn.execute("DELETE FROM sales_data WHERE real_seller=?", (old,))
            deleted += 1; continue
        new = old.replace('_', ' ').strip()
        new = LINKMOM_SPACE_MAP.get(new, new)
        new = SELLER_ALIAS.get(new, new)
        new = resolve_seller(new)
        if old != new:
            conn.execute("UPDATE sales_data SET real_seller=? WHERE real_seller=?", (new, old))
            updated += 1

    # display_record — 링크맘 공백 버전을 공백없는 버전으로 통합
    disp_rows = conn.execute("SELECT DISTINCT seller_name FROM display_record WHERE seller_name!=''").fetchall()
    disp_updated = 0
    for r in disp_rows:
        old = r[0]
        if '고객' in old:
            conn.execute("DELETE FROM display_record WHERE seller_name=?", (old,))
            continue
        new = old.replace('_',' ').strip()
        # 링크맘 공백 정규화 (우선 적용)
        new = LINKMOM_SPACE_MAP.get(new, new)
        new = SELLER_ALIAS.get(new, new)
        new = resolve_seller(new)
        if old != new:
            # UNIQUE 제약 처리: 같은 캠페인+제품이면 점수 합산 후 삭제
            conflicts = conn.execute("""
                SELECT dr_old.id, dr_old.campaign_id, dr_old.product_name, dr_old.score
                FROM display_record dr_old
                WHERE dr_old.seller_name=?
            """, (old,)).fetchall()
            for cf in conflicts:
                cf_id, cf_camp, cf_prod, cf_score = cf
                existing = conn.execute("""
                    SELECT id, score FROM display_record
                    WHERE seller_name=? AND campaign_id=? AND product_name=?
                """, (new, cf_camp, cf_prod)).fetchone()
                if existing:
                    # 더 높은 점수 유지
                    keep_score = max(existing[1] or 0, cf_score or 0)
                    conn.execute("UPDATE display_record SET score=? WHERE id=?", (keep_score, existing[0]))
                    conn.execute("DELETE FROM display_record WHERE id=?", (cf_id,))
                else:
                    conn.execute("UPDATE display_record SET seller_name=? WHERE id=?", (new, cf_id))
            disp_updated += 1

    # branches 지역 자동 배정
    branches = conn.execute("SELECT id, name FROM branches WHERE region='' OR region IS NULL").fetchall()
    region_updated = 0
    for b in branches:
        region = detect_region_from_name(b["name"])
        if region:
            conn.execute("UPDATE branches SET region=? WHERE id=?", (region, b["id"]))
            region_updated += 1

    conn.commit(); conn.close()
    return jsonify({
        "ok": True, "normalized": updated, "deleted": deleted,
        "disp_updated": disp_updated, "region_updated": region_updated
    })

@app.route("/api/admin/merge-branches", methods=["POST"])
@login_required
def merge_branches():
    """띄어쓰기/언더바 차이로 중복된 판매처 통합 — 연간 실적 기준, 모든 연락처 정보 병합"""
    conn = get_db()
    year = str(datetime.now().year)
    branches = [dict(r) for r in conn.execute(
        "SELECT id,name,ceo,ceo_phone,store_manager,store_manager_phone,manager,phone,address,email,region,note,status FROM branches ORDER BY name").fetchall()]

    def normalize_nm(name):
        return name.replace('_','').replace(' ','').replace('(','').replace(')','').lower()

    # 정규화된 이름으로 그룹화
    groups = {}
    for b in branches:
        key = normalize_nm(b['name'])
        groups.setdefault(key, []).append(b)

    merged = 0
    for key, group in groups.items():
        if len(group) < 2: continue
        # 연간 실적 기준으로 대표 선정
        best = None; best_sales = -1
        for b in group:
            sales = conn.execute(
                "SELECT COALESCE(SUM(total),0) FROM sales_data WHERE real_seller=? AND sale_date LIKE ?",
                (b['name'], f"{year}%")).fetchone()[0]
            if sales > best_sales: best_sales = sales; best = b

        # 나머지 브랜치에서 정보 수집하여 best에 병합
        def pick(vals): return next((v for v in vals if v and v.strip()), '')
        for b in group:
            if b['id'] == best['id']: continue
            # 연락처 정보 없는 쪽에서 있는 쪽으로 채우기
            updates = {}
            for field in ['ceo','ceo_phone','store_manager','store_manager_phone','manager','phone','address','email','region']:
                if not best.get(field) and b.get(field):
                    updates[field] = b[field]
            if updates:
                set_clause = ', '.join(f"{k}=?" for k in updates)
                conn.execute(f"UPDATE branches SET {set_clause} WHERE id=?",
                             list(updates.values()) + [best['id']])
                best.update(updates)
            # sales_data real_seller 업데이트
            conn.execute("UPDATE sales_data SET real_seller=? WHERE real_seller=?",
                         (best['name'], b['name']))
            conn.execute("DELETE FROM branches WHERE id=?", (b['id'],))
            merged += 1

    # 지역 자동 배정
    branches_no_region = conn.execute("SELECT id,name FROM branches WHERE region='' OR region IS NULL").fetchall()
    region_updated = 0
    for b in branches_no_region:
        region = detect_region_from_name(b["name"])
        if region:
            conn.execute("UPDATE branches SET region=? WHERE id=?", (region, b["id"]))
            region_updated += 1
    conn.commit(); conn.close()
    return jsonify({"ok": True, "merged": merged, "region_updated": region_updated})

# ── 주별 세부 품목 API ─────────────────────────
@app.route("/api/sales-data/weekly-detail")
@login_required
def sales_weekly_detail():
    week_key = request.args.get("week_key", "")
    seller   = request.args.get("seller",   "").strip()
    conn     = get_db()

    params = [week_key]
    conds  = ["strftime('%Y-%W', sale_date) = ?", "sale_date != ''"]
    if seller:
        conds.append("real_seller = ?")
        params.append(seller)

    where = " AND ".join(conds)

    items = [dict(r) for r in conn.execute(f"""
        SELECT item_name, item_code, item_group,
               SUM(quantity) qty, AVG(unit_price) avg_price, SUM(total) total, COUNT(*) cnt
        FROM sales_data
        WHERE {where}
        GROUP BY item_name ORDER BY total DESC""", params).fetchall()]

    summary = conn.execute(f"""
        SELECT COUNT(*) cnt, SUM(quantity) qty, SUM(total) total,
               MIN(sale_date) date_from, MAX(sale_date) date_to
        FROM sales_data WHERE {where}""", params).fetchone()

    conn.close()
    return jsonify({"items": items, "summary": dict(summary), "week_key": week_key, "seller": seller})

# ── 주별 실적 API ──────────────────────────────
@app.route("/api/sales-data/weekly")
@login_required
def sales_data_weekly():
    year   = request.args.get("year",   str(datetime.now().year))
    month  = request.args.get("month",  "").strip()
    seller = request.args.get("seller", "").strip()
    channel = request.args.get("channel", "").strip()
    conn   = get_db()

    params = []
    conds  = ["sale_date != ''"]

    if month:
        conds.append("sale_date LIKE ?")
        params.append(f"{year}-{month.zfill(2)}%")
    else:
        conds.append("sale_date LIKE ?")
        params.append(f"{year}%")

    if seller:
        conds.append("real_seller = ?")
        params.append(seller)

    if channel:
        conds.append("channel = ?")
        params.append(channel)

    where = " AND ".join(conds)

    rows = [dict(r) for r in conn.execute(f"""
        SELECT
            strftime('%Y-%W', sale_date) AS week_key,
            COUNT(*) cnt,
            SUM(quantity) qty,
            SUM(total) total,
            MIN(sale_date) AS min_date
        FROM sales_data
        WHERE {where} AND sale_date != ''
        GROUP BY week_key
        ORDER BY week_key""", params).fetchall()]
    conn.close()

    # 주차별 일요일~토요일 범위 계산
    from datetime import datetime as dt, timedelta

    def get_week_range(date_str):
        d = dt.strptime(date_str, "%Y-%m-%d")
        wd = d.weekday()  # 0=월
        days_to_sun = (wd + 1) % 7
        sun = d - timedelta(days=days_to_sun)
        sat = sun + timedelta(days=6)
        return sun.strftime("%Y-%m-%d"), sat.strftime("%Y-%m-%d")

    for r in rows:
        try:
            r['week_start'], r['week_end'] = get_week_range(r['min_date'])
        except Exception:
            r['week_start'] = r.get('min_date', '')
            r['week_end']   = ''

    # 선택 월이 있으면 해당 월의 모든 주차를 채움 (데이터 없는 주도 표시)
    if month and rows:
        import calendar
        yr_int = int(year)
        mo_int = int(month)
        # 해당 월의 첫날~마지막날
        first_day = dt(yr_int, mo_int, 1)
        last_day  = dt(yr_int, mo_int, calendar.monthrange(yr_int, mo_int)[1])

        # 해당 월에 포함된 모든 주(일~토) 목록 생성
        all_weeks = {}
        cur = first_day
        while cur <= last_day:
            wk_start, wk_end = get_week_range(cur.strftime("%Y-%m-%d"))
            wk_key = cur.strftime("%Y-%W")
            if wk_key not in all_weeks:
                all_weeks[wk_key] = {'week_key': wk_key, 'week_start': wk_start, 'week_end': wk_end,
                                     'cnt': 0, 'qty': 0, 'total': 0, 'min_date': cur.strftime("%Y-%m-%d")}
            cur += timedelta(days=1)

        # 실제 데이터로 채우기
        data_map = {r['week_key']: r for r in rows}
        for wk_key in all_weeks:
            if wk_key in data_map:
                all_weeks[wk_key] = data_map[wk_key]

        rows = sorted(all_weeks.values(), key=lambda x: x['week_key'])

    return jsonify(rows)

# ── 월별 세부 품목 API ─────────────────────────
@app.route("/api/sales-data/monthly-detail")
@login_required
def sales_monthly_detail():
    year   = request.args.get("year",   str(datetime.now().year))
    month  = request.args.get("month",  "")
    seller = request.args.get("seller", "").strip()
    conn   = get_db()

    params = [f"{year}-{month.zfill(2)}%"] if month else [f"{year}%"]
    where  = "sale_date LIKE ?"
    if seller:
        where  += " AND real_seller=?"
        params.append(seller)

    # 품목별 집계
    items = [dict(r) for r in conn.execute(f"""
        SELECT item_name, item_code, item_group,
               SUM(quantity) qty, AVG(unit_price) avg_price, SUM(total) total, COUNT(*) cnt
        FROM sales_data
        WHERE {where} AND sale_date != ''
        GROUP BY item_name ORDER BY total DESC""", params).fetchall()]

    # 요약
    summary = conn.execute(f"""
        SELECT COUNT(*) cnt, SUM(quantity) qty, SUM(total) total
        FROM sales_data WHERE {where} AND sale_date != ''""", params).fetchone()

    conn.close()
    return jsonify({
        "items": items,
        "summary": dict(summary),
        "year": year, "month": month, "seller": seller
    })

# ── 엑셀 템플릿 다운로드 ───────────────────────
@app.route("/api/template/branches")
@login_required
def template_branches():
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["지사명","지역","담당자","전화","이메일","주소","상태","계약일","수수료율","메모"])
    w.writerow(["서울 강남지사","서울","홍길동","010-1234-5678","example@visang.com","서울시 강남구","운영중","2024-01-01",5.0,"예시 데이터"])
    buf.seek(0)
    return send_file(io.BytesIO(buf.getvalue().encode("utf-8-sig")), mimetype="text/csv",
                     as_attachment=True, download_name="지사_업로드_양식.csv")

@app.route("/api/template/sales")
@login_required
def template_sales():
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["지사명","연도","1월목표","1월실적","2월목표","2월실적","3월목표","3월실적",
                "4월목표","4월실적","5월목표","5월실적","6월목표","6월실적",
                "7월목표","7월실적","8월목표","8월실적","9월목표","9월실적",
                "10월목표","10월실적","11월목표","11월실적","12월목표","12월실적"])
    w.writerow(["서울 강남지사", 2026,
                1000,850, 1200,1100, 1100,980, 1300,1250, 1400,1300, 1200,1150,
                1100,1000, 1300,1200, 1400,1350, 1500,1420, 1600,1500, 1800,1700])
    buf.seek(0)
    return send_file(io.BytesIO(buf.getvalue().encode("utf-8-sig")), mimetype="text/csv",
                     as_attachment=True, download_name="판매부수_업로드_양식.csv")

# ── 엑셀 업로드 (미리보기) ────────────────────
@app.route("/api/upload/branches/preview", methods=["POST"])
@login_required
def upload_branches_preview():
    f = request.files.get("file")
    if not f: return jsonify({"error":"파일이 없습니다"}), 400
    content = f.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    rows, errors = [], []
    REGIONS = ["서울","경기","인천","강원","충북","충남","대전","세종","경북","경남","대구","부산","울산","전북","전남","광주","제주"]
    for i, row in enumerate(reader, 1):
        name = row.get("지사명","").strip()
        region = row.get("지역","").strip()
        if not name:
            errors.append(f"{i}행: 지사명 누락")
            continue
        if region and region not in REGIONS:
            errors.append(f"{i}행 [{name}]: 알 수 없는 지역 '{region}'")
        rows.append({
            "name": name, "region": region,
            "manager": row.get("담당자","").strip(),
            "phone": row.get("전화","").strip(),
            "email": row.get("이메일","").strip(),
            "address": row.get("주소","").strip(),
            "status": row.get("상태","운영중").strip() or "운영중",
            "contract_date": row.get("계약일","").strip(),
            "fee_rate": float(row.get("수수료율",0) or 0),
            "note": row.get("메모","").strip(),
        })
    return jsonify({"rows": rows, "errors": errors, "count": len(rows)})

@app.route("/api/upload/branches/commit", methods=["POST"])
@login_required
def upload_branches_commit():
    data = request.json
    rows = data.get("rows", [])
    mode = data.get("mode", "append")  # append | overwrite
    conn = get_db()
    if mode == "overwrite":
        conn.execute("DELETE FROM branches")
        conn.execute("DELETE FROM sales")
    added = 0
    for r in rows:
        existing = conn.execute("SELECT id FROM branches WHERE name=?", (r["name"],)).fetchone()
        if existing:
            conn.execute("""UPDATE branches SET region=?,manager=?,phone=?,email=?,
                            address=?,status=?,contract_date=?,fee_rate=?,note=? WHERE id=?""",
                (r["region"],r["manager"],r["phone"],r["email"],r["address"],
                 r["status"],r["contract_date"],r["fee_rate"],r["note"],existing["id"]))
        else:
            conn.execute("""INSERT INTO branches(name,region,manager,phone,email,address,status,contract_date,fee_rate,note)
                            VALUES(?,?,?,?,?,?,?,?,?,?)""",
                (r["name"],r["region"],r["manager"],r["phone"],r["email"],
                 r["address"],r["status"],r["contract_date"],r["fee_rate"],r["note"]))
            added += 1
    conn.commit(); conn.close()
    return jsonify({"ok": True, "added": added, "total": len(rows)})

@app.route("/api/upload/sales/preview", methods=["POST"])
@login_required
def upload_sales_preview():
    f = request.files.get("file")
    if not f: return jsonify({"error":"파일이 없습니다"}), 400
    content = f.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    rows, errors = [], []
    conn = get_db()
    for i, row in enumerate(reader, 1):
        name = row.get("지사명","").strip()
        year = row.get("연도","").strip()
        if not name or not year:
            errors.append(f"{i}행: 지사명 또는 연도 누락"); continue
        branch = conn.execute("SELECT id FROM branches WHERE name=?", (name,)).fetchone()
        if not branch:
            errors.append(f"{i}행: '{name}' 지사가 시스템에 없음 (지사 먼저 등록 필요)")
            continue
        months = []
        for m in range(1, 13):
            t = int(row.get(f"{m}월목표", 0) or 0)
            a = int(row.get(f"{m}월실적", 0) or 0)
            months.append({"month": m, "target": t, "actual": a})
        rows.append({"branch_id": branch["id"], "branch_name": name,
                     "year": int(year), "months": months})
    conn.close()
    return jsonify({"rows": rows, "errors": errors, "count": len(rows)})

@app.route("/api/upload/sales/commit", methods=["POST"])
@login_required
def upload_sales_commit():
    data = request.json
    rows = data.get("rows", [])
    conn = get_db()
    for r in rows:
        for m in r["months"]:
            conn.execute("""INSERT INTO sales(branch_id,year,month,target,actual) VALUES(?,?,?,?,?)
                ON CONFLICT(branch_id,year,month) DO UPDATE SET target=excluded.target,actual=excluded.actual""",
                (r["branch_id"], r["year"], m["month"], m["target"], m["actual"]))
    conn.commit(); conn.close()
    return jsonify({"ok": True, "total": len(rows)})

# ── 매장과의 소통 API ────────────────────────────────
@app.route("/api/communication")
@login_required
def api_communication_list():
    """전체 소통 기록 또는 특정 매장의 소통 기록 조회"""
    seller = request.args.get('seller', '').strip()
    conn = get_db()
    if seller:
        rows = [dict(r) for r in conn.execute("""
            SELECT * FROM store_communication WHERE seller_name=? ORDER BY comm_date DESC, id DESC
        """, (seller,)).fetchall()]
    else:
        rows = [dict(r) for r in conn.execute("""
            SELECT * FROM store_communication ORDER BY comm_date DESC, id DESC LIMIT 200
        """).fetchall()]
    conn.close()
    return jsonify(rows)


def polish_memo(raw_text, seller_name='', comm_type='방문'):
    """대충 적은 메모를 영업 전문가 문체의 보고서 형식으로 다듬기 (Claude API 사용, 실패 시 원문 유지)"""
    raw_text = (raw_text or '').strip()
    if not raw_text:
        return raw_text
    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        return raw_text
    try:
        import requests
        prompt = f"""다음은 영업 담당자가 매장 방문/통화 후 대충 적은 메모입니다.
이 내용을 절대 과장하거나 없는 사실을 추가하지 말고, 있는 내용 그대로 영업 보고서에 어울리는
정중하고 간결한 문어체로 다듬어주세요. 불필요한 수식어나 인사말은 넣지 말고, 2~4문장 이내로 작성하세요.
결과는 다듬어진 메모 텍스트만 출력하세요 (설명, 따옴표, 접두사 없이).

매장명: {seller_name}
소통 유형: {comm_type}
원본 메모: {raw_text}"""
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": "claude-sonnet-4-6",
                "max_tokens": 300,
                "messages": [{"role": "user", "content": prompt}],
            },
            timeout=10,
        )
        if resp.status_code == 200:
            data = resp.json()
            text_blocks = [b.get('text','') for b in data.get('content',[]) if b.get('type')=='text']
            polished = ''.join(text_blocks).strip()
            if polished:
                return polished
    except Exception:
        pass
    return raw_text


@app.route("/api/communication/polish", methods=["POST"])
@login_required
def api_communication_polish_preview():
    """메모 다듬기 미리보기 (저장 전 확인용)"""
    d = request.json or {}
    raw_memo = d.get('memo', '').strip()
    seller = d.get('seller_name', '')
    comm_type = d.get('comm_type', '방문')
    if not raw_memo:
        return jsonify({'ok': False, 'msg': '메모를 입력해주세요'}), 400
    polished = polish_memo(raw_memo, seller, comm_type)
    return jsonify({'ok': True, 'polished': polished, 'changed': polished != raw_memo})


@app.route("/api/communication", methods=["POST"])
@login_required
def api_communication_add():
    """소통 기록 추가 — 메모는 자동으로 영업 보고서 문체로 다듬어 저장"""
    d = request.json or {}
    seller = d.get('seller_name', '').strip()
    comm_date = d.get('comm_date', '').strip()
    raw_memo = d.get('memo', '').strip()
    if not seller or not comm_date:
        return jsonify({'ok': False, 'msg': '매장명과 날짜는 필수입니다'}), 400
    comm_type = d.get('comm_type','방문')
    polished = polish_memo(raw_memo, seller, comm_type)
    conn = get_db()
    user_name = session.get('user', {}).get('name', '')
    conn.execute("""INSERT INTO store_communication
        (seller_name, comm_date, comm_type, memo, raw_memo, created_by, created_at)
        VALUES(?,?,?,?,?,?,?)""",
        (seller, comm_date, comm_type, polished, raw_memo,
         user_name, datetime.now().strftime('%Y-%m-%d %H:%M')))
    conn.commit()
    new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.close()
    return jsonify({'ok': True, 'id': new_id, 'polished_memo': polished})


@app.route("/api/communication/<int:cid>", methods=["DELETE"])
@login_required
def api_communication_delete(cid):
    conn = get_db()
    conn.execute("DELETE FROM store_communication WHERE id=?", (cid,))
    conn.commit(); conn.close()
    return jsonify({'ok': True})


@app.route("/api/communication/<int:cid>", methods=["PUT"])
@login_required
def api_communication_update(cid):
    d = request.json or {}
    conn = get_db()
    conn.execute("""UPDATE store_communication SET comm_date=?, comm_type=?, memo=? WHERE id=?""",
        (d.get('comm_date',''), d.get('comm_type','방문'), d.get('memo',''), cid))
    conn.commit(); conn.close()
    return jsonify({'ok': True})


@app.route("/api/communication/report")
@login_required
def api_communication_report():
    """특정 매장의 소통 보고서 — 매출/최근흐름/소통이력 종합"""
    seller = request.args.get('seller', '').strip()
    year = request.args.get('year', str(datetime.now().year))
    if not seller:
        return jsonify({'ok': False, 'msg': '매장명 필요'}), 400

    conn = get_db()

    # 소통 이력
    comm_history = [dict(r) for r in conn.execute("""
        SELECT * FROM store_communication WHERE seller_name=? ORDER BY comm_date DESC
    """, (seller,)).fetchall()]

    # 연간 매출
    total_row = conn.execute("""
        SELECT SUM(total), SUM(quantity), COUNT(*) FROM sales_data
        WHERE real_seller=? AND sale_date LIKE ? AND sale_date!=''
    """, (seller, f"{year}%")).fetchone()
    year_total = total_row[0] or 0
    year_qty   = total_row[1] or 0
    year_cnt   = total_row[2] or 0

    # 월별 흐름
    monthly = [dict(mo=r[0], total=r[1] or 0, qty=r[2] or 0) for r in conn.execute("""
        SELECT CAST(strftime('%m',sale_date) AS INTEGER) mo, SUM(total), SUM(quantity)
        FROM sales_data WHERE real_seller=? AND sale_date LIKE ? AND sale_date!=''
        GROUP BY mo ORDER BY mo
    """, (seller, f"{year}%")).fetchall()]

    # 최근 판매 제품 TOP5 (최근 판매일 기준)
    top_items = [dict(name=normalize_item_name(r[0]), group=r[1], total=r[2] or 0, qty=r[3] or 0, last=r[4])
                 for r in conn.execute("""
        SELECT item_name, item_group, SUM(total), SUM(quantity), MAX(sale_date)
        FROM sales_data WHERE real_seller=? AND sale_date LIKE ? AND sale_date!=''
        GROUP BY item_name ORDER BY SUM(total) DESC LIMIT 8
    """, (seller, f"{year}%")).fetchall()]
    # 브랜드 정규화 + 같은 제품 합산
    merged_items = {}
    for it in top_items:
        b = remap_group(it['group'], '')
        key = it['name']
        if key not in merged_items:
            merged_items[key] = {'name': key, 'brand': b, 'total': 0, 'qty': 0, 'last': it['last']}
        merged_items[key]['total'] += it['total']
        merged_items[key]['qty']   += it['qty']
    top5 = sorted(merged_items.values(), key=lambda x: -x['total'])[:5]

    # 추세 계산
    vals = [m['total'] for m in monthly]
    if len(vals) >= 2:
        pct = round((vals[-1]-vals[-2])/vals[-2]*100, 1) if vals[-2] else 0
        direction = '상승' if pct >= 10 else '하락' if pct <= -10 else '안정'
    else:
        pct = 0; direction = '데이터 부족'

    conn.close()
    return jsonify({
        'ok': True,
        'seller': seller,
        'year': year,
        'comm_history': comm_history,
        'year_total': year_total, 'year_qty': year_qty, 'year_cnt': year_cnt,
        'monthly': monthly, 'top5': top5,
        'trend_pct': pct, 'trend_direction': direction,
    })


@app.route("/api/export/xlsx/communication")
@login_required
def api_export_communication_xlsx():
    """매장과의 소통 엑셀 — 소통이력 + 매출/최근흐름 포함, 심플 보고용"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    seller = request.args.get('seller', '').strip()
    year = request.args.get('year', str(datetime.now().year))

    conn = get_db()
    if seller:
        sellers_to_export = [seller]
    else:
        sellers_to_export = [r[0] for r in conn.execute(
            "SELECT DISTINCT seller_name FROM store_communication ORDER BY seller_name").fetchall()]

    wb = openpyxl.Workbook()
    ws = wb.active
    FNAME = '맑은 고딕'
    def mf(h): return PatternFill("solid", fgColor=h)
    thin = Side(style='thin', color='E5E7EB')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    rgt  = Alignment(horizontal='right', vertical='center')

    first_sheet = True
    for s_name in sellers_to_export:
        if not s_name: continue
        if first_sheet:
            ws.title = s_name[:28]
            first_sheet = False
        else:
            ws = wb.create_sheet(title=s_name[:28])

        # 매장 매출 정보
        total_row = conn.execute("""
            SELECT SUM(total), SUM(quantity), COUNT(*) FROM sales_data
            WHERE real_seller=? AND sale_date LIKE ? AND sale_date!=''
        """, (s_name, f"{year}%")).fetchone()
        year_total = total_row[0] or 0; year_qty = total_row[1] or 0

        monthly = conn.execute("""
            SELECT CAST(strftime('%m',sale_date) AS INTEGER) mo, SUM(total), SUM(quantity)
            FROM sales_data WHERE real_seller=? AND sale_date LIKE ? AND sale_date!=''
            GROUP BY mo ORDER BY mo
        """, (s_name, f"{year}%")).fetchall()

        top_rows = conn.execute("""
            SELECT item_name, item_group, SUM(total) t, SUM(quantity) q
            FROM sales_data WHERE real_seller=? AND sale_date LIKE ? AND sale_date!=''
            GROUP BY item_name ORDER BY t DESC LIMIT 10
        """, (s_name, f"{year}%")).fetchall()
        merged = {}
        for r in top_rows:
            nm = normalize_item_name(r[0]); b = remap_group(r[1], '')
            if nm not in merged: merged[nm] = {'brand': b, 'total': 0, 'qty': 0}
            merged[nm]['total'] += r[2] or 0; merged[nm]['qty'] += r[3] or 0
        top5 = sorted(merged.items(), key=lambda x: -x[1]['total'])[:5]

        comm_rows = conn.execute("""
            SELECT comm_date, comm_type, memo, created_by FROM store_communication
            WHERE seller_name=? ORDER BY comm_date DESC
        """, (s_name,)).fetchall()

        # ── 타이틀 ──
        ws.merge_cells('A1:E1')
        c = ws.cell(row=1, column=1, value=f'{s_name}  —  매장 소통 보고서 ({year}년)')
        c.font=Font(bold=True, size=13, name=FNAME, color='1F2937'); c.alignment=ctr
        ws.row_dimensions[1].height = 26
        ri = 3

        # ── 매출 요약 ──
        c = ws.cell(row=ri, column=1, value='■ 매출 요약'); c.font=Font(bold=True, size=11, name=FNAME); ri += 1
        summary = [('연간 매출', f'{year_total:,}원'), ('연간 판매수량', f'{year_qty:,}개')]
        for label, val in summary:
            ws.cell(row=ri, column=1, value=label).font=Font(size=9, name=FNAME, color='6B7280')
            ws.cell(row=ri, column=2, value=val).font=Font(bold=True, size=10, name=FNAME)
            ri += 1
        ri += 1

        # ── 월별 흐름 ──
        c = ws.cell(row=ri, column=1, value='■ 월별 매출 흐름'); c.font=Font(bold=True, size=11, name=FNAME); ri += 1
        for ci, h in enumerate(['월','매출(원)','수량'], 1):
            c = ws.cell(row=ri, column=ci, value=h); c.font=Font(bold=True, size=9, name=FNAME, color='374151')
            c.fill=mf('F3F4F6'); c.border=bdr; c.alignment=ctr
        ri += 1
        for m in monthly:
            ws.cell(row=ri, column=1, value=f"{m[0]}월").alignment=ctr
            c2 = ws.cell(row=ri, column=2, value=m[1] or 0); c2.number_format='#,##0'; c2.alignment=rgt
            ws.cell(row=ri, column=3, value=m[2] or 0).alignment=ctr
            for ci in range(1,4): ws.cell(row=ri, column=ci).border=bdr
            ri += 1
        ri += 1

        # ── TOP5 제품 ──
        c = ws.cell(row=ri, column=1, value='■ 최근 판매 제품 TOP5'); c.font=Font(bold=True, size=11, name=FNAME); ri += 1
        for ci, h in enumerate(['제품명','브랜드','매출(원)','수량'], 1):
            c = ws.cell(row=ri, column=ci, value=h); c.font=Font(bold=True, size=9, name=FNAME, color='374151')
            c.fill=mf('F3F4F6'); c.border=bdr; c.alignment=ctr
        ri += 1
        for nm, v in top5:
            ws.cell(row=ri, column=1, value=nm).alignment=left
            ws.cell(row=ri, column=2, value=v['brand']).alignment=ctr
            c3 = ws.cell(row=ri, column=3, value=v['total']); c3.number_format='#,##0'; c3.alignment=rgt
            ws.cell(row=ri, column=4, value=v['qty']).alignment=ctr
            for ci in range(1,5): ws.cell(row=ri, column=ci).border=bdr
            ri += 1
        ri += 1

        # ── 소통 이력 ──
        c = ws.cell(row=ri, column=1, value=f'■ 소통 이력 ({len(comm_rows)}건)'); c.font=Font(bold=True, size=11, name=FNAME); ri += 1
        for ci, h in enumerate(['날짜','유형','메모','작성자'], 1):
            c = ws.cell(row=ri, column=ci, value=h); c.font=Font(bold=True, size=9, name=FNAME, color='374151')
            c.fill=mf('F3F4F6'); c.border=bdr; c.alignment=ctr
        ri += 1
        for cr in comm_rows:
            ws.cell(row=ri, column=1, value=cr[0]).alignment=ctr
            ws.cell(row=ri, column=2, value=cr[1]).alignment=ctr
            c3 = ws.cell(row=ri, column=3, value=cr[2]); c3.alignment=left
            ws.cell(row=ri, column=4, value=cr[3]).alignment=ctr
            for ci in range(1,5): ws.cell(row=ri, column=ci).border=bdr
            ws.row_dimensions[ri].height = 26
            ri += 1

        for col, w in zip('ABCDE', [16, 14, 42, 12, 10]):
            ws.column_dimensions[col].width = w

    conn.close()
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f'매장소통_{seller}_{year}.xlsx' if seller else f'매장소통_전체_{year}.xlsx'
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=fname)


# ── 매장 방문 보고서 API ────────────────────────────────
import re as _re_visit

def _copy_sheet_with_style(src_ws, dst_ws):
    """시트를 값+서식(폰트/배경색/테두리/정렬/병합/열너비/행높이)까지 통째로 복사"""
    from copy import copy as _copy_style
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = _copy_style(cell.font)
                new_cell.border = _copy_style(cell.border)
                new_cell.fill = _copy_style(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = _copy_style(cell.protection)
                new_cell.alignment = _copy_style(cell.alignment)
    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width:
            dst_ws.column_dimensions[col_letter].width = dim.width
        if dim.hidden:
            dst_ws.column_dimensions[col_letter].hidden = dim.hidden
    for row_idx, dim in src_ws.row_dimensions.items():
        if dim.height:
            dst_ws.row_dimensions[row_idx].height = dim.height
    for merged_range in src_ws.merged_cells.ranges:
        try: dst_ws.merge_cells(str(merged_range))
        except Exception: pass
    try:
        dst_ws.freeze_panes = src_ws.freeze_panes
    except Exception:
        pass
    try:
        dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    except Exception:
        pass


def _extract_single_sheet_xlsx_b64(src_ws):
    """특정 시트 하나를 서식 그대로 유지한 채 독립된 xlsx 파일(base64)로 추출"""
    import base64
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = (src_ws.title or 'Sheet1')[:31]
    _copy_sheet_with_style(src_ws, new_ws)
    buf = io.BytesIO()
    new_wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode('ascii')


def _parse_visit_report_sheet(ws, source_filename=''):
    """오프라인 매장 방문 보고서 시트를 파싱 — 두 가지 포맷(개별 파일형 / 집계파일 내 시트형) 모두 지원"""
    def cell(r, c):
        try: return ws.cell(r, c).value
        except: return None

    def find_label_row(label_keywords, max_row=20):
        """1열 또는 2열에서 라벨을 찾아 그 행 번호 반환"""
        for ri in range(1, min(max_row, ws.max_row+1)):
            for ci in [2, 1]:
                v = cell(ri, ci)
                if v and any(kw in str(v) for kw in label_keywords):
                    return ri
        return None

    def find_value_near(label_row, label_keywords):
        """해당 라벨이 있는 행에서, 라벨 뒤쪽 열들 중 첫 non-null 값을 찾음"""
        if not label_row: return ''
        for ci in range(1, min(ws.max_column+1, 20)):
            v = cell(label_row, ci)
            if v and any(kw in str(v) for kw in label_keywords):
                for cj in range(ci+1, min(ws.max_column+1, 20)):
                    v2 = cell(label_row, cj)
                    if v2 is not None and str(v2).strip():
                        return v2
        return ''

    # 작성일 / 방문일 (datetime 객체 또는 "5월 7일" 형태 텍스트 모두 지원)
    date_row = find_label_row(['작성일', '방문일'])
    visit_date_raw = find_value_near(date_row, ['방문일']) or find_value_near(date_row, ['작성일'])
    if hasattr(visit_date_raw, 'strftime'):
        visit_date = visit_date_raw.strftime('%Y-%m-%d')
    else:
        raw_str = str(visit_date_raw or '').strip()
        m = _re_visit.search(r'(\d{1,2})\s*월\s*(\d{1,2})\s*일', raw_str)
        if m:
            mo, dy = int(m.group(1)), int(m.group(2))
            # 연도 추정: 파일명에 연도가 있으면 사용, 없으면 현재 연도
            yr_m = _re_visit.search(r'(20\d{2})', source_filename)
            yr = int(yr_m.group(1)) if yr_m else datetime.now().year
            visit_date = f"{yr}-{mo:02d}-{dy:02d}"
        else:
            visit_date = raw_str[:10]

    # 매장명
    store_row = find_label_row(['매장명'])
    store_raw = str(find_value_near(store_row, ['매장명']) or '').strip()
    # "베이비하우스 청라점 (35위/100위) (※ 레카로 인센티브 매장)" 형태에서 순위/부가정보 분리
    store_rank_match = _re_visit.search(r'\((\d+위\s*/\s*\d+위)\)', store_raw)
    store_rank = store_rank_match.group(1) if store_rank_match else ''
    store_name = _re_visit.sub(r'\s*\([^)]*\)', '', store_raw).strip()

    # 담당자 / 작성자
    manager_row = find_label_row(['담당자'])
    manager = str(find_value_near(manager_row, ['담당자']) or '').strip()
    author_row = find_label_row(['작성자'])
    author = str(find_value_near(author_row, ['작성자']) or '').strip()

    # 지역 — 원본 텍스트("경기도" 등)보다 매장명 기반 표준 지역 분류를 우선 사용 (경기북부/남부 등 일관성)
    region = detect_region_from_name(store_name) or ''
    if not region:
        region_row = find_label_row(['지역'])
        region = str(find_value_near(region_row, ['지역']) or '').strip()

    # 브랜드 추정 (매장명 기반)
    brand = ''
    for b in ['베이비하우스', '링크맘', '베이비파크', '베네피아', '베이비세븐', '베이비스토리', '베이비스토어']:
        if b in store_name:
            brand = b; break

    # 직원현황 / 매장규모
    staff_row = find_label_row(['직원수', '직원현황'])
    staff_info = str(find_value_near(staff_row, ['직원수', '직원현황']) or '').strip()
    size_row = find_label_row(['매장규모', '매장현황'])
    store_size = str(find_value_near(size_row, ['매장규모', '매장현황']) or '').strip()

    # 주요 내용 (유모차/카시트/식탁의자/용품/웨건/기타 카테고리별 텍스트)
    content = {}
    content_start = None
    for ri in range(1, min(ws.max_row+1, 100)):
        v = cell(ri, 2)
        if v and ('주요 내용' in str(v) or '주요내용' in str(v)):
            content_start = ri; break
    if content_start:
        cur_cat = None
        stop_labels = ['진열 현황', '마케팅 요청', '영업지원 요청', '요청사항', '타사 프로모션', '후속조치']
        for ri in range(content_start+1, min(content_start+60, ws.max_row+1)):
            label = cell(ri, 2)
            if label and any(sl in str(label) for sl in stop_labels):
                break
            v_cat = cell(ri, 3)
            if v_cat and str(v_cat).strip() in ['유모차','카시트','식탁의자','용품','웨건','기타']:
                cur_cat = str(v_cat).strip()
                txt = cell(ri, 4)
                if txt and str(txt).strip():
                    content.setdefault(cur_cat, []).append(str(txt).strip())
            elif cur_cat:
                # 다음 카테고리 열(3번)이 비어있고, 텍스트가 4번 열에 이어지는 경우
                txt = cell(ri, 4) or cell(ri, 3)
                if txt and str(txt).strip():
                    content.setdefault(cur_cat, []).append(str(txt).strip())

    # 요청사항 (마케팅/영업지원) — 헤더 라벨 행은 제외하고 실제 요청 내용만 수집
    requests_list = []
    HEADER_JUNK = {'요청 내용', '요청자', '목적/배경', '우선순위', '담당부서', '희망 완료일',
                   '진행상태', '요청 여부', '없음', ''}
    for ri in range(1, ws.max_row+1):
        v = cell(ri, 2)
        if v and ('마케팅 요청' in str(v) or '영업지원 요청' in str(v) or
                  (str(v).strip().startswith(('3.','4.')) and '요청사항' in str(v))):
            for rj in range(ri+1, min(ri+15, ws.max_row+1)):
                nxt_label = cell(rj, 2)
                if nxt_label and any(sl in str(nxt_label) for sl in ['타사 프로모션', '후속조치', '영업지원 요청']):
                    break
                content_v = cell(rj, 4) or cell(rj, 3)
                if content_v and str(content_v).strip() and str(content_v).strip() not in HEADER_JUNK:
                    requests_list.append(str(content_v).strip())

    # 타사 프로모션
    promo_start = None
    for ri in range(1, ws.max_row+1):
        v = cell(ri, 2)
        if v and '타사 프로모션' in str(v):
            promo_start = ri; break
    promo_text = ''
    if promo_start:
        parts = []
        for ri in range(promo_start+1, min(promo_start+8, ws.max_row+1)):
            label = cell(ri, 2)
            if label and '후속조치' in str(label):
                break
            for ci in range(2, 6):
                v = cell(ri, ci)
                if v and str(v).strip() not in ('내용',):
                    parts.append(str(v).strip())
        promo_text = '\n'.join(parts)

    # 후속조치 및 메모 — 라벨 행 자체는 제외하고 그 다음 내용만
    followup_start = None
    for ri in range(1, ws.max_row+1):
        v = cell(ri, 2)
        if v and '후속조치' in str(v):
            followup_start = ri; break
    followup_text = ''
    if followup_start:
        parts = []
        for ri in range(followup_start+1, min(followup_start+30, ws.max_row+1)):
            for ci in range(2, 12):
                v = cell(ri, ci)
                if v and str(v).strip() and '후속조치 계획' not in str(v) and '기타 메모' not in str(v):
                    parts.append(str(v).strip())
        followup_text = '\n'.join(parts)

    if not visit_date or not store_name:
        return None

    # 원본 그대로 재현하기 위한 전체 그리드 + 병합셀 캡처 (엑셀 다운로드 시 원본 형태 복원용)
    raw_grid = []
    max_r = min(ws.max_row, 200)
    max_c = min(ws.max_column, 20)
    for ri in range(1, max_r+1):
        row_vals = []
        for ci in range(1, max_c+1):
            v = cell(ri, ci)
            if hasattr(v, 'strftime'):
                v = v.strftime('%Y-%m-%d')
            row_vals.append(v)
        raw_grid.append(row_vals)
    merged_ranges = [str(mr) for mr in ws.merged_cells.ranges]

    return {
        'visit_date': visit_date, 'store_name': store_name, 'brand': brand,
        'region': region or '', 'manager': manager, 'author': author,
        'store_rank': store_rank, 'staff_info': staff_info, 'store_size': store_size,
        'content': content, 'requests': requests_list,
        'followup_text': (promo_text + '\n' + followup_text).strip(),
        'source_filename': source_filename,
        'raw_grid': raw_grid, 'merged_cells': merged_ranges,
        'sheet_title': ws.title,
    }


@app.route("/api/visit-report/upload", methods=["POST"])
@login_required
def api_visit_report_upload():
    """매장 방문 보고서 업로드 — 단일 xlsx, 여러 xlsx, 또는 zip(여러 xlsx 포함) 모두 지원"""
    if 'file' not in request.files:
        return jsonify({'ok': False, 'msg': '파일이 없습니다'}), 400

    files = request.files.getlist('file')
    conn = get_db()
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
    inserted, updated, skipped = 0, 0, 0
    errors = []

    def _process_workbook(wb, fname):
        nonlocal inserted, updated, skipped
        for sh_name in wb.sheetnames:
            if sh_name.strip() in ('매장별 방문현황',):
                continue  # 집계 시트는 건너뜀 (개별 보고서 시트만 파싱)
            ws = wb[sh_name]
            try:
                parsed = _parse_visit_report_sheet(ws, fname)
            except Exception as e:
                errors.append(f"{fname}/{sh_name}: {e}")
                continue
            if not parsed:
                skipped += 1
                continue
            try:
                raw_xlsx_b64 = _extract_single_sheet_xlsx_b64(ws)
            except Exception as e:
                raw_xlsx_b64 = ''
                errors.append(f"{fname}/{sh_name} 서식 저장 실패: {e}")
            existing = conn.execute(
                "SELECT id FROM store_visit_report WHERE visit_date=? AND store_name=?",
                (parsed['visit_date'], parsed['store_name'])).fetchone()
            content_json = json.dumps(parsed['content'], ensure_ascii=False)
            request_json = json.dumps(parsed['requests'], ensure_ascii=False)
            raw_grid_json = json.dumps(parsed.get('raw_grid', []), ensure_ascii=False, default=str)
            merged_cells_json = json.dumps(parsed.get('merged_cells', []), ensure_ascii=False)
            sheet_title = parsed.get('sheet_title', '')
            if existing:
                conn.execute("""UPDATE store_visit_report SET
                    brand=?, region=?, manager=?, author=?, store_rank=?, staff_info=?, store_size=?,
                    content_json=?, request_json=?, followup_text=?, source_filename=?, uploaded_at=?,
                    raw_grid_json=?, merged_cells_json=?, sheet_title=?, raw_xlsx_b64=?
                    WHERE id=?""",
                    (parsed['brand'], parsed['region'], parsed['manager'], parsed['author'],
                     parsed['store_rank'], parsed['staff_info'], parsed['store_size'],
                     content_json, request_json, parsed['followup_text'], parsed['source_filename'],
                     now_str, raw_grid_json, merged_cells_json, sheet_title, raw_xlsx_b64, existing[0]))
                updated += 1
            else:
                conn.execute("""INSERT INTO store_visit_report
                    (visit_date, store_name, brand, region, manager, author, store_rank,
                     staff_info, store_size, content_json, request_json, followup_text,
                     source_filename, uploaded_at, raw_grid_json, merged_cells_json, sheet_title, raw_xlsx_b64)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (parsed['visit_date'], parsed['store_name'], parsed['brand'], parsed['region'],
                     parsed['manager'], parsed['author'], parsed['store_rank'], parsed['staff_info'],
                     parsed['store_size'], content_json, request_json, parsed['followup_text'],
                     parsed['source_filename'], now_str, raw_grid_json, merged_cells_json, sheet_title,
                     raw_xlsx_b64))
                inserted += 1

    for f in files:
        fname = f.filename or ''
        data = f.read()
        try:
            if fname.lower().endswith('.zip'):
                import zipfile
                with zipfile.ZipFile(io.BytesIO(data)) as z:
                    for inner_name in z.namelist():
                        if not inner_name.lower().endswith(('.xlsx', '.xls')):
                            continue
                        inner_data = z.read(inner_name)
                        try:
                            wb = openpyxl.load_workbook(io.BytesIO(inner_data), data_only=True)
                            _process_workbook(wb, inner_name)
                        except Exception as e:
                            errors.append(f"{inner_name}: {e}")
            else:
                wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
                _process_workbook(wb, fname)
        except Exception as e:
            errors.append(f"{fname}: {e}")

    conn.commit(); conn.close()
    return jsonify({'ok': True, 'inserted': inserted, 'updated': updated, 'skipped': skipped,
                     'errors': errors[:10]})


@app.route("/api/visit-report/list")
@login_required
def api_visit_report_list():
    """방문 보고서 목록 조회 — 날짜별/매장별 필터, 정렬, 키워드 검색 지원"""
    store = request.args.get('store', '').strip()
    brand = request.args.get('brand', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    keyword = request.args.get('keyword', '').strip()   # 수정: 보고서 전체 내용 키워드 검색
    group_by = request.args.get('group_by', 'date')  # 'date' or 'store'

    conn = get_db()
    q = "SELECT * FROM store_visit_report WHERE 1=1"
    params = []
    if store: q += " AND store_name LIKE ?"; params.append(f"%{store}%")
    if brand: q += " AND brand=?"; params.append(brand)
    if date_from: q += " AND visit_date>=?"; params.append(date_from)
    if date_to: q += " AND visit_date<=?"; params.append(date_to)
    if keyword:
        # 매장명/담당자/주요내용/요청사항/후속조치/직원현황/매장규모 전체에서 검색
        q += """ AND (
            store_name LIKE ? OR manager LIKE ? OR author LIKE ?
            OR content_json LIKE ? OR request_json LIKE ? OR followup_text LIKE ?
            OR staff_info LIKE ? OR store_size LIKE ?
        )"""
        kw = f"%{keyword}%"
        params += [kw, kw, kw, kw, kw, kw, kw, kw]
    q += " ORDER BY visit_date DESC" if group_by == 'date' else " ORDER BY store_name, visit_date DESC"
    rows = [dict(r) for r in conn.execute(q, params).fetchall()]
    conn.close()

    for r in rows:
        try: r['content'] = json.loads(r.get('content_json') or '{}')
        except: r['content'] = {}
        try: r['requests'] = json.loads(r.get('request_json') or '[]')
        except: r['requests'] = []
        # 수정: 검색어가 어떤 항목에 매치됐는지 프론트에서 하이라이트/미리보기용으로 표시
        if keyword:
            r['keyword_match'] = _find_keyword_context(r, keyword)

    return jsonify(rows)


def _find_keyword_context(report, keyword):
    """검색어가 매치된 위치와 주변 문맥(스니펫)을 찾아 반환 — 검색 결과 미리보기용"""
    kw_lower = keyword.lower()
    fields_to_check = [
        ('매장명', report.get('store_name', '')),
        ('담당자', report.get('manager', '')),
        ('작성자', report.get('author', '')),
        ('직원현황', report.get('staff_info', '')),
        ('매장규모', report.get('store_size', '')),
        ('후속조치/타사프로모션', report.get('followup_text', '')),
    ]
    for cat, texts in (report.get('content') or {}).items():
        fields_to_check.append((f'주요내용({cat})', ' '.join(texts) if isinstance(texts, list) else str(texts)))
    for req in (report.get('requests') or []):
        fields_to_check.append(('요청사항', req))

    for label, text in fields_to_check:
        if text and kw_lower in str(text).lower():
            text = str(text)
            idx = text.lower().find(kw_lower)
            start = max(0, idx - 20)
            end = min(len(text), idx + len(keyword) + 30)
            snippet = ('...' if start > 0 else '') + text[start:end] + ('...' if end < len(text) else '')
            return {'field': label, 'snippet': snippet}
    return {'field': '', 'snippet': ''}


@app.route("/api/visit-report/stores")
@login_required
def api_visit_report_stores():
    """매장별 방문 누적 현황 (요청하신 '매장별 방문현황' 첫 탭 형태 재현)"""
    brand = request.args.get('brand', '').strip()
    conn = get_db()
    q = """SELECT store_name, brand, region, COUNT(*) cnt,
                  GROUP_CONCAT(visit_date) dates, MAX(visit_date) last_date
           FROM store_visit_report WHERE 1=1"""
    params = []
    if brand: q += " AND brand=?"; params.append(brand)
    q += " GROUP BY store_name, brand ORDER BY cnt DESC, store_name"
    rows = [dict(r) for r in conn.execute(q, params).fetchall()]
    conn.close()
    for r in rows:
        dates = sorted((r['dates'] or '').split(','))
        r['dates_list'] = dates
        r['dates'] = ', '.join(d[5:].replace('-', '/') for d in dates if d)
    return jsonify(rows)


@app.route("/api/visit-report/<int:rid>")
@login_required
def api_visit_report_detail(rid):
    """방문 보고서 상세 조회"""
    conn = get_db()
    row = conn.execute("SELECT * FROM store_visit_report WHERE id=?", (rid,)).fetchone()
    conn.close()
    if not row:
        return jsonify({'ok': False, 'msg': '찾을 수 없습니다'}), 404
    r = dict(row)
    try: r['content'] = json.loads(r.get('content_json') or '{}')
    except: r['content'] = {}
    try: r['requests'] = json.loads(r.get('request_json') or '[]')
    except: r['requests'] = []
    return jsonify(r)


@app.route("/api/visit-report/<int:rid>", methods=["DELETE"])
@login_required
def api_visit_report_delete(rid):
    conn = get_db()
    conn.execute("DELETE FROM store_visit_report WHERE id=?", (rid,))
    conn.commit(); conn.close()
    return jsonify({'ok': True})


@app.route("/api/visit-report/brands")
@login_required
def api_visit_report_brands():
    conn = get_db()
    rows = conn.execute("SELECT DISTINCT brand FROM store_visit_report WHERE brand!='' ORDER BY brand").fetchall()
    conn.close()
    return jsonify([r[0] for r in rows])


def _write_dashboard_sheet(ws, rows, FNAME, mf, bdr, ctr, left):
    """매장 방문 현황 대시보드 시트 — 옅은 회색 헤더, A열 여백의 심플한 디자인"""
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    # 브랜드별로 매장 그룹핑 + 매장별 방문 집계
    by_brand_store = {}
    for r in rows:
        brand = r['brand'] or '기타'
        key = (brand, r['store_name'])
        if key not in by_brand_store:
            by_brand_store[key] = {'dates': [], 'sheet_tabs': []}
        by_brand_store[key]['dates'].append(r['visit_date'])
        by_brand_store[key]['sheet_tabs'].append(r.get('sheet_title') or f"{r['store_name']}_{r['visit_date'][5:].replace('-','')}")

    brands = sorted(set(b for b, s in by_brand_store.keys()))
    HEADER_GRAY = 'F3F4F6'  # 옅은 회색 헤더 (진한 파란색 대신 절제된 톤)

    # A열은 여백(spacer)로 비워두고 B열부터 시작
    ws.column_dimensions['A'].width = 2

    ri = 1
    for brand in brands:
        stores_in_brand = [(s, v) for (b, s), v in by_brand_store.items() if b == brand]
        stores_in_brand.sort(key=lambda x: -len(x[1]['dates']))

        ws.merge_cells(f'B{ri}:F{ri}')
        c = ws.cell(row=ri, column=2, value=f"{brand} 매장별 방문 현황")
        c.font = Font(bold=True, size=13, name=FNAME, color='1F2937'); c.alignment = left
        ws.row_dimensions[ri].height = 26
        ri += 1

        headers = ['No.', '매장명', '방문 횟수', '방문일', '시트 탭']
        for ci, h in enumerate(headers, 2):
            c = ws.cell(row=ri, column=ci, value=h)
            c.font = Font(bold=True, size=9, name=FNAME, color='374151'); c.fill = mf(HEADER_GRAY); c.border = bdr; c.alignment = ctr
        ws.row_dimensions[ri].height = 20
        ri += 1

        for i, (store_name, v) in enumerate(stores_in_brand, 1):
            dates_sorted = sorted(v['dates'])
            dates_str = ', '.join(d[5:].replace('-', '/') for d in dates_sorted)
            tabs_str = ', '.join(v['sheet_tabs'])
            row_vals = [i, store_name, f"{len(v['dates'])}회", dates_str, tabs_str]
            for ci, val in enumerate(row_vals, 2):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = Font(size=9, name=FNAME, color='1F2937')
                c.border = bdr
                c.alignment = ctr if ci in (2,4) else left
            ws.row_dimensions[ri].height = 16
            ri += 1
        ri += 2  # 브랜드 간 여백

    for ci, w in zip(range(2,7), [6, 22, 10, 26, 46]):
        ws.column_dimensions[get_column_letter(ci)].width = w


def _write_raw_grid_sheet(ws, raw_grid, merged_cells, FNAME):
    """원본 업로드 파일의 셀 값·병합구조를 최대한 그대로 재현 (원본 그대로 보기용)"""
    from openpyxl.styles import Font, Alignment, PatternFill
    if not raw_grid:
        ws.cell(row=1, column=1, value='원본 데이터가 없습니다')
        return
    for ri, row in enumerate(raw_grid, 1):
        for ci, val in enumerate(row, 1):
            if val is None: continue
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(size=9, name=FNAME)
            c.alignment = Alignment(vertical='top', wrap_text=True)
    for mc in (merged_cells or []):
        try: ws.merge_cells(mc)
        except Exception: pass
    # 첫 행(타이틀), 섹션 헤더(예: "1. 기본 정보")는 굵게 강조
    for ri, row in enumerate(raw_grid[:3], 1):
        for ci in range(1, len(row)+1):
            c = ws.cell(row=ri, column=ci)
            if c.value:
                c.font = Font(bold=True, size=12 if ri==1 else 10, name=FNAME, color='1F2937')
    for ri, row in enumerate(raw_grid, 1):
        v = row[1] if len(row) > 1 else None
        if v and isinstance(v, str) and _re_visit.match(r'^\d\.\s', v):
            for ci in range(1, len(row)+1):
                c = ws.cell(row=ri, column=ci)
                c.font = Font(bold=True, size=10, name=FNAME, color='2563EB')
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 16
    for col_letter in ['C','D','E','F']:
        ws.column_dimensions[col_letter].width = 22


@app.route("/api/export/xlsx/visit-report")
@login_required
def api_export_visit_report_xlsx():
    """방문 보고서 엑셀 다운로드 — 브랜드별 대시보드 요약 + 원본 그대로의 상세 시트 구성"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    store = request.args.get('store', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()

    conn = get_db()
    q = "SELECT * FROM store_visit_report WHERE 1=1"
    params = []
    if store: q += " AND store_name LIKE ?"; params.append(f"%{store}%")
    if date_from: q += " AND visit_date>=?"; params.append(date_from)
    if date_to: q += " AND visit_date<=?"; params.append(date_to)
    q += " ORDER BY brand, visit_date"
    rows = [dict(r) for r in conn.execute(q, params).fetchall()]
    conn.close()

    if not rows:
        return jsonify({'ok': False, 'msg': '조건에 맞는 방문 보고서가 없습니다'}), 404

    FNAME = '맑은 고딕'
    def mf(h): return PatternFill("solid", fgColor=h)
    thin = Side(style='thin', color='E5E7EB')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='top', wrap_text=True)

    wb = openpyxl.Workbook()

    # 1번 시트: 브랜드별 방문현황 대시보드 (스크린샷 스타일)
    ws_dash = wb.active; ws_dash.title = '매장별_방문현황'
    _write_dashboard_sheet(ws_dash, rows, FNAME, mf, bdr, ctr, left)

    # 이후 시트: 각 방문건을 원본 그대로(서식까지) 재현
    used_names = set()
    for r in rows:
        base_name = f"{r['visit_date'][5:].replace('-','')}_{r['store_name']}"[:28]
        sheet_name = base_name
        suffix = 1
        while sheet_name in used_names:
            sheet_name = f"{base_name}_{suffix}"[:31]
            suffix += 1
        used_names.add(sheet_name)
        ws = wb.create_sheet(title=sheet_name)

        raw_xlsx_b64 = r.get('raw_xlsx_b64')
        if raw_xlsx_b64:
            try:
                import base64
                src_bytes = base64.b64decode(raw_xlsx_b64)
                src_wb = openpyxl.load_workbook(io.BytesIO(src_bytes))
                src_ws = src_wb[src_wb.sheetnames[0]]
                _copy_sheet_with_style(src_ws, ws)
                continue
            except Exception:
                pass
        # 서식 원본이 없는 구버전 데이터 — 값만이라도 표시
        try:
            raw_grid = json.loads(r.get('raw_grid_json') or '[]')
            merged_cells = json.loads(r.get('merged_cells_json') or '[]')
        except Exception:
            raw_grid, merged_cells = [], []
        if raw_grid:
            _write_raw_grid_sheet(ws, raw_grid, merged_cells, FNAME)
        else:
            ws.cell(row=1, column=1, value=f"{r['store_name']} — {r['visit_date']}").font = Font(bold=True, size=12, name=FNAME)
            ws.cell(row=3, column=1, value='원본 데이터가 저장되지 않은 보고서입니다. 다시 업로드해주세요.').font = Font(size=9, name=FNAME, color='9CA3AF')

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"방문보고서_{store or '전체'}_{date_from or ''}~{date_to or ''}.xlsx"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=fname)


@app.route("/api/export/xlsx/visit-report/by-store", methods=["POST"])
@login_required
def api_export_visit_report_by_store():
    """선택한 매장들만 원본 그대로 시트 분리하여 엑셀 다운로드"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    d = request.json or {}
    store_names = d.get('store_names', [])
    if not store_names:
        return jsonify({'ok': False, 'msg': '매장을 선택해주세요'}), 400

    conn = get_db()
    placeholders = ','.join('?' for _ in store_names)
    rows = [dict(r) for r in conn.execute(
        f"SELECT * FROM store_visit_report WHERE store_name IN ({placeholders}) ORDER BY store_name, visit_date",
        store_names).fetchall()]
    conn.close()

    if not rows:
        return jsonify({'ok': False, 'msg': '해당 매장의 방문 보고서가 없습니다'}), 404

    FNAME = '맑은 고딕'
    def mf(h): return PatternFill("solid", fgColor=h)
    thin = Side(style='thin', color='E5E7EB')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='top', wrap_text=True)

    wb = openpyxl.Workbook()
    ws_dash = wb.active; ws_dash.title = '매장별_방문현황'
    _write_dashboard_sheet(ws_dash, rows, FNAME, mf, bdr, ctr, left)

    used_names = set()
    for r in rows:
        base_name = f"{r['store_name']}_{r['visit_date'][5:].replace('-','')}"[:28]
        sheet_name = base_name
        suffix = 1
        while sheet_name in used_names:
            sheet_name = f"{base_name}_{suffix}"[:31]; suffix += 1
        used_names.add(sheet_name)
        ws = wb.create_sheet(title=sheet_name)

        raw_xlsx_b64 = r.get('raw_xlsx_b64')
        if raw_xlsx_b64:
            try:
                import base64
                src_bytes = base64.b64decode(raw_xlsx_b64)
                src_wb = openpyxl.load_workbook(io.BytesIO(src_bytes))
                src_ws = src_wb[src_wb.sheetnames[0]]
                _copy_sheet_with_style(src_ws, ws)
                continue
            except Exception:
                pass
        try:
            raw_grid = json.loads(r.get('raw_grid_json') or '[]')
            merged_cells = json.loads(r.get('merged_cells_json') or '[]')
        except Exception:
            raw_grid, merged_cells = [], []
        if raw_grid:
            _write_raw_grid_sheet(ws, raw_grid, merged_cells, FNAME)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"방문보고서_매장별_{'_'.join(store_names[:3])}.xlsx"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=fname)


@app.route("/api/visit-report/requests")
@login_required
def api_visit_report_requests():
    """전체 요청사항 상세 목록 (매장/날짜 컨텍스트 포함)"""
    brand = request.args.get('brand', '').strip()
    conn = get_db()
    q = "SELECT id, visit_date, store_name, brand, manager, request_json FROM store_visit_report WHERE request_json!='[]' AND request_json!=''"
    params = []
    if brand: q += " AND brand=?"; params.append(brand)
    q += " ORDER BY visit_date DESC"
    rows = conn.execute(q, params).fetchall()
    conn.close()

    result = []
    for r in rows:
        try: reqs = json.loads(r[5] or '[]')
        except: reqs = []
        for req in reqs:
            result.append({
                'report_id': r[0], 'visit_date': r[1], 'store_name': r[2],
                'brand': r[3], 'manager': r[4], 'request_text': req,
            })
    return jsonify(result)


@app.route("/api/export/xlsx/visit-report/requests")
@login_required
def api_export_visit_report_requests_xlsx():
    """누적 요청사항 전체를 엑셀로 다운로드"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    conn = get_db()
    rows = conn.execute("""
        SELECT visit_date, store_name, brand, manager, request_json FROM store_visit_report
        WHERE request_json!='[]' AND request_json!='' ORDER BY visit_date DESC""").fetchall()
    conn.close()

    FNAME = '맑은 고딕'
    def mf(h): return PatternFill("solid", fgColor=h)
    thin = Side(style='thin', color='E5E7EB')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '누적_요청사항'
    ws.merge_cells('A1:E1')
    c = ws.cell(row=1, column=1, value='누적 요청사항 전체 현황')
    c.font = Font(bold=True, size=13, name=FNAME, color='1F2937'); c.alignment = ctr
    ws.row_dimensions[1].height = 26

    headers = ['방문일', '매장명', '브랜드', '담당자', '요청 내용']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = Font(bold=True, size=9, name=FNAME, color='374151'); c.fill = mf('F3F4F6'); c.border = bdr; c.alignment = ctr
        ws.column_dimensions[get_column_letter(ci)].width = [12, 20, 12, 10, 60][ci-1]
    ws.row_dimensions[2].height = 20

    ri = 3
    for r in rows:
        try: reqs = json.loads(r[4] or '[]')
        except: reqs = []
        for req in reqs:
            vals = [r[0], r[1], r[2], r[3], req]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.font = Font(size=9, name=FNAME); c.border = bdr
                c.alignment = ctr if ci != 5 else left
            ri += 1

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name='누적_요청사항.xlsx')


# ── 타사 비교 API ────────────────────────────────
COMPETITOR_CATEGORIES = {
    '유모차': ['실버크로스','잉글레시나','뉴나','조이','부가부','시크','미마','스토케','타보','리안','에그','오르빗','오이스터','와이업'],
    '카시트': ['브라이택스','싸이벡스','맥시코시','뉴나','조이','다이치','순성'],
    '식탁의자': ['스토케','야마토야','싸이벡스','이케아','본베베','시디즈','아가드','빼그빼레고','호크베타'],
    '웨건': ['폼포라','킨즈'],
}

# 카테고리 → 자사 item_group 매핑 (sales_data 기준)
CATEGORY_TO_ITEM_GROUP = {
    '유모차':   ['유모차'],
    '카시트':   ['컨버터블카시트','주니어카시트','휴대용부스터'],
    '식탁의자': ['식탁의자','하이체어'],
    '웨건':     [],  # item_name으로 원더폴드 필터링
}

# 자사 제품 유형 분류 (사용자 제공 정확한 매핑)
OUR_PRODUCT_TYPE = {
    '유모차': {
        '데이5': '디럭스', '지오3': '디럭스',
        '허브2': '절충형',
        '에어2': '휴대용',
        '루프트': '휴대용',
        '슈타트듀오': '쌍둥이',
    },
    '카시트': {
        '제논1': '컨버터블', '토론1': '컨버터블', '버디': '컨버터블',
        '벨릭스': '토들러',
        '액시언1': '주니어', '폴디': '주니어',
    },
}
PRODUCT_TYPE_ORDER = {
    '유모차': ['디럭스', '절충형', '휴대용', '쌍둥이'],
    '카시트': ['컨버터블', '주니어', '토들러'],
}


def _naver_search(endpoint, query, display=10, sort='sim'):
    import urllib.request, urllib.parse
    client_id     = os.environ.get('NAVER_CLIENT_ID',     'InqUUQfvWZN1rAZM4whk')
    client_secret = os.environ.get('NAVER_CLIENT_SECRET', 'fXYMLK1N1X')
    url = (f'https://openapi.naver.com/v1/search/{endpoint}.json?query='
           + urllib.parse.quote(query) + f'&display={display}&sort={sort}')
    req = urllib.request.Request(url)
    req.add_header('X-Naver-Client-Id', client_id)
    req.add_header('X-Naver-Client-Secret', client_secret)
    with urllib.request.urlopen(req, timeout=10) as r:
        return json.loads(r.read().decode('utf-8'))


def _strip_tags(s):
    import re as _re6
    return _re6.sub('<[^>]+>', '', s or '')


@app.route("/api/competitor/categories")
@login_required
def api_competitor_categories():
    """품목 카테고리 + 타사 브랜드 목록"""
    return jsonify({cat: brands for cat, brands in COMPETITOR_CATEGORIES.items()})


@app.route("/api/competitor/our-products")
@login_required
def api_competitor_our_products():
    """카테고리별 자사 제품 목록 (판매 데이터 기반) — 브랜드/유형 태그 포함"""
    import re as _re
    category = request.args.get('category', '')
    conn = get_db()
    groups = CATEGORY_TO_ITEM_GROUP.get(category, [])

    if category == '웨건':
        rows = conn.execute("""
            SELECT item_name, item_group, SUM(total) total, SUM(quantity) qty
            FROM sales_data WHERE item_name LIKE '%원더폴드%'
            GROUP BY item_name ORDER BY total DESC
        """).fetchall()
    elif groups:
        placeholders = ','.join('?' for _ in groups)
        rows = conn.execute(f"""
            SELECT item_name, item_group, SUM(total) total, SUM(quantity) qty
            FROM sales_data WHERE item_group IN ({placeholders})
            GROUP BY item_name ORDER BY total DESC
        """, groups).fetchall()
    else:
        rows = []
    conn.close()

    # 카테고리별 제외 키워드 (구형/단종/오분류/타 카테고리 제품)
    EXCLUDE_KEYWORDS = {
        '카시트': ['하이브리드'],
        '유모차': ['원더폴드', '에어cot', '에어Cot', 'AIRCOT', '이지라이프',
                  '에어+', '에어플러스', '허브+', '허브플러스'],  # 단종 제품
    }
    exclude_kw = EXCLUDE_KEYWORDS.get(category, [])
    type_map = OUR_PRODUCT_TYPE.get(category, {})

    products = []
    for r in rows:
        item_name = r[0] or ''
        if any(kw.lower() in item_name.lower() for kw in exclude_kw):
            continue
        brand = remap_group(r[1], item_name)
        norm = normalize_item_name(item_name)
        pname = _re.sub(r'^\[[^\]]+\]', '', norm).split('_')[0].strip()
        products.append({'name': pname, 'brand': brand, 'full_name': norm,
                          'total': r[2] or 0, 'qty': r[3] or 0,
                          'product_type': type_map.get(pname, '')})
    merged = {}
    for p in products:
        key = (p['brand'], p['name'])
        if key not in merged:
            merged[key] = p
        else:
            merged[key]['total'] += p['total']
            merged[key]['qty']   += p['qty']
    result = sorted(merged.values(), key=lambda x: -x['total'])
    return jsonify(result)


@app.route("/api/competitor/product-candidates", methods=["POST"])
@login_required
def api_competitor_product_candidates():
    """타사 브랜드의 구체적 제품명 후보 검색 (쇼핑 검색 기반 제품 라인업 추출)"""
    import re as _re7
    d = request.json or {}
    category = d.get('category', '').strip()
    brand = d.get('brand', '').strip()
    if not category or not brand:
        return jsonify({'ok': False, 'msg': '품목과 브랜드를 선택해주세요'}), 400

    # 이미 캐시된 제품이 있으면 우선 반환
    conn = get_db()
    cached = [dict(r) for r in conn.execute("""
        SELECT product_name, price_text, fetched_at FROM product_research
        WHERE side='competitor' AND category=? AND brand=?
        ORDER BY fetched_at DESC
    """, (category, brand)).fetchall()]
    conn.close()

    candidates = set()
    # 완전 제외 대상 키워드 (판매채널/부가서비스/이벤트성/부속품/후기 관련 — 제품명이 아님)
    NOISE_WORDS = ['네이버', '단독', '대여', '렌탈', '렌트', '제주', '무료', '후방거울', '후방미러',
                   '발판', '스텝', '본사', '판매', '직영', '배송', '무이자', '할부', '사은품',
                   '이벤트', '특가', '쿠폰', '적립', '스토어', '공식몰', '카카오', '톡딜',
                   '로켓배송', '당일발송', '무료배송', '전국', '설치', 'AS', '보증',
                   # 수정5: 추가 요청
                   '유아용', '범용', '전용', '보관', '햇빛가리개', '차양막', '캐노피',
                   '포토후기', '프리미엄', '호환', '악세서리', '액세서리', '거치대', '컵홀더',
                   '매트', '카시트커버', '방수커버', '우비', '비닐커버', '수납백', '수납가방',
                   '중고', '리퍼', '전시', '체험단', '서포터즈', '리뷰단', '기자단',
                   '점', '점포', '매장', '센터', '지점',  # 매장명 자체 노출 방지
                   '정가', '최저가', '가격비교', '할인', '세일']
    # 매장/지점명 패턴 (예: "현대판교점", "이마트양산점" 등 N글자+점/지점/매장 형태)
    STORE_PATTERN = _re7.compile(r'[가-힣A-Za-z]+(점|지점|매장|센터|백화점|아울렛)$')

    try:
        shop = _naver_search('shop', f"{brand} {category}", display=20, sort='sim')
        for item in shop.get('items', []):
            title = _strip_tags(item.get('title',''))
            # 노이즈 단어가 포함된 제목은 통째로 건너뜀 (판매채널 홍보 문구일 가능성 높음)
            if any(nw in title for nw in NOISE_WORDS):
                continue
            # 브랜드명 뒤에 오는 모델명 추출 (예: "실버크로스 니아 유모차" -> "니아")
            t = title.replace(brand, '').strip()
            # 흔한 잡단어 제거
            for junk in [category, '유모차', '카시트', '식탁의자', '웨건', '정품', '공식', '풀박스', '풀세트', '신상', '신제품']:
                t = t.replace(junk, '')
            # 특수문자 정리 후 앞 2단어 정도만 모델명 후보로
            t = _re7.sub(r'[\[\]()【】/|,·\-–]', ' ', t).strip()
            words = [w for w in t.split() if len(w) >= 2][:2]
            if not words:
                continue
            model = ' '.join(words)
            # 최종 검증: 노이즈 단어 포함 여부, 매장명 패턴, 길이
            if not (1 < len(model) <= 20):
                continue
            if any(nw in model for nw in NOISE_WORDS):
                continue
            if any(STORE_PATTERN.match(w) for w in words):
                continue
            candidates.add(model)
    except Exception:
        pass

    return jsonify({
        'ok': True,
        'cached_products': cached,
        'candidates': sorted(candidates)[:12],
    })


def _clean_naver_snippet(s):
    """네이버 검색 스니펫의 중간 생략(...) 흔적을 제거하고, 문장부호 기준 완결된 조각만 남김"""
    import re as _re9
    if not s: return []
    # 네이버는 검색어 주변만 잘라 이어붙이며 "..."로 구간을 표시함 → "..." 기준으로 쪼개고
    # 각 조각에서 완결된 문장(마침표로 끝나는)만 취함
    parts = _re9.split(r'\.{2,}|…', s)
    clean_sentences = []
    for part in parts:
        part = part.strip()
        if not part:
            continue
        # 조각 내에서 마침표/느낌표/물음표로 끝나는 완결 문장만 추출
        sentences = _re9.split(r'(?<=[.!?])\s+', part)
        for sent in sentences:
            sent = sent.strip()
            # 문장 끝에 마침표가 있고, 너무 짧은 조각(단어 파편)이 아닌 것만 채택
            if sent and sent[-1] in '.!?' and 10 <= len(sent) <= 140:
                clean_sentences.append(sent)
    return clean_sentences


def _research_product(side, category, brand, product_name):
    """특정 제품(자사/타사 공용)의 리뷰+설명을 다각도로 방대하게 검색해 장단점/유형 심층 축적"""
    query_base = f"{brand} {product_name}".strip()
    titles, urls = [], []
    review_snippets, con_snippets, pro_snippets, official_snippets, type_snippets = [], [], [], [], []

    try:
        shop = _naver_search('shop', f"{query_base} {category}", display=10, sort='sim')
        for item in shop.get('items', [])[:10]:
            t = _strip_tags(item.get('title',''))
            titles.append(t); urls.append(item.get('link',''))
    except Exception:
        pass

    # 검색 쿼리를 다각도로 확장 (방대한 수집)
    search_plan = [
        ('blog', f"{query_base} 후기", 10, review_snippets),
        ('blog', f"{query_base} 실사용 후기", 8, review_snippets),
        ('blog', f"{query_base} 장점 추천이유", 8, pro_snippets),
        ('blog', f"{query_base} 왜 좋은지", 6, pro_snippets),
        ('blog', f"{query_base} 단점 아쉬운점 불편", 8, con_snippets),
        ('blog', f"{query_base} 후회 고민", 6, con_snippets),
        ('blog', f"{query_base} 특징 스펙 소재", 6, official_snippets),
        ('blog', f"{query_base} 무게 사이즈 접이식", 6, official_snippets),
    ]
    for endpoint, q, disp, bucket in search_plan:
        try:
            res = _naver_search(endpoint, q, display=disp, sort='sim')
            for item in res.get('items', [])[:disp]:
                desc = _strip_tags(item.get('description',''))
                if desc:
                    bucket.extend(_clean_naver_snippet(desc))
        except Exception:
            pass

    total_snippets = review_snippets + pro_snippets + con_snippets + official_snippets
    if not titles and not total_snippets:
        return None

    pros, cons, description = _extract_pros_cons(
        brand, product_name, category, titles,
        review_snippets, pro_snippets, con_snippets, official_snippets)

    # 제품 유형 분류 (자사는 고정 매핑 우선, 없으면 AI/휴리스틱)
    product_type = ''
    if side == 'ours':
        product_type = OUR_PRODUCT_TYPE.get(category, {}).get(product_name, '')
    if not product_type:
        product_type = _classify_product_type(
            category, brand, product_name, titles, total_snippets)

    return {
        'product_name': product_name, 'price_text': '',
        'pros': pros, 'cons': cons, 'description': description,
        'review_snippets': total_snippets[:20],
        'source_titles': titles[:6], 'source_urls': urls[:6],
        'product_type': product_type,
    }


def _classify_product_type(category, brand, product_name, titles, snippets):
    """리뷰/설명 텍스트를 근거로 제품 유형 분류 (유모차: 디럭스/절충형/휴대용/쌍둥이, 카시트: 컨버터블/주니어/토들러)"""
    type_options = PRODUCT_TYPE_ORDER.get(category)
    if not type_options:
        return ''

    combined = ' '.join(titles[:6] + snippets[:10])
    if not combined.strip():
        return ''

    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if api_key:
        try:
            import requests
            prompt = f"""'{brand} {product_name}' ({category})의 검색 결과를 보고, 이 제품이 다음 유형 중 어디에 해당하는지 하나만 고르세요: {', '.join(type_options)}

판단 기준:
- 유모차: 디럭스(신생아부터 사용 가능한 대형 고급형), 절충형(중형, 아기~유아 모두 사용), 휴대용(경량 접이식, 여행용), 쌍둥이(2인용)
- 카시트: 컨버터블(신생아~유아 겸용, 회전형 포함), 주니어(체구가 큰 유아~어린이용, 부스터 겸용), 토들러(영유아 전용 카시트)

검색 결과: {combined[:1200]}

가장 적합한 유형 하나만 정확히 그 단어로만 답하세요 (예: "디럭스"). 판단이 어려우면 "미분류"라고 답하세요."""
            resp = requests.post(
                "https://api.anthropic.com/v1/messages",
                headers={"x-api-key": api_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
                json={"model": "claude-sonnet-4-6", "max_tokens": 20,
                      "messages": [{"role": "user", "content": prompt}]},
                timeout=12,
            )
            if resp.status_code == 200:
                data = resp.json()
                text_blocks = [b.get('text','') for b in data.get('content',[]) if b.get('type')=='text']
                answer = ''.join(text_blocks).strip()
                for opt in type_options:
                    if opt in answer:
                        return opt
        except Exception:
            pass

    # Fallback: 키워드 기반 휴리스틱
    KEYWORD_MAP = {
        '유모차': {
            '디럭스': ['디럭스', '신생아', '하이엔드', '프리미엄', '고급형'],
            '절충형': ['절충형', '세미디럭스', '중형'],
            '휴대용': ['휴대용', '경량', '초경량', '여행용', '기내용', '접이식', '가벼운'],
            '쌍둥이': ['쌍둥이', '트윈', '2인용', '듀오'],
        },
        '카시트': {
            '컨버터블': ['컨버터블', '회전형', '360도', '신생아겸용'],
            '주니어': ['주니어', '어린이', '체구가 큰', '부스터겸용'],
            '토들러': ['토들러', '영유아', '유아전용'],
        },
    }
    kw_map = KEYWORD_MAP.get(category, {})
    scores = {opt: 0 for opt in type_options}
    for opt, kws in kw_map.items():
        for kw in kws:
            scores[opt] += combined.count(kw)
    best = max(scores.items(), key=lambda x: x[1])
    return best[0] if best[1] > 0 else ''


def _extract_pros_cons(brand, product_name, category, titles,
                         review_snippets=None, pro_snippets=None, con_snippets=None, official_snippets=None):
    """수집된 검색결과(후기/장점/단점/공식정보 각각 분리, 이미 완결 문장으로 정제됨)에서 TOP3 장단점을
    '짧고 명확한 핵심 문구'로 요약 (Claude API 우선, 실패 시 테마 기반 키워드 분석)
    반환 형식: pros/cons는 [{"tag":str, "point":str, "reason":str}, ...] 구조
    point는 "작고 가볍다" 같이 5~15자 내외의 핵심 요약 문구, reason은 그 근거 설명"""
    review_snippets = review_snippets or []
    pro_snippets = pro_snippets or []
    con_snippets = con_snippets or []
    official_snippets = official_snippets or []

    all_snippets = titles + review_snippets + pro_snippets + con_snippets + official_snippets
    if not any(all_snippets):
        return [], [], ''

    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if api_key:
        try:
            import requests
            prompt = f"""당신은 유아용품 업계 15년차 리서치/데이터 분석 전문가입니다. 이 분석은 매장 영업 현장에서 즉시 활용할 요약 자료이므로, 장황한 문장이 아니라 "딱 봐도 이해되는 짧고 명확한 핵심 문구"로 정리해야 합니다.
'{brand} {product_name}' ({category})에 대해 인터넷에서 방대하게 수집한 원문 자료를 드립니다.

[제품/판매 정보]
{' / '.join(titles[:8])}

[일반 사용후기 — 총 {len(review_snippets)}개]
{' / '.join(review_snippets[:15]) or '(수집된 후기 없음)'}

[장점/추천이유 관련 — 총 {len(pro_snippets)}개]
{' / '.join(pro_snippets[:12]) or '(수집된 정보 없음)'}

[단점/아쉬운점 관련 — 총 {len(con_snippets)}개]
{' / '.join(con_snippets[:12]) or '(수집된 정보 없음)'}

[제품 특징/스펙 관련 — 총 {len(official_snippets)}개]
{' / '.join(official_snippets[:12]) or '(수집된 정보 없음)'}

작성 원칙 (매우 중요 — 반드시 지킬 것):
1. 방대한 원문을 전부 읽고, 반복적으로 나타나는 "핵심 테마"를 파악하세요 (예: 무게, 접이식, 시트감, 소재, 디자인, 안전성, 수납, 조립, 회전기능, 바퀴 등).
2. point는 절대 원문 문장을 그대로 쓰지 말고, 테마를 압축한 짧은 핵심 문구로 작성하세요. 길이는 5~15자 내외.
   - 나쁜 예 (금지): "결국 큰맘 먹고 줄즈 에어2를 구매하게 되었습니다" ← 원문 그대로, 의미 없음
   - 좋은 예: "작고 가볍다" / "폴딩이 쉽다" / "해먹시트가 편하다" / "통기성이 좋다" / "색상이 예쁘다"
3. reason에는 point를 뒷받침하는 근거를 한 문장으로 (원문에서 확인된 구체적 사실 인용).
4. 장점 TOP3, 단점 TOP3는 원문 전체에서 "가장 많이 반복되는 테마" 순서로 선정하세요.
5. 블로그 인사말, 구매 결정 서사("~ 하다가 결국 구매했습니다" 류), 광고 해시태그는 테마로 취급하지 마세요.
6. 단점 정보가 부족하면, 이 카테고리 제품이 일반적으로 갖는 한계를 전문가 관점에서 짧게 짚어주세요.
7. 태그는 [소재/디자인] [기능성] [무게/휴대성] [안전성] [내구성] [편의성] [브랜드신뢰도] 중 선택하세요.
8. 가격은 언급하지 마세요.

다음 JSON 형식으로만 답하세요 (설명 없이 JSON만):
{{
  "pros": [{{"tag":"태그", "point":"5~15자 핵심 문구", "reason":"근거 한 문장"}}],
  "cons": [{{"tag":"태그", "point":"5~15자 핵심 문구", "reason":"근거 한 문장"}}],
  "description": "제품 특징을 압축한 1~2문장"
}}
pros/cons는 각각 최대 3개, 테마 언급 빈도 높은 순."""
            resp = requests.post(
                "https://api.anthropic.com/v1/messages",
                headers={"x-api-key": api_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
                json={"model": "claude-sonnet-4-6", "max_tokens": 1200,
                      "messages": [{"role": "user", "content": prompt}]},
                timeout=25,
            )
            if resp.status_code == 200:
                data = resp.json()
                text_blocks = [b.get('text','') for b in data.get('content',[]) if b.get('type')=='text']
                raw = ''.join(text_blocks).strip().replace('```json','').replace('```','').strip()
                parsed = json.loads(raw)
                pros = parsed.get('pros', [])[:3]
                cons = parsed.get('cons', [])[:3]
                pros = [p if isinstance(p, dict) else {'tag':'', 'point':str(p), 'reason':''} for p in pros]
                cons = [c if isinstance(c, dict) else {'tag':'', 'point':str(c), 'reason':''} for c in cons]
                return pros, cons, parsed.get('description', '')
        except Exception:
            pass

    # Fallback: API 키가 없을 때 — 테마(속성) 사전 기반 키워드 매칭 + 감성 분류로 짧은 핵심 문구 생성
    return _theme_based_fallback(all_snippets, review_snippets, pro_snippets, con_snippets, official_snippets)


# 유아용품 공통 속성 테마 사전 — 리서치 전문가 관점에서 자주 언급되는 속성 16종
PRODUCT_THEMES = [
    {'id':'weight',    'kw':['가볍','가벼운','가벼워','경량','무게가 적','무게 적'], 'neg_kw':['무겁','무거워','무게가 있','묵직'],
     'pro':'작고 가볍다', 'con':'무게가 있는 편이다', 'tag':'무게/휴대성'},
    {'id':'fold',      'kw':['폴딩','접이','원터치','한손으로 접','접기 쉽','접는게 쉽'], 'neg_kw':['접기 어렵','접는게 불편','폴딩이 불편','접이가 힘'],
     'pro':'폴딩이 쉽다', 'con':'접이식 조작이 불편하다', 'tag':'기능성'},
    {'id':'seat',      'kw':['해먹시트','시트감','착석감','앉는 느낌이 편','시트가 편'], 'neg_kw':['시트가 좁','시트가 딱딱','착석감이 별로'],
     'pro':'해먹시트가 편하다', 'con':'시트 착석감이 아쉽다', 'tag':'편의성'},
    {'id':'breath',    'kw':['통풍','통기성','매쉬','시원하','바람이 잘 통'], 'neg_kw':['덥다','통풍이 안','땀이 차'],
     'pro':'통기성이 좋다', 'con':'통풍이 다소 아쉽다', 'tag':'소재/디자인'},
    {'id':'design',    'kw':['디자인이 예쁘','색상이 예쁘','예뻐서','디자인 만족','색감이 좋'], 'neg_kw':['디자인이 아쉽','색상 선택지가 적','칙칙'],
     'pro':'디자인·색감이 예쁘다', 'con':'색상 선택지가 제한적이다', 'tag':'소재/디자인'},
    {'id':'durability','kw':['튼튼','내구성이 좋','견고','단단하'], 'neg_kw':['내구성이 아쉽','약하다','금방 헐거워'],
     'pro':'프레임이 튼튼하다', 'con':'내구성이 다소 아쉽다', 'tag':'내구성'},
    {'id':'safety',    'kw':['안전벨트','안전하','충돌 보호','안전성이 좋'], 'neg_kw':['안전벨트가 불편','안전성이 아쉽'],
     'pro':'안전 기능이 우수하다', 'con':'안전 기능이 다소 미흡하다', 'tag':'안전성'},
    {'id':'storage',   'kw':['수납공간이 넉넉','바구니가 크','수납이 좋','짐 많이'], 'neg_kw':['수납공간이 작','수납이 부족','바구니가 작'],
     'pro':'수납공간이 넉넉하다', 'con':'수납공간이 부족하다', 'tag':'편의성'},
    {'id':'rotation',  'kw':['360도 회전','회전이 부드럽','방향전환이 쉽'], 'neg_kw':['회전이 뻑뻑','방향전환이 불편'],
     'pro':'360도 회전이 편리하다', 'con':'회전 조작이 뻑뻑하다', 'tag':'기능성'},
    {'id':'assembly',  'kw':['조립이 간단','조립이 쉽','설치가 쉬'], 'neg_kw':['조립이 복잡','조립이 어렵','설치가 오래'],
     'pro':'조립이 간단하다', 'con':'조립이 복잡한 편이다', 'tag':'편의성'},
    {'id':'wheel',     'kw':['바퀴가 부드럽','주행감이 좋','승차감이 좋'], 'neg_kw':['바퀴가 작','승차감이 아쉽','덜컹거려'],
     'pro':'바퀴 주행감이 좋다', 'con':'바퀴 승차감이 아쉽다', 'tag':'기능성'},
    {'id':'handle',    'kw':['손잡이 높이조절','손잡이가 편'], 'neg_kw':['손잡이가 불편','높이조절이 안'],
     'pro':'손잡이 높이조절이 편리하다', 'con':'손잡이 조작이 불편하다', 'tag':'편의성'},
    {'id':'size',      'kw':['부피가 작','콤팩트','크기가 작아 보관'], 'neg_kw':['부피가 커','크기가 커서 보관','자리를 많이'],
     'pro':'부피가 작아 보관이 쉽다', 'con':'부피가 커서 보관이 불편하다', 'tag':'무게/휴대성'},
    {'id':'service',   'kw':['as가 빠르','as 응대가 좋','서비스가 친절'], 'neg_kw':['as가 느리','문의 응대가 아쉽'],
     'pro':'A/S 대응이 빠르다', 'con':'A/S 대응이 느린 편이다', 'tag':'브랜드신뢰도'},
    {'id':'material',  'kw':['소재가 좋','원단이 고급','촉감이 좋'], 'neg_kw':['소재가 아쉽','원단이 얇','촉감이 별로'],
     'pro':'소재/원단 품질이 좋다', 'con':'소재가 다소 아쉽다', 'tag':'소재/디자인'},
    {'id':'space',     'kw':['공간을 적게','실내에서도 편','좁은 곳에서도'], 'neg_kw':['공간을 많이 차지','좁은 곳에서 불편'],
     'pro':'공간을 적게 차지한다', 'con':'실내 공간을 많이 차지한다', 'tag':'무게/휴대성'},
]


def _theme_based_fallback(all_snippets, review_snippets, pro_snippets, con_snippets, official_snippets):
    """리서치 전문가 방식의 테마 사전 매칭 — 원문을 그대로 노출하지 않고, 감지된 속성을 짧은 핵심 문구로 변환.
    긍정/부정 키워드 등장 빈도를 세어 TOP3 테마를 선정한다."""
    combined_pro_pool = ' '.join(pro_snippets + review_snippets + official_snippets)
    combined_con_pool = ' '.join(con_snippets + review_snippets)

    pro_scores, con_scores = [], []
    for theme in PRODUCT_THEMES:
        pro_hits = sum(combined_pro_pool.count(kw) for kw in theme['kw'])
        con_hits = sum(combined_con_pool.count(kw) for kw in theme['neg_kw'])
        con_hits += sum(combined_pro_pool.count(kw) for kw in theme['neg_kw'])
        if pro_hits > 0:
            pro_scores.append((pro_hits, theme))
        if con_hits > 0:
            con_scores.append((con_hits, theme))

    pro_scores.sort(key=lambda x: -x[0])
    con_scores.sort(key=lambda x: -x[0])

    def _to_items(scores, key):
        items = []
        for hits, theme in scores[:3]:
            items.append({
                'tag': theme['tag'],
                'point': theme[key],
                'reason': f"수집된 자료 {hits}건에서 관련 언급이 확인됨",
            })
        return items

    pros = _to_items(pro_scores, 'pro')
    cons = _to_items(con_scores, 'con')

    import re as _re_module
    BOILERPLATE_PATTERNS = [
        r'안녕하세요', r'^오늘은', r'소개해드릴게요', r'소개해드리겠습니다', r'포스팅',
        r'추천드립니다$', r'추천드려요$', r'가져왔습니다', r'^구매했어요', r'^구매했습니다',
        r'해시태그', r'^<<.*>>$', r'^#', r'결국.*이다!?$', r'최종 선택',
        r'큰맘 먹고', r'구매하게 되었습니다', r'구매하게 됐습니다', r'구입하게 되었습니다',
        r'결정하게 되었습니다', r'선택하게 되었습니다', r'고민 끝에', r'고민하다가',
    ]
    def _is_boilerplate(s):
        return any(_re_module.search(pat, s) for pat in BOILERPLATE_PATTERNS)
    clean_desc_pool = [s for s in (official_snippets + review_snippets) if not _is_boilerplate(s) and len(s) >= 12]
    description = clean_desc_pool[0] if clean_desc_pool else ''

    return pros, cons, description


@app.route("/api/competitor/fetch-product", methods=["POST"])
@login_required
def api_competitor_fetch_product():
    """특정 타사 제품(브랜드+모델명)의 상세 리서치 수집 및 저장"""
    d = request.json or {}
    category = d.get('category', '').strip()
    brand = d.get('brand', '').strip()
    product_name = d.get('product_name', '').strip()
    if not (category and brand and product_name):
        return jsonify({'ok': False, 'msg': '품목/브랜드/제품명이 필요합니다'}), 400

    result = _research_product('competitor', category, brand, product_name)
    if not result:
        return jsonify({'ok': False, 'msg': '검색 결과가 없습니다. 제품명을 다르게 시도해보세요'}), 404

    conn = get_db()
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
    conn.execute("""INSERT INTO product_research
        (side, category, brand, product_name, price_text, pros, cons, description,
         review_snippets, source_titles, source_urls, fetched_at, product_type)
        VALUES('competitor',?,?,?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(side, category, brand, product_name) DO UPDATE SET
        price_text=excluded.price_text, pros=excluded.pros, cons=excluded.cons,
        description=excluded.description, review_snippets=excluded.review_snippets,
        source_titles=excluded.source_titles, source_urls=excluded.source_urls,
        fetched_at=excluded.fetched_at, product_type=excluded.product_type""",
        (category, brand, product_name, result['price_text'],
         json.dumps(result['pros'], ensure_ascii=False), json.dumps(result['cons'], ensure_ascii=False),
         result['description'], json.dumps(result['review_snippets'], ensure_ascii=False),
         ' | '.join(result['source_titles']), ' | '.join(result['source_urls']), now_str,
         result.get('product_type','')))
    conn.commit(); conn.close()

    return jsonify({'ok': True, **result, 'fetched_at': now_str})


@app.route("/api/competitor/fetch-our-product", methods=["POST"])
@login_required
def api_competitor_fetch_our_product():
    """자사 제품의 외부 리뷰 기반 장단점 리서치 수집 및 저장"""
    d = request.json or {}
    category = d.get('category', '').strip()
    brand = d.get('brand', '').strip()
    product_name = d.get('product_name', '').strip()
    if not (category and brand and product_name):
        return jsonify({'ok': False, 'msg': '품목/브랜드/제품명이 필요합니다'}), 400

    result = _research_product('ours', category, brand, product_name)
    if not result:
        # 검색 결과가 없어도 내부 브랜드 강점으로 최소한의 데이터 구성
        base_strength = OUR_BRAND_STRENGTHS.get(brand, '')
        result = {'product_name': product_name, 'price_text': '',
                   'pros': [base_strength] if base_strength else [],
                   'cons': [], 'description': '', 'review_snippets': [],
                   'source_titles': [], 'source_urls': [],
                   'product_type': OUR_PRODUCT_TYPE.get(category, {}).get(product_name, '')}

    conn = get_db()
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
    conn.execute("""INSERT INTO product_research
        (side, category, brand, product_name, price_text, pros, cons, description,
         review_snippets, source_titles, source_urls, fetched_at, product_type)
        VALUES('ours',?,?,?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(side, category, brand, product_name) DO UPDATE SET
        price_text=excluded.price_text, pros=excluded.pros, cons=excluded.cons,
        description=excluded.description, review_snippets=excluded.review_snippets,
        source_titles=excluded.source_titles, source_urls=excluded.source_urls,
        fetched_at=excluded.fetched_at, product_type=excluded.product_type""",
        (category, brand, product_name, result['price_text'],
         json.dumps(result['pros'], ensure_ascii=False), json.dumps(result['cons'], ensure_ascii=False),
         result['description'], json.dumps(result['review_snippets'], ensure_ascii=False),
         ' | '.join(result['source_titles']), ' | '.join(result['source_urls']), now_str,
         result.get('product_type','')))
    conn.commit(); conn.close()

    return jsonify({'ok': True, **result, 'fetched_at': now_str})


@app.route("/api/competitor/research/delete", methods=["POST"])
@login_required
def api_competitor_research_delete():
    """수집된 제품 리서치 삭제"""
    d = request.json or {}
    side = d.get('side', '').strip()
    category = d.get('category', '').strip()
    brand = d.get('brand', '').strip()
    product_name = d.get('product_name', '').strip()
    if not all([side, category, brand, product_name]):
        return jsonify({'ok': False, 'msg': '필요한 정보가 부족합니다'}), 400
    conn = get_db()
    conn.execute("""DELETE FROM product_research
        WHERE side=? AND category=? AND brand=? AND product_name=?""",
        (side, category, brand, product_name))
    conn.commit(); conn.close()
    return jsonify({'ok': True})


OUR_BRAND_STRENGTHS = {
    '줄즈':      'SNS 바이럴 확산력과 트렌디한 색상 구성, 재구매율이 높은 실용적 유모차',
    '레카로':    '독일 안전 인증 기반 신뢰도, 카시트 전문 브랜드로서의 전문성',
    '엔픽스':    '국내 브랜드 특유의 가성비와 A/S 접근성, 국내 육아 환경에 최적화',
    '카오스':    '프리미엄 하이체어 라인, 원목 소재와 오래 쓸 수 있는 확장성',
    '원더폴드':  '웨건 시장 국내 독점적 포지션, 다인승 특화 설계',
    'ABC디자인': '유럽 감성 디자인과 패밀리 브랜드 이미지, 소재 고급감',
    '타프토이즈': '유아 발달단계별 완구 전문성, 선물/사은품 수요 대응',
}


@app.route("/api/competitor/research")
@login_required
def api_competitor_research_get():
    """저장된 제품 리서치 데이터 조회 (자사/타사)"""
    side = request.args.get('side', '')
    category = request.args.get('category', '')
    brand = request.args.get('brand', '')
    conn = get_db()
    q = "SELECT * FROM product_research WHERE 1=1"
    params = []
    if side: q += " AND side=?"; params.append(side)
    if category: q += " AND category=?"; params.append(category)
    if brand: q += " AND brand=?"; params.append(brand)
    q += " ORDER BY fetched_at DESC"
    rows = [dict(r) for r in conn.execute(q, params).fetchall()]
    conn.close()
    for r in rows:
        try: r['pros'] = json.loads(r.get('pros') or '[]')
        except: r['pros'] = []
        try: r['cons'] = json.loads(r.get('cons') or '[]')
        except: r['cons'] = []
    return jsonify(rows)


@app.route("/api/competitor/compare", methods=["POST"])
@login_required
def api_competitor_compare():
    """자사 제품 vs 타사 제품(구체적 모델) 비교 분석 생성 — 축적된 리서치 데이터 기반"""
    d = request.json or {}
    category = d.get('category', '').strip()
    our_brand = d.get('our_brand', '').strip()
    our_product = d.get('our_product', '').strip()
    competitor_brand = d.get('competitor_brand', '').strip()
    competitor_product = d.get('competitor_product', '').strip()

    if not all([category, our_brand, our_product, competitor_brand, competitor_product]):
        return jsonify({'ok': False, 'msg': '자사 제품과 타사 제품을 모두 선택해주세요'}), 400

    conn = get_db()
    our_row = conn.execute("""SELECT * FROM product_research
        WHERE side='ours' AND category=? AND brand=? AND product_name=?""",
        (category, our_brand, our_product)).fetchone()
    comp_row = conn.execute("""SELECT * FROM product_research
        WHERE side='competitor' AND category=? AND brand=? AND product_name=?""",
        (category, competitor_brand, competitor_product)).fetchone()
    conn.close()

    if not comp_row:
        return jsonify({'ok': False, 'msg': '먼저 타사 제품 정보를 불러와주세요'}), 400

    our_data = dict(our_row) if our_row else {
        'pros': '[]', 'cons': '[]', 'description': '', 'price_text': ''
    }
    comp_data = dict(comp_row)

    comparison = generate_comparison(
        category, our_brand, our_product, our_data,
        competitor_brand, competitor_product, comp_data)

    conn = get_db()
    conn.execute("""INSERT INTO competitor_comparison
        (category, our_product, competitor_brand, competitor_product, comparison_text, created_at)
        VALUES(?,?,?,?,?,?)""",
        (category, our_product, competitor_brand, competitor_product,
         json.dumps(comparison, ensure_ascii=False), datetime.now().strftime('%Y-%m-%d %H:%M')))
    conn.commit(); conn.close()

    return jsonify({'ok': True, 'comparison': comparison})


@app.route("/api/competitor/type-options")
@login_required
def api_competitor_type_options():
    """카테고리 내에서 리서치가 축적된 제품 유형 목록 반환 (휴대용/절충형/디럭스 등)"""
    category = request.args.get('category', '')
    type_order = PRODUCT_TYPE_ORDER.get(category, [])
    conn = get_db()
    rows = conn.execute("""
        SELECT DISTINCT product_type FROM product_research
        WHERE category=? AND product_type!=''
    """, (category,)).fetchall()
    conn.close()
    found_types = set(r[0] for r in rows)
    # 정해진 순서대로, 실제 데이터가 있는 유형만 반환
    ordered = [t for t in type_order if t in found_types]
    # 순서에 없는 유형도 혹시 있으면 추가
    ordered += [t for t in found_types if t not in ordered]
    return jsonify(ordered)


@app.route("/api/competitor/type-compare")
@login_required
def api_competitor_type_compare():
    """제품 유형(휴대용/절충형/디럭스 등) 단위로 자사 vs 모든 타사 제품 일괄 비교"""
    category = request.args.get('category', '')
    product_type = request.args.get('product_type', '')
    if not category or not product_type:
        return jsonify({'ok': False, 'msg': '품목과 유형을 선택해주세요'}), 400

    conn = get_db()
    our_rows = [dict(r) for r in conn.execute("""
        SELECT * FROM product_research WHERE side='ours' AND category=? AND product_type=?
        ORDER BY brand, product_name
    """, (category, product_type)).fetchall()]
    comp_rows = [dict(r) for r in conn.execute("""
        SELECT * FROM product_research WHERE side='competitor' AND category=? AND product_type=?
        ORDER BY brand, product_name
    """, (category, product_type)).fetchall()]
    conn.close()

    def _parse_row(r):
        try: r['pros'] = json.loads(r.get('pros') or '[]')
        except: r['pros'] = []
        try: r['cons'] = json.loads(r.get('cons') or '[]')
        except: r['cons'] = []
        r['pros'] = [p if isinstance(p, dict) else {'tag':'','point':str(p),'reason':''} for p in r['pros']]
        r['cons'] = [c if isinstance(c, dict) else {'tag':'','point':str(c),'reason':''} for c in r['cons']]
        return r

    our_rows = [_parse_row(r) for r in our_rows]
    comp_rows = [_parse_row(r) for r in comp_rows]

    if not our_rows and not comp_rows:
        return jsonify({'ok': False, 'msg': '이 유형에 리서치된 제품이 없습니다. 먼저 개별 제품 리서치를 수집해주세요'}), 404

    return jsonify({'ok': True, 'category': category, 'product_type': product_type,
                     'our_products': our_rows, 'competitor_products': comp_rows})


@app.route("/api/competitor/export-xlsx")
@login_required
def api_competitor_export_xlsx():
    """비교 리포트 엑셀 다운로드 — 단일 제품 비교(mode=single) 또는 유형별 일괄 비교(mode=type)"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    mode = request.args.get('mode', 'single')
    category = request.args.get('category', '')
    FNAME = '맑은 고딕'
    def mf(h): return PatternFill("solid", fgColor=h)
    thin = Side(style='thin', color='E5E7EB')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    wb = openpyxl.Workbook()
    ws = wb.active

    def _parse_items(raw):
        try: items = json.loads(raw or '[]')
        except: items = []
        return [it if isinstance(it, dict) else {'tag':'','point':str(it),'reason':''} for it in items]

    def _write_items_cell(ws, ri, ci, items, label_color):
        lines = []
        for it in items:
            tag = f"[{it.get('tag')}] " if it.get('tag') else ''
            lines.append(f"• {tag}{it.get('point','')}")
            if it.get('reason'):
                lines.append(f"   → {it.get('reason')}")
        c = ws.cell(row=ri, column=ci, value='\n'.join(lines) if lines else '수집된 정보 없음')
        c.font = Font(size=9, name=FNAME, color='1F2937')
        c.alignment = left
        c.border = bdr
        return c

    if mode == 'single':
        our_brand = request.args.get('our_brand','')
        our_product = request.args.get('our_product','')
        competitor_brand = request.args.get('competitor_brand','')
        competitor_product = request.args.get('competitor_product','')

        conn = get_db()
        our_row = conn.execute("""SELECT * FROM product_research
            WHERE side='ours' AND category=? AND brand=? AND product_name=?""",
            (category, our_brand, our_product)).fetchone()
        comp_row = conn.execute("""SELECT * FROM product_research
            WHERE side='competitor' AND category=? AND brand=? AND product_name=?""",
            (category, competitor_brand, competitor_product)).fetchone()
        conn.close()
        our_data = dict(our_row) if our_row else {'pros':'[]','cons':'[]','description':''}
        comp_data = dict(comp_row) if comp_row else {'pros':'[]','cons':'[]','description':''}

        comparison = generate_comparison(category, our_brand, our_product, our_data,
                                          competitor_brand, competitor_product, comp_data)

        ws.title = f'{our_product[:10]}_vs_{competitor_product[:10]}'[:31]
        ws.merge_cells('A1:B1')
        c = ws.cell(row=1, column=1, value=f'{our_brand} {our_product}  vs  {competitor_brand} {competitor_product}   ({category} 비교 리포트)')
        c.font = Font(bold=True, size=13, name=FNAME, color='1F2937'); c.alignment = ctr
        ws.row_dimensions[1].height = 26

        ri = 2
        if comparison.get('summary'):
            ws.merge_cells(f'A{ri}:B{ri}')
            c = ws.cell(row=ri, column=1, value=f"핵심 요약: {comparison['summary']}")
            c.font = Font(bold=True, size=10, name=FNAME, color='2563EB'); c.fill = mf('EFF6FF'); c.alignment = left
            ws.row_dimensions[ri].height = 30
            ri += 2

        headers = [f'{our_brand} {our_product}', f'{competitor_brand} {competitor_product}']
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=ri, column=ci, value=h)
            c.font = Font(bold=True, size=10, name=FNAME, color='374151'); c.fill = mf('F3F4F6'); c.border = bdr; c.alignment = ctr
        ws.row_dimensions[ri].height = 20
        ri += 1

        ws.cell(row=ri, column=1, value='장점').font = Font(bold=True, size=9, name=FNAME, color='16A34A')
        r0 = ri
        ws.merge_cells(f'A{ri}:A{ri}')
        _write_items_cell(ws, ri, 1, comparison.get('our_pros',[]), 'green')
        _write_items_cell(ws, ri, 2, comparison.get('competitor_pros',[]), 'green')
        ws.row_dimensions[ri].height = 120
        ri += 1

        _write_items_cell(ws, ri, 1, comparison.get('our_cons',[]), 'red')
        _write_items_cell(ws, ri, 2, comparison.get('competitor_cons',[]), 'red')
        ws.row_dimensions[ri].height = 90
        ri += 2

        ws.merge_cells(f'A{ri}:B{ri}')
        c = ws.cell(row=ri, column=1, value='영업 실전 포인트')
        c.font = Font(bold=True, size=11, name=FNAME, color='1F2937'); ri += 1
        for i, p in enumerate(comparison.get('selling_points', []), 1):
            ws.merge_cells(f'A{ri}:B{ri}')
            c = ws.cell(row=ri, column=1, value=f"{i}. {p}")
            c.font = Font(size=9, name=FNAME, color='374151'); c.alignment = left
            ws.row_dimensions[ri].height = 34
            ri += 1

        ws.column_dimensions['A'].width = 55
        ws.column_dimensions['B'].width = 55
        fname = f'타사비교_{our_product}_vs_{competitor_product}.xlsx'

    else:  # mode == 'type' — 유형별 일괄 비교
        product_type = request.args.get('product_type', '')
        conn = get_db()
        our_rows = [dict(r) for r in conn.execute("""
            SELECT * FROM product_research WHERE side='ours' AND category=? AND product_type=?
            ORDER BY brand, product_name""", (category, product_type)).fetchall()]
        comp_rows = [dict(r) for r in conn.execute("""
            SELECT * FROM product_research WHERE side='competitor' AND category=? AND product_type=?
            ORDER BY brand, product_name""", (category, product_type)).fetchall()]
        conn.close()

        all_rows = our_rows + comp_rows
        ws.title = f'{category}_{product_type}'[:31]
        ncol = len(all_rows)
        if ncol == 0:
            ws.cell(row=1, column=1, value='데이터 없음')
        else:
            ws.merge_cells(f'A1:{get_column_letter(ncol+1)}1')
            c = ws.cell(row=1, column=1, value=f'{category} · {product_type} 유형 일괄 비교 리포트')
            c.font = Font(bold=True, size=13, name=FNAME, color='1F2937'); c.alignment = ctr
            ws.row_dimensions[1].height = 26

            ri = 3
            ws.cell(row=ri, column=1, value='').fill = mf('F3F4F6')
            for ci, r in enumerate(all_rows, 2):
                side_label = '자사' if r['side'] == 'ours' else '경쟁사'
                c = ws.cell(row=ri, column=ci, value=f"[{side_label}] {r['brand']} {r['product_name']}")
                c.font = Font(bold=True, size=10, name=FNAME, color='FFFFFF' if r['side']=='ours' else '374151')
                c.fill = mf('2563EB' if r['side']=='ours' else 'F3F4F6')
                c.border = bdr; c.alignment = ctr
                ws.column_dimensions[get_column_letter(ci)].width = 40
            ws.row_dimensions[ri].height = 22
            ri += 1

            c = ws.cell(row=ri, column=1, value='장점')
            c.font = Font(bold=True, size=9, name=FNAME, color='16A34A'); c.border = bdr
            for ci, r in enumerate(all_rows, 2):
                _write_items_cell(ws, ri, ci, _parse_items(r.get('pros')), 'green')
            ws.row_dimensions[ri].height = 140
            ri += 1

            c = ws.cell(row=ri, column=1, value='단점')
            c.font = Font(bold=True, size=9, name=FNAME, color='DC2626'); c.border = bdr
            for ci, r in enumerate(all_rows, 2):
                _write_items_cell(ws, ri, ci, _parse_items(r.get('cons')), 'red')
            ws.row_dimensions[ri].height = 110
            ri += 1

            c = ws.cell(row=ri, column=1, value='제품 설명')
            c.font = Font(bold=True, size=9, name=FNAME, color='6B7280'); c.border = bdr
            for ci, r in enumerate(all_rows, 2):
                c2 = ws.cell(row=ri, column=ci, value=r.get('description','') or '—')
                c2.font = Font(size=9, name=FNAME, color='6B7280'); c2.alignment = left; c2.border = bdr
            ws.row_dimensions[ri].height = 50

            ws.column_dimensions['A'].width = 12
        fname = f'타사비교_{category}_{product_type}_일괄비교.xlsx'

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=fname)


def generate_comparison(category, our_brand, our_product, our_data,
                          competitor_brand, competitor_product, comp_data):
    """축적된 상세 리서치 데이터를 기반으로 제품 단위 심층 비교 분석 생성 (이유 포함 구조)"""
    api_key = os.environ.get('ANTHROPIC_API_KEY', '')

    def _load_items(raw_json):
        try: items = json.loads(raw_json or '[]')
        except: items = []
        return [it if isinstance(it, dict) else {'tag':'', 'point':str(it), 'reason':''} for it in items]

    our_pros_raw  = _load_items(our_data.get('pros'))
    our_cons_raw  = _load_items(our_data.get('cons'))
    comp_pros_raw = _load_items(comp_data.get('pros'))
    comp_cons_raw = _load_items(comp_data.get('cons'))
    try: our_reviews = json.loads(our_data.get('review_snippets') or '[]')
    except: our_reviews = []
    try: comp_reviews = json.loads(comp_data.get('review_snippets') or '[]')
    except: comp_reviews = []

    fallback_strength = OUR_BRAND_STRENGTHS.get(our_brand, '국내 유통망과 사후관리 대응력')

    def _fmt_items(items):
        return '; '.join(f"[{it.get('tag','')}] {it.get('point','')} (이유: {it.get('reason','')})" for it in items) or '(없음)'

    if api_key:
        try:
            import requests
            prompt = f"""당신은 유아용품 업계 15년차 상품기획/영업 전문가이자 마케터입니다. 이 자료는 실제 매장 영업 현장에서 고객 응대용으로 쓰입니다.
아래 실제 수집된 리뷰·설명 데이터를 근거로 두 제품을 심층 비교 분석해주세요.
정형화된 뻔한 비교("A/S가 좋다", "국내 유통이 강점이다" 같은 상투적 표현 금지)가 아니라, 실제 데이터에 기반한 구체적이고 날카로운 비교를 원합니다.

[품목] {category}

[자사 제품] {our_brand} {our_product}
- 제품 설명: {our_data.get('description','')}
- 기존 분석 장점: {_fmt_items(our_pros_raw)}
- 기존 분석 단점: {_fmt_items(our_cons_raw)}
- 원본 리뷰/설명 조각들: {' / '.join(our_reviews[:8])}

[경쟁 제품] {competitor_brand} {competitor_product}
- 제품 설명: {comp_data.get('description','')}
- 기존 분석 장점: {_fmt_items(comp_pros_raw)}
- 기존 분석 단점: {_fmt_items(comp_cons_raw)}
- 원본 리뷰/설명 조각들: {' / '.join(comp_reviews[:8])}

작성 원칙 (매우 중요 — 반드시 "왜 그런지" 이유를 포함):
1. 각 제품마다 TOP3 장점, TOP3 단점만 선별하세요. 산발적 언급보다 여러 자료에서 공통되는 포인트를 우선하세요.
2. 각 항목은 point(무엇이 좋은지/아쉬운지)와 reason(왜 그런지, 어떤 상황·어떤 고객에게 실질적 이점/불편이 되는지)을 반드시 함께 작성하세요. "추천합니다"처럼 이유 없는 결론만 쓰지 마세요.
3. 태그는 [소재/디자인] [기능성] [무게/휴대성] [안전성] [내구성] [편의성] [브랜드신뢰도] 중 선택하세요.
4. 실제 데이터에 없는 내용은 지어내지 마세요. 완결된 문장으로 다듬는 것은 필수입니다.
5. 가격은 절대 언급하지 마세요.
6. 영업 포인트는 반드시 위에서 도출한 TOP3 장단점(특히 reason)을 근거로, "고객이 경쟁사 제품을 언급하며 망설일 때 이렇게 응대하라"는 구체적 실전 화법으로 작성하세요. 일반론 금지, 왜 그렇게 응대해야 하는지 이유가 드러나야 합니다.
7. 원본 리뷰 조각 중 블로그 인사말("안녕하세요"), 홍보 해시태그, "~추천드립니다", "구매했어요" 같은 광고성/서두 문구는 절대 장점·단점 내용으로 쓰지 마세요.
8. 같은 제품의 장점과 단점이 서로 겹치는 내용이 되지 않게 하고, 자사와 경쟁사의 장점끼리도 동일한 문구가 반복되지 않게 각각 고유한 포인트로 작성하세요.

다음 JSON 형식으로만 답하세요 (설명 없이 JSON만):
{{
  "summary": "누가 봐도 한눈에 이해되는 비교 결론 한 문장",
  "competitor_pros": [{{"tag":"태그","point":"무엇이 좋은지","reason":"왜 좋은지 구체적으로"}}],
  "competitor_cons": [{{"tag":"태그","point":"무엇이 아쉬운지","reason":"왜 아쉬운지 구체적으로"}}],
  "our_pros": [{{"tag":"태그","point":"무엇이 좋은지","reason":"왜 좋은지 구체적으로"}}],
  "our_cons": [{{"tag":"태그","point":"무엇이 아쉬운지(최대2개)","reason":"왜 아쉬운지, 정직하게"}}],
  "selling_points": ["이번 비교의 구체적 장단점과 이유를 인용한 실전 응대 화법 3~5개"]
}}"""
            resp = requests.post(
                "https://api.anthropic.com/v1/messages",
                headers={"x-api-key": api_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
                json={"model": "claude-sonnet-4-6", "max_tokens": 1600,
                      "messages": [{"role": "user", "content": prompt}]},
                timeout=25,
            )
            if resp.status_code == 200:
                data = resp.json()
                text_blocks = [b.get('text','') for b in data.get('content',[]) if b.get('type')=='text']
                raw = ''.join(text_blocks).strip().replace('```json','').replace('```','').strip()
                parsed = json.loads(raw)
                required = ['competitor_pros','competitor_cons','our_pros','selling_points']
                if all(k in parsed for k in required):
                    parsed.setdefault('our_cons', [])
                    parsed.setdefault('summary', '')
                    return parsed
        except Exception:
            pass

    # Fallback: 원본 데이터(이미 point+reason 구조) 그대로 노출
    return {
        'summary': f"{our_brand} {our_product}와 {competitor_brand} {competitor_product}의 수집된 리뷰 데이터를 비교했습니다.",
        'competitor_pros': comp_pros_raw[:3] or [{'tag':'','point':f"{competitor_brand} {competitor_product}에 대한 장점 정보가 충분히 수집되지 않았습니다",'reason':''}],
        'competitor_cons': comp_cons_raw[:3] or [{'tag':'','point':"단점 정보가 충분히 수집되지 않았습니다",'reason':''}],
        'our_pros': our_pros_raw[:3] or [{'tag':'','point':fallback_strength,'reason':''}],
        'our_cons': our_cons_raw[:2],
        'selling_points': [
            "실물 체험 및 매장 상담을 통해 고객이 직접 비교해보고 확신을 갖도록 안내해주세요.",
            "국내 정식 유통 채널을 통한 신속한 A/S 대응 가능 여부를 강조해주세요.",
        ],
    }




# ── SNS 활용 매장 API ────────────────────────────────
@app.route("/api/sns/search", methods=["POST"])
@login_required
def api_sns_search():
    """네이버 블로그 검색으로 매장별 블로그 현황 자동 분석"""
    import urllib.request, urllib.parse, os, re
    from datetime import datetime as dt2

    sellers = request.json.get('sellers', [])
    if not sellers:
        return jsonify({'ok': False, 'msg': '매장명 없음'}), 400

    client_id     = os.environ.get('NAVER_CLIENT_ID',     'InqUUQfvWZN1rAZM4whk')
    client_secret = os.environ.get('NAVER_CLIENT_SECRET', 'fXYMLK1N1X')

    def strip_tags(s): return re.sub('<[^>]+>', '', s or '')

    def naver_endpoint(endpoint, query, display=20, sort='date'):
        url = (f'https://openapi.naver.com/v1/search/{endpoint}.json?query='
               + urllib.parse.quote(query) + f'&display={display}&sort={sort}')
        req = urllib.request.Request(url)
        req.add_header('X-Naver-Client-Id',     client_id)
        req.add_header('X-Naver-Client-Secret', client_secret)
        with urllib.request.urlopen(req, timeout=10) as r:
            return json.loads(r.read().decode('utf-8'))

    def naver_blog(query, display=20, sort='date'):
        """수정1: 네이버 블로그 + 카페(커뮤니티 후기) + 웹문서(인스타/유튜브 등 외부 링크 포함) 통합 검색
        — 네이버는 인스타그램 공식 검색 API를 제공하지 않으므로, 카페·웹문서 검색으로 커버리지를 넓힘"""
        total = 0
        all_items = []
        for endpoint, per_display in [('blog', display), ('cafearticle', max(display//2, 5)), ('webkr', max(display//2, 5))]:
            try:
                res = naver_endpoint(endpoint, query, display=per_display, sort=sort if endpoint=='blog' else 'sim')
                total += res.get('total', 0)
                for item in res.get('items', []):
                    item['_source'] = endpoint
                    all_items.append(item)
            except Exception:
                continue
        # 블로그 우선, 날짜 최신순 정렬 시도 (postdate 없는 카페/웹 항목은 뒤로)
        all_items.sort(key=lambda x: x.get('postdate', ''), reverse=True)
        return {'total': total, 'items': all_items}

    def parse_date(s):
        try: return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
        except: return ''

    def calc_score(total, latest, recent_30d, has_product):
        score = 0
        # 총 게시글 수 (최대 30점)
        if total >= 200: score += 30
        elif total >= 100: score += 25
        elif total >= 50:  score += 20
        elif total >= 20:  score += 15
        elif total >= 5:   score += 10
        elif total >= 1:   score += 5
        # 최근 글 날짜 (최대 40점)
        if latest:
            try:
                days = (dt2.now() - dt2.strptime(latest, '%Y-%m-%d')).days
                if   days <= 7:   score += 40
                elif days <= 14:  score += 35
                elif days <= 30:  score += 28
                elif days <= 60:  score += 20
                elif days <= 90:  score += 12
                elif days <= 180: score += 6
                elif days <= 365: score += 2
            except: pass
        # 30일 이내 게시글 수 (최대 20점)
        if   recent_30d >= 20: score += 20
        elif recent_30d >= 10: score += 16
        elif recent_30d >= 5:  score += 12
        elif recent_30d >= 3:  score += 8
        elif recent_30d >= 1:  score += 4
        # 제품 관련 포스팅 (10점)
        if has_product: score += 10
        return min(score, 100)

    def grade(score):
        if score >= 80: return 'A'
        if score >= 60: return 'B'
        if score >= 40: return 'C'
        if score >= 20: return 'D'
        return 'E'

    PRODUCT_KEYWORDS = ['엔픽스','줄즈','레카로','원더폴드','카오스','타프토이즈',
                        '유모차','카시트','보행기','웨건','하이체어','유아용품']
    # 자사 브랜드 키워드 (수정3: 전체 자사 브랜드)
    OUR_BRANDS = [
        '엔픽스', 'ENFIX', 'enfix',
        '줄즈', 'Joie', 'joie', 'JOIE',
        '레카로', 'Recaro', 'RECARO',
        '원더폴드', 'Wonderfold', 'WONDERFOLD',
        '카오스', 'Kaos',
        'ABC디자인', 'ABC Design', 'abc디자인',
        '타프토이즈', 'Taf Toys', 'TAFTOYS',
    ]

    results = []
    conn_insta = get_db()  # 수정1: 기등록된 인스타 인증 게시물 점수를 조회해 네이버 점수와 합산
    insta_bonus_map = {}
    for r in conn_insta.execute("SELECT seller_name, SUM(score) FROM instagram_post GROUP BY seller_name").fetchall():
        insta_bonus_map[r[0]] = r[1] or 0
    # 수정3: 매장에 등록된 공식 블로그 URL — 해당 도메인 글만 "자사 발행"으로 신뢰도 높게 인정
    own_blog_url_map = {}
    for r in conn_insta.execute("SELECT seller_name, blog_url FROM sns_info WHERE blog_url!=''").fetchall():
        own_blog_url_map[r[0]] = r[1]
    conn_insta.close()

    for seller in sellers[:50]:
        res = {'seller_name': seller, 'ok': False, 'error': '',
               'blog_total': 0, 'blog_latest': '', 'blog_recent_30d': 0,
               'blog_score': 0, 'blog_grade': 'E', 'blog_platform': '',
               'blog_has_product_post': 0, 'blog_recent_titles': '', 'blog_recent_links': '', 'blog_keywords': '',
               'blog_product_promo': '', 'blog_promo_count': 0, 'blog_promo_latest': '', 'instagram_handle': ''}
        try:
            clean = seller.replace('_', ' ').strip()
            d = naver_blog(clean, 20, 'date')
            total = d.get('total', 0)
            items = d.get('items', [])

            # 수정: 인스타그램 게시물을 더 적극적으로 포착하기 위한 전용 검색 (webkr에 instagram.com 도메인 힌트 추가)
            try:
                insta_d = naver_endpoint('webkr', f'{clean} instagram.com', display=10, sort='sim')
                for item in insta_d.get('items', []):
                    if 'instagram.com' in item.get('link',''):
                        item['_source'] = 'webkr'
                        items.append(item)
            except Exception:
                pass

            # 수정1: 인스타그램 프로필(계정) 링크 탐지 — /p/, /reel/, /stories/ 등 게시물 링크가 아닌
            # 프로필 URL(instagram.com/계정명/) 패턴을 찾아 매장 인스타 계정으로 저장
            instagram_handle = ''
            NON_PROFILE_PATHS = ('p', 'reel', 'reels', 'stories', 'explore', 'accounts', 'tv')
            for item in items:
                link = item.get('link', '')
                m = re.search(r'instagram\.com/([a-zA-Z0-9_.]+)/?', link)
                if m and m.group(1) not in NON_PROFILE_PATHS:
                    instagram_handle = m.group(1)
                    break

            # 최신 날짜
            dates = [parse_date(i.get('postdate','')) for i in items if i.get('postdate')]
            latest = dates[0] if dates else ''

            # 30일 이내 글 수
            now = dt2.now()
            recent_30d = 0
            for i in items:
                pd = parse_date(i.get('postdate',''))
                if pd:
                    try:
                        if (now - dt2.strptime(pd, '%Y-%m-%d')).days <= 30:
                            recent_30d += 1
                    except: pass

            # 제품 관련 포스팅 여부
            all_text = ' '.join(strip_tags(i.get('title','')) + strip_tags(i.get('description',''))
                                for i in items)
            has_product = any(kw in all_text for kw in PRODUCT_KEYWORDS)

            # 자사 제품 홍보 포스팅 감지 (수정3: 손님 후기 게시물 제외, 매장이 직접 발행한 글만)
            import json as _json
            REVIEW_PATTERNS = ['후기', '내돈내산', '다녀왔', '방문기', '솔직후기', '직접 사용', '내가 사용',
                                '내돈산', '리뷰', '써보니', '써봤', '사용기', '구매 후기', '구입 후기']
            own_domain = own_blog_url_map.get(seller, '')  # 등록된 공식 블로그가 있으면 그 도메인만 신뢰
            promo_posts = []
            for item in items:
                title = strip_tags(item.get('title',''))
                desc  = strip_tags(item.get('description',''))
                link  = item.get('link','')
                pdate = parse_date(item.get('postdate',''))
                combined = title + ' ' + desc
                found_brands = [b for b in OUR_BRANDS if b in combined]
                if not found_brands: continue
                if own_domain:
                    # 공식 블로그가 등록돼 있으면 그 도메인 글만 인정 (가장 신뢰도 높은 필터)
                    if own_domain not in link: continue
                else:
                    # 미등록 시 손님 후기 패턴 제외로 필터링
                    if any(p in combined for p in REVIEW_PATTERNS): continue
                promo_posts.append({
                    'date':   pdate,
                    'title':  title[:50],
                    'brands': found_brands,
                    'link':   link,
                    'is_instagram': 'instagram.com' in link,
                })
            promo_count  = len(promo_posts)
            promo_latest = promo_posts[0]['date'] if promo_posts else ''
            promo_json   = _json.dumps(promo_posts[:10], ensure_ascii=False)

            # 플랫폼 파악 (수정1: 블로그/카페/웹문서 통합 소스 기반, 인스타·유튜브 링크는 어느 소스에서든 감지)
            platforms = []
            for i in items:
                link = i.get('link', '')
                src = i.get('_source', 'blog')
                if 'instagram' in link:      platforms.append('인스타그램')
                elif 'youtube' in link:      platforms.append('유튜브')
                elif 'blog.naver' in link:   platforms.append('네이버블로그')
                elif 'tistory' in link:      platforms.append('티스토리')
                elif 'brunch' in link:       platforms.append('브런치')
                elif src == 'cafearticle':   platforms.append('네이버카페')
                elif src == 'webkr':         platforms.append('웹문서')
                else:                        platforms.append('기타블로그')
            from collections import Counter
            platform_str = ', '.join(f"{k}({v})" for k,v in Counter(platforms).most_common(5))

            # 최근 제목 5개 + 링크 (수정6: 클릭 시 원문으로 이동)
            titles = [strip_tags(i.get('title',''))[:40] for i in items[:5]]
            links  = [i.get('link','') for i in items[:5]]

            # 키워드 추출 (제목에서)
            all_titles = ' '.join(strip_tags(i.get('title','')) for i in items[:10])
            found_kws = [kw for kw in PRODUCT_KEYWORDS if kw in all_titles]

            # 자사 홍보 포스팅이 있으면 +15점 가산
            score = calc_score(total, latest, recent_30d, has_product)
            if promo_count >= 5:   score = min(score + 15, 100)
            elif promo_count >= 2: score = min(score + 10, 100)
            elif promo_count >= 1: score = min(score + 5,  100)

            # 수정1: 인스타그램 인증 게시물 점수를 별도 항목이 아닌 하나의 통합 점수로 합산
            score = min(score + insta_bonus_map.get(seller, 0), 100)

            res.update({
                'ok': True,
                'blog_total': total,
                'blog_latest': latest,
                'blog_recent_30d': recent_30d,
                'blog_has_product_post': 1 if has_product else 0,
                'blog_platform': platform_str,
                'blog_recent_titles': ' | '.join(titles),
                'blog_recent_links': ' | '.join(links),
                'instagram_handle': instagram_handle,
                'blog_keywords': ', '.join(found_kws),
                'blog_score': score,
                'blog_grade': grade(score),
                'blog_product_promo': promo_json,
                'blog_promo_count': promo_count,
                'blog_promo_latest': promo_latest,
            })
        except urllib.error.HTTPError as e:
            res['error'] = f'HTTP {e.code}'
        except Exception as e:
            res['error'] = str(e)[:60]
        results.append(res)

    return jsonify({'results': results, 'ok': True})


@app.route("/api/sns/save-search", methods=["POST"])
@login_required
def api_sns_save_search():
    """검색 결과 DB 저장 — UPSERT 방식, 단일 커밋으로 속도 개선"""
    results = request.json.get('results', [])
    conn = get_db()
    updated = 0
    now = datetime.now().strftime('%Y-%m-%d %H:%M')

    # 마이그레이션: 필요한 컬럼 없으면 추가
    try:
        existing_cols = [r[1] for r in conn.execute("PRAGMA table_info(sns_info)").fetchall()]
        new_cols = {
            'blog_platform': 'TEXT DEFAULT ""', 'blog_total_posts': 'INTEGER DEFAULT 0',
            'blog_latest_date': 'TEXT DEFAULT ""', 'blog_recent_30d': 'INTEGER DEFAULT 0',
            'blog_recent_titles': 'TEXT DEFAULT ""', 'blog_recent_links': 'TEXT DEFAULT ""',
            'blog_keywords': 'TEXT DEFAULT ""',
            'blog_has_product_post': 'INTEGER DEFAULT 0', 'blog_grade': 'TEXT DEFAULT ""',
            'blog_score': 'INTEGER DEFAULT 0', 'last_searched': 'TEXT DEFAULT ""',
            'blog_product_promo': 'TEXT DEFAULT ""', 'blog_promo_count': 'INTEGER DEFAULT 0',
            'blog_promo_latest': 'TEXT DEFAULT ""', 'instagram_handle': 'TEXT DEFAULT ""',
        }
        for col, typ in new_cols.items():
            if col not in existing_cols:
                conn.execute(f"ALTER TABLE sns_info ADD COLUMN {col} {typ}")
        conn.commit()
    except Exception as e:
        app.logger.error(f"sns migration error: {e}")

    # 기존 데이터 미리 로드 (SELECT 루프 방지 → 속도 개선)
    existing = {r[0] for r in conn.execute("SELECT seller_name FROM sns_info").fetchall()}

    conn.execute("BEGIN")  # 트랜잭션 시작
    for r in results:
        if not r.get('ok'): continue
        seller = r.get('seller_name','')
        if not seller: continue
        try:
            vals = (
                r.get('blog_total', r.get('blog_total_posts', 0)),
                r.get('blog_latest', r.get('blog_latest_date', '')),
                r.get('blog_recent_30d', 0),
                r.get('blog_has_product_post', 0),
                r.get('blog_platform', ''),
                r.get('blog_recent_titles', ''),
                r.get('blog_recent_links', ''),
                r.get('blog_keywords', ''),
                r.get('blog_score', 0),
                r.get('blog_grade', ''),
                r.get('blog_product_promo', ''),
                r.get('blog_promo_count', 0),
                r.get('blog_promo_latest', ''),
                r.get('instagram_handle', ''),
                now, now,
            )
            if seller in existing:
                conn.execute("""UPDATE sns_info SET
                    blog_total_posts=?, blog_latest_date=?, blog_recent_30d=?,
                    blog_has_product_post=?, blog_platform=?, blog_recent_titles=?,
                    blog_recent_links=?, blog_keywords=?, blog_score=?, blog_grade=?,
                    blog_product_promo=?, blog_promo_count=?, blog_promo_latest=?,
                    instagram_handle=?, last_searched=?, updated_at=?
                    WHERE seller_name=?""", (*vals, seller))
            else:
                conn.execute("""INSERT INTO sns_info
                    (blog_total_posts, blog_latest_date, blog_recent_30d, blog_has_product_post,
                     blog_platform, blog_recent_titles, blog_recent_links, blog_keywords, blog_score, blog_grade,
                     blog_product_promo, blog_promo_count, blog_promo_latest, instagram_handle, last_searched, updated_at,
                     seller_name)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (*vals, seller))
                existing.add(seller)
            updated += 1
        except Exception as e:
            app.logger.error(f"sns save error {seller}: {e}")
    conn.execute("COMMIT")
    conn.close()
    return jsonify({'ok': True, 'updated': updated})


@app.route("/api/sns/list")
@login_required
def api_sns_list():
    year = request.args.get("year", str(datetime.now().year))
    conn = get_db()
    sellers = [r[0] for r in conn.execute(
        f"SELECT DISTINCT real_seller FROM sales_data "
        f"WHERE real_seller!='' AND sale_date LIKE '{year}%' ORDER BY real_seller").fetchall()]
    sns_map = {r['seller_name']: dict(r) for r in conn.execute("SELECT * FROM sns_info").fetchall()}
    sales_map = {r[0]: r[1] for r in conn.execute(
        f"SELECT real_seller, SUM(total) FROM sales_data "
        f"WHERE sale_date LIKE '{year}%' AND real_seller!='' GROUP BY real_seller").fetchall()}
    # 수정1+2: 인스타그램 인증 게시물 — 건수 + 누적 조회수 (점수는 이미 blog_score에 통합되어 있으므로 별도 가산 없음)
    insta_map = {}
    for r in conn.execute("""SELECT seller_name, COUNT(*) cnt, SUM(score) bonus, SUM(COALESCE(view_count,0)) views
                              FROM instagram_post GROUP BY seller_name""").fetchall():
        insta_map[r[0]] = {'cnt': r[1], 'bonus': r[2] or 0, 'views': r[3] or 0}
    conn.close()
    result = []
    for s in sellers:
        info = sns_map.get(s, {})
        insta = insta_map.get(s, {'cnt': 0, 'bonus': 0, 'views': 0})
        # 수정1: blog_score는 이미 "전체 분석" 시점에 인스타 인증 점수까지 합산된 통합 점수
        total_score = info.get('blog_score', info.get('sns_score', 0))
        result.append({
            'seller_name':           s,
            'blog_url':              info.get('blog_url',''),
            'blog_name':             info.get('blog_name',''),
            'blog_platform':         info.get('blog_platform',''),
            'blog_total_posts':      info.get('blog_total_posts', 0),
            'blog_latest_date':      info.get('blog_latest_date',''),
            'blog_recent_30d':       info.get('blog_recent_30d', 0),
            'blog_recent_titles':    info.get('blog_recent_titles',''),
            'blog_recent_links':     info.get('blog_recent_links',''),
            'blog_keywords':         info.get('blog_keywords',''),
            'blog_score':   total_score,
            'blog_grade':   info.get('blog_grade',''),
            'last_searched':info.get('last_searched', info.get('last_checked','')),
            'memo':         info.get('memo',''),
            'year_sales':   sales_map.get(s, 0),
            'blog_product_promo':  info.get('blog_product_promo',''),
            'blog_promo_count':    info.get('blog_promo_count', 0),
            'blog_promo_latest':   info.get('blog_promo_latest',''),
            'instagram_handle':     info.get('instagram_handle',''),
            'instagram_post_count': insta['cnt'],
            'instagram_views':      insta['views'],
            'total_score':          total_score,
        })
    result.sort(key=lambda x: -x['year_sales'])
    return jsonify(result)


# ── 인스타그램 피드/릴스 인증 & 가산점 API ────────────────────────
OUR_BRAND_LIST = ['엔픽스','줄즈','레카로','원더폴드','카오스','ABC디자인','타프토이즈']

@app.route("/api/sns/instagram/list")
@login_required
def api_instagram_list():
    """등록된 인스타그램 인증 게시물 목록 (전체 또는 특정 매장)"""
    seller = request.args.get('seller', '').strip()
    conn = get_db()
    if seller:
        rows = [dict(r) for r in conn.execute(
            "SELECT * FROM instagram_post WHERE seller_name=? ORDER BY posted_date DESC, id DESC",
            (seller,)).fetchall()]
    else:
        rows = [dict(r) for r in conn.execute(
            "SELECT * FROM instagram_post ORDER BY created_at DESC LIMIT 200").fetchall()]
    conn.close()
    return jsonify(rows)


@app.route("/api/sns/instagram/add", methods=["POST"])
@login_required
def api_instagram_add():
    """매장의 자사 제품 인스타 피드/릴스 인증 등록 → 자동 가산점 부여
    점수 기준: 릴스(영상) 8점, 피드(사진) 5점 — 브랜드 태그가 명확히 확인된 경우 기본값, 필요시 수동 조정 가능"""
    d = request.json or {}
    seller = d.get('seller_name', '').strip()
    post_url = d.get('post_url', '').strip()
    post_type = d.get('post_type', '피드').strip()
    brand = d.get('brand', '').strip()
    product_name = d.get('product_name', '').strip()
    posted_date = d.get('posted_date', '').strip()
    note = d.get('note', '').strip()

    if not seller or not post_url:
        return jsonify({'ok': False, 'msg': '매장명과 게시물 링크는 필수입니다'}), 400
    if 'instagram.com' not in post_url:
        return jsonify({'ok': False, 'msg': '인스타그램 게시물 링크(instagram.com)만 등록 가능합니다'}), 400

    score = d.get('score')
    if score is None:
        score = 8 if post_type == '릴스' else 5
    score = int(score)

    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    registered_by = session.get('user', {}).get('name', '')

    conn = get_db()
    try:
        conn.execute("""INSERT INTO instagram_post
            (seller_name, post_url, post_type, brand, product_name, score, note, registered_by, posted_date, created_at)
            VALUES(?,?,?,?,?,?,?,?,?,?)""",
            (seller, post_url, post_type, brand, product_name, score, note, registered_by,
             posted_date or now[:10], now))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'ok': False, 'msg': '이미 등록된 게시물 링크입니다'}), 400
    conn.close()
    return jsonify({'ok': True, 'score': score})


def _parse_view_count(raw):
    """조회수 텍스트를 정수로 변환 — '144', '2.2만', '1.2천' 등 지원. 파싱 불가 시 (None, 원본텍스트) 반환"""
    if raw is None:
        return None, ''
    if isinstance(raw, (int, float)):
        return int(raw), str(raw)
    s = str(raw).strip()
    if not s:
        return None, ''
    import re as _re_vc
    m = _re_vc.match(r'^([\d.]+)\s*(만|천)?$', s)
    if m:
        num = float(m.group(1))
        unit = m.group(2)
        if unit == '만': num *= 10000
        elif unit == '천': num *= 1000
        return int(num), s
    # 순수 숫자(콤마 포함)
    m2 = _re_vc.match(r'^[\d,]+$', s)
    if m2:
        try: return int(s.replace(',', '')), s
        except: pass
    # 파싱 불가한 텍스트(예: "릴스가 아닌 피드 업로드") — 노트로만 보존
    return None, s


@app.route("/api/sns/instagram/upload", methods=["POST"])
@login_required
def api_instagram_upload():
    """SNS 협업 캠페인 엑셀 업로드 — 참여(O) 표시된 매장을 자동으로 인스타그램 인증 등록 + 가산점 부여
    지원 형식: 업체구분/거래처코드/거래처명/실적용거래처명 + (참여/URL/조회수) 반복 컬럼 블록 (1차, 2차, ... 다회차 협업 지원)"""
    import re as _re_up

    if 'file' not in request.files:
        return jsonify({'ok': False, 'msg': '파일이 없습니다'}), 400
    f = request.files['file']
    fname = f.filename or ''
    data = f.read()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:
        return jsonify({'ok': False, 'msg': f'파일을 읽을 수 없습니다: {e}'}), 400

    ws = wb[wb.sheetnames[0]]
    raw_rows = list(ws.iter_rows(values_only=True))

    # 캠페인명(타이틀) 탐색 — "※"로 시작하는 행
    campaign_name = ''
    for row in raw_rows[:6]:
        for c in row:
            if c and isinstance(c, str) and '※' in c:
                campaign_name = c.replace('※', '').strip()
                break
        if campaign_name: break
    if not campaign_name:
        campaign_name = fname.rsplit('.', 1)[0]

    # 브랜드 추정: 시트명 우선 (예: "2026_레카로" → "레카로"), 실패 시 파일명에서 추정
    sheet_name = ws.title
    brand_guess = ''
    for b in ['줄즈','레카로','ABC디자인','원더폴드','카오스','엔픽스','타프토이즈']:
        if b in sheet_name or b in campaign_name or b in fname:
            brand_guess = b; break

    # 제품명 추정: 파일명 첫 토큰 (예: "토론_레드닷_..." → "토론")
    product_guess = ''
    fname_tokens = _re_up.split(r'[_\-]', fname.rsplit('.', 1)[0])
    if fname_tokens and fname_tokens[0] and fname_tokens[0] not in ('xlsx',):
        product_guess = fname_tokens[0]

    # 헤더 행(실적용거래처명 포함) 탐색
    header_row_idx = None
    for i, row in enumerate(raw_rows[:15]):
        vals = [str(c).strip() for c in row if c is not None]
        if any('실적용거래처명' in v for v in vals):
            header_row_idx = i; break
    if header_row_idx is None:
        return jsonify({'ok': False, 'msg': '"실적용거래처명" 헤더를 찾을 수 없습니다. 파일 형식을 확인해주세요'}), 400

    header = list(raw_rows[header_row_idx])
    subheader = list(raw_rows[header_row_idx+1]) if header_row_idx+1 < len(raw_rows) else []
    max_col = len(header)

    def hv(row, ci): return str(row[ci]).strip() if ci < len(row) and row[ci] is not None else ''

    # 매장명 컬럼(실적용거래처명) 위치
    seller_col = next((ci for ci in range(max_col) if '실적용거래처명' in hv(header, ci)), None)
    if seller_col is None:
        return jsonify({'ok': False, 'msg': '매장명 컬럼을 찾을 수 없습니다'}), 400

    # 참여/URL/조회수 컬럼 그룹(회차) 자동 탐지
    rounds = []
    consumed = set()
    for ci in range(seller_col+1, max_col):
        if ci in consumed: continue
        if hv(subheader, ci) == '참여':
            participate_col = ci
            url_col = ci+1
            view_col = None
            for vc in range(ci+1, min(ci+4, max_col)):
                if hv(header, vc) == '조회수':
                    view_col = vc; break
            rounds.append({'label': hv(header, ci) or f'{len(rounds)+1}차',
                            'participate_col': participate_col, 'url_col': url_col, 'view_col': view_col})
            consumed.update([participate_col, url_col])
            if view_col: consumed.add(view_col)
        elif hv(header, ci) == '조회수' and ci not in consumed:
            url_col = ci-1
            if url_col not in consumed and url_col > seller_col:
                rounds.append({'label': hv(header, url_col) or f'{len(rounds)+1}차',
                                'participate_col': None, 'url_col': url_col, 'view_col': ci})
                consumed.update([url_col, ci])

    if not rounds:
        return jsonify({'ok': False, 'msg': '참여/URL 컬럼 구조를 인식하지 못했습니다'}), 400

    conn = get_db()
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    registered_by = session.get('user', {}).get('name', '') + ' (엑셀 업로드)'

    total_rows = 0
    matched = 0
    registered = 0
    skipped_dup = 0
    unmatched_stores = []

    data_start = header_row_idx + 2
    for row in raw_rows[data_start:]:
        if seller_col >= len(row): continue
        seller_raw = row[seller_col]
        if not seller_raw: continue
        total_rows += 1

        # 실적용거래처명 정규화 (언더바 → 공백, 별칭 매핑 재적용)
        seller_norm = str(seller_raw).replace('_', ' ').strip()
        seller_norm = SELLER_ALIAS.get(seller_norm, seller_norm)
        seller_norm = resolve_seller(seller_norm)
        matched += 1

        for rnd in rounds:
            pcol, ucol, vcol = rnd['participate_col'], rnd['url_col'], rnd['view_col']
            url_val = row[ucol] if ucol < len(row) else None
            if not url_val: continue
            # 참여 컬럼이 있으면 O 표시 확인, 없으면 URL 존재 자체를 참여로 간주
            if pcol is not None:
                p_val = str(row[pcol]).strip() if pcol < len(row) and row[pcol] is not None else ''
                if p_val.upper() != 'O': continue

            url_str = str(url_val).strip()
            if 'instagram.com' not in url_str: continue

            view_raw = row[vcol] if vcol is not None and vcol < len(row) else None
            view_count, view_raw_text = _parse_view_count(view_raw)

            post_type = '릴스' if '/reel/' in url_str else '피드'
            note = ''
            if view_raw_text and view_count is None:
                note = view_raw_text
                if '피드' in view_raw_text: post_type = '피드'
                elif '릴스' in view_raw_text: post_type = '릴스'

            score = 8 if post_type == '릴스' else 5

            try:
                conn.execute("""INSERT INTO instagram_post
                    (seller_name, post_url, post_type, brand, product_name, score, note,
                     registered_by, posted_date, created_at, view_count, view_count_raw, campaign_name)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (seller_norm, url_str, post_type, brand_guess, product_guess, score,
                     f"{rnd['label']} 협업" + (f' — {note}' if note else ''),
                     registered_by, now[:10], now, view_count, view_raw_text, campaign_name))
                conn.commit()
                registered += 1
            except sqlite3.IntegrityError:
                # 이미 등록된 링크 — 조회수만 최신화
                conn.execute("""UPDATE instagram_post SET view_count=?, view_count_raw=?
                    WHERE seller_name=? AND post_url=?""",
                    (view_count, view_raw_text, seller_norm, url_str))
                conn.commit()
                skipped_dup += 1

    conn.close()
    return jsonify({
        'ok': True, 'campaign_name': campaign_name, 'brand': brand_guess, 'product_name': product_guess,
        'rounds_detected': [r['label'] for r in rounds],
        'total_rows': total_rows, 'registered': registered, 'updated_existing': skipped_dup,
    })


@app.route("/api/sns/instagram/<int:pid>", methods=["DELETE"])
@login_required
def api_instagram_delete(pid):
    conn = get_db()
    conn.execute("DELETE FROM instagram_post WHERE id=?", (pid,))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

# ══════════════════════════════════════════════════════
# ── 업무 탭 API ──────────────────────────────────────
# ══════════════════════════════════════════════════════

MAJOR_REGION_MAP = {
    # 대권역 매핑: 세부 지역 → 수도권/충청권/영남권/호남권/강원기타
    '서울': '수도권', '경기북부': '수도권', '경기남부': '수도권', '인천': '수도권',
    '대전': '충청권', '충북': '충청권', '충남': '충청권', '세종': '충청권',
    '부산': '영남권', '대구': '영남권', '울산': '영남권', '경북': '영남권', '경남': '영남권',
    '광주': '호남권', '전북': '호남권', '전남': '호남권',
    '강원': '강원/기타', '제주': '강원/기타', '기타': '강원/기타',
}
MAJOR_REGION_ORDER = ['수도권', '충청권', '영남권', '호남권', '강원/기타']


def _major_region(seller_name):
    small = detect_region_from_name(seller_name) or '기타'
    return MAJOR_REGION_MAP.get(small, '강원/기타')


@app.route("/api/work/managers")
@login_required
def api_work_managers():
    """담당자 목록 (branches.manager 기준, 매장 수 포함)"""
    conn = get_db()
    rows = conn.execute("""
        SELECT manager, COUNT(*) cnt FROM branches
        WHERE manager IS NOT NULL AND manager != '' AND status='운영중'
        GROUP BY manager ORDER BY cnt DESC
    """).fetchall()
    conn.close()
    return jsonify([{'manager': r[0], 'store_count': r[1]} for r in rows])


@app.route("/api/work/stores-by-manager")
@login_required
def api_work_stores_by_manager():
    """담당자별 매장 리스트 (지역/최근매출 포함)"""
    manager = request.args.get('manager', '').strip()
    year = request.args.get('year', str(datetime.now().year))
    conn = get_db()
    q = "SELECT name, region, phone FROM branches WHERE status='운영중'"
    params = []
    if manager:
        q += " AND manager=?"; params.append(manager)
    q += " ORDER BY name"
    stores = conn.execute(q, params).fetchall()

    sales_map = {r[0]: r[1] for r in conn.execute(
        f"SELECT real_seller, SUM(total) FROM sales_data WHERE sale_date LIKE '{year}%' GROUP BY real_seller").fetchall()}
    conn.close()

    result = []
    for name, region, phone in stores:
        result.append({
            'name': name, 'region': region or detect_region_from_name(name) or '',
            'major_region': _major_region(name), 'phone': phone or '',
            'year_sales': sales_map.get(name, 0),
        })
    result.sort(key=lambda x: -x['year_sales'])
    return jsonify(result)


def _work_kpi_data(year, month, manager):
    """월간 목표 vs 실적 KPI 계산 (API/엑셀 export 공용 헬퍼)"""
    conn = get_db()

    # 담당 매장 목록
    q = "SELECT name FROM branches WHERE status='운영중'"
    params = []
    if manager: q += " AND manager=?"; params.append(manager)
    managed_stores = [r[0] for r in conn.execute(q, params).fetchall()]
    managed_set = set(managed_stores)
    total_managed = len(managed_stores) or 1

    ym = f"{year}-{month:02d}"

    # 1) 방문 매장 수: store_visit_report 기준 실제 방문 매장 (담당자 필드 매칭)
    visit_q = "SELECT DISTINCT store_name FROM store_visit_report WHERE visit_date LIKE ?"
    visit_params = [f"{ym}%"]
    if manager: visit_q += " AND manager=?"; visit_params.append(manager)
    visited_stores = [r[0] for r in conn.execute(visit_q, visit_params).fetchall()]
    visit_actual = len([s for s in visited_stores if not manager or s in managed_set])
    visit_target = max(round(total_managed * 0.8), 1)  # 담당 매장의 80% 방문을 목표로 산정

    # 2) 신규 진열 확보: display_record 중 이달 applied_date가 있고 has_display=1인 매장 수 (담당 매장 한정)
    disp_q = """SELECT DISTINCT dr.seller_name FROM display_record dr
                WHERE dr.has_display=1 AND dr.applied_date LIKE ?"""
    disp_params = [f"{ym}%"]
    disp_rows = [r[0] for r in conn.execute(disp_q, disp_params).fetchall()]
    disp_actual = len([s for s in disp_rows if not manager or s in managed_set])
    disp_target = max(round(total_managed * 0.15), 3)  # 담당 매장의 약 15% 신규 진열 목표

    # 3) 프로모션 참여 매장: 이달 기간이 걸치는 캠페인에 has_display=1인 담당 매장 수
    promo_q = """SELECT DISTINCT dr.seller_name FROM display_record dr
                 JOIN display_campaign dc ON dr.campaign_id=dc.id
                 WHERE dr.has_display=1 AND dc.period_start<=? AND dc.period_end>=?"""
    month_end = f"{year}-{month:02d}-31"
    promo_rows = [r[0] for r in conn.execute(promo_q, (month_end, f"{ym}-01")).fetchall()]
    promo_actual = len([s for s in promo_rows if not manager or s in managed_set])
    promo_target = max(round(total_managed * 0.2), 3)

    # 4) 발주 전환율: 이달 매출 발생 담당 매장 비율
    sales_q = f"SELECT DISTINCT real_seller FROM sales_data WHERE sale_date LIKE ? AND total>0"
    sales_rows = [r[0] for r in conn.execute(sales_q, (f"{ym}%",)).fetchall()]
    ordered_actual = len([s for s in sales_rows if not manager or s in managed_set])
    conversion_actual = round(ordered_actual / total_managed * 100, 1)
    conversion_target = 70.0  # 업계 통상 목표치

    # 전월 실적 (증감 비교용)
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    prev_ym = f"{prev_year}-{prev_month:02d}"
    prev_visit = len(set(r[0] for r in conn.execute(
        "SELECT DISTINCT store_name FROM store_visit_report WHERE visit_date LIKE ?"
        + (" AND manager=?" if manager else ""),
        [f"{prev_ym}%"] + ([manager] if manager else [])).fetchall()
        if not manager or r[0] in managed_set))

    conn.close()

    def _item(key, label, target, actual, prev, unit=''):
        rate = round(actual / target * 100, 1) if target else 0
        delta = actual - prev
        return {'key': key, 'label': label, 'target': target, 'actual': actual, 'rate': rate,
                'prev': prev, 'delta': delta, 'unit': unit, 'is_manual': False}

    items = [
        _item('visit_count', '방문 매장 수', visit_target, visit_actual, prev_visit, '개'),
        _item('new_display', '신규 진열 확보', disp_target, disp_actual, 0, '건'),
        _item('promo_join', '프로모션 참여 매장', promo_target, promo_actual, 0, '개'),
        _item('conversion_rate', '발주 전환율', conversion_target, conversion_actual, 0, '%'),
    ]

    # 수정2-1: 수기 입력값이 있으면 자동계산값을 덮어씀 (목표·실적 모두 사용자가 직접 관리 가능)
    conn2 = get_db()
    manual_rows = conn2.execute(
        "SELECT item_key, target, actual FROM work_kpi_manual WHERE year=? AND month=? AND manager=?",
        (year, month, manager or '전체')).fetchall()
    conn2.close()
    manual_map = {r[0]: {'target': r[1], 'actual': r[2]} for r in manual_rows}
    for it in items:
        m = manual_map.get(it['key'])
        if m:
            if m['target'] is not None:
                it['target'] = m['target']; it['is_manual'] = True
            if m['actual'] is not None:
                it['actual'] = m['actual']; it['is_manual'] = True
            it['rate'] = round(it['actual'] / it['target'] * 100, 1) if it['target'] else 0

    return {'ok': True, 'items': items, 'total_managed': total_managed}


@app.route("/api/work/kpi")
@login_required
def api_work_kpi():
    year = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))
    manager = request.args.get('manager', '').strip()
    return jsonify(_work_kpi_data(year, month, manager))


@app.route("/api/work/kpi", methods=["POST"])
@login_required
def api_work_kpi_save():
    """월간 목표/실적 수기 입력 저장 (수정2-1)"""
    d = request.json or {}
    year = int(d.get('year')); month = int(d.get('month'))
    manager = d.get('manager', '전체').strip() or '전체'
    item_key = d.get('item_key', '').strip()
    target = d.get('target', None)
    actual = d.get('actual', None)
    if not item_key:
        return jsonify({'ok': False, 'msg': '항목 정보가 필요합니다'}), 400
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    conn = get_db()
    conn.execute("""INSERT INTO work_kpi_manual (year, month, manager, item_key, target, actual, updated_at)
        VALUES(?,?,?,?,?,?,?)
        ON CONFLICT(year, month, manager, item_key) DO UPDATE SET
        target=excluded.target, actual=excluded.actual, updated_at=excluded.updated_at""",
        (year, month, manager, item_key,
         float(target) if target not in (None, '') else None,
         float(actual) if actual not in (None, '') else None, now))
    conn.commit(); conn.close()
    return jsonify({'ok': True})


def _work_brand_perf_data(year, month, manager):
    """브랜드별 실적 계산 — 목표매출은 제품별로 '작년 동월 실적 × 1.05(5% 성장)'을 산정한 뒤 브랜드 단위로 합산 (API/엑셀 export 공용)"""
    conn = get_db()

    q = "SELECT name FROM branches WHERE status='운영중'"
    params = []
    if manager: q += " AND manager=?"; params.append(manager)
    managed_stores = [r[0] for r in conn.execute(q, params).fetchall()]

    def _product_sales(y, m, stores):
        """제품별(item_name) 매출 집계"""
        if not stores:
            return {}
        placeholders = ','.join('?' for _ in stores)
        rows = conn.execute(f"""
            SELECT item_group, item_name, SUM(total) t FROM sales_data
            WHERE sale_date LIKE ? AND real_seller IN ({placeholders})
            GROUP BY item_group, item_name""",
            [f"{y}-{m:02d}%"] + stores).fetchall()
        out = {}
        for grp, name, total in rows:
            out[(grp, name)] = {'brand': remap_group(grp, name), 'total': total or 0}
        return out

    actual_products = _product_sales(year, month, managed_stores)
    last_year_products = _product_sales(year - 1, month, managed_stores)  # 작년 동월

    # 실적: 브랜드별 합산
    actual_map = {}
    for (grp, name), v in actual_products.items():
        actual_map[v['brand']] = actual_map.get(v['brand'], 0) + v['total']

    # 목표: 제품별로 작년 동월 실적 × 1.05 산정 후 브랜드 단위로 합산
    target_map = {}
    products_without_history = 0
    for (grp, name), v in actual_products.items():
        last_year_val = last_year_products.get((grp, name), {}).get('total', 0)
        products_without_history += (1 if last_year_val == 0 else 0)
        target_map[v['brand']] = target_map.get(v['brand'], 0) + round(last_year_val * 1.05)
    # 작년엔 있었지만 올해 실적이 아직 없는 제품도 목표에는 반영 (성장 목표는 유지되어야 하므로)
    for (grp, name), v in last_year_products.items():
        if (grp, name) not in actual_products:
            b = v['brand']
            target_map[b] = target_map.get(b, 0) + round(v['total'] * 1.05)

    conn.close()

    all_brands = set(actual_map.keys()) | set(target_map.keys()) | set(BRAND_ORDER)
    result = []
    for b in BRAND_ORDER + sorted(all_brands - set(BRAND_ORDER)):
        if b not in all_brands: continue
        target = target_map.get(b, 0)
        actual = actual_map.get(b, 0)
        rate = round(actual / target * 100, 1) if target else (100.0 if actual > 0 else 0)
        result.append({'brand': b, 'target': target, 'actual': actual, 'rate': rate})
    return {'ok': True, 'items': result,
                     'basis': f'{year-1}년 {month}월 실적 × 1.05 (전년 동기 대비 5% 성장 목표)'}


@app.route("/api/work/brand-performance")
@login_required
def api_work_brand_performance():
    year = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))
    manager = request.args.get('manager', '').strip()
    return jsonify(_work_brand_perf_data(year, month, manager))


def _work_coverage_data(year, month, manager):
    """매장 방문 커버리지 계산 (API/엑셀 export 공용) — 담당 매장수는 자동 계산, 방문 매장수는 수기 입력값 반영, 담당 매장 세부 리스트 포함"""
    conn = get_db()

    q = "SELECT name FROM branches WHERE status='운영중'"
    params = []
    if manager: q += " AND manager=?"; params.append(manager)
    stores = [r[0] for r in conn.execute(q, params).fetchall()]

    sales_map = {r[0]: r[1] for r in conn.execute(
        f"SELECT real_seller, SUM(total) FROM sales_data WHERE sale_date LIKE '{year}-{month:02d}%' GROUP BY real_seller").fetchall()}

    region_stores = {}  # mr -> [매장명, ...]
    for s in stores:
        mr = _major_region(s)
        region_stores.setdefault(mr, []).append(s)

    visited_q = "SELECT region, visited_count, note FROM work_visit_coverage WHERE year=? AND month=? AND manager=?"
    visited_map = {r[0]: {'visited': r[1], 'note': r[2]} for r in
                   conn.execute(visited_q, (year, month, manager or '전체')).fetchall()}
    conn.close()

    result = []
    for mr in MAJOR_REGION_ORDER:
        store_list = sorted(region_stores.get(mr, []))
        managed = len(store_list)
        if managed == 0 and mr not in visited_map: continue
        v = visited_map.get(mr, {'visited': 0, 'note': ''})
        rate = round(v['visited'] / managed * 100, 1) if managed else 0
        # 수정2-3: 담당 매장 세부 리스트 (매장명 + 이달 매출)
        store_detail = [{'name': s, 'sales': sales_map.get(s, 0)} for s in store_list]
        result.append({'region': mr, 'managed': managed, 'visited': v['visited'],
                        'rate': rate, 'note': v['note'] or '', 'stores': store_detail})
    return {'ok': True, 'items': result}


@app.route("/api/work/visit-coverage")
@login_required
def api_work_visit_coverage():
    year = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))
    manager = request.args.get('manager', '').strip()
    return jsonify(_work_coverage_data(year, month, manager))


@app.route("/api/work/visit-coverage", methods=["POST"])
@login_required
def api_work_visit_coverage_save():
    """방문 매장수 수기 입력 저장"""
    d = request.json or {}
    year = int(d.get('year')); month = int(d.get('month'))
    manager = d.get('manager', '전체').strip() or '전체'
    region = d.get('region', '').strip()
    visited = int(d.get('visited', 0))
    note = d.get('note', '').strip()
    if not region:
        return jsonify({'ok': False, 'msg': '권역 정보가 필요합니다'}), 400
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    conn = get_db()
    conn.execute("""INSERT INTO work_visit_coverage (year, month, manager, region, visited_count, note, updated_at)
        VALUES(?,?,?,?,?,?,?)
        ON CONFLICT(year, month, manager, region) DO UPDATE SET
        visited_count=excluded.visited_count, note=excluded.note, updated_at=excluded.updated_at""",
        (year, month, manager, region, visited, note, now))
    conn.commit(); conn.close()
    return jsonify({'ok': True})


@app.route("/api/work/promotion/upload", methods=["POST"])
@login_required
def api_work_promotion_upload():
    """행사·프로모션 공지 엑셀 업로드 → 자유 텍스트 파싱해서 캘린더 항목으로 자동 요약 등록"""
    import re as _re_promo

    if 'file' not in request.files:
        return jsonify({'ok': False, 'msg': '파일이 없습니다'}), 400
    f = request.files['file']
    fname = f.filename or ''
    data = f.read()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:
        return jsonify({'ok': False, 'msg': f'파일을 읽을 수 없습니다: {e}'}), 400

    now = datetime.now()
    events = []

    date_pattern = _re_promo.compile(
        r'(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일부터\s*(\d{1,2})일까지\s*(?:/\s*(.+))?')

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 채널/대상 추정 (시트명 기준: "백화점_서양" → 서양네트웍스 백화점 채널, "백화점_링크맘" → 링크맘)
        target_channel = sheet_name.replace('_', ' ')

        # 브랜드/행사명 타이틀 (2행 부근 "* ~~ 행사" 텍스트)
        title_text = ''
        for row in ws.iter_rows(min_row=1, max_row=4, values_only=True):
            for c in row:
                if c and isinstance(c, str) and c.strip().startswith('*'):
                    title_text = c.strip().lstrip('*').strip()
                    break
            if title_text: break

        # 브랜드 감지: 타이틀에서 우선, 없으면 시트 전체 텍스트(예: "* 줄즈 공급 관련") 스캔
        brand_guess = ''
        for b in ['줄즈','레카로','ABC디자인','원더폴드','카오스','엔픽스','타프토이즈']:
            if b in title_text: brand_guess = b; break
        if not brand_guess:
            for row in ws.iter_rows(values_only=True):
                for c in row:
                    if not c or not isinstance(c, str): continue
                    for b in ['줄즈','레카로','ABC디자인','원더폴드','카오스','엔픽스','타프토이즈']:
                        if b in c: brand_guess = b; break
                    if brand_guess: break
                if brand_guess: break

        # 준비물/POP 추정: 사은품 컬럼 텍스트들 수집
        prep_items = set()
        for row in ws.iter_rows(values_only=True):
            for c in row:
                if c and isinstance(c, str) and ('커버' in c or '모기장' in c or '범퍼' in c or '인서트' in c) and len(c) < 40:
                    prep_items.add(c.strip())

        # 수정2-4: 상품별 가격/공급조건/사은품 상세표 추출 (클릭 시 세부 확인용)
        price_table = []
        header_row_idx = None
        all_rows = list(ws.iter_rows(values_only=True))
        for i, row in enumerate(all_rows):
            texts = [str(c).strip() for c in row if c]
            if any('상품명' in t for t in texts):
                header_row_idx = i
                break
        if header_row_idx is not None:
            # 상품명 다음 행부터 데이터 (헤더가 1~2줄일 수 있어 상품명 있는 행+1, +2 모두 시도)
            for row in all_rows[header_row_idx+1: header_row_idx+30]:
                vals = [c for c in row if c is not None]
                if not vals: continue
                first_text = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                if not first_text or first_text.startswith(('-', '*', '※')) or '상품명' in first_text:
                    continue
                # 숫자형 값들(가격) + 문자열(사은품) 분리
                nums = [c for c in row if isinstance(c, (int, float))]
                texts_in_row = [str(c).strip() for c in row if isinstance(c, str) and c.strip() and c.strip() != first_text]
                gift = next((t for t in texts_in_row if len(t) < 60 and ('커버' in t or '모기장' in t or '범퍼' in t or '인서트' in t or '없음' in t)), '')
                if nums:
                    entry = {
                        'product': first_text,
                        'consumer_price': nums[0] if len(nums) > 0 else None,
                        'event_price': nums[2] if len(nums) > 2 else (nums[-1] if nums else None),
                        'supply_price': nums[-2] if len(nums) >= 2 else None,
                        'gift': gift,
                    }
                    # 공급가격 정보가 없는 행은 별도 안내문/각주가 잘못 잡힌 것일 가능성이 높아 제외
                    if entry['supply_price'] is not None and len(nums) >= 3:
                        price_table.append(entry)

        # 기간/행사명 패턴 탐색 (전체 시트 텍스트에서)
        found_any = False
        for row in ws.iter_rows(values_only=True):
            for c in row:
                if not c or not isinstance(c, str): continue
                m = date_pattern.search(c)
                if m:
                    yy, mm, d1, d2, ev_name = m.groups()
                    period_start = f"{yy}-{int(mm):02d}-{int(d1):02d}"
                    period_end = f"{yy}-{int(mm):02d}-{int(d2):02d}"
                    events.append({
                        'year': int(yy), 'month': int(mm),
                        'period_start': period_start, 'period_end': period_end,
                        'brand': brand_guess, 'event_name': (ev_name or title_text or sheet_name).strip(),
                        'target_channel': target_channel,
                        'prep_items': ', '.join(sorted(prep_items))[:200],
                        'status': '예정',
                        'price_table': price_table,
                    })
                    found_any = True
        if not found_any and title_text:
            # 날짜 패턴을 못 찾았어도 타이틀 정보는 등록 (기간 미상으로)
            events.append({
                'year': now.year, 'month': now.month,
                'period_start': '', 'period_end': '',
                'brand': brand_guess, 'event_name': title_text,
                'target_channel': target_channel,
                'prep_items': ', '.join(sorted(prep_items))[:200],
                'status': '예정',
                'price_table': price_table,
            })

    if not events:
        return jsonify({'ok': False, 'msg': '행사 정보를 인식하지 못했습니다. 파일 형식을 확인해주세요'}), 400

    conn = get_db()
    now_str = now.strftime('%Y-%m-%d %H:%M')
    inserted = 0
    for ev in events:
        conn.execute("""INSERT INTO work_promotion
            (year, month, period_start, period_end, brand, event_name, target_channel, prep_items, status, source_filename, created_at, detail_json)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
            (ev['year'], ev['month'], ev['period_start'], ev['period_end'], ev['brand'],
             ev['event_name'], ev['target_channel'], ev['prep_items'], ev['status'], fname, now_str,
             json.dumps(ev.get('price_table', []), ensure_ascii=False)))
        inserted += 1
    conn.commit(); conn.close()

    return jsonify({'ok': True, 'inserted': inserted, 'events': events})


@app.route("/api/work/promotion/list")
@login_required
def api_work_promotion_list():
    year = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))
    conn = get_db()
    rows = [dict(r) for r in conn.execute(
        "SELECT * FROM work_promotion WHERE year=? AND month=? ORDER BY period_start",
        (year, month)).fetchall()]
    conn.close()
    for r in rows:
        try: r['price_table'] = json.loads(r.get('detail_json') or '[]')
        except Exception: r['price_table'] = []
    return jsonify(rows)


@app.route("/api/work/promotion/<int:pid>", methods=["DELETE"])
@login_required
def api_work_promotion_delete(pid):
    conn = get_db()
    conn.execute("DELETE FROM work_promotion WHERE id=?", (pid,))
    conn.commit(); conn.close()
    return jsonify({'ok': True})


@app.route("/api/work/promotion/<int:pid>/status", methods=["POST"])
@login_required
def api_work_promotion_status(pid):
    status = (request.json or {}).get('status', '예정')
    conn = get_db()
    conn.execute("UPDATE work_promotion SET status=? WHERE id=?", (status, pid))
    conn.commit(); conn.close()
    return jsonify({'ok': True})


@app.route("/api/work/retro")
@login_required
def api_work_retro():
    year = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))
    manager = request.args.get('manager', '전체').strip() or '전체'
    conn = get_db()
    row = conn.execute("SELECT * FROM work_retro WHERE year=? AND month=? AND manager=?",
                        (year, month, manager)).fetchone()
    conn.close()
    if row:
        return jsonify({'ok': True, 'keep_text': row['keep_text'], 'problem_text': row['problem_text'],
                         'try_text': row['try_text'], 'updated_at': row['updated_at']})
    return jsonify({'ok': True, 'keep_text': '', 'problem_text': '', 'try_text': '', 'updated_at': ''})


@app.route("/api/work/retro", methods=["POST"])
@login_required
def api_work_retro_save():
    d = request.json or {}
    year = int(d.get('year')); month = int(d.get('month'))
    manager = d.get('manager', '전체').strip() or '전체'
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    conn = get_db()
    conn.execute("""INSERT INTO work_retro (year, month, manager, keep_text, problem_text, try_text, updated_at)
        VALUES(?,?,?,?,?,?,?)
        ON CONFLICT(year, month, manager) DO UPDATE SET
        keep_text=excluded.keep_text, problem_text=excluded.problem_text,
        try_text=excluded.try_text, updated_at=excluded.updated_at""",
        (year, month, manager, d.get('keep_text',''), d.get('problem_text',''), d.get('try_text',''), now))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

@app.route("/api/export/xlsx/work")
@login_required
def api_export_work_xlsx():
    """업무 탭 엑셀 다운로드 — 기존 앱과 동일한 절제된 스타일 (맑은 고딕, A열 여백, 흰 배경)"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    year = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))
    manager = request.args.get('manager', '').strip()

    FNAME = '맑은 고딕'
    def mf(h): return PatternFill("solid", fgColor=h)
    thin = Side(style='thin', color='E5E7EB')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    rgt  = Alignment(horizontal='right', vertical='center')

    wb = openpyxl.Workbook()
    mgr_label = manager or '전체'
    title_suffix = f"{year}년 {month}월 · {mgr_label}"

    def _title(ws, text, span_to='F'):
        ws.column_dimensions['A'].width = 2
        ws.merge_cells(f'B1:{span_to}1')
        c = ws.cell(row=1, column=2, value=text)
        c.font = Font(bold=True, size=13, name=FNAME, color='1F2937'); c.alignment = ctr
        ws.row_dimensions[1].height = 26

    # ── 시트1: 월간 목표 vs 실적 ──
    ws1 = wb.active; ws1.title = '월간목표vs실적'
    _title(ws1, f'월간 목표 vs 실적  ({title_suffix})', 'F')
    hdrs = ['항목', '목표', '실적', '달성률', '전월대비']
    for ci, h in enumerate(hdrs, 2):
        c = ws1.cell(row=3, column=ci, value=h)
        c.font = Font(bold=True, size=9, name=FNAME, color='374151'); c.fill = mf('F3F4F6'); c.border = bdr; c.alignment = ctr
    ws1.row_dimensions[3].height = 20
    kpi_data = _work_kpi_data(year, month, manager)
    ri = 4
    for it in kpi_data['items']:
        vals = [it['label'], f"{it['target']}{it['unit']}", f"{it['actual']}{it['unit']}",
                f"{it['rate']}%", f"{'+' if it['delta']>=0 else ''}{it['delta']}" if it['delta'] else '-']
        for ci, v in enumerate(vals, 2):
            c = ws1.cell(row=ri, column=ci, value=v); c.font = Font(size=9, name=FNAME); c.border = bdr
            c.alignment = left if ci == 2 else ctr
        ri += 1
    for ci, w in zip(range(2,7), [20,14,14,12,12]):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ── 시트2: 브랜드별 실적 ──
    ws2 = wb.create_sheet('브랜드별실적')
    _title(ws2, f'브랜드별 실적  ({title_suffix})  · 목표=전년동월×1.05', 'E')
    hdrs2 = ['브랜드', '목표매출(원)', '실적매출(원)', '달성률']
    for ci, h in enumerate(hdrs2, 2):
        c = ws2.cell(row=3, column=ci, value=h)
        c.font = Font(bold=True, size=9, name=FNAME, color='374151'); c.fill = mf('F3F4F6'); c.border = bdr; c.alignment = ctr
    ws2.row_dimensions[3].height = 20
    brand_data = _work_brand_perf_data(year, month, manager)
    ri = 4
    for it in brand_data['items']:
        c = ws2.cell(row=ri, column=2, value=it['brand']); c.font = Font(size=9, name=FNAME, bold=True); c.border = bdr; c.alignment = left
        c2 = ws2.cell(row=ri, column=3, value=it['target']); c2.number_format = '#,##0'; c2.font = Font(size=9, name=FNAME); c2.border = bdr; c2.alignment = rgt
        c3 = ws2.cell(row=ri, column=4, value=it['actual']); c3.number_format = '#,##0'; c3.font = Font(size=9, name=FNAME); c3.border = bdr; c3.alignment = rgt
        c4 = ws2.cell(row=ri, column=5, value=f"{it['rate']}%")
        c4.font = Font(size=9, name=FNAME, bold=True, color='16A34A' if it['rate']>=100 else 'DC2626' if it['rate']<70 else 'D97706')
        c4.border = bdr; c4.alignment = ctr
        ri += 1
    for ci, w in zip(range(2,6), [16,18,18,12]):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ── 시트3: 매장 방문 커버리지 (세부 매장 리스트 포함) ──
    ws3 = wb.create_sheet('매장방문커버리지')
    _title(ws3, f'매장 방문 커버리지  ({title_suffix})', 'F')
    cov_data = _work_coverage_data(year, month, manager)
    ri = 3
    hdrs3 = ['권역', '담당매장수', '방문매장수', '방문율', '비고']
    for ci, h in enumerate(hdrs3, 2):
        c = ws3.cell(row=ri, column=ci, value=h)
        c.font = Font(bold=True, size=9, name=FNAME, color='374151'); c.fill = mf('F3F4F6'); c.border = bdr; c.alignment = ctr
    ws3.row_dimensions[ri].height = 20
    ri += 1
    for it in cov_data['items']:
        vals = [it['region'], it['managed'], it['visited'], f"{it['rate']}%", it['note']]
        for ci, v in enumerate(vals, 2):
            c = ws3.cell(row=ri, column=ci, value=v); c.font = Font(size=9, name=FNAME); c.border = bdr
            c.alignment = left if ci in (2,6) else ctr
        ri += 1
        # 세부 매장 리스트 (수정2-3)
        if it['stores']:
            ws3.merge_cells(f'B{ri}:F{ri}')
            store_names = ', '.join(f"{s['name']}({s['sales']:,}원)" for s in it['stores'])
            c = ws3.cell(row=ri, column=2, value=f"   ㄴ 담당 매장: {store_names}")
            c.font = Font(size=8, name=FNAME, color='6B7280'); c.alignment = left
            ws3.row_dimensions[ri].height = 26
            ri += 1
    for ci, w in zip(range(2,7), [12,12,12,10,50]):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    # ── 시트4: 프로모션·행사 캘린더 ──
    ws4 = wb.create_sheet('프로모션캘린더')
    _title(ws4, f'프로모션 · 행사 캘린더  ({title_suffix})', 'G')
    conn = get_db()
    promo_rows = [dict(r) for r in conn.execute(
        "SELECT * FROM work_promotion WHERE year=? AND month=? ORDER BY period_start", (year, month)).fetchall()]
    conn.close()
    hdrs4 = ['기간', '브랜드', '행사명', '대상채널', '준비물/사은품', '상태']
    for ci, h in enumerate(hdrs4, 2):
        c = ws4.cell(row=3, column=ci, value=h)
        c.font = Font(bold=True, size=9, name=FNAME, color='374151'); c.fill = mf('F3F4F6'); c.border = bdr; c.alignment = ctr
    ws4.row_dimensions[3].height = 20
    ri = 4
    for r in promo_rows:
        period = f"{r['period_start']}~{r['period_end']}" if r['period_start'] else '기간 미정'
        vals = [period, r['brand'], r['event_name'], r['target_channel'], r['prep_items'], r['status']]
        for ci, v in enumerate(vals, 2):
            c = ws4.cell(row=ri, column=ci, value=v); c.font = Font(size=9, name=FNAME); c.border = bdr
            c.alignment = left if ci in (4,6) else ctr
        ri += 1
    for ci, w in zip(range(2,8), [20,10,20,16,36,10]):
        ws4.column_dimensions[get_column_letter(ci)].width = w

    # ── 시트5: 월간 회고 ──
    ws5 = wb.create_sheet('월간회고')
    _title(ws5, f'월간 회고 (KPT)  ({title_suffix})', 'C')
    conn = get_db()
    retro_row = conn.execute("SELECT * FROM work_retro WHERE year=? AND month=? AND manager=?",
                              (year, month, manager or '전체')).fetchone()
    conn.close()
    retro = dict(retro_row) if retro_row else {'keep_text':'', 'problem_text':'', 'try_text':''}
    labels = [('Keep (잘한 점)', retro.get('keep_text','')), ('Problem (아쉬운 점)', retro.get('problem_text','')),
              ('Try (다음 달 시도)', retro.get('try_text',''))]
    ri = 3
    for label, text in labels:
        ws5.cell(row=ri, column=2, value=label).font = Font(bold=True, size=10, name=FNAME, color='1F2937')
        ri += 1
        ws5.merge_cells(f'B{ri}:C{ri}')
        c = ws5.cell(row=ri, column=2, value=text or '(내용 없음)')
        c.font = Font(size=9, name=FNAME); c.alignment = left
        ws5.row_dimensions[ri].height = 60
        ri += 2
    ws5.column_dimensions['B'].width = 20
    ws5.column_dimensions['C'].width = 50

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f'업무_{year}년{month}월_{mgr_label}.xlsx'
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=fname)




@app.route("/api/sns/save-memo", methods=["POST"])
@login_required
def api_sns_save_memo():
    """메모 + 블로그 URL/이름 수동 저장"""
    d    = request.json or {}
    name = d.get('seller_name','').strip()
    if not name: return jsonify({'ok':False}), 400
    now  = datetime.now().strftime('%Y-%m-%d %H:%M')
    conn = get_db()
    conn.execute("""INSERT INTO sns_info(seller_name,blog_url,blog_name,memo,updated_at)
        VALUES(?,?,?,?,?)
        ON CONFLICT(seller_name) DO UPDATE SET
        blog_url=CASE WHEN ?!='' THEN ? ELSE blog_url END,
        blog_name=CASE WHEN ?!='' THEN ? ELSE blog_name END,
        memo=excluded.memo, updated_at=excluded.updated_at""",
        (name, d.get('blog_url',''), d.get('blog_name',''), d.get('memo',''), now,
         d.get('blog_url',''), d.get('blog_url',''),
         d.get('blog_name',''), d.get('blog_name','')))
    conn.commit(); conn.close()
    return jsonify({'ok': True})


# ── 매장 방문 일정 API ───────────────────────────────
@app.route("/api/visit/schedule")
@login_required
def api_visit_schedule():
    year  = request.args.get("year",  str(datetime.now().year))
    month = request.args.get("month", "")
    conn  = get_db()
    cond  = f"visit_date LIKE '{year}-{month.zfill(2)}%'" if month else f"visit_date LIKE '{year}%'"
    rows  = [dict(r) for r in conn.execute(
        f"SELECT * FROM visit_schedule WHERE {cond} ORDER BY visit_date, priority").fetchall()]
    conn.close()
    return jsonify(rows)

@app.route("/api/visit/generate", methods=["POST"])
@login_required
def api_visit_generate():
    """데이터 기반 방문 일정 자동 생성"""
    from datetime import datetime as dt2, timedelta
    import calendar as cal_mod

    d     = request.json or {}
    year  = int(d.get('year',  dt2.now().year))
    month = int(d.get('month', dt2.now().month))

    conn = get_db()

    # 수동 입력 일정 보호 (is_manual=1 유지)
    conn.execute("""DELETE FROM visit_schedule
        WHERE visit_date LIKE ? AND is_manual=0""",
        (f"{year}-{month:02d}%",))

    # 매장별 데이터 분석
    sellers_data = [dict(r) for r in conn.execute(f"""
        SELECT real_seller,
               SUM(total) total,
               COUNT(*) cnt,
               MAX(sale_date) last_sale,
               COUNT(DISTINCT item_group) brand_cnt
        FROM sales_data
        WHERE real_seller!='' AND sale_date LIKE '{year}%'
        GROUP BY real_seller ORDER BY total DESC""").fetchall()]

    # 전체 합계
    total_all = sum(s['total'] or 0 for s in sellers_data) or 1

    # 재고 데이터 (있으면)
    stock_items = {r[0]:r[1] for r in conn.execute(
        "SELECT item_name, quantity FROM stock_data WHERE quantity>0").fetchall()}

    # 기존 수동 일정 날짜 파악 (충돌 방지)
    manual_dates = set(r[0] for r in conn.execute(
        f"SELECT DISTINCT visit_date FROM visit_schedule "
        f"WHERE visit_date LIKE '{year}-{month:02d}%' AND is_manual=1").fetchall())

    schedules = []

    for s in sellers_data:
        seller  = s['real_seller']
        total   = s['total'] or 0
        pct     = total / total_all * 100
        last    = s['last_sale'] or ''
        brands  = s['brand_cnt'] or 0

        # 등급 산정
        if pct >= 10:    grade='A'; base_freq=14   # 2주 1회
        elif pct >= 5:   grade='B'; base_freq=21   # 3주 1회
        elif pct >= 2:   grade='C'; base_freq=28   # 4주 1회
        elif pct >= 0.5: grade='D'; base_freq=42   # 6주 1회
        else:            grade='E'; base_freq=60   # 2달 1회

        # 최근 방문 여부 고려
        days_since_last = 999
        if last:
            try:
                days_since_last = (dt2.now() - dt2.strptime(last, '%Y-%m-%d')).days
            except: pass

        # 우선순위 결정
        priority = 1 if grade=='A' else 2 if grade=='B' else 3 if grade=='C' else 4

        # 타프토이즈 미취급 → 우선순위 상향
        taft_sold = conn.execute("""
            SELECT COUNT(DISTINCT item_name) FROM sales_data
            WHERE real_seller=? AND item_group='TAFTOYS'""", (seller,)).fetchone()[0]
        if taft_sold == 0: priority = max(1, priority-1)

        # 방문 이유 생성
        reasons = []
        if grade in ('A','B'): reasons.append(f"{grade}등급 핵심 거래처 정기 방문")
        if taft_sold == 0: reasons.append("타프토이즈 신규 도입 제안")
        if brands <= 2: reasons.append("취급 브랜드 확대 논의")
        if days_since_last > 45: reasons.append(f"장기 미방문 ({days_since_last}일)")
        reason = ' / '.join(reasons[:2]) if reasons else "정기 방문"

        # 체크포인트
        check_points = []
        check_points.append(f"재고 현황 확인 ({s.get('cnt',0)}건 판매 이력)")
        if taft_sold == 0:
            check_points.append("타프토이즈 라인업 제안 (재고 있는 인기 제품 중심)")
        if brands <= 3:
            check_points.append(f"취급 브랜드 {brands}개 → 확대 가능성 논의")
        check_points.append("진열 상태 점검 및 POP 교체")
        check_points.append("사장님 VOC 청취")

        schedules.append({
            'seller_name':  seller,
            'visit_type':   'auto',
            'reason':       reason,
            'priority':     priority,
            'status':       'planned',
            'check_points': '\n'.join(f"□ {cp}" for cp in check_points),
            'is_manual':    0,
            'region':       detect_region_from_name(seller),  # 지역 태그
            'created_at':   dt2.now().strftime('%Y-%m-%d %H:%M'),
            'updated_at':   dt2.now().strftime('%Y-%m-%d %H:%M'),
        })

    # ── 동선 최적화: 지역별 묶어서 날짜 배정 ──────────────
    # 지역 우선순위 (서울/경기 → 인천 → 충청 → 부산/경상 → 전라 → 강원 → 제주)
    REGION_ORDER = ['서울','경기','인천','경기북부','경기남부','수도권',
                    '충청','대전','충남','충북',
                    '부산','경상','대구','울산','경남','경북',
                    '전라','광주','전남','전북',
                    '강원',
                    '제주',
                    '']  # 지역 미상 → 마지막

    def region_order_key(region):
        r = (region or '').strip()
        for i, rk in enumerate(REGION_ORDER):
            if rk and rk in r: return i
        return len(REGION_ORDER)

    # 우선순위 → 지역 순서로 정렬
    schedules.sort(key=lambda x: (x['priority'], region_order_key(x.get('region',''))))

    # 영업일 목록 생성 (주말 제외, 하루 4개 매장)
    from collections import defaultdict
    import calendar as cal_mod
    _, last_day = cal_mod.monthrange(year, month)
    workdays = []
    for d_num in range(1, last_day+1):
        wd = dt2(year, month, d_num).weekday()
        if wd < 5:  # 월~금
            workdays.append(f"{year}-{month:02d}-{d_num:02d}")

    # 수동 일정 날짜별 카운트
    date_count = defaultdict(int)
    for md in manual_dates: date_count[md] += 1

    # 같은 지역 묶기 → 같은 날 배정
    SLOTS_PER_DAY = 4
    workday_idx = 0

    # 지역별 그룹 생성
    region_groups = defaultdict(list)
    for sc in schedules:
        rg = detect_region_from_name(sc['seller_name']) or '기타'
        # 지역 대분류
        if any(k in rg for k in ['서울','경기','인천']): rg_key='수도권'
        elif any(k in rg for k in ['충청','대전','충남','충북']): rg_key='충청권'
        elif any(k in rg for k in ['부산','경상','대구','울산','경남','경북']): rg_key='경상권'
        elif any(k in rg for k in ['전라','광주','전남','전북']): rg_key='전라권'
        elif '강원' in rg: rg_key='강원권'
        elif '제주' in rg: rg_key='제주권'
        else: rg_key='기타'
        region_groups[rg_key].append(sc)

    # 지역 방문 순서 (이동 동선)
    ROUTE_ORDER = ['수도권','충청권','경상권','전라권','강원권','제주권','기타']

    for region_key in ROUTE_ORDER:
        group = region_groups.get(region_key, [])
        if not group: continue
        # 같은 지역은 연속된 날에 배정
        for sc in group:
            # 빈 슬롯 있는 영업일 찾기
            while workday_idx < len(workdays) and date_count[workdays[workday_idx]] >= SLOTS_PER_DAY:
                workday_idx += 1
            if workday_idx >= len(workdays):
                workday_idx = len(workdays) - 1  # 마지막 날에 몰아서
            chosen_date = workdays[workday_idx]
            sc['visit_date'] = chosen_date
            date_count[chosen_date] += 1
            # 이 슬롯이 꽉 차면 다음 날로
            if date_count[chosen_date] >= SLOTS_PER_DAY:
                workday_idx += 1

    # DB 저장
    inserted = 0
    for s in schedules:
        conn.execute("""INSERT INTO visit_schedule
            (visit_date,seller_name,visit_type,reason,priority,status,check_points,is_manual,created_at,updated_at)
            VALUES(?,?,?,?,?,?,?,?,?,?)""",
            (s['visit_date'],s['seller_name'],s['visit_type'],s['reason'],
             s['priority'],s['status'],s['check_points'],0,s['created_at'],s['updated_at']))
        inserted += 1

    conn.commit(); conn.close()
    return jsonify({'ok': True, 'inserted': inserted, 'year': year, 'month': month})

@app.route("/api/visit/save", methods=["POST"])
@login_required
def api_visit_save():
    """방문 일정 수동 저장/수정"""
    from datetime import datetime as dt2
    d    = request.json or {}
    conn = get_db()
    now  = dt2.now().strftime('%Y-%m-%d %H:%M')
    vid  = d.get('id')
    if vid:  # 수정
        conn.execute("""UPDATE visit_schedule SET
            visit_date=?, seller_name=?, reason=?, priority=?, status=?,
            check_points=?, result_memo=?, is_manual=1, updated_at=?
            WHERE id=?""",
            (d['visit_date'], d['seller_name'], d.get('reason',''), d.get('priority',2),
             d.get('status','planned'), d.get('check_points',''),
             d.get('result_memo',''), now, vid))
    else:  # 신규
        conn.execute("""INSERT INTO visit_schedule
            (visit_date,seller_name,visit_type,reason,priority,status,check_points,result_memo,is_manual,created_at,updated_at)
            VALUES(?,?,?,?,?,?,?,?,1,?,?)""",
            (d['visit_date'], d['seller_name'], 'manual', d.get('reason',''),
             d.get('priority',2), d.get('status','planned'),
             d.get('check_points',''), d.get('result_memo',''), now, now))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

@app.route("/api/visit/delete", methods=["POST"])
@login_required
def api_visit_delete():
    vid = (request.json or {}).get('id')
    if not vid: return jsonify({'ok':False}), 400
    conn = get_db(); conn.execute("DELETE FROM visit_schedule WHERE id=?", (vid,))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

# ── 재고 현황 API ────────────────────────────────────
@app.route("/api/report/template/upload", methods=["POST"])
@login_required
def api_report_template_upload():
    """보고서 양식 xlsx 업로드 — 구조 분석 후 저장"""
    from datetime import datetime as dt2
    if 'file' not in request.files:
        return jsonify({'ok':False,'msg':'파일 없음'}), 400
    file  = request.files['file']
    ttype = request.form.get('type', 'weekly')  # weekly / visit
    data  = file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        # 양식 구조 추출 (모든 셀 내용)
        structure = []
        for row in ws.iter_rows(values_only=True):
            row_vals = [str(c or '') for c in row]
            if any(v.strip() for v in row_vals):
                structure.append(row_vals)

        conn = get_db()
        import json as json_mod
        conn.execute("""INSERT INTO report_template (template_name, template_type, columns, uploaded_at)
            VALUES(?,?,?,?)
            ON CONFLICT DO NOTHING""", ('', ttype, '', dt2.now().strftime('%Y-%m-%d %H:%M')))
        # UPDATE 방식
        conn.execute("""DELETE FROM report_template WHERE template_type=?""", (ttype,))
        conn.execute("""INSERT INTO report_template (template_name, template_type, columns, uploaded_at)
            VALUES(?,?,?,?)""",
            (file.filename, ttype, json_mod.dumps(structure, ensure_ascii=False),
             dt2.now().strftime('%Y-%m-%d %H:%M')))
        conn.commit(); conn.close()

        return jsonify({'ok': True, 'type': ttype, 'rows': len(structure),
                        'preview': structure[:5]})  # 앞 5행 미리보기
    except Exception as e:
        return jsonify({'ok': False, 'msg': str(e)}), 500


@app.route("/api/report/generate", methods=["POST"])
@login_required
def api_report_generate():
    """보고서 양식에 맞게 내용 자동 생성"""
    import json as json_mod, re
    d      = request.json or {}
    rtype  = d.get('type', 'weekly')
    input_text = d.get('input_text', '')
    seller = d.get('seller', '')
    date   = d.get('date', '')
    week_start = d.get('week_start', '')
    next_plan  = d.get('next_plan', '')
    brand_note = d.get('brand_note', '')

    # 양식 불러오기
    conn = get_db()
    tmpl_row = conn.execute("SELECT columns FROM report_template WHERE template_type=?", (rtype,)).fetchone()

    # 매장 데이터
    seller_data = None
    if seller:
        year = datetime.now().year
        rows = conn.execute("""SELECT SUM(total) total, COUNT(*) cnt,
            COUNT(DISTINCT strftime('%m',sale_date)) months
            FROM sales_data WHERE real_seller=? AND sale_date LIKE ?""",
            (seller, f"{year}%")).fetchone()
        if rows: seller_data = dict(rows)

    conn.close()

    # 자동 분류
    SELF_BRANDS  = ['줄즈','레카로','엔픽스','타프토이즈','원더폴드','카오스','ABC','abc디자인','ENFIX']
    RIVAL_BRANDS = ['그라코','에르고베이비','컴비','조이','페도','마클라렌','퀴니','부가부','스토케','요요']
    SPECIAL_BRNDS= ['부가부','스토케','요요']

    lines = [l.strip() for l in re.split(r'[.!?\n]+', input_text) if l.strip()]
    self_items  = [l for l in lines if any(b.lower() in l.lower() for b in SELF_BRANDS)]
    rival_items = [l for l in lines if any(b.lower() in l.lower() for b in RIVAL_BRANDS)]
    order_items = [l for l in lines if re.search(r'발주|주문|개.*받|건.*받', l)]
    visit_items = [l for l in lines if re.search(r'방문|들렀|들러|방문함', l)]
    other_items = [l for l in lines if l not in self_items+rival_items+order_items+visit_items]

    # 특이사항 (부가부/스토케/요요)
    special_items = [l for l in lines if any(b in l for b in SPECIAL_BRNDS)]
    if brand_note: special_items.append(brand_note)
    special_str = '\n'.join(f"  • {l}" for l in special_items) if special_items else '  -'

    total_sales = seller_data['total'] if seller_data else 0
    months_cnt  = seller_data['months'] if seller_data else 1
    avg_monthly = int(total_sales / max(months_cnt, 1)) if total_sales else 0

    now = datetime.now()

    if tmpl_row and tmpl_row[0]:
        # 양식 있으면 구조 기반으로 채우기
        structure = json_mod.loads(tmpl_row[0])
        filled = []
        for row in structure:
            new_row = []
            for cell in row:
                c = cell
                # 치환 패턴
                c = c.replace('{{매장명}}', seller or '').replace('{{방문일}}', date or '')
                c = c.replace('{{보고주간}}', week_start or '').replace('{{작성일}}', now.strftime('%Y-%m-%d'))
                c = c.replace('{{누적매출}}', f"{total_sales:,}원" if total_sales else '-')
                c = c.replace('{{월평균매출}}', f"{avg_monthly:,}원" if avg_monthly else '-')
                c = c.replace('{{자사주요내용}}', '\n'.join(f"• {l}" for l in self_items) or '-')
                c = c.replace('{{타사주요내용}}', '\n'.join(f"• {l}" for l in rival_items) or '-')
                c = c.replace('{{발주내역}}', '\n'.join(f"• {l}" for l in order_items) or '-')
                c = c.replace('{{방문내용}}', '\n'.join(f"• {l}" for l in visit_items) or '-')
                c = c.replace('{{특이사항}}', special_str)
                c = c.replace('{{다음주계획}}', next_plan or '-')
                c = c.replace('{{기타}}', '\n'.join(f"• {l}" for l in other_items) or '-')
                # 빈 칸(___) 채우기
                if set(c.strip()) <= {'_'} and c.strip():
                    c = '(작성 필요)'
                new_row.append(c)
            filled.append(new_row)
        return jsonify({'ok': True, 'type': rtype, 'structure': filled, 'has_template': True})

    else:
        # 양식 없으면 텍스트 형식으로 생성
        if rtype == 'weekly':
            wd = datetime.strptime(week_start, '%Y-%m-%d') if week_start else now
            we = wd + __import__('datetime').timedelta(days=6) if week_start else now
            text = f"""주간 업무 보고서
{'─'*50}
보고 기간: {wd.strftime('%Y.%m.%d')} ~ {we.strftime('%Y.%m.%d')}
작성일: {now.strftime('%Y.%m.%d')}
{'─'*50}

【1. 매장 방문 현황】
{chr(10).join(f"  • {l}" for l in visit_items) or "  -"}

【2. 발주/수주 현황】
{chr(10).join(f"  • {l}" for l in order_items) or "  -"}

【3. 자사 브랜드 주요 내용】
{chr(10).join(f"  • {l}" for l in self_items) or "  -"}

【4. 타사/경쟁사 동향】
{chr(10).join(f"  • {l}" for l in rival_items) or "  -"}

【5. 다음 주 계획】
{chr(10).join(f"  • {l.strip()}" for l in (next_plan or '').split('\n') if l.strip()) or "  -"}
{'─'*50}"""
        else:
            text = f"""매장 방문 보고서
{'='*50}
【기본 정보】
매장명     : {seller}
방문일     : {date}
누적매출   : {total_sales:,}원 ({now.year}년)
월평균매출 : {avg_monthly:,}원
{'─'*50}

【자사 주요 내용】
{chr(10).join(f"  • {l}" for l in self_items) or "  -"}

【발주/수주 내역】
{chr(10).join(f"  • {l}" for l in order_items) or "  -"}

【타사/경쟁사 주요 내용】
{chr(10).join(f"  • {l}" for l in rival_items) or "  -"}

【브랜드별 특이사항 (부가부/스토케/요요)】
{special_str}

【기타 특이사항】
{chr(10).join(f"  • {l}" for l in other_items) or "  -"}
{'='*50}
작성일: {now.strftime('%Y.%m.%d')}"""

        return jsonify({'ok': True, 'type': rtype, 'text': text, 'has_template': False})


@app.route("/api/report/export", methods=["POST"])
@login_required
def api_report_export():
    """보고서 엑셀 다운로드"""
    import json as json_mod
    d = request.json or {}
    rtype = d.get('type', 'weekly')
    structure = d.get('structure')  # 양식 기반
    text = d.get('text')            # 텍스트 기반

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '주간업무보고' if rtype=='weekly' else '매장방문보고서'

    thin = Side(style='thin', color='CCCCCC')
    bdr  = Border(left=thin,right=thin,top=thin,bottom=thin)

    if structure:
        for ri, row in enumerate(structure, 1):
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.alignment = Alignment(wrap_text=True, vertical='top')
                if val and val != '(작성 필요)':
                    c.border = bdr
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = 25
    elif text:
        for ri, line in enumerate(text.split('\n'), 1):
            c = ws.cell(row=ri, column=1, value=line)
            c.alignment = Alignment(wrap_text=True)
        ws.column_dimensions['A'].width = 80

    fname = f"{'주간업무보고' if rtype=='weekly' else '매장방문보고서'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


@app.route("/api/stock/upload", methods=["POST"])
@login_required
def api_stock_upload():
    """재고 xlsx 업로드 — 컬럼 자동 감지"""
    from datetime import datetime as dt2
    if 'file' not in request.files:
        return jsonify({'ok':False,'msg':'파일 없음'}), 400
    file = request.files['file']
    data = file.read()
    try:
        wb  = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        ws  = wb.active
        rows_parsed = []
        headers = []
        raw_rows = list(ws.iter_rows(values_only=True))

        # 헤더 행 찾기 (첫 5행 중에서 가장 많은 비어있지 않은 값이 있는 행)
        header_row_idx = 0
        for i, row in enumerate(raw_rows[:5]):
            non_empty = sum(1 for c in row if c is not None and str(c).strip())
            if non_empty >= 2:
                header_row_idx = i
                break

        headers = [str(c or '').strip() for c in raw_rows[header_row_idx]]
        app.logger.info(f"재고 헤더: {headers}")

        # 컬럼 인덱스 찾기 (다양한 이름 허용)
        NAME_KEYS  = ['제품명','품목명','상품명','item_name','name','품명','제품','상품','아이템']
        QTY_KEYS   = ['수량','재고수량','재고','quantity','qty','잔여수량','현재고','재고량']
        GROUP_KEYS = ['품목그룹','그룹','브랜드','카테고리','item_group','group','분류']

        def find_col(keys, hdrs):
            for k in keys:
                for i, h in enumerate(hdrs):
                    if k.lower() in h.lower(): return i
            return -1

        name_idx  = find_col(NAME_KEYS,  headers)
        qty_idx   = find_col(QTY_KEYS,   headers)
        group_idx = find_col(GROUP_KEYS, headers)

        # 못 찾으면 위치로 추정 (0=이름, 1=그룹, 마지막=수량)
        if name_idx == -1:  name_idx = 0
        if qty_idx == -1:   qty_idx  = len(headers)-1
        if group_idx == -1: group_idx = 1 if len(headers) > 2 else -1

        for row in raw_rows[header_row_idx+1:]:
            if not any(row): continue
            vals = list(row)
            item_name  = str(vals[name_idx] or '').strip()  if name_idx < len(vals)  else ''
            quantity   = vals[qty_idx]    if qty_idx < len(vals)  else 0
            item_group = str(vals[group_idx] or '').strip() if group_idx >= 0 and group_idx < len(vals) else ''

            if not item_name: continue
            try: quantity = int(float(str(quantity or 0).replace(',','')))
            except: quantity = 0

            rows_parsed.append({'item_name': item_name, 'item_group': item_group, 'quantity': quantity})

        batch = dt2.now().strftime('%Y%m%d%H%M%S')
        conn  = get_db()
        conn.execute("DELETE FROM stock_data")
        for r in rows_parsed:
            conn.execute("INSERT INTO stock_data (item_name,item_group,quantity,upload_date,upload_batch) VALUES(?,?,?,?,?)",
                (r['item_name'], r['item_group'], r['quantity'], dt2.now().strftime('%Y-%m-%d'), batch))
        conn.commit(); conn.close()

        # 미리보기 (처음 3개)
        preview = rows_parsed[:3]
        return jsonify({'ok': True, 'rows': len(rows_parsed), 'batch': batch,
                        'headers_detected': {'name': headers[name_idx] if name_idx < len(headers) else '',
                                             'qty': headers[qty_idx] if qty_idx < len(headers) else '',
                                             'group': headers[group_idx] if group_idx >= 0 and group_idx < len(headers) else ''},
                        'preview': preview})
    except Exception as e:
        import traceback; app.logger.error(traceback.format_exc())
        return jsonify({'ok': False, 'msg': str(e)}), 500

@app.route("/api/stock/list")
@login_required
def api_stock_list():
    conn = get_db()
    rows = [dict(r) for r in conn.execute(
        "SELECT * FROM stock_data WHERE quantity>0 ORDER BY item_group, item_name").fetchall()]
    conn.close()

    # 중복 제거 + 브랜드/모델/컬러 분리
    merged = {}
    for r in rows:
        brand = remap_group(r['item_group'], r['item_name'])
        norm  = normalize_item_name(r['item_name'])

        # 컬러/옵션 추출 (언더바 이후)
        import re
        color_match = re.search(r'_(.+)$', r['item_name'])
        color = color_match.group(1).strip() if color_match else ''
        # 한정판 표시
        is_limited = '한정판' in r['item_name'] or '한정' in r['item_name']

        key = (brand, norm, color)
        if key not in merged:
            merged[key] = {
                'brand': brand,
                'item_name_norm': norm,
                'color': color,
                'is_limited': is_limited,
                'quantity': 0,
                'item_group': r['item_group'],
                'upload_date': r.get('upload_date',''),
            }
        merged[key]['quantity'] += r['quantity']

    result = sorted(merged.values(), key=lambda x: (x['brand'], x['item_name_norm'], x['color']))
    return jsonify(result)

@app.route("/api/export/xlsx/display")
@login_required
def api_export_display():
    """현재 선택한 품목그룹에 맞는 진열/판매 현황 엑셀 — 거래처코드 포함, 컬러별 교차표"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import re, json as _json

    year       = request.args.get("year",  str(datetime.now().year))
    mode       = request.args.get("mode",  "display")   # display / sales
    brand_sel  = request.args.get("brand", "")          # 선택된 브랜드명 (한글)

    conn = get_db()

    # ── 거래처코드 매핑 (없어도 동작) ──────────────────
    try:
        code_rows = [dict(r) for r in conn.execute("SELECT * FROM seller_code").fetchall()]
    except Exception:
        code_rows = []
    seller_to_code = {r['real_seller']: r for r in code_rows if r.get('real_seller')}

    # 수정4: sales_data에 실제로 캡처된 거래처코드(엑셀 원본 AC열)를 우선 사용
    real_code_map = {}
    for r in conn.execute("""
        SELECT real_seller, trade_code, COUNT(*) cnt FROM sales_data
        WHERE real_seller!='' AND trade_code!=''
        GROUP BY real_seller, trade_code ORDER BY cnt DESC""").fetchall():
        if r[0] not in real_code_map:
            real_code_map[r[0]] = r[1]
    has_real_codes = len(real_code_map) > 0

    GROUP_ORDER_LIST = ['베이비파크','베이비하우스','링크맘','기타']
    def seller_group(seller):
        info = seller_to_code.get(seller, {})
        g = info.get('group_name','')
        if not g:
            nm = (seller or '').lower()
            if '베이비파크' in nm: g='베이비파크'
            elif '베이비하우스' in nm: g='베이비하우스'
            elif '링크맘' in nm or '베네피아' in nm: g='링크맘'
            else: g='기타'
        return g

    all_sellers_raw = [r[0] for r in conn.execute(
        "SELECT DISTINCT real_seller FROM sales_data WHERE real_seller!='' ORDER BY real_seller").fetchall()]
    all_sellers = sorted(all_sellers_raw, key=lambda s: (
        GROUP_ORDER_LIST.index(seller_group(s)) if seller_group(s) in GROUP_ORDER_LIST else 99, s))

    # 수정2: 제품별관리 엑셀에서 생략해야 할 매장 목록 (더 이상 거래 없는 폐업/오류 매장 등)
    EXPORT_EXCLUDE_STORES = {
        '드림오피스 청주점', '베이비플러스 남양주점', '베이비플러스 부천점', '베이비플러스 부평점',
        '베하위례점', '베하하남미사점', '육아대장 평촌점', '주식회사 더케이앤피', '주식회사 동화',
        '한토이 경기광주점', '베네피아 구로점', '베네피아 안양점',
        # 수정2(신규): 중복/오류 매장 추가 제외
        '남양주베이비하우스', '베이비하우스 부산동래', '베이비하우스 안양',
        '베이비하우스 전라광주', '베이비하우스 천안',
    }
    all_sellers = [s for s in all_sellers if s not in EXPORT_EXCLUDE_STORES]

    # 수정3: "OO" / "OO점" 중복 매장 자동 통합 — "점"이 붙은 쪽(실거래 매장 표기)으로 병합
    def _merge_duplicate_stores(sellers):
        """매장명이 'X'와 'X점'처럼 접미사만 다른 경우, '점'이 붙은 쪽을 대표로 채택하고
        나머지는 그 매장으로 흡수되도록 매핑 테이블 반환: {원본매장명: 대표매장명}"""
        by_base = {}
        for s in sellers:
            base = s[:-1].strip() if s.endswith('점') else s.strip()
            by_base.setdefault(base, []).append(s)
        merge_map = {}
        for base, variants in by_base.items():
            if len(variants) < 2:
                continue
            with_jeom = [v for v in variants if v.endswith('점')]
            canonical = with_jeom[0] if with_jeom else variants[0]
            for v in variants:
                if v != canonical:
                    merge_map[v] = canonical
        return merge_map

    seller_merge_map = _merge_duplicate_stores(all_sellers)
    if seller_merge_map:
        all_sellers = [s for s in all_sellers if s not in seller_merge_map]

    # ── 브랜드에 맞는 탭(모델) 목록 동적 생성 ────────
    if not brand_sel:
        return jsonify({'ok': False, 'msg': '브랜드를 선택하세요'}), 400

    # 해당 브랜드의 모든 제품 조회
    rows_all = conn.execute("""
        SELECT item_group, item_name, SUM(quantity) qty
        FROM sales_data
        WHERE sale_date LIKE ? AND real_seller!=''
        GROUP BY item_group, item_name
        ORDER BY item_name
    """, (f"{year}%",)).fetchall()

    # 브랜드 필터 + 모델별 색상 수집
    brand_items = {}   # normalize_name → {colors: [...], pattern: ..., group: ...}
    for r in rows_all:
        brand = remap_group(r[0], r[1])
        if brand != brand_sel: continue
        norm  = normalize_item_name(r[1])
        color = re.sub(r'.*?_', '', r[1], count=1) if '_' in r[1] else '기본'
        color = re.sub(r'\s*\([^)]*\)', '', color).strip()
        # 수정5: 캐노피형을 별도 라인으로 분리하지 않고 기본 모델(예: 토론1)로 통합
        # 수정4,5,6: 브랜드별 커스텀 라벨 규칙(원더폴드/카오스) 우선 적용, 없으면 기본(대괄호 제거) 라벨 사용
        custom_label = get_custom_product_label(brand_sel, r[1])
        norm_key = custom_label if custom_label else norm
        if norm_key not in brand_items:
            brand_items[norm_key] = {
                'label': custom_label if custom_label else get_product_display_label(brand_sel, norm),
                'norm': norm,
                'is_canopy': False,
                'colors': [],
                'color_set': set(),
                'group': r[0],
            }
        if color and color not in brand_items[norm_key]['color_set']:
            brand_items[norm_key]['colors'].append(color)
            brand_items[norm_key]['color_set'].add(color)

    if not brand_items:
        conn.close()
        return jsonify({'ok': False, 'msg': f'{brand_sel} 데이터 없음'}), 404

    now_str  = datetime.now().strftime('%Y.%m.%d')

    # ══════════════════════════════════════════════════════
    # 수정1: mode='sales' — 단일 시트, 월별×제품라인 매트릭스 (색상 구분 없이 총 수량)
    # ══════════════════════════════════════════════════════
    if mode == 'sales':
        FNAME = '맑은 고딕'
        def mf(h): return PatternFill("solid", fgColor=h)
        thin   = Side(style='thin', color='D9D9D9')
        dashed = Side(style='dashed', color='9CA3AF')  # 수정6: 월 구분용 점선
        bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)
        bdr_month_end = Border(left=thin, right=dashed, top=thin, bottom=thin)  # 월 마지막(합계) 컬럼
        ctr    = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_a = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 제품라인 목록 (brand_items의 모델명 기준, 색상/캐노피 무관 통합 — 수정5)
        def _product_order_key(k):
            label = brand_items[k]['label']
            order = PRODUCT_LABEL_ORDER.get(brand_sel)
            if brand_sel == '원더폴드':
                wf_order = ['L2','L4','W2','W4','W2슈퍼맨','W4슈퍼맨','W2폭스','W4폭스']
                return (wf_order.index(label) if label in wf_order else 99, label)
            if brand_sel == '카오스':
                ks_order = ['오크','비치','리사이클']
                return (ks_order.index(label) if label in ks_order else 99, label)
            if order:
                return (order.index(label) if label in order else 99, label)
            return (99, label)
        product_lines = sorted(brand_items.keys(), key=_product_order_key)
        product_labels = [brand_items[k]['label'] for k in product_lines]

        import re as _re3
        def extract_base2(nm):
            m = _re3.match(r'^\[([^\]]+)\](.+)$', nm.strip())
            return m.group(2).strip() if m else nm.strip()

        def match_product_line(item_name):
            """item_name(색상·캐노피 포함)을 제품라인 중 하나로 매칭 — 캐노피형도 기본 모델에 통합
            원더폴드/카오스는 커스텀 라벨 규칙(수정5,6)으로 우선 매칭"""
            custom_label = get_custom_product_label(brand_sel, item_name)
            if custom_label:
                return custom_label if custom_label in product_lines else None
            for pk in product_lines:
                base = extract_base2(brand_items[pk]['norm'])
                if base in item_name:
                    return pk
            return None

        # 월별×매장별×제품라인 데이터 조회 (수량 + 금액 함께)
        rows_monthly = conn.execute("""
            SELECT real_seller, item_name, CAST(strftime('%m',sale_date) AS INTEGER) mo,
                   SUM(quantity) qty, SUM(total) amt
            FROM sales_data
            WHERE sale_date LIKE ? AND real_seller!=''
            GROUP BY real_seller, item_name, mo
        """, (f"{year}%",)).fetchall()

        # 매장별 채널(오프라인/백화점) 조회 — 수정4
        seller_channel = {}
        for r in conn.execute("""
            SELECT real_seller, (SELECT channel FROM sales_data sd2 WHERE sd2.real_seller=sd1.real_seller
                   GROUP BY channel ORDER BY COUNT(*) DESC LIMIT 1) ch
            FROM sales_data sd1 WHERE real_seller!='' GROUP BY real_seller""").fetchall():
            seller_channel[r[0]] = r[1] or '오프라인'

        # qty_map[(seller, month, product_line)] = qty,  amt_map: 금액
        qty_map, amt_map = {}, {}
        months_with_data = set()
        for seller, item_name, mo, qty, amt in rows_monthly:
            brand_of_item = remap_group('', item_name) if not item_name.startswith('[') else remap_group('X', item_name)
            if brand_of_item != brand_sel: continue
            pk = match_product_line(item_name)
            if not pk: continue
            # 수정3: 중복 매장(점 유무) 통합 — 병합 대상이면 대표 매장명으로 흡수
            seller_canon = seller_merge_map.get(seller, seller)
            if seller_canon in EXPORT_EXCLUDE_STORES: continue
            key = (seller_canon, mo, pk)
            qty_map[key] = qty_map.get(key, 0) + (qty or 0)
            amt_map[key] = amt_map.get(key, 0) + (amt or 0)
            months_with_data.add(mo)

        months = sorted(months_with_data) if months_with_data else list(range(1, 13))
        n_prod = len(product_lines)
        col_start = 2
        # 수정4: 실제 거래처코드가 있으면 컬럼 유지, 전혀 없으면 컬럼 자체를 제거
        fixed_hdrs = ['업체구분', '거래처코드', '거래처명', '실적용거래처명'] if has_real_codes else ['업체구분', '거래처명', '실적용거래처명']
        n_fixed = len(fixed_hdrs)
        month_block_width = n_prod + 1

        offline_sellers = [s for s in all_sellers if seller_channel.get(s, '오프라인') != '백화점']
        dept_sellers    = [s for s in all_sellers if seller_channel.get(s, '오프라인') == '백화점']

        HDR_BG = mf('F2F2F2')  # 수정1: 흰색, 배경1, 5% 더 어둡게
        BLACK  = '000000'      # 수정1: 검정, 텍스트1
        dotted = Side(style='hair', color='808080')  # 수정6: 더 촘촘한 점선(hair 스타일)

        def _build_matrix_sheet(ws, sheet_title_text, sellers, value_map, number_format):
            ws.column_dimensions['A'].width = 2
            total_cols = n_fixed + len(months) * month_block_width
            ws.merge_cells(f"B2:{get_column_letter(min(2+n_fixed+month_block_width-1, total_cols+1))}2")
            c = ws.cell(row=2, column=2, value=sheet_title_text)
            c.font = Font(bold=True, size=12, name=FNAME, color=BLACK); c.alignment = left_a
            ws.row_dimensions[2].height = 22

            for ci, h in enumerate(fixed_hdrs, col_start):
                c = ws.cell(row=3, column=ci, value=h)
                c.font = Font(bold=True, size=9, name=FNAME, color=BLACK); c.fill = HDR_BG; c.alignment = ctr
                ws.merge_cells(start_row=3, start_column=ci, end_row=4, end_column=ci)

            col = col_start + n_fixed
            for mo in months:
                end_col = col + n_prod
                ws.merge_cells(f"{get_column_letter(col)}3:{get_column_letter(end_col)}3")
                c = ws.cell(row=3, column=col, value=f"{mo}월")
                c.font = Font(bold=True, size=10, name=FNAME, color=BLACK); c.fill = HDR_BG; c.alignment = ctr
                for pi, plabel in enumerate(product_labels):
                    c2 = ws.cell(row=4, column=col+pi, value=plabel)
                    c2.font = Font(size=9, name=FNAME, color=BLACK); c2.fill = HDR_BG; c2.alignment = ctr
                c3 = ws.cell(row=4, column=end_col, value='합계')
                c3.font = Font(bold=True, size=9, name=FNAME, color=BLACK); c3.fill = HDR_BG; c3.alignment = ctr
                col += month_block_width
            ws.row_dimensions[3].height = 20; ws.row_dimensions[4].height = 18

            def _write_product_group_subtotal(grp_name, grp_seller_list):
                nonlocal ri
                ws.cell(row=ri, column=col_start, value=f"{grp_name} 소계")
                ws.merge_cells(start_row=ri, start_column=col_start, end_row=ri, end_column=col_start+n_fixed-1)
                tcell = ws.cell(row=ri, column=col_start)
                tcell.font = Font(bold=True, size=9, name=FNAME, color=BLACK); tcell.fill = HDR_BG; tcell.alignment = ctr
                col2 = col_start + n_fixed
                for mo in months:
                    grp_total = 0
                    for pi, pk in enumerate(product_lines):
                        pv = sum(value_map.get((s, mo, pk), 0) for s in grp_seller_list)
                        c2 = ws.cell(row=ri, column=col2+pi, value=pv if pv else 0)
                        c2.font = Font(size=9, name=FNAME, color=BLACK); c2.fill = HDR_BG; c2.alignment = ctr
                        if number_format: c2.number_format = number_format
                        grp_total += pv
                    cs2 = ws.cell(row=ri, column=col2+n_prod, value=grp_total if grp_total else 0)
                    cs2.font = Font(bold=True, size=9, name=FNAME, color=BLACK); cs2.fill = HDR_BG; cs2.alignment = ctr
                    if number_format: cs2.number_format = number_format
                    col2 += month_block_width
                ws.row_dimensions[ri].height = 15
                ri += 1

            ri = 5; prev_grp = None
            first_data_row = 5
            current_grp_sellers = []
            for seller in sellers:
                grp = seller_group(seller)
                if prev_grp is not None and grp != prev_grp and current_grp_sellers:
                    _write_product_group_subtotal(prev_grp, current_grp_sellers)
                    current_grp_sellers = []
                current_grp_sellers.append(seller)
                info = seller_to_code.get(seller, {})
                # 수정4: 실제 캡처된 거래처코드 우선, 없으면 매핑 테이블 값
                code = real_code_map.get(seller, '') or info.get('code', '')
                orig = info.get('orig_name', seller)
                gv = grp if grp != prev_grp else ''
                prev_grp = grp
                fixed_vals = [gv, code, orig, seller] if has_real_codes else [gv, orig, seller]
                for ci, v in zip(range(col_start, col_start+n_fixed), fixed_vals):
                    c = ws.cell(row=ri, column=ci, value=v or None)
                    c.font = Font(size=9, name=FNAME, color=BLACK, bold=(ci==col_start))
                    c.alignment = ctr if ci in (col_start, col_start+1) else left_a

                col = col_start + n_fixed
                for mo in months:
                    mo_total = 0
                    for pi, pk in enumerate(product_lines):
                        val = value_map.get((seller, mo, pk), 0)
                        c = ws.cell(row=ri, column=col+pi, value=val if val else 0)
                        c.font = Font(size=9, name=FNAME, color=BLACK); c.alignment = ctr
                        if number_format: c.number_format = number_format
                        mo_total += val
                    cs = ws.cell(row=ri, column=col+n_prod, value=mo_total if mo_total else 0)
                    cs.font = Font(bold=True, size=9, name=FNAME, color=BLACK); cs.alignment = ctr
                    if number_format: cs.number_format = number_format
                    col += month_block_width
                ws.row_dimensions[ri].height = 15
                ri += 1
            if current_grp_sellers:
                _write_product_group_subtotal(prev_grp, current_grp_sellers)
            last_data_row = ri - 1

            # 수정5: 하단 총합계 행
            total_row = ri
            ws.cell(row=total_row, column=col_start, value='총합계')
            ws.merge_cells(start_row=total_row, start_column=col_start, end_row=total_row, end_column=col_start+n_fixed-1)
            tc = ws.cell(row=total_row, column=col_start)
            tc.font = Font(bold=True, size=10, name=FNAME, color=BLACK); tc.fill = HDR_BG; tc.alignment = ctr

            col = col_start + n_fixed
            for mo in months:
                grand_total = 0
                for pi, pk in enumerate(product_lines):
                    col_sum = sum(value_map.get((s, mo, pk), 0) for s in sellers)
                    c = ws.cell(row=total_row, column=col+pi, value=col_sum if col_sum else 0)
                    c.font = Font(bold=True, size=9, name=FNAME, color=BLACK); c.fill = HDR_BG; c.alignment = ctr
                    if number_format: c.number_format = number_format
                    grand_total += col_sum
                cs = ws.cell(row=total_row, column=col+n_prod, value=grand_total if grand_total else 0)
                cs.font = Font(bold=True, size=9, name=FNAME, color=BLACK); cs.fill = HDR_BG; cs.alignment = ctr
                if number_format: cs.number_format = number_format
                col += month_block_width
            ws.row_dimensions[total_row].height = 18
            last_border_row = total_row

            # 수정1: 데이터 테두리 — 안쪽만 점선 (외곽선 없음)
            last_col = col_start + n_fixed + len(months)*month_block_width - 1
            for r_ in range(3, last_border_row+1):
                for c_ in range(col_start, last_col+1):
                    cell = ws.cell(row=r_, column=c_)
                    left_side  = dotted if c_ > col_start else None
                    right_side = dotted if c_ < last_col else None
                    top_side   = dotted if r_ > 3 else None
                    bottom_side= dotted if r_ < last_border_row else None
                    cell.border = Border(left=left_side, right=right_side, top=top_side, bottom=bottom_side)

            # 수정3: 숫자 컬럼(제품라인+합계) 너비 통일 — 콤마 포함 숫자가 '###'로 안 보이도록 충분히 확보
            num_col_width = 11 if number_format else 8
            ws.column_dimensions['B'].width = 12
            if has_real_codes:
                ws.column_dimensions['C'].width = 14
                ws.column_dimensions['D'].width = 22
                ws.column_dimensions['E'].width = 24
            else:
                ws.column_dimensions['C'].width = 22
                ws.column_dimensions['D'].width = 24
            col = col_start + n_fixed
            for mo in months:
                for pi in range(n_prod):
                    ws.column_dimensions[get_column_letter(col+pi)].width = num_col_width
                ws.column_dimensions[get_column_letter(col+n_prod)].width = num_col_width
                col += month_block_width
            ws.freeze_panes = get_column_letter(col_start+n_fixed) + '5'

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # 시트1: 판매수량 (오프라인)
        ws_qty = wb.create_sheet(f'판매수량_{brand_sel}'[:31])
        _build_matrix_sheet(ws_qty, f"※ 오프라인 판매수량_{brand_sel}_{year}", offline_sellers, qty_map, None)

        # 시트2: 판매금액 (오프라인) — 수정1
        ws_amt = wb.create_sheet(f'판매금액_{brand_sel}'[:31])
        _build_matrix_sheet(ws_amt, f"※ 오프라인 판매금액_{brand_sel}_{year}", offline_sellers, amt_map, '#,##0')

        # 시트3: 백화점 (수정4) — 가이아/서양네트웍스 등 백화점 채널 매장만 별도 시트로 분리
        if dept_sellers:
            ws_dept = wb.create_sheet('백화점'[:31])
            _build_matrix_sheet(ws_dept, f"※ 백화점 판매수량_{brand_sel}_{year}", dept_sellers, qty_map, None)

        # 시트4: 기초 데이터 (수정7) — 요약 데이터의 근거가 된 원본 판매 데이터 그대로 (검증용)
        ws_raw = wb.create_sheet('기초데이터')
        ws_raw.column_dimensions['A'].width = 2
        raw_hdrs = ['일자', '거래처명', '실적용거래처명', '거래처코드', '품목명', '수량', '단가', '공급가액', '부가세', '합계', '채널']
        for ci, h in enumerate(raw_hdrs, 2):
            c = ws_raw.cell(row=2, column=ci, value=h)
            c.font = Font(bold=True, size=9, name=FNAME, color=BLACK); c.fill = HDR_BG; c.alignment = ctr
        ws_raw.row_dimensions[2].height = 20
        raw_rows = conn.execute("""
            SELECT sale_date, seller_name, real_seller, trade_code, item_name,
                   quantity, unit_price, supply_price, vat, total, channel
            FROM sales_data
            WHERE sale_date LIKE ? AND real_seller!=''
            ORDER BY sale_date, real_seller""", (f"{year}%",)).fetchall()
        ri_raw = 3
        for r in raw_rows:
            item_brand = remap_group('', r[4]) if not r[4].startswith('[') else remap_group('X', r[4])
            if item_brand != brand_sel: continue
            for ci, v in enumerate(r, 2):
                c = ws_raw.cell(row=ri_raw, column=ci, value=v)
                c.font = Font(size=8, name=FNAME, color=BLACK)
                c.alignment = ctr if ci in (2,5,6,7) else left_a
                if ci in (8,9,10,11): c.number_format = '#,##0'; c.alignment = Alignment(horizontal='right')
            ri_raw += 1
        for ci, w in zip(range(2,13), [11,20,20,14,26,8,10,11,10,11,9]):
            ws_raw.column_dimensions[get_column_letter(ci)].width = w
        ws_raw.freeze_panes = 'B3'

        conn.close()
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        fname = f"판매현황_{brand_sel}_{year}.xlsx"
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=fname)

    # ══════════════════════════════════════════════════════
    # mode='display' — 기존 로직 유지 (모델별 탭, 색상별 교차표) — 버튼 라벨만 "판매 현황(색상별)"로 변경 (수정2)
    # ══════════════════════════════════════════════════════
    HDR_FILL = PatternFill("solid", fgColor="D9E1F2")
    GRP_FILL = PatternFill("solid", fgColor="F2F2F2")
    HIT_FILL = PatternFill("solid", fgColor="E2EFDA")
    SUM_FILL = PatternFill("solid", fgColor="BDD7EE")
    WHT_FILL = PatternFill("solid", fgColor="FFFFFF")
    thin     = Side(style='thin', color='CCCCCC')
    bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_a   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    bold10   = lambda: Font(bold=True, size=10, name='맑은 고딕')
    norm9    = lambda: Font(size=9, name='맑은 고딕')

    now_str  = datetime.now().strftime('%Y.%m.%d')
    mode_lbl = '진열 현황' if mode == 'display' else '판매 현황'

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for norm_key, tab_info in sorted(brand_items.items()):
        tab_label  = tab_info['label'].replace('[','').replace(']','').strip()[:31]
        norm_name  = tab_info['norm']
        is_canopy  = tab_info['is_canopy']
        colors     = tab_info['colors']
        ws = wb.create_sheet(title=tab_label)

        # 데이터 조회 — 해당 모델 (캐노피형 포함, 기본 모델로 통합)
        # norm_name 예: '[줄즈]에어2' → base: '에어2'
        import re as _re2
        def extract_base(nm):
            """[브랜드]모델명 → 모델명만 추출"""
            m = _re2.match(r'^\[([^\]]+)\](.+)$', nm.strip())
            return m.group(2).strip() if m else nm.strip()

        base = extract_base(norm_name)
        rows_data = conn.execute("""
            SELECT real_seller, item_name, SUM(quantity) qty
            FROM sales_data
            WHERE item_name LIKE ? AND sale_date LIKE ? AND real_seller!=''
            GROUP BY real_seller, item_name
        """, (f"%{base}%", f"{year}%")).fetchall()

        def extract_color(iname):
            m = re.search(r'_([^_\(]+)', iname)
            c = m.group(1).strip() if m else '기본'
            return re.sub(r'\s*\([^)]*\)', '', c).strip()

        # 실제 데이터에서 색상 재추출 (순서 유지)
        real_colors = []
        seen_c = set()
        for r in rows_data:
            c = extract_color(r[1])
            if c not in seen_c:
                real_colors.append(c)
                seen_c.add(c)
        if not real_colors: real_colors = ['(데이터없음)']

        data_map = {}
        for r in rows_data:
            c2 = extract_color(r[1])
            k2 = (r[0], c2)
            data_map[k2] = data_map.get(k2, 0) + (r[2] or 0)

        n_color = len(real_colors)
        sum_col = 6 + n_color   # col6~(5+n_color) = 색상, col(6+n_color) = 합계

        # 행1: 타이틀
        last_col = get_column_letter(sum_col)
        ws.merge_cells(f"B1:{last_col}1")
        tc = ws.cell(row=1, column=2, value=f"거래처별 {tab_label} {mode_lbl}  ▶  {year}년  |  {now_str}")
        tc.font = Font(bold=True, size=12, name='맑은 고딕')
        tc.fill = WHT_FILL; tc.alignment = center
        ws.row_dimensions[1].height = 26; ws.row_dimensions[2].height = 6

        # 행3: 고정 헤더
        for ci, h in enumerate(['업체구분','거래처코드','거래처명','실적용거래처명'], 2):
            c3 = ws.cell(row=3, column=ci, value=h)
            c3.font = bold10(); c3.fill = HDR_FILL; c3.alignment = center; c3.border = bdr

        # 행3: 제품명 병합
        # 색상: 6열~(5+n_color)열, 합계: (6+n_color)열
        # 브랜드명 헤더 merge: F3 ~ (5+n_color)열
        if n_color > 1:
            ws.merge_cells(f"F3:{get_column_letter(5+n_color)}3")
        c3p = ws.cell(row=3, column=6, value=tab_label)
        c3p.font = bold10(); c3p.fill = SUM_FILL; c3p.alignment = center; c3p.border = bdr

        # 합계 헤더
        c3s = ws.cell(row=3, column=sum_col, value='합계')
        c3s.font = bold10(); c3s.fill = HDR_FILL; c3s.alignment = center; c3s.border = bdr

        # 행4: 색상 헤더
        for ci2, color in enumerate(real_colors, 6):
            c4 = ws.cell(row=4, column=ci2, value=color)
            c4.font = Font(bold=True, size=9, name='맑은 고딕')
            c4.fill = GRP_FILL; c4.alignment = center; c4.border = bdr
        ws.row_dimensions[3].height = 22; ws.row_dimensions[4].height = 18

        # 데이터 행
        ri = 5
        prev_grp = None
        color_totals = {c: 0 for c in real_colors}
        grand_total = 0

        for seller in all_sellers:
            grp  = seller_group(seller)
            info = seller_to_code.get(seller, {})
            code = info.get('code','')
            orig = info.get('orig_name', seller)

            c_grp = ws.cell(row=ri, column=2)
            if grp != prev_grp:
                c_grp.value = grp; c_grp.font = bold10(); prev_grp = grp
            c_grp.fill = GRP_FILL; c_grp.alignment = center; c_grp.border = bdr

            ws.cell(row=ri, column=3, value=code or '').font = norm9()
            ws.cell(row=ri, column=3).alignment = center; ws.cell(row=ri, column=3).border = bdr
            ws.cell(row=ri, column=4, value=orig).font = norm9()
            ws.cell(row=ri, column=4).alignment = left_a; ws.cell(row=ri, column=4).border = bdr
            ws.cell(row=ri, column=5, value=seller).font = norm9()
            ws.cell(row=ri, column=5).alignment = left_a; ws.cell(row=ri, column=5).border = bdr

            row_total = 0
            for ci3, color in enumerate(real_colors, 6):
                qty = data_map.get((seller, color), 0)
                cv = ws.cell(row=ri, column=ci3, value=qty if qty else None)
                cv.font = norm9(); cv.alignment = center; cv.border = bdr
                if qty:
                    cv.fill = HIT_FILL
                    color_totals[color] = color_totals.get(color, 0) + qty
                    row_total += qty

            cs = ws.cell(row=ri, column=sum_col, value=row_total if row_total else None)
            cs.font = Font(bold=True, size=9, name='맑은 고딕')
            cs.alignment = center; cs.border = bdr
            if row_total:
                cs.fill = SUM_FILL; grand_total += row_total
            ws.row_dimensions[ri].height = 15
            ri += 1

        # 합계 행
        for ci4 in range(2, sum_col+1):
            ct = ws.cell(row=ri, column=ci4)
            ct.fill = HDR_FILL; ct.border = bdr; ct.font = bold10(); ct.alignment = center
        ws.cell(row=ri, column=2).value = '총 합계'
        for ci5, color in enumerate(real_colors, 6):
            ws.cell(row=ri, column=ci5).value = color_totals.get(color, 0) or None
        ws.cell(row=ri, column=sum_col).value = grand_total or None
        ws.row_dimensions[ri].height = 20

        # 열 너비
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 24
        ws.column_dimensions['E'].width = 26
        for ci6 in range(6, 6+n_color):
            ws.column_dimensions[get_column_letter(ci6)].width = max(12, len(real_colors[ci6-6])+2)
        ws.column_dimensions[get_column_letter(sum_col)].width = 8
        ws.freeze_panes = 'B5'

    conn.close()
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    mode_str  = '진열현황' if mode == 'display' else '판매현황'
    brand_str = brand_sel.replace(' ','_')
    fname = f"거래처별_{brand_str}_{mode_str}_{year}.xlsx"
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=fname)


# ── 행사 및 진열 신청 API ──────────────────────────
# ── 행사/진열 신청 API (v2 — 기한 기반 점수) ────────────────────
@app.route("/api/display/campaigns")
@login_required
def api_display_campaigns():
    """캠페인 목록 + 업로드 이력"""
    conn = get_db()
    campaigns = [dict(r) for r in conn.execute(
        "SELECT * FROM display_campaign ORDER BY id DESC").fetchall()]
    for c in campaigns:
        c['uploads'] = [dict(r) for r in conn.execute(
            "SELECT * FROM display_upload WHERE campaign_id=? ORDER BY upload_seq",
            (c['id'],)).fetchall()]
        c['total_records'] = conn.execute(
            "SELECT COUNT(*) FROM display_record WHERE campaign_id=? AND has_display=1",
            (c['id'],)).fetchone()[0]
    conn.close()
    return jsonify(campaigns)

@app.route("/api/display/campaign/create", methods=["POST"])
@login_required
def api_display_campaign_create():
    """새 캠페인 생성"""
    d = request.json or {}
    from datetime import datetime as dt2
    conn = get_db()
    conn.execute("""INSERT INTO display_campaign
        (campaign_name, brand, event_type, period_start, period_end,
         score_in_period, score_out_period, created_at)
        VALUES(?,?,?,?,?,?,?,?)""",
        (d.get('campaign_name',''), d.get('brand',''),
         d.get('event_type','display'),
         d.get('period_start',''), d.get('period_end',''),
         int(d.get('score_in_period', 5)), int(d.get('score_out_period', 2)),
         dt2.now().strftime('%Y-%m-%d %H:%M')))
    cid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.commit(); conn.close()
    return jsonify({'ok': True, 'id': cid})

@app.route("/api/display/campaign/update", methods=["POST"])
@login_required
def api_display_campaign_update():
    """캠페인 기간/점수 수정"""
    d = request.json or {}
    conn = get_db()
    conn.execute("""UPDATE display_campaign SET
        campaign_name=?, period_start=?, period_end=?,
        score_in_period=?, score_out_period=?
        WHERE id=?""",
        (d.get('campaign_name',''), d.get('period_start',''), d.get('period_end',''),
         int(d.get('score_in_period',5)), int(d.get('score_out_period',2)), d.get('id')))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

@app.route("/api/display/campaign/delete", methods=["POST"])
@login_required
def api_display_campaign_delete():
    cid = (request.json or {}).get('id')
    if not cid: return jsonify({'ok':False}), 400
    conn = get_db()
    conn.execute("DELETE FROM display_record WHERE campaign_id=?", (cid,))
    conn.execute("DELETE FROM display_upload WHERE campaign_id=?", (cid,))
    conn.execute("DELETE FROM display_campaign WHERE id=?", (cid,))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

@app.route("/api/display/upload", methods=["POST"])
@login_required
def api_display_upload():
    """진열/행사 엑셀 업로드
    - 시트명 자동 파싱 → 연도/제품명 추출
    - 합계 컬럼 자동 감지 → 진열 여부 판단
    - 캠페인의 기한 내/외 기준으로 점수 자동 부여
    - 같은 캠페인+매장+제품 조합은 최초 진열 점수 유지 (누적)
    """
    from datetime import datetime as dt2
    import re as _re

    if 'file' not in request.files:
        return jsonify({'ok': False, 'msg': '파일 없음'}), 400

    campaign_id = request.form.get('campaign_id')
    if not campaign_id:
        return jsonify({'ok': False, 'msg': '캠페인을 선택해주세요'}), 400

    file = request.files['file']
    data = file.read()
    now_str = dt2.now().strftime('%Y-%m-%d %H:%M')
    today = dt2.now().strftime('%Y-%m-%d')

    conn = get_db()

    # 캠페인 정보 로드
    campaign = conn.execute(
        "SELECT * FROM display_campaign WHERE id=?", (campaign_id,)).fetchone()
    if not campaign:
        conn.close()
        return jsonify({'ok': False, 'msg': '캠페인 없음'}), 404
    campaign = dict(campaign)

    # 점수 결정 (기한 내 여부)
    p_start = campaign.get('period_start','')
    p_end   = campaign.get('period_end','')
    if p_start and p_end and p_start <= today <= p_end:
        base_score = campaign.get('score_in_period', 5)
        period_label = f"기한 내 ({p_start}~{p_end})"
    elif p_end and today > p_end:
        base_score = campaign.get('score_out_period', 2)
        period_label = f"기한 후 (+{base_score}점)"
    else:
        base_score = campaign.get('score_in_period', 5)
        period_label = "기한 내"

    # 이번 업로드 seq 계산
    upload_seq = (conn.execute(
        "SELECT MAX(upload_seq) FROM display_upload WHERE campaign_id=?",
        (campaign_id,)).fetchone()[0] or 0) + 1

    # 엑셀 파싱
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:
        conn.close()
        return jsonify({'ok': False, 'msg': f'엑셀 파싱 실패: {e}'}), 500

    # real_seller 매핑
    # 실적용거래처명 컬럼을 그대로 사용 — 언더바만 공백으로 변환
    # 링크맘 공백 정규화 매핑 (공백 있는 버전 → 공백 없는 버전)
    LINKMOM_NORM = {
        '링크맘 경기 광주점':   '링크맘 경기광주점',
        '링크맘 대구 달성점':   '링크맘 대구달성점',
        '링크맘 대구 성서점':   '링크맘 대구성서점',
        '링크맘 대구 수성점':   '링크맘 대구수성점',
        '링크맘 파주 직영점':   '링크맘 파주직영점',
        '링크맘 의정부 민락점': '링크맘 의정부민락점',
        '링크맘 의정부 직영점': '링크맘 의정부직영점',
    }

    def clean_seller_name(raw):
        if not raw: return ''
        s = str(raw).strip()
        s = s.replace('_', ' ').strip()
        if s in ('합계', '총 합계', '총합계', '소계', ''): return ''
        # 링크맘 공백 정규화
        s = LINKMOM_NORM.get(s, s)
        return s

    total_inserted = 0
    total_updated  = 0
    sheet_results  = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        raw_rows = list(ws.iter_rows(values_only=True))

        # 시트명에서 제품명 추출: "2026_벨릭스" → "벨릭스"
        sheet_clean = _re.sub(r'^20\d{2}_', '', sheet_name).strip()
        # 행사 시트 판별
        is_event_sheet = '행사' in sheet_clean
        actual_type = 'event' if is_event_sheet else 'display'

        # 헤더 행 찾기 (업체구분/거래처코드 등 포함된 행)
        header_row_idx = None
        for i, row in enumerate(raw_rows[:8]):
            vals = [str(c).strip() for c in row if c is not None]
            if any(k in vals for k in ['업체구분','거래처코드','거래처명']):
                header_row_idx = i
                break

        if header_row_idx is None:
            sheet_results.append({'sheet': sheet_name, 'product': sheet_clean, 'inserted': 0, 'skipped': 0, 'msg': '헤더 없음'})
            continue

        headers = [str(c or '').strip() for c in raw_rows[header_row_idx]]

        # 색상 헤더 행 (헤더 바로 다음 줄에 색상명이 있는 구조) — Sheet1 대체 탐색에도 사용하므로 먼저 확보
        color_row_idx = header_row_idx + 1
        color_headers = []
        if color_row_idx < len(raw_rows):
            color_headers = [str(c or '').strip() for c in raw_rows[color_row_idx]]

        # "Sheet1" 같은 일반 시트명이면, 헤더 행 → 색상 헤더 행 → 캠페인명 순으로 실제 제품명을 탐색
        JUNK_HEADER_VALS = ('업체구분','거래처코드','거래처명','실적용거래처명','합계','No','번호','')
        if _re.match(r'^(sheet|시트)\s*\d*$', sheet_clean, _re.IGNORECASE):
            found = ''
            for h in headers:
                if h and h.strip() not in JUNK_HEADER_VALS:
                    found = h.strip(); break
            if not found:
                for h in color_headers:
                    if h and h.strip() not in JUNK_HEADER_VALS and h.strip() not in ('합계','총합계'):
                        found = h.strip(); break
            if not found:
                # 최종 fallback: 캠페인명(브랜드 제외한 순수 이름)을 제품명으로 사용
                camp_name_clean = _re.sub(r'\s*(진열|행사)?\s*신청.*$', '', campaign.get('campaign_name','') or '').strip()
                found = camp_name_clean or campaign.get('brand','') or f"제품_{sheet_name}"
            sheet_clean = found

        # 컬럼 인덱스 파악
        def find_col(keys, hdrs):
            for k in keys:
                for i, h in enumerate(hdrs):
                    if k == h.strip(): return i
            return -1

        NAME_IDX  = find_col(['실적용거래처명','거래처명'], headers)
        CODE_IDX  = find_col(['거래처코드'], headers)
        SUM_IDX   = find_col(['합계'], headers)

        if NAME_IDX == -1: NAME_IDX = 3   # 기본값
        if SUM_IDX  == -1: SUM_IDX  = len(headers)-2

        # 색상 컬럼 범위: NAME_IDX+1 ~ SUM_IDX-1 (색상명이 있는 컬럼만)
        color_cols = []
        for ci in range(NAME_IDX + 1, SUM_IDX):
            cname = color_headers[ci] if ci < len(color_headers) else ''
            if cname and cname not in ('합계','총합계'):
                color_cols.append((ci, cname))

        # 업로드 이력 저장 — sheet_name이 달라도 같은 product_name(제품)이면 갱신 (Sheet1, 진열현황 등 중복 방지)
        existing_upload = conn.execute(
            "SELECT id FROM display_upload WHERE campaign_id=? AND product_name=?",
            (campaign_id, sheet_clean)).fetchone()
        if existing_upload:
            upload_id = existing_upload[0]
            conn.execute("UPDATE display_upload SET sheet_name=?,upload_seq=?,upload_date=?,upload_at=? WHERE id=?",
                (sheet_name, upload_seq, today, now_str, upload_id))
        else:
            conn.execute("""INSERT INTO display_upload
                (campaign_id, sheet_name, product_name, upload_seq, upload_date, upload_at)
                VALUES(?,?,?,?,?,?)""",
                (campaign_id, sheet_name, sheet_clean, upload_seq, today, now_str))
            upload_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        # 데이터 행 처리 (헤더 아래 1줄은 색상 행일 수 있으므로 스킵)
        data_start = header_row_idx + 2
        inserted = skipped = 0

        for row in raw_rows[data_start:]:
            if not any(row): continue

            # 매장명
            raw_name = row[NAME_IDX] if NAME_IDX < len(row) else None
            if not raw_name or str(raw_name).strip() in ('', '합계', '총 합계', '소계'): continue

            # 합계
            raw_sum = row[SUM_IDX] if SUM_IDX < len(row) else 0
            try:
                qty = int(float(str(raw_sum or 0).replace(',','')))
            except:
                qty = 0

            # 색상별 수량 추출 (스누즈2 등 색상별 진열현황 표시용)
            color_qty = {}
            for ci, cname in color_cols:
                if ci < len(row) and row[ci]:
                    try:
                        cv = int(float(str(row[ci]).replace(',','')))
                        if cv > 0: color_qty[cname] = cv
                    except: pass
            color_json = json.dumps(color_qty, ensure_ascii=False) if color_qty else ''

            has_display = 1 if qty > 0 else 0
            seller = clean_seller_name(raw_name)
            if not seller: continue

            # 누적 처리: 이미 has_display=1 기록이 있으면 점수 유지 (덮어쓰지 않음)
            existing = conn.execute("""SELECT id, has_display, score, is_manual FROM display_record
                WHERE campaign_id=? AND seller_name=? AND product_name=?""",
                (campaign_id, seller, sheet_clean)).fetchone()

            if existing:
                ex_id, ex_has, ex_score, ex_manual = existing
                if ex_manual:
                    skipped += 1  # 수동 수정된 건 건드리지 않음
                    continue
                if ex_has == 1 and has_display == 0:
                    skipped += 1  # 이미 진열 완료인데 이번에 0이면 유지
                    continue
                if ex_has == 0 and has_display == 1:
                    # 이번에 처음 진열 확인 → 점수 부여 + 신청일 기록 (기존 신청일 없을 때만)
                    conn.execute("""UPDATE display_record SET
                        has_display=1, quantity=?, score=?, upload_id=?, upload_date=?, updated_at=?, color_detail=?,
                        applied_date=CASE WHEN applied_date='' OR applied_date IS NULL THEN ? ELSE applied_date END
                        WHERE id=?""",
                        (qty, base_score, upload_id, today, now_str, color_json, today, ex_id))
                    total_updated += 1
                else:
                    # 수량만 업데이트 (신청일 등 다른 정보는 건드리지 않음)
                    conn.execute("UPDATE display_record SET quantity=?,updated_at=?,color_detail=? WHERE id=?",
                        (qty, now_str, color_json, ex_id))
                    skipped += 1
            else:
                # 신규 등록 — 진열 확인된 경우에만 신청일 기록
                code = str(row[CODE_IDX]).strip() if CODE_IDX >= 0 and CODE_IDX < len(row) and row[CODE_IDX] else ''
                conn.execute("""INSERT INTO display_record
                    (campaign_id, upload_id, seller_name, seller_code, product_name,
                     quantity, has_display, score, upload_date, updated_at, color_detail, applied_date)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (campaign_id, upload_id, seller, code, sheet_clean,
                     qty, has_display,
                     base_score if has_display else 0,
                     today, now_str, color_json,
                     today if has_display else ''))
                inserted += 1

        total_inserted += inserted
        sheet_results.append({
            'sheet': sheet_name, 'product': sheet_clean,
            'inserted': inserted, 'updated': total_updated,
            'skipped': skipped, 'type': actual_type
        })

    conn.commit(); conn.close()
    return jsonify({
        'ok': True, 'campaign_id': campaign_id,
        'upload_seq': upload_seq,
        'base_score': base_score, 'period_label': period_label,
        'total_inserted': total_inserted,
        'sheets': sheet_results
    })


@app.route("/api/display/scores")
@login_required
def api_display_scores():
    """캠페인별 매장 점수 집계"""
    campaign_id = request.args.get('campaign_id')
    year = request.args.get('year', str(datetime.now().year))
    conn = get_db()

    # 업체구분 정렬 헬퍼
    GROUP_ORDER = ['베이비하우스', '링크맘', '베이비파크', '베이비세븐', '베이비스토리', '베이비스토어']
    def seller_group_key(name):
        for i, g in enumerate(GROUP_ORDER):
            if g in name: return (i, name)
        return (len(GROUP_ORDER), name)

    # 판매처 관리(branches)에서 담당자 매핑 — name 정확 일치 + 공백 제거 매칭
    branch_managers = {}
    try:
        for r in conn.execute("SELECT name, manager FROM branches WHERE manager IS NOT NULL AND manager!=''").fetchall():
            branch_managers[r[0]] = r[1]
    except Exception:
        pass
    branch_managers_nospace = {k.replace(' ',''): v for k, v in branch_managers.items()}

    def get_manager(seller_name):
        if seller_name in branch_managers:
            return branch_managers[seller_name]
        ns = seller_name.replace(' ', '')
        if ns in branch_managers_nospace:
            return branch_managers_nospace[ns]
        return ''

    if campaign_id:
        # 특정 캠페인 점수 — id 포함, color_detail/visit_done/call_done/note/applied_date 추가
        records = [dict(r) for r in conn.execute("""
            SELECT id, seller_name, product_name, has_display, quantity, score, is_manual, upload_date,
                   color_detail, visit_done, call_done, note, applied_date
            FROM display_record WHERE campaign_id=? ORDER BY seller_name, product_name
        """, (campaign_id,)).fetchall()]
    else:
        # 전체 캠페인 합산
        # 수정3: 같은 매장이 같은 제품(product_name)으로 여러 캠페인에 중복 기록되어 있으면
        # 점수를 더하지 않고 "더 높은 점수" 하나만 채택 (이미 참여한 매장이 재업로드로 중복 집계되는 것 방지)
        records = [dict(r) for r in conn.execute("""
            WITH best_per_product AS (
                SELECT dr.seller_name, dr.product_name,
                       MAX(dr.score) AS best_score,
                       (SELECT dc2.campaign_name FROM display_record dr2
                        JOIN display_campaign dc2 ON dr2.campaign_id=dc2.id
                        WHERE dr2.seller_name=dr.seller_name AND dr2.product_name=dr.product_name
                              AND dr2.has_display=1
                        ORDER BY dr2.score DESC LIMIT 1) AS best_campaign_name,
                       MAX(dr.has_display) AS has_display
                FROM display_record dr
                WHERE dr.has_display=1
                GROUP BY dr.seller_name, dr.product_name
            )
            SELECT seller_name,
                   SUM(best_score) total_score,
                   COUNT(*) display_cnt,
                   GROUP_CONCAT(best_campaign_name || '§' || product_name || '§' || best_score, '|') detail_raw
            FROM best_per_product
            GROUP BY seller_name ORDER BY total_score DESC
        """).fetchall()]

    # 판매 데이터 — 연매출 조회용 (실적용거래처명 → real_seller 매핑 포함)
    sales_map = {}
    # 직접 매칭
    for r in conn.execute(f"SELECT real_seller, SUM(total) FROM sales_data WHERE sale_date LIKE '{year}%' AND real_seller!='' GROUP BY real_seller").fetchall():
        sales_map[r[0]] = r[1]
    # 공백 제거 매칭 (링크맘 대구 성서점 vs 링크맘 대구성서점)
    sales_map_nospace = {k.replace(' ',''): v for k, v in sales_map.items()}

    def get_sales(seller_name):
        v = sales_map.get(seller_name, 0)
        if not v:
            v = sales_map_nospace.get(seller_name.replace(' ',''), 0)
        return v or 0

    all_sellers = [r[0] for r in conn.execute(
        f"SELECT DISTINCT real_seller FROM sales_data WHERE real_seller!='' AND sale_date LIKE '{year}%' ORDER BY real_seller"
    ).fetchall()]
    conn.close()

    if campaign_id:
        # 캠페인별 상세: 매장 × 제품 형태, 업체구분 순서로 정렬
        seller_map = {}
        for r in records:
            s = r['seller_name']
            if s not in seller_map: seller_map[s] = {'products': [], 'total_score': 0, 'total_qty': 0}
            # color_detail JSON 파싱
            try:
                r['color_detail_parsed'] = json.loads(r.get('color_detail') or '{}')
            except Exception:
                r['color_detail_parsed'] = {}
            seller_map[s]['products'].append(r)
            seller_map[s]['total_score'] += r['score'] or 0
            seller_map[s]['total_qty']   += r['quantity'] or 0

        # 진열 레코드에 있는 매장만 사용 (판매 데이터와 무관)
        # 업체구분 순서로 정렬

        result = [{'seller_name':s, **v, 'year_sales': get_sales(s), 'manager': get_manager(s)}
                  for s,v in seller_map.items()]
        # 업체구분 정렬
        result.sort(key=lambda x: (seller_group_key(x['seller_name'])[0],
                                   x['seller_name']))
        return jsonify(result)
    else:
        # 전체 합산: 진열 기록 있는 매장 + 판매 데이터 매장 포함
        all_display_sellers = {r['seller_name'] for r in records}
        merged_sellers = sorted(
            set(all_sellers) | all_display_sellers,
            key=lambda s: (seller_group_key(s)[0], s)
        )
        score_map = {r['seller_name']: r for r in records}
        result = []
        for s in merged_sellers:
            info = score_map.get(s, {})
            detail_raw = info.get('detail_raw','')
            detail_items = []
            if detail_raw:
                for part in detail_raw.split('|'):
                    parts = part.split('§')
                    if len(parts) >= 3:
                        try:
                            camp, prod, sc = parts[0], parts[1], parts[2]
                            detail_items.append({'campaign':camp, 'product':prod, 'score':int(sc)})
                        except: pass
                    elif len(parts) == 2:
                        try:
                            detail_items.append({'campaign': parts[0], 'product': '', 'score': int(parts[1])})
                        except: pass
            result.append({
                'seller_name':  s,
                'total_score':  info.get('total_score', 0) or 0,
                'display_cnt':  info.get('display_cnt', 0) or 0,
                'detail_items': detail_items,
                'year_sales':   get_sales(s),
                'manager':      get_manager(s),
            })

        # 동점 순위 계산 (dense rank — 동점 같은 순위, 다음 점수는 다음 순위)
        result.sort(key=lambda x: -x['total_score'])
        rank = 0; prev_score = None
        for r in result:
            sc = r['total_score']
            if sc != prev_score:
                rank += 1
                prev_score = sc
            r['rank'] = rank

        # A/B/C/D 등급 — 순위 기준
        for r in result:
            rk = r['rank']
            r['grade'] = 'A' if rk <= 25 else 'B' if rk <= 50 else 'C' if rk <= 75 else 'D' if rk <= 100 else 'E'

        return jsonify(result)


@app.route("/api/display/record/update", methods=["POST"])
@login_required
def api_display_record_update():
    """수동 점수/메모 수정 — id 직접 또는 updates 배열"""
    d = request.json or {}
    conn = get_db()
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    updated = 0

    # 단일 id 방식
    rid = d.get('id')
    if rid:
        conn.execute("""UPDATE display_record SET score=?,has_display=?,memo=?,is_manual=1,updated_at=?
            WHERE id=?""",
            (int(d.get('score',0)), int(d.get('has_display',0)), d.get('memo',''), now, rid))
        updated = conn.execute("SELECT changes()").fetchone()[0]

    # 배열 방식 (여러 레코드 한번에)
    for item in d.get('updates', []):
        item_rid = item.get('id')
        if item_rid:
            conn.execute("""UPDATE display_record SET score=?,has_display=?,memo=?,is_manual=1,updated_at=?
                WHERE id=?""",
                (int(item.get('score',0)), int(item.get('has_display',0)), item.get('memo',''), now, item_rid))
            updated += 1

    conn.commit(); conn.close()
    return jsonify({'ok': True, 'updated': updated})


@app.route("/api/display/record/followup", methods=["POST"])
@login_required
def api_display_record_followup():
    """방문/전화 체크 및 비고 업데이트 (점수와 무관, 가벼운 토글)"""
    d = request.json or {}
    rid = d.get('id')
    if not rid:
        return jsonify({'ok': False, 'msg': 'id 필요'}), 400
    conn = get_db()
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    fields, params = [], []
    if 'visit_done' in d:
        fields.append('visit_done=?'); params.append(int(d['visit_done']))
    if 'call_done' in d:
        fields.append('call_done=?'); params.append(int(d['call_done']))
    if 'note' in d:
        fields.append('note=?'); params.append(d['note'])
    if not fields:
        conn.close()
        return jsonify({'ok': False, 'msg': '변경할 필드 없음'}), 400
    fields.append('updated_at=?'); params.append(now)
    params.append(rid)
    conn.execute(f"UPDATE display_record SET {','.join(fields)} WHERE id=?", params)
    conn.commit(); conn.close()
    return jsonify({'ok': True})


@app.route("/api/display/export/ranking")
@login_required
def api_display_export_ranking():
    """전체 합산 랭킹 엑셀 — 순위 없는 매장 포함, 흰 배경, 쉼표 숫자"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    year = request.args.get('year', str(datetime.now().year))

    conn = get_db()
    # 수정3: 점수 있는 매장 — 같은 제품에 중복 참여 기록이 있으면 더 높은 점수만 채택 (합산 아님)
    records = [dict(r) for r in conn.execute("""
        WITH best_per_product AS (
            SELECT dr.seller_name, dr.product_name,
                   MAX(dr.score) AS best_score,
                   (SELECT dc2.campaign_name FROM display_record dr2
                    JOIN display_campaign dc2 ON dr2.campaign_id=dc2.id
                    WHERE dr2.seller_name=dr.seller_name AND dr2.product_name=dr.product_name
                          AND dr2.has_display=1
                    ORDER BY dr2.score DESC LIMIT 1) AS best_campaign_name
            FROM display_record dr
            WHERE dr.has_display=1
            GROUP BY dr.seller_name, dr.product_name
        )
        SELECT seller_name, SUM(best_score) total_score,
               COUNT(*) display_cnt,
               GROUP_CONCAT(best_campaign_name || '|' || product_name || '|' || best_score, '§') detail_raw
        FROM best_per_product
        GROUP BY seller_name ORDER BY total_score DESC
    """).fetchall()]
    # 점수 없는 매장 (진열 기록 있는 모든 매장)
    all_record_sellers = {r['seller_name'] for r in records}
    no_score = [dict(r) for r in conn.execute("""
        SELECT DISTINCT seller_name FROM display_record WHERE has_display=0
    """).fetchall() if r['seller_name'] not in all_record_sellers]

    sales_map = {r[0]:r[1] for r in conn.execute(
        f"SELECT real_seller, SUM(total) FROM sales_data WHERE sale_date LIKE '{year}%' AND real_seller!='' GROUP BY real_seller"
    ).fetchall()}
    # 공백 제거 매칭
    sales_nospace = {k.replace(' ',''):v for k,v in sales_map.items()}
    def get_sales(s):
        return sales_map.get(s, sales_nospace.get(s.replace(' ',''), 0)) or 0

    conn.close()

    # 동점 순위 (dense rank)
    rank_val = 0; prev_sc = None
    for r in records:
        sc = r.get('total_score',0) or 0
        if sc != prev_sc: rank_val += 1; prev_sc = sc
        r['rank'] = rank_val
        r['grade'] = 'A' if rank_val<=25 else 'B' if rank_val<=50 else 'C' if rank_val<=75 else 'D' if rank_val<=100 else 'E'

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = '전체합산랭킹'

    def mf(hex_): return PatternFill("solid", fgColor=hex_)
    thin = Side(style='thin', color='E5E7EB')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    right_a = Alignment(horizontal='right', vertical='center')
    GRADE_COLOR = {'A':'EFF6FF','B':'F0FDF4','C':'FFFBEB','D':'F9FAFB'}

    # A열은 여백(spacer)로 비워두고 B열부터 시작 — 심플하고 깔끔한 레이아웃
    ws.column_dimensions['A'].width = 2

    ws.merge_cells('B1:H1')
    c = ws.cell(row=1,column=2,value=f'행사 및 진열 신청 전체 합산 랭킹 ({year}년)')
    c.font=Font(bold=True,size=13,name='맑은 고딕',color='1F2937'); c.fill=mf('FFFFFF'); c.alignment=ctr
    ws.row_dimensions[1].height=28

    hdrs = ['순위','등급','매장명','누적 점수','참여 수','참여 행사/진열 내역','연매출']
    widths = [8, 8, 26, 12, 10, 55, 18]
    for ci,(h,w) in enumerate(zip(hdrs,widths),2):
        c = ws.cell(row=2,column=ci,value=h)
        c.font=Font(bold=True,size=10,name='맑은 고딕',color='374151'); c.fill=mf('F3F4F6'); c.border=bdr; c.alignment=ctr
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[2].height=20

    ri = 3
    for r in records:
        grade = r.get('grade','E'); score = r.get('total_score',0) or 0
        # 수정1: 점수 표시 없이 참여한 행사/진열명만 리스트로 (누적점수·참여수는 별도 컬럼으로 이미 구분)
        camp_names = []
        for item in (r.get('detail_raw','') or '').split('§'):
            if item.count('|') >= 2:
                try:
                    camp, prod, sc = item.split('|', 2)
                    camp_names.append(camp.strip())
                except: pass
            elif '|' in item:
                try:
                    prod, sc = item.split('|', 1)
                    camp_names.append(prod.strip())
                except: pass
        # 중복 제거 (순서 유지)
        seen = set(); unique_names = []
        for n in camp_names:
            if n not in seen:
                seen.add(n); unique_names.append(n)
        row_fill = mf(GRADE_COLOR.get(grade,'FFFFFF'))
        sales_val = get_sales(r['seller_name'])
        vals = [r.get('rank',''), grade, r['seller_name'], score,
                r.get('display_cnt',0), ' / '.join(unique_names[:6]), sales_val]
        for ci,v in enumerate(vals,2):
            c=ws.cell(row=ri,column=ci,value=v)
            c.font=Font(size=9,name='맑은 고딕',color='1F2937'); c.border=bdr
            c.alignment=ctr if ci not in (4,7) else left
            if grade in GRADE_COLOR and score>0: c.fill=row_fill
            if ci==8:
                c.number_format='#,##0'; c.alignment=right_a
        ws.row_dimensions[ri].height=15; ri+=1

    # 점수 없는 매장도 포함 (구분선 추가)
    if no_score:
        c = ws.cell(row=ri,column=2,value='— 미참여 매장 —')
        ws.merge_cells(f'B{ri}:H{ri}')
        c.font=Font(size=9,name='맑은 고딕',color='9CA3AF',italic=True)
        c.fill=mf('F9FAFB'); c.alignment=ctr
        ws.row_dimensions[ri].height=13; ri+=1
        for r in no_score:
            seller = r['seller_name']
            for ci,v in enumerate(['—', '—', seller, 0, 0, '미참여', get_sales(seller)],2):
                c=ws.cell(row=ri,column=ci,value=v)
                c.font=Font(size=9,name='맑은 고딕',color='9CA3AF'); c.border=bdr
                c.alignment=ctr if ci!=4 else left
                if ci==8: c.number_format='#,##0'; c.alignment=right_a
            ws.row_dimensions[ri].height=14; ri+=1

    ws.freeze_panes='B3'
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=f'행사진열_전체랭킹_{year}.xlsx')


@app.route("/api/display/campaign/<int:campaign_id>/fix-product-names", methods=["POST"])
@login_required
def api_display_fix_product_names(campaign_id):
    """캠페인 내에서 'Sheet1' 같은 일반 시트명으로 잘못 저장된 제품명을
    올바른 제품명(다른 정상 제품명 또는 캠페인명)으로 병합 정리"""
    import re as _re_fix

    conn = get_db()
    campaign = conn.execute("SELECT * FROM display_campaign WHERE id=?", (campaign_id,)).fetchone()
    if not campaign:
        conn.close()
        return jsonify({'ok': False, 'msg': '캠페인을 찾을 수 없습니다'}), 404
    campaign = dict(campaign)

    all_products = [r[0] for r in conn.execute(
        "SELECT DISTINCT product_name FROM display_upload WHERE campaign_id=?", (campaign_id,)).fetchall()]

    JUNK_PATTERN = _re_fix.compile(r'^(sheet|시트)\s*\d*$', _re_fix.IGNORECASE)
    junk_products = [p for p in all_products if JUNK_PATTERN.match(p or '')]
    normal_products = [p for p in all_products if not JUNK_PATTERN.match(p or '')]

    if not junk_products:
        conn.close()
        return jsonify({'ok': True, 'merged': 0, 'msg': '정리할 항목이 없습니다'})

    # 병합 대상 결정: 정상 제품명이 정확히 1개면 그것으로, 여러 개/없으면 캠페인명으로
    if len(normal_products) == 1:
        target_name = normal_products[0]
    else:
        camp_name_clean = _re_fix.sub(r'\s*(진열|행사)?\s*신청.*$', '', campaign.get('campaign_name','') or '').strip()
        target_name = camp_name_clean or campaign.get('brand','') or '통합제품'

    merged_count = 0
    for junk in junk_products:
        if junk == target_name:
            continue
        # display_record 병합 — 매장별로 겹치면 더 높은 점수/수량 유지, UNIQUE(campaign_id,seller_name,product_name) 제약 고려
        junk_records = conn.execute(
            "SELECT id, seller_name, has_display, quantity, score, color_detail, applied_date FROM display_record "
            "WHERE campaign_id=? AND product_name=?", (campaign_id, junk)).fetchall()
        for jr in junk_records:
            jr_id, seller, has_d, qty, score, color_detail, applied = jr
            existing = conn.execute(
                "SELECT id, has_display, quantity, score, applied_date FROM display_record "
                "WHERE campaign_id=? AND seller_name=? AND product_name=?",
                (campaign_id, seller, target_name)).fetchone()
            if existing:
                ex_id, ex_has, ex_qty, ex_score, ex_applied = existing
                new_qty = (ex_qty or 0) + (qty or 0)
                new_has = max(ex_has or 0, has_d or 0)
                new_score = max(ex_score or 0, score or 0)
                new_applied = ex_applied or applied
                conn.execute("UPDATE display_record SET has_display=?, quantity=?, score=?, applied_date=? WHERE id=?",
                             (new_has, new_qty, new_score, new_applied, ex_id))
                conn.execute("DELETE FROM display_record WHERE id=?", (jr_id,))
            else:
                conn.execute("UPDATE display_record SET product_name=? WHERE id=?", (target_name, jr_id))
            merged_count += 1

        # display_upload 정리 — junk 업로드 이력을 target으로 통합(중복이면 삭제)
        junk_uploads = conn.execute(
            "SELECT id FROM display_upload WHERE campaign_id=? AND product_name=?", (campaign_id, junk)).fetchall()
        target_upload = conn.execute(
            "SELECT id FROM display_upload WHERE campaign_id=? AND product_name=?", (campaign_id, target_name)).fetchone()
        for ju in junk_uploads:
            if target_upload:
                conn.execute("DELETE FROM display_upload WHERE id=?", (ju[0],))
            else:
                conn.execute("UPDATE display_upload SET product_name=? WHERE id=?", (target_name, ju[0]))

    conn.commit(); conn.close()
    return jsonify({'ok': True, 'merged': merged_count, 'target_name': target_name,
                     'cleaned_products': junk_products})


@app.route("/api/display/export/campaign")
@login_required
def api_display_export_campaign():
    """특정 캠페인 매장별 발주 현황 엑셀 다운로드 — 깔끔한 보고용 디자인"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    campaign_id = request.args.get('campaign_id')
    year = request.args.get('year', str(datetime.now().year))
    if not campaign_id:
        return jsonify({'ok':False,'msg':'campaign_id 필요'}), 400

    conn = get_db()
    camp_row = conn.execute("SELECT * FROM display_campaign WHERE id=?", (campaign_id,)).fetchone()
    campaign = dict(camp_row) if camp_row else {}
    records = [dict(r) for r in conn.execute("""
        SELECT seller_name, product_name, has_display, quantity, score, upload_date,
               color_detail, visit_done, call_done, note
        FROM display_record WHERE campaign_id=? ORDER BY seller_name, product_name
    """, (campaign_id,)).fetchall()]
    products = [r[0] for r in conn.execute(
        "SELECT DISTINCT product_name FROM display_upload WHERE campaign_id=? ORDER BY upload_seq", (campaign_id,)).fetchall()]

    # 판매처 관리(branches)에서 담당자 매핑
    branch_managers = {}
    try:
        for r in conn.execute("SELECT name, manager FROM branches WHERE manager IS NOT NULL AND manager!=''").fetchall():
            branch_managers[r[0]] = r[1]
    except Exception:
        pass
    branch_managers_nospace = {k.replace(' ',''): v for k, v in branch_managers.items()}
    def get_manager(s):
        return branch_managers.get(s, branch_managers_nospace.get(s.replace(' ',''), ''))

    conn.close()

    camp_name = campaign.get('campaign_name','캠페인')
    score_in  = campaign.get('score_in_period',5)
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = (camp_name[:28] + f'+{score_in}pt')[:31]

    # ── 절제된 컬러 팔레트 (보고용) ──────────────
    def mf(hex_): return PatternFill("solid", fgColor=hex_)
    NAVY    = '1F2937'   # 헤더 타이틀 (진한 회색-네이비, 과하지 않게)
    LGRAY   = 'F3F4F6'   # 그룹 구분 행
    HGRAY   = 'E5E7EB'   # 헤더 배경
    BORDERC = 'D1D5DB'
    thin=Side(style='thin',color=BORDERC); bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    ctr=Alignment(horizontal='center',vertical='center')
    left=Alignment(horizontal='left',vertical='center')
    FNAME='맑은 고딕'

    EXTRA_COLS = ['발주 수량','방문 및 설명\n(1차)','미신청 후 전화\n(2차)','비고']
    # A열은 여백(spacer)으로 비워두고 B열부터 시작 — 깔끔한 레이아웃
    ws.column_dimensions['A'].width = 2
    COL0 = 2  # 실제 컨텐츠 시작 컬럼(B)
    col_count = COL0 - 1 + 3 + len(products) + len(EXTRA_COLS)

    # 매장별 집계
    seller_map={}
    for r in records:
        s=r['seller_name']
        if s not in seller_map: seller_map[s]={'prods':{},'total_qty':0}
        try:
            r['color_detail_parsed'] = json.loads(r.get('color_detail') or '{}')
        except Exception:
            r['color_detail_parsed'] = {}
        seller_map[s]['prods'][r['product_name']]=r
        seller_map[s]['total_qty']+=r['quantity'] or 0

    # 수정: 전체 매장 = 업로드된(엑셀에 있는) 매장 수만 기준
    total_seller_cnt  = len(seller_map)
    participating_cnt = sum(1 for v in seller_map.values() if v['total_qty']>0)
    total_qty_sum     = sum(v['total_qty'] for v in seller_map.values())

    # ── 상단 타이틀 + KPI (절제된 색상) ─────────
    ws.merge_cells(f'B1:{get_column_letter(col_count)}1')
    kpi_text = f"{camp_name} ({year}년)   ·   전체 {total_seller_cnt}개 매장 중 {participating_cnt}개 매장 진열 신청   ·   총 발주수량 {total_qty_sum:,}개"
    c=ws.cell(row=1,column=COL0,value=kpi_text)
    c.font=Font(bold=True,size=12,name=FNAME,color='FFFFFF'); c.fill=mf(NAVY); c.alignment=ctr
    ws.row_dimensions[1].height=28

    period=f"기한: {campaign.get('period_start','')} ~ {campaign.get('period_end','')}"
    ws.merge_cells(f'B2:{get_column_letter(col_count)}2')
    c=ws.cell(row=2,column=COL0,value=period); c.font=Font(size=9,name=FNAME,color='6B7280'); c.alignment=left
    ws.row_dimensions[2].height=16

    # ── 헤더 ──────────────────────────────────
    headers = ['구분','매장명','담당자'] + products + EXTRA_COLS
    for offset,h in enumerate(headers):
        ci = COL0 + offset
        c=ws.cell(row=3,column=ci,value=h)
        c.font=Font(bold=True,size=9,name=FNAME,color='374151')
        c.fill=mf(HGRAY); c.border=bdr
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    ws.row_dimensions[3].height=30

    ws.column_dimensions[get_column_letter(COL0)].width=12
    ws.column_dimensions[get_column_letter(COL0+1)].width=22
    ws.column_dimensions[get_column_letter(COL0+2)].width=12
    for pi in range(len(products)):
        ws.column_dimensions[get_column_letter(COL0+3+pi)].width=max(16,len(products[pi])+4)
    extra_start = COL0+3+len(products)
    ws.column_dimensions[get_column_letter(extra_start)].width=10
    ws.column_dimensions[get_column_letter(extra_start+1)].width=14
    ws.column_dimensions[get_column_letter(extra_start+2)].width=14
    ws.column_dimensions[get_column_letter(extra_start+3)].width=20

    GRPS=['베이비하우스','링크맘','베이비파크']
    def grp(n): return next((g for g in GRPS if g in n),'기타')
    sellers=sorted(seller_map.keys(),key=lambda s:(GRPS.index(grp(s)) if grp(s) in GRPS else 99,s))

    ri=4; prev_g=''
    grp_total_qty = {}
    for seller in sellers:
        g=grp(seller)
        first_in_group = (g != prev_g)
        if first_in_group:
            prev_g=g
        info=seller_map[seller]
        manager = get_manager(seller)
        # 구분: 그룹 첫 매장에만 표시 (반복 제거)
        row=[g if first_in_group else '', seller, manager]
        visit_done = call_done = 0; note_val = ''
        for p in products:
            pd=info['prods'].get(p)
            if pd and pd.get('has_display'):
                colors = pd.get('color_detail_parsed') or {}
                if colors:
                    cell_val = ', '.join(f"{c_} {q_}개" for c_, q_ in colors.items())
                else:
                    qty=pd.get('quantity',0) or 0
                    cell_val=f"{qty}개" if qty>0 else "진열"
            elif pd:
                cell_val='—'
            else:
                cell_val=''
            row.append(cell_val)
            if pd:
                visit_done = visit_done or pd.get('visit_done',0)
                call_done  = call_done  or pd.get('call_done',0)
                if pd.get('note'): note_val = pd.get('note')
        row.append(info['total_qty'])
        row.append('완료' if visit_done else '')
        row.append('완료' if call_done else '')
        row.append(note_val)

        for offset,v in enumerate(row):
            ci = COL0 + offset
            c=ws.cell(row=ri,column=ci,value=v); c.font=Font(size=9,name=FNAME,color='1F2937'); c.border=bdr
            c.alignment=ctr if offset!=1 else left
            if offset==0 and v:  # 구분 강조 (그룹 첫 행)
                c.font=Font(bold=True,size=9,name=FNAME,color='374151')
            if ci==extra_start:  # 발주 수량만 약하게 강조
                c.font=Font(bold=True,size=9,name=FNAME,color='1F2937')
        ws.row_dimensions[ri].height=15; ri+=1
        grp_total_qty[g] = grp_total_qty.get(g,0) + info['total_qty']

    # ── 하단 합계 (절제된 디자인) ────────────────
    grand_qty = sum(grp_total_qty.values()) or 1
    ws.merge_cells(f'{get_column_letter(COL0)}{ri}:{get_column_letter(COL0+2)}{ri}')
    c=ws.cell(row=ri,column=COL0,value='합계'); c.font=Font(bold=True,size=10,name=FNAME,color='374151')
    c.fill=mf(LGRAY); c.alignment=ctr; c.border=bdr
    for ci in range(COL0+3, COL0+3+len(products)):
        c=ws.cell(row=ri,column=ci,value=''); c.fill=mf(LGRAY); c.border=bdr
    c=ws.cell(row=ri,column=extra_start,value=grand_qty)
    c.font=Font(bold=True,size=10,name=FNAME,color='374151'); c.fill=mf(LGRAY); c.alignment=ctr; c.border=bdr
    ws.merge_cells(f'{get_column_letter(extra_start+1)}{ri}:{get_column_letter(col_count)}{ri}')
    c=ws.cell(row=ri,column=extra_start+1,value=f'발주수량 100%  ({grand_qty:,}개)')
    c.font=Font(bold=True,size=9,name=FNAME,color='6B7280'); c.fill=mf(LGRAY); c.alignment=left; c.border=bdr
    ws.row_dimensions[ri].height=20
    ri += 1

    # 업체구분별 비율 표
    if len(grp_total_qty) > 1:
        ws.merge_cells(f'{get_column_letter(COL0)}{ri}:{get_column_letter(COL0+2)}{ri}')
        c=ws.cell(row=ri,column=COL0,value='구분별 비중'); c.font=Font(bold=True,size=9,name=FNAME,color='9CA3AF')
        c.alignment=left
        ri += 1
        for g_name, g_qty in sorted(grp_total_qty.items(), key=lambda x:-x[1]):
            pct = round(g_qty/grand_qty*100,1)
            ws.merge_cells(f'{get_column_letter(COL0)}{ri}:{get_column_letter(COL0+2)}{ri}')
            c=ws.cell(row=ri,column=COL0,value=f'  {g_name}   {g_qty:,}개 ({pct}%)')
            c.font=Font(size=9,name=FNAME,color='9CA3AF'); c.alignment=left
            ri += 1

    ws.freeze_panes=f'{get_column_letter(COL0)}4'
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    safe=campaign.get('campaign_name','캠페인').replace('/','_')
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=f'진열신청_{safe}_{year}.xlsx')


@app.route("/api/display/events")
@login_required
def api_display_events():
    conn = get_db()
    rows = [dict(r) for r in conn.execute(
        "SELECT * FROM display_event ORDER BY event_type, upload_order, id").fetchall()]
    conn.close()
    return jsonify(rows)

# Render/gunicorn 실행 시 자동 초기화
init_db()

if __name__ == "__main__":
    import webbrowser, threading
    threading.Timer(1.2, lambda: webbrowser.open("http://localhost:5001")).start()
    app.run(debug=False, port=5001)
